# -*- coding: utf-8 -*-
import openpyxl, warnings
warnings.filterwarnings('ignore')
dl = r'C:\Users\User\AppData\Local\Temp\test_asia_v2.xlsx'
wb = openpyxl.load_workbook(dl, data_only=True)
ws = wb['工作表3']
checks = [
    ('B5','資金用途','II-3購買3C產品'),
    ('B9','姓名','王小明'),
    ('D9','身分證','J122562348'),
    ('F9','出生','1989/09/16'),
    ('B10','婚姻','未婚'),
    ('D10','學歷','高中/職'),
    ('B11','發證日','2025/05/20'),
    ('D11','發證地','苗縣'),
    ('F11','狀態','換發'),
    ('B12','戶籍市','苗栗縣'),
    ('C12','戶籍區','頭份市'),
    ('D12','戶籍址','大營路97巷1弄10號'),
    ('B14','居住','親屬'),
    ('D14','居住年','10'),
    ('B15','電話','0900048459'),
    ('B16','Email','a0900048459@gmail.com'),
    ('B17','公司','冠軍建材股份有限公司'),
    ('D17','行業','製造業'),
    ('G17','職務','技術與工程'),
    ('B18','區碼','037'),
    ('C18','號碼','561761'),
    ('G18','年資','5.6'),
    ('H18','月薪','5.8'),
    ('B19','公司市','苗栗縣'),
    ('C19','公司區','造橋鄉'),
    ('D19','公司址','乳姑山2號'),
    ('B21','聯絡人','吳騰萬'),
    ('D21','關係','父母'),
    ('B25','電話','0911797470'),
]
ok=fail=0
for ref,label,exp in checks:
    val=str(ws[ref].value) if ws[ref].value else '(empty)'
    if val==exp:
        ok+=1
    else:
        fail+=1
        print(f'  FAIL {ref}({label}): expect={exp}, got={val}')
print(f'{ok}/{ok+fail} OK')
# Check reference data intact
ws4 = wb['工作表4']
print(f'Ref: A1={ws4["A1"].value}, AN1={ws4["AN1"].value}')
ws2 = wb['汽車廠牌車款參照表']
print(f'Car: B2={ws2["B2"].value}')
