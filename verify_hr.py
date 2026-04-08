# -*- coding: utf-8 -*-
import openpyxl, warnings
warnings.filterwarnings('ignore')
wb = openpyxl.load_workbook(r'C:\Users\User\AppData\Local\Temp\test_hr.xlsx', data_only=True)
ws = wb.worksheets[0]
checks = [
    ('C11','姓名','陳耀晨'),('F11','身分證','H123407309'),
    ('C12','生日','075/04/03'),('F12','發證日','112/03/10 換'),
    ('C13','婚姻','已婚'),('F13','發證地','桃市'),
    ('C14','學歷','專科、大學'),('F14','電話','0953-119943'),
    ('C15','戶籍','桃園市大溪區文化路103巷11號三樓'),
    ('H16','住家備註','同戶籍'),
    ('C17','LINE','0953-119943'),('F17','資金用途','家用'),
    ('H17','公司電話','03-4026689'),
    ('C18','公司','宏羚公司'),('G18','職稱','工程師'),
    ('I18','年資','1年8月'),
    ('I19','月薪','4.5萬'),
    ('C23','聯絡人1','謝汶芮'),('F23','聯絡人2','紀景淳'),
    ('C25','電話1','0976-276827'),('F25','電話2','0923-500672'),
    ('D22','知情1','知情'),('G22','知情2','保密'),
    ('C37','銀行','台灣土地銀行'),('C38','分行','平鎮分行'),
    ('C42','商品','安卓手機'),('F42','型號','OPPO A77'),
    ('C20','電信','台灣大哥大'),
    ('G19','行業','製造業'),
]
ok=fail=0
for ref,label,exp in checks:
    val=str(ws[ref].value) if ws[ref].value else '(empty)'
    if val==exp: ok+=1
    else: fail+=1; print(f'  FAIL {ref}({label}): expect={exp}, got={val}')
print(f'{ok}/{ok+fail} OK')
