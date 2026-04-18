from fastapi import FastAPI, Request, BackgroundTasks
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse
import uvicorn
import requests as req_lib
import os
import re
import sqlite3
import json
import io
from datetime import datetime
import uuid
import secrets
import pathlib
import hmac
import hashlib
import base64
from html import escape as _html_escape
from typing import Optional, List, Dict, Any


def h(val) -> str:
    """HTML escape helper — 防止 XSS（含引號）

    修復 Bug 4：加 quote=True，escape 單引號和雙引號，防止 inline JS/HTML attribute 注入
    修復 Bug 13：用 is None 判斷，避免 0/False 被當空值
    """
    if val is None:
        return ""
    return _html_escape(str(val), quote=True)

app = FastAPI()

CHANNEL_ACCESS_TOKEN = os.getenv("CHANNEL_ACCESS_TOKEN", "")
LINE_CHANNEL_SECRET = os.getenv("LINE_CHANNEL_SECRET", "")
BACKUP_ENABLED = os.getenv("BACKUP_ENABLED", "").lower() == "true"
GOOGLE_SERVICE_ACCOUNT_JSON = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "")
GOOGLE_DRIVE_FOLDER_ID = os.getenv("GOOGLE_DRIVE_FOLDER_ID", "")
BACKUP_KEEP_DAILY = int(os.getenv("BACKUP_KEEP_DAILY", "30"))
BACKUP_KEEP_MONTHLY = int(os.getenv("BACKUP_KEEP_MONTHLY", "12"))
A_GROUP_ID = os.getenv("A_GROUP_ID", "Cb3579e75c94437ed22aafc7b1f6aecdd")
B_GROUP_ID = "Cd14f3ee775f1d9f5cfdafb223173cbef"
C_GROUP_ID = "C1a647fcb29a74842eceeb18e7a53823d"
DB_PATH = os.getenv("DB_PATH", "/var/data/loan_system.db")

# 申請書範本自訂上傳目錄（與 DB 同目錄，保證 Render 持久磁碟保留）
TEMPLATES_DIR = os.path.join(os.path.dirname(DB_PATH) or ".", "templates")
try:
    os.makedirs(TEMPLATES_DIR, exist_ok=True)
except Exception:
    pass

# 11 個 Excel 申請書方案（名稱, 內建範本檔名）— 與 PLAN_TEMPLATE_MAP 對齊
APPLICATION_PLAN_LIST = [
    ("亞太商品", "亞太商品範本.xlsx"),
    ("亞太機車15萬", "亞太15萬機車範本.xlsx"),
    ("亞太機車25萬", "亞太15萬機車範本.xlsx"),
    ("亞太工會機車", "亞太工會範本.xlsx"),
    ("和裕機車", "和裕維力貸機車範本).xlsx"),
    ("和裕商品", "和裕維力貸商品範本.xlsx"),
    ("第一", "第一申請書範本.xlsx"),
    ("貸救補", "貸就補範本.xlsx"),
    ("21機車12萬", "21機車申請書範本xlsx.xlsx"),
    ("21機車25萬", "21機25萬範本.xlsx"),
    ("21商品", "21商品範本.xlsx"),
]

def _plan_to_safe_key(plan_name: str) -> str:
    """把中文方案名轉為安全檔名（防 path traversal）"""
    return plan_name.replace("/", "_").replace("\\", "_").replace("..", "_")

def get_custom_template_path(plan_name: str) -> str:
    """回傳自訂範本絕對路徑（不檢查是否存在）"""
    return os.path.join(TEMPLATES_DIR, f"{_plan_to_safe_key(plan_name)}.xlsx")

def get_template_mapping_path(plan_name: str) -> str:
    """回傳映射 JSON 路徑"""
    return os.path.join(TEMPLATES_DIR, f"{_plan_to_safe_key(plan_name)}.mapping.json")

def get_template_prev_path(plan_name: str) -> str:
    """回傳上一版範本路徑（用於異動比對）"""
    return os.path.join(TEMPLATES_DIR, f"{_plan_to_safe_key(plan_name)}.prev.xlsx")

# ==========================================================
# 範本欄位系統（階段 2：可自訂映射）
# ==========================================================

# 中文欄位字典（顯示給使用者選擇）
FIELD_LABELS = {
    # 基本個人資料
    "customer_name": "姓名",
    "id_no": "身分證字號",
    "birth_date": "出生日期",
    "phone": "行動電話",
    "email": "電子信箱",
    "line_id": "LINE ID",
    "marriage": "婚姻狀態",
    "education": "最高學歷",
    # 身分證
    "id_issue_date": "發證日期",
    "id_issue_place": "發證地",
    "id_issue_type": "換補發類別（初發/補發/換發）",
    # 戶籍地址
    "reg_city": "戶籍縣市",
    "reg_district": "戶籍鄉鎮區",
    "reg_address": "戶籍詳細地址",
    "reg_phone": "戶籍電話",
    "reg_full_address": "戶籍完整地址（縣市+區+地址）",
    # 現居地址
    "live_city": "現居縣市",
    "live_district": "現居鄉鎮區",
    "live_address": "現居詳細地址",
    "live_phone": "現住電話",
    "live_same_as_reg": "現居同戶籍（0/1）",
    "live_full_address": "現居完整地址（自動判斷同戶籍）",
    "live_status": "居住狀況",
    "live_years": "現居年數",
    "live_months": "現居月數",
    # 公司
    "company": "公司名稱（簡版）",
    "company_name_detail": "公司名稱（詳細版）",
    "company_role": "職稱",
    "company_phone_area": "公司電話區碼",
    "company_phone_num": "公司電話號碼",
    "company_phone_ext": "公司電話分機",
    "company_full_phone": "公司完整電話（區碼-號碼）",
    "company_years": "工作年資（年）",
    "company_months": "工作年資（月）",
    "company_salary": "月薪",
    "company_city": "公司縣市",
    "company_district": "公司鄉鎮區",
    "company_address": "公司詳細地址",
    "company_full_address": "公司完整地址",
    # 聯絡人
    "contact1_name": "聯絡人1 姓名",
    "contact1_relation": "聯絡人1 關係",
    "contact1_phone": "聯絡人1 電話",
    "contact1_known": "聯絡人1 知情狀態",
    "contact2_name": "聯絡人2 姓名",
    "contact2_relation": "聯絡人2 關係",
    "contact2_phone": "聯絡人2 電話",
    "contact2_known": "聯絡人2 知情狀態",
    "carrier": "電信業者",
    "fb": "客戶 Facebook",
    # 車輛（機車/汽車）
    "adminb_brand": "廠牌",
    "vehicle_plate": "牌照號碼",
    "adminb_vehicle_type": "車輛型式",
    "adminb_engine_no": "引擎號碼",
    "adminb_body_no": "車身號碼",
    "adminb_mfg_date": "出廠年月",
    "adminb_displacement": "排氣量",
    "adminb_color": "車身顏色",
    # adminB 其他
    "adminb_fund_use": "資金用途（亞太分類）",
    "adminb_industry": "行業類別（亞太用）",
    "adminb_role": "職務（亞太用）",
    "adminb_hr_industry": "行業類別（和裕用）",
    "adminb_hr_role": "職務（和裕用）",
    "adminb_bank": "撥款銀行",
    "adminb_branch": "撥款分行",
    "adminb_product": "商品名稱（和裕）",
    "adminb_model": "商品型號（和裕）",
    "adminb_product_name": "商品名稱（貸救補）",
    "adminb_product_model": "商品型號（貸救補）",
    "adminb_contact_time": "可照會時間",
    # 其他
    "approved_amount": "核准金額",
    # 特殊
    "__CLEAR__": "清空（填空白，不保留示範值）",
    "__KEEP__": "保留原值（不要動這格）",
}

# 每個方案主填寫表名稱對照（用於 sheet 匹配 fallback）
PRIMARY_SHEET_NAMES = {
    "亞太商品": ["進件表格", "工作表3"],
    "亞太機車15萬": ["進件表格", "工作表3"],
    "亞太機車25萬": ["進件表格", "工作表3"],
    "亞太工會機車": ["進件表格", "工作表3"],
    "和裕機車": ["和裕維力貸"],
    "和裕商品": ["和裕維力貸"],
    "第一": ["申請書"],
    "貸救補": ["進件申請書"],
    "21機車12萬": ["工作表1"],
    "21機車25萬": ["工作表1"],
    "21商品": ["工作表1"],
}

# 預設映射模板（使用者首次建立映射時可套用）
# 結構：{plan: {sheet_name: {cell: field_key}}}
# 特殊值：_rule: "clear_all" 表示該 sheet 所有對應儲存格清空
DEFAULT_MAPPINGS = {
    "亞太商品": {
        "進件表格": {
            "B5": "adminb_fund_use",
            "B9": "customer_name", "D9": "id_no", "F9": "birth_date",
            "B10": "marriage", "D10": "education",
            "B11": "id_issue_date", "D11": "id_issue_place", "F11": "id_issue_type",
            "B12": "reg_city", "C12": "reg_district", "D12": "reg_address",
            "B13": "live_city", "C13": "live_district", "D13": "live_address",
            "B14": "live_status", "D14": "live_years",
            "B15": "phone", "B16": "email",
            "B17": "company_name_detail",
            "E17": "adminb_industry", "G17": "adminb_role",
            "B18": "company_phone_area", "C18": "company_phone_num", "E18": "company_phone_ext",
            "G18": "company_years", "H18": "company_salary",
            "B19": "company_city", "C19": "company_district", "D19": "company_address",
            "B21": "contact1_name", "D21": "contact1_relation",
            "B25": "contact1_phone",
        },
        "擔保品資訊": {
            "B2": "__CLEAR__", "B3": "__CLEAR__", "B4": "__CLEAR__",
            "B5": "__CLEAR__", "B6": "__CLEAR__", "B7": "__CLEAR__", "B8": "__CLEAR__",
        }
    },
    "亞太機車15萬": {
        "進件表格": {
            "B5": "adminb_fund_use",
            "B7": "adminb_vehicle_type", "D7": "adminb_engine_no",
            "F7": "adminb_displacement", "H7": "adminb_color",
            "B9": "customer_name", "D9": "id_no", "F9": "birth_date",
            "B10": "marriage", "D10": "education",
            "B11": "id_issue_date", "D11": "id_issue_place", "F11": "id_issue_type",
            "B12": "reg_city", "C12": "reg_district", "D12": "reg_address",
            "B13": "live_city", "C13": "live_district", "D13": "live_address",
            "B14": "live_status", "D14": "live_years",
            "B15": "phone", "B16": "email",
            "B17": "company_name_detail",
            "E17": "adminb_industry", "G17": "adminb_role",
            "B18": "company_phone_area", "C18": "company_phone_num", "E18": "company_phone_ext",
            "G18": "company_years", "H18": "company_salary",
            "B19": "company_city", "C19": "company_district", "D19": "company_address",
            "B21": "contact1_name", "D21": "contact1_relation",
            "B25": "contact1_phone",
            "K2": "adminb_brand", "K3": "vehicle_plate",
        },
        "擔保品資訊": {
            "B2": "adminb_brand", "B3": "vehicle_plate",
            "B4": "adminb_vehicle_type", "B5": "adminb_engine_no",
            "B6": "adminb_body_no", "B7": "adminb_mfg_date",
            "B8": "adminb_displacement",
        }
    },
    "和裕機車": {
        "和裕維力貸": {
            "C11": "customer_name", "F11": "id_no",
            "C12": "birth_date", "F12": "id_issue_date",
            "C13": "marriage", "F13": "id_issue_place",
            "C14": "education", "F14": "phone",
            "C15": "reg_full_address", "G15": "reg_phone",
            "C16": "live_full_address", "G16": "live_phone",
            "C17": "line_id",
            "C18": "company_name_detail", "G18": "company_role",
            "H17": "company_full_phone",
            "C19": "company_full_address",
            "I18": "company_years", "I19": "company_salary",
            "C20": "carrier",
            "G19": "adminb_hr_industry",
            "D22": "contact1_known", "G22": "contact2_known",
            "C23": "contact1_name", "F23": "contact2_name",
            "C24": "contact1_relation", "F24": "contact2_relation",
            "C25": "contact1_phone", "F25": "contact2_phone",
            "C37": "adminb_bank", "C38": "adminb_branch",
            "C42": "adminb_product", "F42": "adminb_model",
        },
        "擔保品資訊": {
            "B2": "adminb_brand", "B3": "vehicle_plate",
            "B4": "adminb_vehicle_type", "B5": "adminb_engine_no",
            "B6": "adminb_body_no", "B7": "adminb_mfg_date",
            "B8": "adminb_displacement",
        }
    },
    "21機車12萬": {
        "工作表1": {
            "C3": "customer_name", "E3": "id_no", "J3": "phone",
            "C4": "birth_date", "E4": "id_issue_date",
            "F4": "id_issue_place", "G4": "id_issue_type",
            "J4": "live_phone",
            "C6": "reg_city", "D6": "reg_district", "E6": "reg_address",
            "C7": "live_city", "D7": "live_district", "E7": "live_address",
            "C8": "email",
            "C9": "company_name_detail",
            "C10": "company_role", "E10": "company_salary",
            "G10": "company_years", "H10": "company_months",
            "C17": "contact1_name", "E17": "contact1_relation",
            "H17": "contact1_phone", "K17": "contact1_known",
            "C18": "contact2_name", "E18": "contact2_relation",
            "H18": "contact2_phone", "K18": "contact2_known",
        },
        "擔保品資訊": {
            "B2": "adminb_brand", "B3": "vehicle_plate",
            "B4": "adminb_vehicle_type", "B5": "adminb_engine_no",
            "B6": "adminb_body_no", "B7": "adminb_mfg_date",
            "B8": "adminb_displacement",
        }
    },
}
# 相近方案共用預設映射
DEFAULT_MAPPINGS["亞太機車25萬"] = DEFAULT_MAPPINGS["亞太機車15萬"]
DEFAULT_MAPPINGS["亞太工會機車"] = DEFAULT_MAPPINGS["亞太機車15萬"]
DEFAULT_MAPPINGS["和裕商品"] = {
    "和裕維力貸": dict(DEFAULT_MAPPINGS["和裕機車"]["和裕維力貸"]),
    # 和裕商品沒有擔保品資訊需求
}
DEFAULT_MAPPINGS["21機車25萬"] = DEFAULT_MAPPINGS["21機車12萬"]
DEFAULT_MAPPINGS["21商品"] = {
    "工作表1": dict(DEFAULT_MAPPINGS["21機車12萬"]["工作表1"])
    # 21商品無擔保品
}


def _build_reverse_field_map(plan_name: str) -> dict:
    """從 DEFAULT_MAPPINGS 主表反向產生 {field_key: 預設原始 cell}。
    僅使用「主表」（如進件表格、工作表1、申請書等）建立反向映射，
    避免跨 sheet 衝突（例如進件表格 B7=車輛型式 vs 擔保品資訊 B7=出廠年月）。
    """
    out = {}
    plan_defaults = DEFAULT_MAPPINGS.get(plan_name, {})
    primary_aliases = set(PRIMARY_SHEET_NAMES.get(plan_name, []))
    # 找主表：優先 PRIMARY_SHEET_NAMES 內的名稱，否則第一個 sheet
    primary_sheet = None
    for sn in plan_defaults.keys():
        if sn in primary_aliases:
            primary_sheet = sn
            break
    if primary_sheet is None and plan_defaults:
        primary_sheet = next(iter(plan_defaults.keys()))
    if primary_sheet is None:
        return out
    cell_map = plan_defaults.get(primary_sheet, {})
    for cell, field in cell_map.items():
        if not field or field in ("__CLEAR__", "__KEEP__"):
            continue
        if field not in out:
            out[field] = cell
    return out


def compute_field_value(field_key: str, r: dict, plan_name: str = "",
                        processed_cells: dict = None, reverse_map: dict = None) -> str:
    """
    回傳該 field 在該 plan 下要填入的值。
    優先順序：__CLEAR__/__KEEP__ → 衍生欄位 → plan-aware 處理過的值 → 原始 DB 值
    """
    if field_key is None or field_key == "":
        return None
    if field_key == "__KEEP__":
        return None
    if field_key == "__CLEAR__":
        return ""

    def v(k):
        x = r.get(k, "")
        if x is None:
            return ""
        return str(x).strip()

    # 衍生欄位（不需 plan formatter）
    if field_key == "reg_full_address":
        return (v("reg_city") + v("reg_district") + v("reg_address")).strip()
    if field_key == "live_full_address":
        if v("live_same_as_reg") == "1":
            return (v("reg_city") + v("reg_district") + v("reg_address")).strip()
        # 縣市+區相同 → 當成同戶籍，用戶籍地
        if v("live_city") and v("live_city") == v("reg_city") and v("live_district") == v("reg_district"):
            return (v("reg_city") + v("reg_district") + v("reg_address")).strip()
        return (v("live_city") + v("live_district") + v("live_address")).strip()
    if field_key == "company_full_address":
        return (v("company_city") + v("company_district") + v("company_address")).strip()
    if field_key == "company_full_phone":
        a = v("company_phone_area")
        n = v("company_phone_num")
        if a == "mobile":
            a = ""
        return (a + "-" + n) if (a and n) else n

    # plan-aware 處理：從現有 _build_cell_map 取得處理過的值
    # 必須有 processed_cells（_build_cell_map 結果）+ reverse_map
    if processed_cells and reverse_map and field_key in reverse_map:
        original_cell = reverse_map[field_key]
        if original_cell in processed_cells:
            val = processed_cells[original_cell]
            if val is None:
                # 預設邏輯回 None 表示不動 → 我們改為清空 / 用 raw
                return v(field_key)
            return str(val) if not isinstance(val, str) else val

    # Fallback：直接取 DB 值
    return v(field_key)


def scan_xlsx_structure(xlsx_bytes: bytes) -> dict:
    """
    掃描 xlsx 結構，回傳所有 sheet 的儲存格資訊（含下拉選單）。
    回傳格式：
    {
        "sheets": [
            {
                "name": "進件表格",
                "state": "visible",
                "cells": [
                    {"ref": "A1", "row": 1, "col": "A", "value": "申請書", "type": "text"},
                    {"ref": "B5", "row": 5, "col": "B", "value": "", "type": "empty",
                     "dropdown": ["選項1","選項2"]},
                ]
            },
            ...
        ]
    }
    """
    import zipfile as _zf
    import re as _re

    z = _zf.ZipFile(io.BytesIO(xlsx_bytes))
    try:
        wb_xml = z.read("xl/workbook.xml").decode("utf-8", errors="replace")
        rels_xml = z.read("xl/_rels/workbook.xml.rels").decode("utf-8", errors="replace")
    except KeyError:
        return {"sheets": [], "error": "不是合法 xlsx"}

    # shared strings
    ss_texts = []
    try:
        ss_xml = z.read("xl/sharedStrings.xml").decode("utf-8", errors="replace")
        for m in _re.finditer(r'<si>(.*?)</si>', ss_xml, _re.DOTALL):
            si = m.group(1)
            ts = _re.findall(r'<t[^>]*>([^<]*)</t>', si)
            ss_texts.append("".join(ts))
    except KeyError:
        pass

    # sheet 列表
    rel_map = dict(_re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', rels_xml))
    sheets_info = []
    for m in _re.finditer(r'<sheet\b[^/]*?/>', wb_xml):
        block = m.group(0)
        nm = _re.search(r'name="([^"]+)"', block)
        rm = _re.search(r'r:id="([^"]+)"', block)
        sm = _re.search(r'state="([^"]+)"', block)
        if not (nm and rm):
            continue
        sheet_name = nm.group(1)
        rid = rm.group(1)
        state = sm.group(1) if sm else "visible"
        target = rel_map.get(rid, "")
        if not target:
            continue
        path = "xl/" + target.lstrip("/")
        try:
            sheet_xml = z.read(path).decode("utf-8", errors="replace")
        except KeyError:
            continue

        # 解析所有儲存格
        cells = []
        for cm in _re.finditer(
            r'<c\s+r="([A-Z]+)(\d+)"(?:\s+[^>]*?)?\s*(?:\s+t="(\w+)")?[^>]*?>(?:\s*<f[^>]*>([^<]*)</f>)?\s*(?:<v>([^<]*)</v>|<is><t[^>]*>([^<]*)</t></is>)?\s*</c>',
            sheet_xml
        ):
            col = cm.group(1); row = int(cm.group(2))
            t = cm.group(3)
            formula = cm.group(4)
            v_content = cm.group(5)
            inline_t = cm.group(6)

            if inline_t is not None:
                val = inline_t
                ctype = "text"
            elif t == "s" and v_content is not None:
                try:
                    val = ss_texts[int(v_content)]
                    ctype = "text"
                except (ValueError, IndexError):
                    val = f"ss#{v_content}"
                    ctype = "text"
            elif v_content is not None:
                val = v_content
                ctype = "number" if not formula else "formula"
            elif formula:
                val = ""
                ctype = "formula"
            else:
                val = ""
                ctype = "empty"

            cells.append({
                "ref": f"{col}{row}",
                "row": row, "col": col,
                "value": val,
                "type": ctype,
                "has_formula": bool(formula),
            })

        # 下拉選單 (dataValidations)
        for dv in _re.finditer(
            r'<dataValidation\s+type="list"[^>]*?sqref="([^"]+)"[^>]*>(.*?)</dataValidation>',
            sheet_xml, _re.DOTALL
        ):
            sqref = dv.group(1)
            inner = dv.group(2)
            f1 = _re.search(r'<formula1>([^<]+)</formula1>', inner)
            if not f1:
                continue
            raw = f1.group(1)
            options = []
            if raw.startswith('"') and raw.endswith('"'):
                # 內嵌選項 "選項1,選項2"
                options = [o.strip() for o in raw.strip('"').split(",")]
            # 套用到所有符合的 cell（sqref 可能是 "B5 B6" 或 "B5:B10"）
            targets = set()
            for part in sqref.split():
                if ":" in part:
                    a, b = part.split(":")
                    am = _re.match(r'([A-Z]+)(\d+)', a)
                    bm = _re.match(r'([A-Z]+)(\d+)', b)
                    if am and bm and am.group(1) == bm.group(1):
                        for rr in range(int(am.group(2)), int(bm.group(2)) + 1):
                            targets.add(f"{am.group(1)}{rr}")
                else:
                    targets.add(part)
            for c in cells:
                if c["ref"] in targets:
                    c["dropdown"] = options

        sheets_info.append({
            "name": sheet_name,
            "state": state,
            "cells": cells,
        })

    z.close()
    return {"sheets": sheets_info}


SIBLING_PLAN_GROUPS = [
    # 同一群組內的方案共用映射框架（座標不變、欄位對應一致）
    {"亞太機車15萬", "亞太機車25萬", "亞太工會機車"},
    {"和裕機車", "和裕商品"},
    {"21機車12萬", "21機車25萬", "21商品"},
]

def load_template_mapping(plan_name: str) -> dict:
    """讀取 {plan}.mapping.json，若不存在則回傳同群組 sibling 的映射作為 fallback"""
    path = get_template_mapping_path(plan_name)
    if os.path.isfile(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    # 自己沒映射 → 找同群組 sibling 的
    for group in SIBLING_PLAN_GROUPS:
        if plan_name in group:
            for sibling in group:
                if sibling == plan_name:
                    continue
                sib_path = get_template_mapping_path(sibling)
                if os.path.isfile(sib_path):
                    try:
                        with open(sib_path, "r", encoding="utf-8") as f:
                            return json.load(f)
                    except Exception:
                        pass
    return {}


def get_inherited_mapping_source(plan_name: str) -> str:
    """若該方案的映射是繼承自 sibling，回傳 sibling 名稱；否則回傳空字串。"""
    if os.path.isfile(get_template_mapping_path(plan_name)):
        return ""
    for group in SIBLING_PLAN_GROUPS:
        if plan_name in group:
            for sibling in group:
                if sibling != plan_name and os.path.isfile(get_template_mapping_path(sibling)):
                    return sibling
    return ""


def save_template_mapping(plan_name: str, mapping: dict) -> None:
    os.makedirs(TEMPLATES_DIR, exist_ok=True)
    path = get_template_mapping_path(plan_name)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(mapping, f, ensure_ascii=False, indent=2)


def get_default_mapping(plan_name: str) -> dict:
    """取得預設映射（做為首次編輯時的起點）"""
    return DEFAULT_MAPPINGS.get(plan_name, {})


def _strip_sheet_colors(xlsx_bytes: bytes, target_sheet_name: str) -> bytes:
    """
    把指定工作表的所有儲存格 fill 改為無色（白色），並清除該工作表的條件格式。
    保留字體、邊框、對齊、數字格式。
    用於上傳範本時自動清除「擔保品資訊」分頁的黃底等視覺干擾。
    """
    import zipfile as _zf
    import re as _re

    try:
        z_in = _zf.ZipFile(io.BytesIO(xlsx_bytes))
    except Exception:
        return xlsx_bytes

    try:
        wb_xml = z_in.read('xl/workbook.xml').decode('utf-8')
        rels_xml = z_in.read('xl/_rels/workbook.xml.rels').decode('utf-8')
    except KeyError:
        z_in.close()
        return xlsx_bytes

    rel_map = dict(_re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', rels_xml))
    target_path = None
    for m in _re.finditer(r'<sheet\b[^/]*?/>', wb_xml):
        block = m.group(0)
        nm = _re.search(r'name="([^"]+)"', block)
        rm = _re.search(r'r:id="([^"]+)"', block)
        if nm and rm and nm.group(1) == target_sheet_name:
            target_path = 'xl/' + rel_map[rm.group(1)].lstrip('/')
            break

    if not target_path or target_path not in z_in.namelist():
        z_in.close()
        return xlsx_bytes

    sheet_xml = z_in.read(target_path).decode('utf-8')
    try:
        styles_xml = z_in.read('xl/styles.xml').decode('utf-8')
    except KeyError:
        z_in.close()
        return xlsx_bytes

    # 解析 cellXfs
    xfs_m = _re.search(r'(<cellXfs[^>]*>)(.*?)(</cellXfs>)', styles_xml, _re.DOTALL)
    if not xfs_m:
        z_in.close()
        return xlsx_bytes
    xfs_open = xfs_m.group(1)
    xfs_body = xfs_m.group(2)
    xfs_close = xfs_m.group(3)

    xf_blocks = _re.findall(r'<xf\s[^/]*(?:/>|>(?:[^<]|<(?!/xf>))*</xf>)', xfs_body, _re.DOTALL)

    # 找出哪些 fontId 的字體有非預設色（紅、藍等），等下要把它們也標準化
    fonts_m = _re.search(r'<fonts[^>]*>(.*?)</fonts>', styles_xml, _re.DOTALL)
    colored_fonts = set()
    if fonts_m:
        font_blocks = _re.findall(r'<font>(.*?)</font>', fonts_m.group(1), _re.DOTALL)
        for fi, fb in enumerate(font_blocks):
            cm = _re.search(r'<color\s+rgb="([^"]+)"', fb)
            if cm:
                rgb = cm.group(1).upper()
                # 任何指定 RGB（非 theme/auto）都視為「有顏色」，包括 FF000000 黑色
                # 但黑色 (FF000000 / 000000) 不算干擾
                if rgb not in ("FF000000", "000000"):
                    colored_fonts.add(fi)

    # 對該 sheet 內所有 cell 的 s="X"，建立 X -> 新 index 映射
    # 兩個目標：(1) fillId=0  (2) fontId 非彩色字體
    s_old_to_new = {}
    new_xfs = []
    for s_match in _re.finditer(r'<c\s+r="[A-Z]+\d+"\s+s="(\d+)"', sheet_xml):
        old_s = int(s_match.group(1))
        if old_s in s_old_to_new or old_s >= len(xf_blocks):
            continue
        old_xf = xf_blocks[old_s]
        fill_m = _re.search(r'fillId="(\d+)"', old_xf)
        font_m = _re.search(r'fontId="(\d+)"', old_xf)
        need_new = False
        new_xf = old_xf
        if fill_m and fill_m.group(1) != '0':
            new_xf = _re.sub(r'fillId="\d+"', 'fillId="0"', new_xf)
            new_xf = _re.sub(r'\s+applyFill="1"', '', new_xf)
            need_new = True
        if font_m and int(font_m.group(1)) in colored_fonts:
            # 把 fontId 改成 0（預設黑字）
            new_xf = _re.sub(r'fontId="\d+"', 'fontId="0"', new_xf)
            new_xf = _re.sub(r'\s+applyFont="1"', '', new_xf)
            need_new = True
        if need_new:
            new_idx = len(xf_blocks) + len(new_xfs)
            new_xfs.append(new_xf)
            s_old_to_new[old_s] = new_idx
        else:
            s_old_to_new[old_s] = old_s

    # 更新 styles.xml
    if new_xfs:
        new_count = len(xf_blocks) + len(new_xfs)
        new_xfs_open = _re.sub(r'count="\d+"', f'count="{new_count}"', xfs_open)
        new_styles_xml = (styles_xml[:xfs_m.start()] +
                          new_xfs_open + xfs_body + ''.join(new_xfs) + xfs_close +
                          styles_xml[xfs_m.end():])
    else:
        new_styles_xml = styles_xml

    # 更新 sheet xml: 替換 s 屬性
    def repl_s(mm):
        old_s = int(mm.group(1))
        new_s = s_old_to_new.get(old_s, old_s)
        if new_s == old_s:
            return mm.group(0)
        return _re.sub(r's="\d+"', f's="{new_s}"', mm.group(0))
    new_sheet_xml = _re.sub(r'<c\s+r="[A-Z]+\d+"\s+s="(\d+)"', repl_s, sheet_xml)

    # 移除條件格式（普通 + x14 擴充）
    new_sheet_xml = _re.sub(r'<conditionalFormatting[^>]*>.*?</conditionalFormatting>',
                             '', new_sheet_xml, flags=_re.DOTALL)
    new_sheet_xml = _re.sub(r'<x14:conditionalFormattings[^>]*>.*?</x14:conditionalFormattings>',
                             '', new_sheet_xml, flags=_re.DOTALL)
    # 清空殘留 ext / extLst
    new_sheet_xml = _re.sub(r'<ext\s[^>]*>\s*</ext>', '', new_sheet_xml)
    new_sheet_xml = _re.sub(r'<extLst>\s*</extLst>', '', new_sheet_xml)

    # 重新打包
    out_buf = io.BytesIO()
    z_out = _zf.ZipFile(out_buf, 'w', _zf.ZIP_DEFLATED)
    for item in z_in.infolist():
        if item.filename == target_path:
            z_out.writestr(item, new_sheet_xml.encode('utf-8'))
        elif item.filename == 'xl/styles.xml' and new_xfs:
            z_out.writestr(item, new_styles_xml.encode('utf-8'))
        else:
            z_out.writestr(item, z_in.read(item.filename))
    z_out.close()
    z_in.close()
    return out_buf.getvalue()

# 日報三段結構
REPORT_SECTION_1 = [
    "麻吉", "和潤", "中租", "裕融", "21汽車", "亞太", "創鉅", "21",
    "第一", "合信", "興達", "和裕", "鄉民", "喬美",
    "分貝汽車", "分貝機車", "貸救補", "預付手機分期", "融易", "手機分期", "鼎多",
    "送件",
]
REPORT_SECTION_4 = ["待撥款"]
REPORT_SECTION_2 = [
    "銀行", "零卡", "商品貸", "代書", "當舖專案", "核准",
]
REPORT_SECTION_3 = [
    "房地", "新鑫", "核准(房地)",
]
# 合併（用於其他地方參考）
REPORT_SECTIONS = REPORT_SECTION_1 + REPORT_SECTION_2 + REPORT_SECTION_3 + REPORT_SECTION_4

# 公司辨識（解析訊息用）
COMPANY_LIST = [
    # 和裕方案
    "和裕商品", "和裕機車", "和裕機",
    # 亞太方案（含帶金額/工會版本）
    "亞太商品", "亞太機車", "亞太工會", "亞太汽車",
    "亞太工15", "亞太工", "亞太機",
    "亞太25", "亞太15",
    "預付手機原融", "預付手機分期",
    "手機分期", "貸救補",
    "合信機車", "合信手機", "合信二輪",
    # 分貝方案
    "分貝汽車", "分貝機車", "分貝汽", "分貝機",
    # 21方案
    "21汽車", "21機車", "21機25萬", "21機25", "21汽", "21機", "21商品",
    # 創鉅方案
    "創鉅手機", "創鉅機車", "創鉅手", "創鉅機",
    # 麻吉方案
    "麻吉機車", "麻吉手機", "麻吉機", "麻吉手",
    "喬美40", "喬美14",
    "興達機車", "興達機",
    "第一機車", "第一商品",
    "鄉民", "喬美", "麻吉", "亞太",
    "和裕", "第一", "合信", "興達", "中租", "裕融", "裕榮", "創鉅", "和潤",
    "銀行", "零卡", "銀角", "商品貸", "代書", "當舖", "融易",
    "慢點付", "分期趣", "鼎多", "預付手機", "21",
    "新鑫", "房地", "土地",
    "分貝", "鄉", "銀", "C", "商", "代", "研", "當",
]

STATUS_WORDS = [
    "婉拒", "核准", "核準", "待核准", "附條件", "等保書",
    "補件", "補資料", "退件", "不承作", "無額度", "無法承作", "照會",
    "保密", "NA", "待撥款", "可送", "缺資料", "無可知情", "聯絡人皆可知情",
    "補行照", "補照會", "補照片", "補時段", "補案件資料", "補聯徵", "補保人",
    "已補", "待補", "轉", "申覆",
    "派對保", "委對收", "對好", "對保完成", "不收不簽",
    "已撥款", "今日撥款", "今日已撥款", "已加撥", "核准已撥",
    "預計排撥", "排撥",
]

ACTION_KEYWORDS = [
    "結案", "刪掉", "不追了", "全部不送", "已撥款結案",
    "補件", "補資料", "缺資料", "婉拒", "核准", "核準", "待核准", "附條件",
    "照會", "退件", "等保書",
    "不承作", "無額度", "無法承作", "待撥款",
    "補行照", "補照會", "補照片", "補時段", "補案件資料",
    "補聯徵", "補保人", "保密", "無可知情", "聯絡人皆可知情", "已補",
    "補案件", "補", "待補", "重啟", "再辦",
    "派對保", "委對收", "對好", "對保完成", "不收不簽",
    "已撥款", "今日撥款", "今日已撥款", "已加撥", "核准已撥",
    "預計排撥", "排撥",
]

DELETE_KEYWORDS = [
    "結案", "刪掉", "不追了", "全部不送", "已撥款結案",
    "違約金結案", "已支付違約金", "違約金已支付",
    "已收到違約金",  # Bug 14: 正確字
    "以收到違約金",  # Bug 14: 錯字版本，向後相容
    "收到違約金", "違約金已收", "違約金",
]
BLOCK_KEYWORDS = ["鼎信", "禾基"]

# 專案別名對照
# ============================
# 公司別名對照表（完整版）
# ============================
COMPANY_ALIAS = {
    # 亞太相關
    "熊速貸": "亞太商品",
    "工會機車動擔": "亞太工會",
    "機車動擔設定": "亞太機車",
    "亞太(工會)": "亞太工會",
    "亞太(工)": "亞太工會",
    "亞太工會15萬": "亞太工會",
    # 和裕相關
    "維力商品貸": "和裕商品",
    "維力機車專": "和裕機車",
    "維力機車貸": "和裕機車",
    "維力": "和裕",
    # 21 全形/異體字
    "廿一": "21",
    "二十一世紀": "21",
    "２１": "21",
    # 21 車型縮寫（canonical 對應 PLAN_INFO）
    "21": "21商品",
    "21機": "21機車12萬",
    "21機25": "21機車25萬",
    "21機12萬": "21機車12萬",
    "21機25萬": "21機車25萬",
    # 送件順序常用單字簡稱
    "鄉": "鄉民",
    "銀": "銀行",
    "C": "零卡",
    "商": "商品貸",
    "代": "代書",
    "當": "當舖",
    "喬": "喬美",
    "麻": "麻吉機車",
    "分": "分貝機車",
    "研": "商品貸",  # 研=商
    "鼎多": "喬美",  # 鼎多=喬美
    # 興達/貸救補/歐
    "興機": "興達機車",
    "貸10": "貸救補",
    "貸救": "貸救補",
    "貸就補": "貸救補",  # 錯字向下相容
    "歐": "商品貸",
    # 新新專/維力
    "新新專": "和裕",
    "新新": "和裕",
    # 一路發
    "一路發汽車貸款": "裕融",
    "一路發代償專案": "裕融",
    # 英文縮寫
    "TAC": "裕融",
    "EGO": "第一",
    # 融資房地
    "新鑫": "房地",
    # 銀行類 → 統一歸到「銀行」
    "中信": "銀行", "中國信託": "銀行",
    "台新": "銀行", "國泰": "銀行", "富邦": "銀行", "玉山": "銀行",
    "永豐": "銀行", "新光": "銀行", "凱基": "銀行", "元大": "銀行",
    "兆豐": "銀行", "華南": "銀行",
    "聯邦": "銀行", "遠東": "銀行", "王道": "銀行", "星展": "銀行",
    "渣打": "銀行", "匯豐": "銀行", "花旗": "銀行", "樂天": "銀行",
    "合庫": "銀行", "土銀": "銀行", "台銀": "銀行", "彰銀": "銀行",
    "板信": "銀行", "陽信": "銀行", "京城": "銀行", "安泰": "銀行",
    "日盛": "銀行", "上海商銀": "銀行", "台中銀": "銀行", "高雄銀": "銀行",
    "三信商銀": "銀行", "大眾": "銀行", "台灣企銀": "銀行", "台企銀": "銀行",
    # 零卡類 → 統一歸到「零卡」
    "大哥付": "零卡", "幫你付": "零卡", "遠信": "零卡", "月付大人": "零卡",
    "銀角": "零卡", "慢點付": "零卡", "先享後付": "零卡",
    "PI錢包": "零卡", "PI": "零卡", "FULA": "零卡", "分期趣": "零卡",
    "刷卡換現": "零卡", "信用卡換現": "零卡", "信用卡": "零卡",
}

# 方案簡稱→(正式名稱, 金額期數描述) 照會同時送件用
PLAN_INFO = {
    "亞太商品": ("亞太商品", "12萬/30期"), "亞太12": ("亞太商品", "12萬/30期"), "亞太": ("亞太商品", "12萬/30期"),
    "亞太機車15萬": ("亞太機車15萬", "15萬/36期"), "亞太15": ("亞太機車15萬", "15萬/36期"),
    "亞太機車25萬": ("亞太機車25萬", "25萬/48期"), "亞太25": ("亞太機車25萬", "25萬/48期"),
    "亞太機25萬": ("亞太機車25萬", "25萬/48期"),
    "亞太機": ("亞太機車15萬", "15萬/36期"), "亞太機車": ("亞太機車15萬", "15萬/36期"),
    "亞太汽車": ("亞太汽車", ""), "亞太汽": ("亞太汽車", ""),
    "亞太工會機車": ("亞太工會機車", "15萬"), "亞太工會": ("亞太工會機車", "15萬"), "亞太工": ("亞太工會機車", "15萬"),
    "和裕機車": ("和裕機車", "15萬/24期"), "和裕機": ("和裕機車", "15萬/24期"),
    "和裕商品": ("和裕商品", "12萬/24期"), "和裕": ("和裕商品", "12萬/24期"),
    "21機車12萬": ("21機車12萬", "12萬/24期"), "21機": ("21機車12萬", "12萬/24期"),
    "21機車25萬": ("21機車25萬", "25萬/48期"), "21機25": ("21機車25萬", "25萬/48期"),
    "21商品": ("21商品", "12萬/24期"), "21": ("21商品", "12萬/24期"),
    "第一": ("第一", "30萬/24期"),
    "貸救補": ("貸救補", "10萬/24期"), "貸10": ("貸救補", "10萬/24期"), "貸就補": ("貸救補", "10萬/24期"),
    "麻吉機車": ("麻吉機車", "10萬/24期"), "麻吉機": ("麻吉機車", "10萬/24期"),
    "麻吉手機": ("麻吉手機", "10萬/24期"), "麻吉手": ("麻吉手機", "10萬/24期"),
    "喬美": ("喬美", "14萬/30期"), "鼎多": ("喬美", "14萬/30期"),
    "分貝機車": ("分貝機車", ""), "分貝機": ("分貝機車", ""),
    "分貝汽車": ("分貝汽車", ""), "分貝汽": ("分貝汽車", ""),
    "21汽車": ("21汽車", ""), "21汽": ("21汽車", ""),
    "鄉民": ("鄉民貸", ""), "鄉": ("鄉民貸", ""),
    "銀行": ("銀行", ""), "銀": ("銀行", ""),
    "零卡": ("零卡", ""), "C": ("零卡", ""),
    "商品貸": ("商品貸", ""), "商": ("商品貸", ""),
    "代書": ("代書", ""), "代": ("代書", ""),
    "當舖": ("當舖", ""), "當": ("當舖", ""),
    "房地": ("房地", ""), "新鑫": ("房地", ""),
}

APPROVAL_EXCLUDE_KEYWORDS = ["初估", "待補", "照會", "金主初估", "需補", "補資料才"]

IGNORE_NAME_SET = {
    "信用不良", "不需要了", "不用了", "不要了", "結案", "補件", "核准", "婉拒",
    "申覆", "申請", "待審", "初估",
    "照會", "等保書", "待撥款", "缺資料", "補資料", "資料補", "補來", "退件",
    "助理", "AI助理", "先生", "小姐", "無可知情", "聯絡人皆可知情", "下一家",
    "可知情", "聯絡人", "皆可知情", "不可知情", "無空間", "信用卡",
    "機車貸", "汽車貸", "商品貸", "無貸款", "合照後", "後補", "來補",
    "空間", "貸款", "申請", "提供", "保證", "聯徵", "繳息",
    "NA", "RE", "JCIC", "ID", "CCIS", "TAC", "EGO",
}

CHINESE_NAME_RE = re.compile(r"[\u4e00-\u9fff]{2,6}")
ID_RE = re.compile(r"[A-Z][A-Z0-9]\d{8}")

# 支援有無 - 的日期格式，如 115/3/2廖俊宏 或 115/3/2-廖俊宏
DATE_NAME_ID_INLINE_RE = re.compile(
    r"^\s*(\d{2,6}/\d{1,2}/\d{1,2})\s*[-－]?\s*([\u4e00-\u9fff]{2,6})\s*([A-Z][A-Z0-9]\d{8})",
    re.IGNORECASE,
)
DATE_NAME_ONLY_RE = re.compile(
    r"^\s*(\d{2,6}/\d{1,2}/\d{1,2})\s*[-－]?\s*([\u4e00-\u9fff]{2,6})(?:\s*$|\s*[-－/\u4e00-\u9fff])",
    re.IGNORECASE,
)
# 短日期：3/2廖俊宏 或 3/2-廖俊宏（有無-都支援）
SHORT_DATE_NAME_RE = re.compile(
    r"^\s*(\d{1,2}/\d{1,2})\s*[-－]?\s*([\u4e00-\u9fff]{2,6})"
)
# 送件順序格式：4/1-高郡惠-喬美/亞太/和裕
ROUTE_ORDER_RE = re.compile(
    r"^\s*(\d{1,4}/\d{1,2}(?:/\d{1,2})?)\s*[-－]\s*([\u4e00-\u9fff]{2,6})\s*[-－]\s*((?:[^\s/\n]+/)*[^\s/\n@]+)(?:\s*@AI)?\s*$",
    re.IGNORECASE,
)
# 單公司核准/婉拒格式：03/04-黃娫柔-房地核准20萬 / 03/04-黃娫柔-新鑫核准20萬
# 用 negative lookbehind 排除「待核准」「待核準」誤判為核准
SINGLE_APPROVAL_RE = re.compile(
    r"^\s*(\d{1,4}/\d{1,2}(?:/\d{1,2})?)\s*[-－]\s*([\u4e00-\u9fff]{2,6})\s*[-－]\s*([^\s/\n@-]+?)(?<!待)(核准|核準|婉拒)(\d+(?:\.\d+)?萬)?(?:\s*@AI)?\s*$",
    re.IGNORECASE,
)

# ⭐ 照會注意事項偵測（等同已送件）
NOTIFICATION_TRIGGER_RE = re.compile(r"照會注意事項")
EXTRA_COMPANY_RE = re.compile(r"[+＋]\s*([\u4e00-\u9fff0-9]{1,8}?)\s*一起")
# 轉送格式：8/5-戴君哲-轉21 或 8/11-林曉薇-轉麻吉 6/18
TRANSFER_RE = re.compile(
    r"^\s*(\d{1,4}/\d{1,2}(?:/\d{1,2})?)\s*[-－]\s*([\u4e00-\u9fff]{2,6})\s*[-－]\s*轉\s*([^\s@]+)(?:\s.*)?(?:\s*@AI)?\s*$",
    re.IGNORECASE,
)


def is_notification_briefing(text: str) -> bool:
    """是否為照會注意事項格式：第一行姓名（2-4中文字）+ 第二行「照會注意事項」"""
    if not text:
        return False
    lines = [l.strip() for l in text.strip().splitlines() if l.strip()]
    if len(lines) < 2:
        return False
    # 第一行：純中文姓名 2-4 字
    name = lines[0]
    if not re.fullmatch(r"[\u4e00-\u9fff]{2,6}", name):
        return False
    # 第二行：照會注意事項
    if "照會注意事項" not in lines[1]:
        return False
    return True


def extract_extra_company(text: str) -> str:
    """從照會訊息末尾抓「+XXX一起」的額外公司名（原始字串，呼叫處再正規化）"""
    if not text:
        return ""
    m = EXTRA_COMPANY_RE.search(text)
    return m.group(1).strip() if m else ""


def parse_notification_fields(text: str) -> dict:
    """從照會注意事項訊息提取客戶欄位資料"""
    fields = {}
    # 居住年數：居住5年
    m = re.search(r"居住\s*(\d+)\s*年", text)
    if m:
        fields["live_years"] = m.group(1)
    # 居住狀況：父母名下/自有/租屋/配偶
    for status in ["父母", "自有", "租屋", "配偶", "親屬", "宿舍"]:
        if status in text:
            fields["live_status"] = status + ("名下" if "名下" in text else "")
            break
    # 年資：工作年資6年 / 年資6年
    m = re.search(r"(?:工作)?年資\s*(\d+)\s*年", text)
    if m:
        fields["company_years"] = m.group(1)
    # 月薪：月薪10萬 / 月薪50000
    m = re.search(r"月薪\s*(\d+(?:\.\d+)?)\s*萬", text)
    if m:
        fields["company_salary"] = str(int(float(m.group(1)) * 10000))
    else:
        m = re.search(r"月薪\s*(\d{4,})", text)
        if m:
            fields["company_salary"] = m.group(1)
    # 送件金額/期數：14萬/30期 或 10萬 30期（注意：送件金額 ≠ 核准金額）
    m = re.search(r"(\d+(?:\.\d+)?)\s*萬\s*/?\s*(\d+)\s*期", text)
    if m:
        fields["notify_amount"] = m.group(1)
        fields["notify_period"] = m.group(2)
    # 學歷：大學畢 / 高中 / 專科
    for edu, val in [("大學", "專科/大學"), ("專科", "專科/大學"), ("高中", "高中/職"),
                     ("高職", "高中/職"), ("研究所", "研究所以上"), ("碩士", "研究所以上")]:
        if edu in text:
            fields["education"] = val
            break
    # 資金用途：資金用途：家用
    m = re.search(r"資金用途\s*[：:]\s*(\S+)", text)
    if m:
        fields["fund_use"] = m.group(1)
    return fields


def handle_notification_briefing(block_text: str, source_group_id: str, reply_token: str) -> Optional[str]:
    """處理照會注意事項 — 等同已送件，提取欄位資料更新客戶"""
    lines = block_text.strip().splitlines()
    if not lines:
        return None
    # 第一行是客戶姓名
    name = lines[0].strip()
    if not name:
        return None
    rows = find_active_by_name(name)
    same = [r for r in rows if r["source_group_id"] == source_group_id]
    target = same[0] if same else (rows[0] if rows else None)
    if not target:
        return f"⚠️ 照會注意事項：找不到客戶「{name}」"

    # 提取欄位
    fields = parse_notification_fields(block_text)

    # 更新客戶：標記已送件 + 寫入欄位
    # 照會是送件前的動作，金額 → notify_amount（送件金額，非核准金額）
    # target 是 sqlite3.Row，用 [] 訪問
    sec = target["current_company"] or target["company"] or "送件"
    update_fields = {
        "text": block_text,
        "from_group_id": source_group_id,
        "report_section": sec,
    }
    if fields.get("notify_amount"):
        update_fields["notify_amount"] = fields["notify_amount"]
        if fields.get("notify_period"):
            update_fields["notify_period"] = fields["notify_period"]

    update_customer(target["case_id"], **update_fields)

    # 把照會中解析出的欄位寫入 DB（直接 SQL 更新非核心欄位）
    db_updates = {}
    if fields.get("live_years"):
        db_updates["live_years"] = fields["live_years"]
    if fields.get("live_status"):
        db_updates["live_status"] = fields["live_status"]
    if fields.get("company_years"):
        db_updates["company_years"] = fields["company_years"]
    if fields.get("company_salary"):
        db_updates["company_salary"] = fields["company_salary"]
    if fields.get("education"):
        db_updates["education"] = fields["education"]
    if db_updates:
        with db_conn(commit=True) as conn:
            cur = conn.cursor()
            set_clause = ", ".join(f"{k}=?" for k in db_updates)
            vals = list(db_updates.values()) + [target["case_id"]]
            cur.execute(f"UPDATE customers SET {set_clause} WHERE case_id=?", vals)

    # 回覆
    parsed_info = []
    if fields.get("live_years"):
        parsed_info.append(f"居住{fields['live_years']}年")
    if fields.get("company_years"):
        parsed_info.append(f"年資{fields['company_years']}年")
    if fields.get("company_salary"):
        parsed_info.append(f"月薪{fields['company_salary']}")
    if fields.get("notify_amount"):
        per = fields.get("notify_period", "")
        parsed_info.append(f"送件金額{fields['notify_amount']}萬/{per}期" if per else f"送件金額{fields['notify_amount']}萬")
    if fields.get("education"):
        parsed_info.append(f"學歷{fields['education']}")

    info_str = "、".join(parsed_info) if parsed_info else ""
    msg = f"📋 已收到照會：{name}"
    if info_str:
        msg += f"\n📝 已記錄：{info_str}"
    return msg


# ⭐ 對保 4 子步驟偵測
def detect_pairing_substep(text: str) -> str:
    """偵測對保子步驟，回傳子狀態名稱（空字串=不是對保訊息）

    回傳值：
    - "派對保" - 8a 行政發派對保（含「派對保」「辦理方案」+「對保地區」）
    - "委對收" - 8b 對保員接單回覆「委對收」
    - "約時間" - 8c 對保員回時間+地點
    - "對好"   - 8d 對保完成回報（「對好」「對保完成」「不收不簽」）
    """
    if not text:
        return ""
    if "對好" in text or "對保完成" in text or "不收不簽" in text or "不簽不收" in text:
        return "對好"
    if "委對收" in text:
        return "委對收"
    if "派對保" in text or ("辦理方案" in text and "對保地區" in text):
        return "派對保"
    if ("對保時間" in text or "對保地點" in text or
        ("時間" in text and "地點" in text and len(text) < 100)):
        return "約時間"
    return ""


# =========================
# 基本工具
# =========================
def now_iso() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def short_id() -> str:
    return str(uuid.uuid4())[:8]


def get_conn():
    """取得 DB 連線（向後相容，需手動 close）"""
    pathlib.Path(DB_PATH).parent.mkdir(parents=True, exist_ok=True)
    # Bug 16: timeout 從 10 加到 30 秒，給 BEGIN IMMEDIATE 排隊更多時間
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


from contextlib import contextmanager

@contextmanager
def db_conn(commit: bool = False):
    """DB 連線 context manager（Bug 5/6/16 修復）

    保證：
    - 例外時連線一定會 close（Bug 5 連線洩漏）
    - commit=True 時用 BEGIN IMMEDIATE 立刻拿寫鎖，防 race（Bug 6/16）
    - 例外時自動 rollback
    """
    conn = get_conn()
    try:
        if commit:
            # Bug 16: 切換為手動交易模式，立刻拿寫鎖（避免 read-modify-write race）
            conn.isolation_level = None
            conn.execute("BEGIN IMMEDIATE")
        yield conn
        if commit:
            conn.execute("COMMIT")
    except Exception:
        try:
            if commit:
                conn.execute("ROLLBACK")
        except Exception:
            pass
        raise
    finally:
        try:
            conn.close()
        except Exception:
            pass


# =========================
# Bug 16: 姓名並發鎖
# =========================
import threading as _threading

_name_locks: Dict[str, "_threading.Lock"] = {}
_name_locks_guard = _threading.Lock()


def get_name_lock(name: str):
    """取得姓名鎖（同名訊息排隊處理）

    用法：
        with get_name_lock("王小明"):
            # 同名訊息會排隊
            customer = find_active_by_name(...)
            update_customer(...)

    沒姓名時返回 dummy 鎖（不阻塞任何東西）
    """
    name = (name or "").strip()
    if not name:
        return _DummyLock()
    with _name_locks_guard:
        lock = _name_locks.get(name)
        if lock is None:
            lock = _threading.Lock()
            _name_locks[name] = lock
        return lock


class _DummyLock:
    """空鎖，給沒姓名的訊息用"""
    def __enter__(self):
        return self
    def __exit__(self, *args):
        pass
    def acquire(self, *args, **kwargs):
        return True
    def release(self):
        pass


def normalize_id_no(id_no) -> str:
    """正規化身分證字號（Bug 12 修復）

    處理：
    - None → ""
    - 大小寫不一致 → 全大寫
    - 前後空白 → 去除
    """
    return (id_no or "").strip().upper()


def extract_first_line(text: str) -> str:
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    return lines[0] if lines else ""





def normalize_ai_text(text: str) -> str:
    """統一化文字：全形轉半形、去空白"""
    text = (text or "")
    # 全形＠→@、全形空格→半形
    text = text.replace("＠", "@").replace("　", " ")
    # 全形英文 Ａ-Ｚ ａ-ｚ → 半形
    result = ""
    for c in text:
        code = ord(c)
        if 0xFF01 <= code <= 0xFF5E:
            result += chr(code - 0xFEE0)
        else:
            result += c
    return result


def has_ai_trigger(text: str) -> bool:
    """支援全形半形大小寫，如 @AI / ＠AI / @ai / @ A I / ＠ＡＩ"""
    normalized = re.sub(r"\s+", "", normalize_ai_text(text))
    return bool(re.search(r"@ai|#ai", normalized, re.IGNORECASE))


def strip_ai_trigger(text: str) -> str:
    """去掉所有 @AI 觸發詞 + LINE @標記其他人"""
    text = normalize_ai_text(text)
    # 先去掉 @AI
    text = re.sub(r"(@ai助理|#ai助理|@ai|#ai)", "", text, flags=re.IGNORECASE)
    # 再去掉其他 LINE @標記（@人名，但不是 @ai）
    text = re.sub(r"@[^\s@]{1,20}", "", text)
    return text.strip()


def is_blocked(text: str) -> bool:
    return any(w in (text or "") for w in BLOCK_KEYWORDS)


def is_closed_text(text: str) -> bool:
    return any(w in (text or "") for w in DELETE_KEYWORDS)


def contains_status_word(text: str) -> bool:
    return any(w in (text or "") for w in STATUS_WORDS)


def has_business_action_word(text: str) -> bool:
    return any(w in (text or "") for w in ACTION_KEYWORDS)


def extract_id_no(text: str) -> str:
    m = ID_RE.search((text or "").upper())
    return m.group(0) if m else ""


def extract_company(text: str) -> str:
    if not text:
        return ""
    # 先把身分證/居留證號碼移除（避免 U121558670 的 21 被抓成公司）
    text = re.sub(r"[A-Z][A-Z0-9]\d{8}", "", text, flags=re.IGNORECASE)
    # 只看狀態關鍵字之前的部分（狀態後面都是備註原因，不抓）
    search_text = text
    for kw in ["核准", "核準", "婉拒", "申覆失敗", "補件", "待核准", "照會", "撥款"]:
        idx = search_text.find(kw)
        if idx > 0:
            search_text = search_text[:idx]
            break
    # 只看第一行（多行訊息後面通常是備註）
    search_text = search_text.splitlines()[0] if search_text else ""
    # 找出所有匹配及其位置，取最前面出現的
    matches = []
    for c in COMPANY_LIST:
        idx = search_text.find(c)
        if idx >= 0:
            matches.append((idx, c, c))
    for alias, real in COMPANY_ALIAS.items():
        idx = search_text.find(alias)
        if idx >= 0:
            matches.append((idx, alias, real))
    if not matches:
        return ""
    # 按位置排序，同位置時長度長的優先（亞太商品 > 亞太）
    matches.sort(key=lambda x: (x[0], -len(x[1])))
    return matches[0][2]


def get_group_name(group_id: str) -> str:
    if not group_id:
        return "未知群組"
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT group_name FROM groups WHERE group_id=?", (group_id,))
    row = cur.fetchone()
    conn.close()
    return row["group_name"] if row else "未知群組"


def get_all_group_names() -> Dict[str, str]:
    """一次載入所有群組名稱，避免 N+1 查詢"""
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT group_id, group_name FROM groups")
    result = {r["group_id"]: r["group_name"] for r in cur.fetchall()}
    conn.close()
    return result


def get_sales_group_ids() -> List[str]:
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT group_id FROM groups WHERE group_type='SALES_GROUP' AND is_active=1")
    rows = cur.fetchall()
    conn.close()
    return [r["group_id"] for r in rows]


def get_admin_group_ids() -> List[str]:
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT group_id FROM groups WHERE group_type='ADMIN_GROUP' AND is_active=1")
    rows = cur.fetchall()
    conn.close()
    return [r["group_id"] for r in rows]


# =========================
# 送件順序工具
# =========================
def parse_route_order_line(line: str) -> Dict:
    """解析「4/1-高郡惠-喬美/亞太/和裕」，回傳 dict 或 {}"""
    # Bug 1: 攔截「轉XXX」格式
    if re.match(r"^\s*\d{1,4}/\d{1,2}(?:/\d{1,2})?\s*[-－]\s*[\u4e00-\u9fff]{2,6}\s*[-－]\s*轉", line.strip()):
        return {}
    m = ROUTE_ORDER_RE.match(line.strip())
    if not m:
        return {}
    companies_str = m.group(3).strip()
    raw_companies = [COMPANY_ALIAS.get(c.strip(), c.strip()) for c in companies_str.split("/") if c.strip()]
    if not raw_companies:
        return {}
    # 去重（防手滑打兩次同一家），記錄重複數量
    seen = set()
    companies = []
    for c in raw_companies:
        if c not in seen:
            seen.add(c)
            companies.append(c)
    dupe_count = len(raw_companies) - len(companies)
    return {"date": m.group(1), "name": m.group(2), "companies": companies, "dupe_count": dupe_count}


def is_route_order_line(line: str) -> bool:
    return bool(parse_route_order_line(line))


# 送件金額/期數末尾偵測（送件金額 ≠ 核准金額）
# 僅在 @AI 觸發的訊息末尾抓，避免誤判日期（4/17）
_NOTIFY_TAIL_RE = re.compile(
    r"[-\s]\s*(\d+(?:\.\d+)?)\s*萬?\s*[/／]\s*(\d+)\s*期?\s*@AI\s*$",
    re.IGNORECASE
)


def extract_notify_amount_period(text: str):
    """從 @AI 訊息末尾抓「金額/期數」送件值，回傳 (amount, period) 或 (None, None)"""
    if not text:
        return None, None
    m = _NOTIFY_TAIL_RE.search(text.strip())
    if m:
        return m.group(1), m.group(2)
    return None, None


def make_route_json(companies: List[str], current_index: int = 0, history: List = None) -> str:
    return json.dumps({"order": companies, "current_index": current_index, "history": history or []}, ensure_ascii=False)


def parse_route_json(route_plan: str) -> Dict:
    if not route_plan:
        return {"order": [], "current_index": 0, "history": []}
    try:
        data = json.loads(route_plan)
        if isinstance(data, dict) and "order" in data:
            return data
    except Exception:
        pass
    return {"order": [], "current_index": 0, "history": []}


def get_current_company(route_plan: str) -> str:
    data = parse_route_json(route_plan)
    order, idx = data.get("order", []), data.get("current_index", 0)
    return order[idx] if order and 0 <= idx < len(order) else ""


def get_next_company(route_plan: str) -> str:
    data = parse_route_json(route_plan)
    order, idx = data.get("order", []), data.get("current_index", 0)
    next_idx = idx + 1
    return order[next_idx] if next_idx < len(order) else ""


def advance_route(route_plan: str, status: str, amount: str = "", disbursed: str = "") -> str:
    # Bug 7: current_index 邊界檢查，防越界後靜默失敗
    data = parse_route_json(route_plan)
    order = data.get("order", []) or []
    idx = data.get("current_index", 0)
    if not isinstance(idx, int) or idx < 0:
        idx = 0
    history = data.get("history", []) or []
    current = order[idx] if 0 <= idx < len(order) else ""
    if current:
        entry = {"company": current, "status": status, "date": now_iso()[:10]}
        if amount:
            entry["amount"] = amount
        if disbursed:
            entry["disbursed"] = disbursed
        history.append(entry)
    # 推進但不超過 order 長度
    data["current_index"] = min(idx + 1, len(order))
    data["history"] = history
    return json.dumps(data, ensure_ascii=False)


def get_all_approved(route_plan: str) -> list:
    """從 history 取得所有核准記錄"""
    data = parse_route_json(route_plan)
    history = data.get("history", [])
    return [h for h in history if h.get("status") in ("核准", "待撥款", "撥款") and h.get("amount")]

def get_all_disbursed(route_plan: str) -> list:
    """從 history 取得所有撥款記錄"""
    data = parse_route_json(route_plan)
    history = data.get("history", [])
    return [h for h in history if h.get("disbursed") or h.get("disbursement_date")]

def get_total_approved_amount(route_plan: str) -> str:
    """計算所有核准金額總計"""
    approved = get_all_approved(route_plan)
    if not approved:
        return ""
    if len(approved) == 1:
        return approved[0].get("amount", "")
    # 多家核准，顯示各家
    parts = []
    for h in approved:
        co = h.get("company", "")
        amt = h.get("amount", "")
        if amt:
            parts.append(f"{co}{amt}")
    return " + ".join(parts)

def get_amount_from_history(route_plan: str, company: str) -> str:
    """從 history 找指定公司的核准金額（支援模糊比對）"""
    data = parse_route_json(route_plan)
    history = data.get("history", [])
    # 先精確比對
    for h in reversed(history):
        if h.get("company", "") == company and h.get("amount"):
            return h["amount"]
    # 再模糊比對（和裕 能匹配 和裕商品/和裕機車，亞太 能匹配 亞太商品/亞太機車）
    for h in reversed(history):
        hc = h.get("company", "")
        if h.get("amount") and (company in hc or hc in company):
            return h["amount"]
    return ""


def update_company_amount_in_history(route_plan: str, company: str, amount: str) -> str:
    """在 history 裡更新或新增指定公司的金額記錄（支援模糊比對）"""
    data = parse_route_json(route_plan)
    history = data.get("history", [])
    # 精確比對
    for h in reversed(history):
        if h.get("company", "") == company:
            h["amount"] = amount
            data["history"] = history
            return json.dumps(data, ensure_ascii=False)
    # 模糊比對
    for h in reversed(history):
        hc = h.get("company", "")
        if company in hc or hc in company:
            h["amount"] = amount
            data["history"] = history
            return json.dumps(data, ensure_ascii=False)
    # 找不到就新增一筆
    history.append({"company": company, "status": "核准", "amount": amount, "date": now_iso()[:10]})
    data["history"] = history
    return json.dumps(data, ensure_ascii=False)


def set_disbursed_in_history(route_plan: str, company: str, disb_date: str) -> str:
    """在 history 裡設定指定公司的撥款日期"""
    data = parse_route_json(route_plan)
    history = data.get("history", [])
    found = False
    for h in reversed(history):
        hc = h.get("company", "")
        if hc == company or company in hc or hc in company:
            h["disbursed"] = disb_date
            found = True
            break
    if not found:
        history.append({"company": company, "status": "撥款", "date": disb_date, "disbursed": disb_date})
    data["history"] = history
    return json.dumps(data, ensure_ascii=False)


def advance_route_to(route_plan: str, target: str, status: str):
    """轉到指定公司，回傳 (新json, ok, err)"""
    data = parse_route_json(route_plan)
    order, idx, history = data.get("order", []), data.get("current_index", 0), data.get("history", [])
    target_idx = next((i for i, c in enumerate(order) if target in c or c in target), None)
    if target_idx is None:
        return route_plan, False, f"找不到 {target} 在送件順序中"
    for i in range(idx, target_idx):
        history.append({"company": order[i], "status": status, "date": now_iso()[:10]})
    data["current_index"] = target_idx
    data["history"] = history
    return json.dumps(data, ensure_ascii=False), True, ""


# =========================
# 名稱解析
# =========================
def parse_header_fields(text: str) -> Dict[str, str]:
    text = normalize_ai_text(text)  # 全形轉半形（支援全形數字和斜線）
    lines = [x.strip() for x in (text or "").splitlines() if x.strip()]
    first = lines[0] if lines else ""
    compact = re.sub(r"\s+", "", text or "")
    result = {"date": "", "name": "", "id_no": ""}

    m = DATE_NAME_ID_INLINE_RE.search(first)
    if m:
        result["date"], result["name"], result["id_no"] = m.group(1), m.group(2), m.group(3).upper()
        return result
    m = DATE_NAME_ID_INLINE_RE.search(compact)
    if m:
        result["date"], result["name"], result["id_no"] = m.group(1), m.group(2), m.group(3).upper()
        return result
    m = DATE_NAME_ONLY_RE.search(first)
    if m:
        result["date"], result["name"] = m.group(1), m.group(2)
        if len(lines) >= 2:
            id_m = ID_RE.search(lines[1].upper())
            if id_m:
                result["id_no"] = id_m.group(0)
        return result
    m = SHORT_DATE_NAME_RE.search(first)
    if m:
        result["date"], result["name"] = m.group(1), m.group(2)
        if len(lines) >= 2:
            id_m = ID_RE.search(lines[1].upper())
            if id_m:
                result["id_no"] = id_m.group(0)
        return result

    result["id_no"] = extract_id_no(text)
    first_line = extract_first_line(text)
    first_line = re.split(r"[:：]|->", first_line, maxsplit=1)[0].strip()
    first_line = re.split(
        r"結案|補件|補資料|婉拒|核准|照會|退件|等保書|缺資料|待撥款|保密|無可知情|聯絡人皆可知情",
        first_line, maxsplit=1,
    )[0].strip()
    if "｜" in first_line:
        first_line = first_line.split("｜", 1)[0].strip()
    m_name = CHINESE_NAME_RE.search(first_line)
    if m_name:
        result["name"] = m_name.group(0)
    return result


def extract_name(text: str) -> str:
    name = parse_header_fields(text).get("name", "")
    if not name or name in IGNORE_NAME_SET:
        return ""
    # 避免公司名被誤當客戶姓名（如單獨一行「喬美」「麻吉10」「裕融100萬/72期」）
    # COMPANY_LIST + COMPANY_ALIAS keys + COMPANY_ALIAS values 都視為公司簡稱
    all_companies = set(COMPANY_LIST) | set(COMPANY_ALIAS.keys()) | set(COMPANY_ALIAS.values())
    if name in all_companies:
        return ""
    return name


def looks_like_new_case_block(block: str) -> bool:
    f = parse_header_fields(block)
    return bool(f.get("date") and f.get("name") and f.get("id_no"))


def is_format_trigger(block: str) -> bool:
    first = extract_first_line(block)
    if "｜" in first and len([p for p in first.split("｜") if p.strip()]) >= 2:
        return True
    return "->" in first


def looks_like_case_start(line: str) -> bool:
    line = line.strip()
    if not line or re.match(r"^[\[【(（]\d+[\]】)）]", line):
        return False
    if line.startswith("@") or line.startswith("#") or line in {"助理", "AI助理"}:
        return False
    compact = re.sub(r"\s+", "", line)
    if DATE_NAME_ID_INLINE_RE.search(compact) or DATE_NAME_ONLY_RE.search(line):
        return True
    if SHORT_DATE_NAME_RE.search(line) or is_route_order_line(line) or is_format_trigger(line):
        return True
    return False


def split_multi_cases(text: str) -> List[str]:
    text = (text or "").strip()
    if not text:
        return []
    text = re.sub(r"\n\s*/\s*\n", "\n<<<SPLIT>>>\n", text)
    text = re.sub(r"\n\s*\n+", "\n<<<SPLIT>>>\n", text)
    raw_parts = [p.strip() for p in text.split("<<<SPLIT>>>") if p.strip()]
    final_blocks: List[str] = []
    for part in raw_parts:
        lines = [l.rstrip() for l in part.splitlines() if l.strip()]
        if not lines:
            continue
        current: List[str] = []
        for line in lines:
            if looks_like_case_start(line) and current:
                final_blocks.append("\n".join(current).strip())
                current = []
            current.append(line.strip())
        if current:
            final_blocks.append("\n".join(current).strip())
    return [b for b in final_blocks if b.strip()]


# =========================
# LINE API
# =========================
def push_text(to_group_id: str, text: str):
    """推送訊息。超過 4900 字會自動分段推送，每段都加「(N/M)」標記"""
    if not CHANNEL_ACCESS_TOKEN:
        return False, "未設定 CHANNEL_ACCESS_TOKEN"
    text = text or ""
    # 超長時分段（每段 4900 字）
    if len(text) > 4900:
        segments = []
        # 優先在換行分段，避免切半
        remaining = text
        while len(remaining) > 4900:
            cut = remaining.rfind("\n", 0, 4900)
            if cut < 3500:  # 沒找到合適換行，強切
                cut = 4900
            segments.append(remaining[:cut])
            remaining = remaining[cut:].lstrip("\n")
        if remaining:
            segments.append(remaining)
        total = len(segments)
        last_ok, last_err = True, ""
        for i, seg in enumerate(segments, 1):
            tag = f"({i}/{total}) " if total > 1 else ""
            try:
                resp = req_lib.post(
                    "https://api.line.me/v2/bot/message/push",
                    headers={"Authorization": f"Bearer {CHANNEL_ACCESS_TOKEN}", "Content-Type": "application/json"},
                    json={"to": to_group_id, "messages": [{"type": "text", "text": (tag + seg)[:4900]}]},
                    timeout=10,
                )
                if resp.status_code != 200:
                    last_ok = False
                    last_err = f"HTTP {resp.status_code}"
            except Exception as e:
                last_ok = False
                last_err = str(e)
        return last_ok, last_err
    try:
        resp = req_lib.post(
            "https://api.line.me/v2/bot/message/push",
            headers={"Authorization": f"Bearer {CHANNEL_ACCESS_TOKEN}", "Content-Type": "application/json"},
            json={"to": to_group_id, "messages": [{"type": "text", "text": text[:4900]}]},
            timeout=10,
        )
        return (200 <= resp.status_code < 300), "" if 200 <= resp.status_code < 300 else f"LINE push失敗({resp.status_code})"
    except Exception as e:
        return False, str(e)


def reply_text(reply_token: str, text: str):
    if not CHANNEL_ACCESS_TOKEN or reply_token == "TEST":
        return
    try:
        req_lib.post(
            "https://api.line.me/v2/bot/message/reply",
            headers={"Authorization": f"Bearer {CHANNEL_ACCESS_TOKEN}", "Content-Type": "application/json"},
            json={"replyToken": reply_token, "messages": [{"type": "text", "text": text[:4900]}]},
            timeout=10,
        )
    except Exception:
        pass


def make_quick_reply_item(label: str, text: str):
    return {"type": "action", "action": {"type": "message", "label": label[:20], "text": text}}


def push_text_with_buttons(to_group_id: str, text: str, items):
    """推送訊息 + Quick Reply 按鈕（用於主動通知需要復原等操作的群組）"""
    if not CHANNEL_ACCESS_TOKEN:
        return False, "未設定 CHANNEL_ACCESS_TOKEN"
    if len(items) > 13:
        items = items[:13]
    try:
        req_lib.post(
            "https://api.line.me/v2/bot/message/push",
            headers={"Authorization": f"Bearer {CHANNEL_ACCESS_TOKEN}", "Content-Type": "application/json"},
            json={"to": to_group_id, "messages": [{"type": "text", "text": (text or "")[:4900], "quickReply": {"items": items}}]},
            timeout=10,
        )
        return True, ""
    except Exception as e:
        return False, str(e)


def reply_quick_reply(reply_token: str, text: str, items):
    if not CHANNEL_ACCESS_TOKEN or reply_token == "TEST":
        return
    # LINE Quick Reply 上限 13 個按鈕，超過警告 + 截斷
    if len(items) > 13:
        print(f"[reply_quick_reply] 警告：按鈕數 {len(items)} 超過 13，已截斷")
        items = items[:13]
    try:
        req_lib.post(
            "https://api.line.me/v2/bot/message/reply",
            headers={"Authorization": f"Bearer {CHANNEL_ACCESS_TOKEN}", "Content-Type": "application/json"},
            json={"replyToken": reply_token, "messages": [{"type": "text", "text": text[:4900], "quickReply": {"items": items}}]},
            timeout=10,
        )
    except Exception:
        pass


# =========================
# 資料庫
# =========================
def ensure_column(cur, table: str, column: str, definition: str):
    """加欄位（DB migration 用）

    Bug 7: 加 whitelist 防 SQL injection。
    table/column 必須是合法識別字（字母+底線+數字，不能以數字開頭）
    definition 必須是 SQLite 合法型別字串
    """
    if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", table):
        raise ValueError(f"非法的 table 名稱：{table!r}")
    if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", column):
        raise ValueError(f"非法的 column 名稱：{column!r}")
    # definition 嚴格白名單：擋 SQL injection。要新增型別請加到這個 set 裡
    _ALLOWED_DEFS = {
        # TEXT
        "TEXT", "TEXT NOT NULL", "TEXT DEFAULT ''", "TEXT DEFAULT '' NOT NULL",
        # INTEGER
        "INTEGER", "INTEGER NOT NULL", "INTEGER DEFAULT 0", "INTEGER DEFAULT 0 NOT NULL",
        "INTEGER DEFAULT 1", "INTEGER DEFAULT 1 NOT NULL",
        # REAL
        "REAL", "REAL NOT NULL", "REAL DEFAULT 0", "REAL DEFAULT 0 NOT NULL",
        # NUMERIC / BLOB / TIMESTAMP
        "NUMERIC", "BLOB",
        "TIMESTAMP", "DATETIME", "DATE",
    }
    if definition not in _ALLOWED_DEFS:
        raise ValueError(
            f"非法的 definition：{definition!r}\n"
            f"請使用以下其中一種，或在 ensure_column 的 _ALLOWED_DEFS 新增：\n"
            f"{sorted(_ALLOWED_DEFS)}"
        )
    cur.execute(f"PRAGMA table_info({table})")
    cols = [row[1] for row in cur.fetchall()]
    if column not in cols:
        cur.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")


def init_db():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""CREATE TABLE IF NOT EXISTS groups (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        group_id TEXT UNIQUE NOT NULL, group_name TEXT,
        group_type TEXT NOT NULL, is_active INTEGER DEFAULT 1, created_at TEXT NOT NULL
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS customers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        case_id TEXT UNIQUE NOT NULL, customer_name TEXT NOT NULL,
        id_no TEXT, source_group_id TEXT NOT NULL,
        company TEXT, route_plan TEXT, current_company TEXT, report_section TEXT,
        last_update TEXT, status TEXT NOT NULL DEFAULT 'ACTIVE',
        created_at TEXT NOT NULL, updated_at TEXT NOT NULL
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS case_logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        case_id TEXT NOT NULL, customer_name TEXT NOT NULL,
        id_no TEXT, company TEXT, message_text TEXT NOT NULL,
        from_group_id TEXT NOT NULL, created_at TEXT NOT NULL
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS pending_actions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        action_id TEXT UNIQUE NOT NULL, action_type TEXT NOT NULL,
        payload TEXT NOT NULL, created_at TEXT NOT NULL
    )""")
    # 自動補欄位，不破壞舊資料
    for col, defn in [
        ("route_plan", "TEXT"),
        ("current_company", "TEXT"),
        ("report_section", "TEXT"),
        ("approved_amount", "TEXT"),       # 核准金額
        ("disbursement_date", "TEXT"),     # 撥款日期
        ("birth_date", "TEXT"),
        ("phone", "TEXT"),
        ("email", "TEXT"),
        ("line_id", "TEXT"),
        ("marriage", "TEXT"),
        ("education", "TEXT"),
        ("id_issue_date", "TEXT"),
        ("id_issue_place", "TEXT"),
        ("id_issue_type", "TEXT"),
        ("reg_city", "TEXT"),
        ("reg_district", "TEXT"),
        ("reg_address", "TEXT"),
        ("reg_phone", "TEXT"),
        ("live_city", "TEXT"),
        ("live_district", "TEXT"),
        ("live_address", "TEXT"),
        ("live_phone", "TEXT"),
        ("live_same_as_reg", "TEXT"),
        ("live_status", "TEXT"),
        ("live_years", "TEXT"),
        ("live_months", "TEXT"),
        ("company_name_detail", "TEXT"),
        ("company_industry", "TEXT"),
        ("company_role", "TEXT"),
        ("company_phone_area", "TEXT"),
        ("company_phone_num", "TEXT"),
        ("company_phone_ext", "TEXT"),
        ("company_years", "TEXT"),
        ("company_salary", "TEXT"),
        ("company_city", "TEXT"),
        ("company_district", "TEXT"),
        ("company_address", "TEXT"),
        ("bank_name", "TEXT"),
        ("bank_branch", "TEXT"),
        ("bank_account", "TEXT"),
        ("bank_holder", "TEXT"),
        ("contact1_name", "TEXT"),
        ("contact1_relation", "TEXT"),
        ("contact1_phone", "TEXT"),
        ("contact1_known", "TEXT"),
        ("contact2_name", "TEXT"),
        ("contact2_relation", "TEXT"),
        ("contact2_phone", "TEXT"),
        ("contact2_known", "TEXT"),
        ("eval_fund_need", "TEXT"),
        ("eval_sent_3m", "TEXT"),
        ("eval_labor_ins", "TEXT"),
        ("eval_salary_transfer", "TEXT"),
        ("eval_alert", "TEXT"),
        ("eval_vehicle", "TEXT"),
        ("eval_property", "TEXT"),
        ("eval_credit_card", "TEXT"),
        ("eval_fine", "TEXT"),
        ("eval_fuel_tax", "TEXT"),
        ("eval_late", "TEXT"),
        ("eval_late_days", "TEXT"),
        ("debt_list", "TEXT"),
        ("eval_note", "TEXT"),
        ("signature_img", "TEXT"),
        ("selected_plans", "TEXT"),
        ("product_type", "TEXT"),
        ("product_model", "TEXT"),
        ("product_imei", "TEXT"),
        ("vehicle_plate", "TEXT"),
        ("vehicle_type", "TEXT"),
        ("vehicle_owner", "TEXT"),
        ("sales_name", "TEXT"),
        ("created_by_role", "TEXT"),
        ("adminb_selected_plans", "TEXT"),
        ("adminb_contact_time", "TEXT"),
        ("adminb_bank", "TEXT"),
        ("adminb_branch", "TEXT"),
        ("adminb_account", "TEXT"),
        ("adminb_industry", "TEXT"),
        ("adminb_brand", "TEXT"),
        ("adminb_hr_role", "TEXT"),
        ("adminb_hr_industry", "TEXT"),
        ("signature_applicant", "TEXT"),
        ("signature_legal_rep", "TEXT"),
        ("adminb_role", "TEXT"),
        ("adminb_vehicle_type", "TEXT"),
        ("adminb_engine_no", "TEXT"),
        ("adminb_displacement", "TEXT"),
        ("adminb_color", "TEXT"),
        ("adminb_body_no", "TEXT"),
        ("adminb_mfg_date", "TEXT"),
        ("vehicle_duration", "TEXT"),
        ("vehicle_road_reg", "TEXT"),
        ("vehicle_space", "TEXT"),
        ("adminb_fund_use", "TEXT"),
        ("adminb_product", "TEXT"),
        ("adminb_model", "TEXT"),
        ("adminb_product_name", "TEXT"),
        ("adminb_product_model", "TEXT"),
        ("adminb_21car_project", "TEXT"),
        ("adminb_21car_price", "TEXT"),
        ("adminb_21car_ref_src", "TEXT"),
        ("adminb_21car_ref_price", "TEXT"),
        ("adminb_21car_rate", "TEXT"),
        ("adminb_21car_amount", "TEXT"),
        ("adminb_21car_period", "TEXT"),
        ("adminb_21car_monthly", "TEXT"),
        ("adminb_21car_fund", "TEXT"),
        ("adminb_21car_hascc", "TEXT"),
        ("adminb_mj_brand", "TEXT"),
        ("adminb_mj_model", "TEXT"),
        ("adminb_ry_amount", "TEXT"),
        ("adminb_ry_period", "TEXT"),
        ("notify_amount", "TEXT"),
        ("notify_period", "TEXT"),
        ("adminb_credit_bank", "TEXT"),
        ("adminb_credit_no", "TEXT"),
        ("adminb_credit_exp", "TEXT"),
        ("adminb_credit_limit", "TEXT"),
        ("adminb_credit_late", "TEXT"),
        ("adminb_credit_pay", "TEXT"),
        ("eval_license", "TEXT"),
        ("eval_law", "TEXT"),
        ("carrier", "TEXT"),
        ("fb", "TEXT"),
        ("company_months", "TEXT"),
        ("concurrent_companies", "TEXT"),
        ("company_status", "TEXT"),
        ("signing_area", "TEXT"),
        ("signing_salesperson", "TEXT"),
        ("signing_company", "TEXT"),
        ("signing_time", "TEXT"),
        ("signing_location", "TEXT"),
    ]:
        ensure_column(cur, "customers", col, defn)
    # groups 表新增業務群對應欄位
    ensure_column(cur, "groups", "linked_sales_group_id", "TEXT")
    ensure_column(cur, "groups", "password_hash", "TEXT")
    # case_logs 加快照欄位（用於 @AI 姓名 還原）
    ensure_column(cur, "case_logs", "snapshot_json", "TEXT")
    # settings 表
    cur.execute("""CREATE TABLE IF NOT EXISTS settings (
        key TEXT PRIMARY KEY NOT NULL,
        value TEXT NOT NULL,
        updated_at TEXT NOT NULL
    )""")
    # login_attempts 表
    cur.execute("""CREATE TABLE IF NOT EXISTS login_attempts (
        identifier TEXT PRIMARY KEY NOT NULL,
        attempts INTEGER DEFAULT 0,
        locked_until TEXT DEFAULT '',
        updated_at TEXT NOT NULL
    )""")
    # sessions 表（安全 session token）
    cur.execute("""CREATE TABLE IF NOT EXISTS sessions (
        token TEXT PRIMARY KEY NOT NULL,
        role TEXT NOT NULL,
        group_id TEXT DEFAULT '',
        created_at TEXT NOT NULL,
        expires_at TEXT NOT NULL
    )""")
    # form_tokens 表（客戶自助填單一次性連結）
    cur.execute("""CREATE TABLE IF NOT EXISTS form_tokens (
        token TEXT PRIMARY KEY NOT NULL,
        group_id TEXT NOT NULL,
        note TEXT DEFAULT '',
        created_at TEXT NOT NULL,
        expires_at TEXT NOT NULL,
        used_at TEXT DEFAULT '',
        case_id TEXT DEFAULT '',
        revoked_at TEXT DEFAULT ''
    )""")
    # 索引
    cur.execute("CREATE INDEX IF NOT EXISTS idx_cust_id_no ON customers(id_no)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_cust_name ON customers(customer_name)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_cust_group ON customers(source_group_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_cust_status ON customers(status)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_logs_case ON case_logs(case_id)")
    conn.commit()
    conn.close()
    init_settings()



def seed_groups():
    conn = get_conn()
    cur = conn.cursor()
    now = now_iso()
    for row in [
        (A_GROUP_ID, "A群", "A_GROUP", 1, now),
        (B_GROUP_ID, "B群", "SALES_GROUP", 1, now),
        (C_GROUP_ID, "C群", "SALES_GROUP", 1, now),
    ]:
        cur.execute("INSERT OR IGNORE INTO groups (group_id,group_name,group_type,is_active,created_at) VALUES (?,?,?,?,?)", row)
    conn.commit()
    conn.close()


def init_settings():
    """初始化密碼（Bug 3 修復）

    優先順序：
    1. 已存在密碼 → 不動
    2. 環境變數有設 → 用環境變數的密碼
    3. 都沒有 → 產生隨機密碼並印出 log（不再用硬編碼預設值）
    """
    import secrets as _secrets
    env_var_map = {
        "admin_pw": "ADMIN_PASSWORD",
        "adminB_pw": "ADMINB_PASSWORD",
        "report_pw": "REPORT_PASSWORD",
        "vba_secret": "VBA_SECRET",
    }
    conn = get_conn(); cur = conn.cursor()
    for key, env_var in env_var_map.items():
        cur.execute("SELECT value FROM settings WHERE key=?", (key,))
        if cur.fetchone():
            continue  # 已有密碼，跳過
        env_pw = os.getenv(env_var, "")
        if env_pw:
            pw_to_set = env_pw
            try:
                print(f"[init_settings] {key}: loaded from env {env_var}")
            except Exception:
                pass
        else:
            pw_to_set = _secrets.token_urlsafe(16)
            try:
                print(f"[init_settings] WARNING {key} env not set, generated random password: {pw_to_set}")
                print(f"[init_settings] WARNING please save it and set env var {env_var}")
            except Exception:
                pass
        cur.execute("INSERT INTO settings (key,value,updated_at) VALUES (?,?,?)",
            (key, hash_pw(pw_to_set), now_iso()))
    conn.commit(); conn.close()


# =========================
# DB CRUD
# =========================
def create_customer_record(name, id_no, company, source_group_id, text,
                            route_plan="", current_company="", report_section="") -> str:
    """建立客戶（Bug 5/6/12 修復：context manager + transaction + id_no normalize）"""
    id_no = normalize_id_no(id_no)
    now = now_iso()
    with db_conn(commit=True) as conn:
        cur = conn.cursor()
        # 先找有沒有行政A已填的 PENDING 客戶（同身分證號）
        if id_no:
            cur.execute("SELECT * FROM customers WHERE id_no=? AND status='PENDING' ORDER BY updated_at DESC LIMIT 1", (id_no,))
            pending = cur.fetchone()
        else:
            pending = None
        if pending:
            case_id = pending["case_id"]
            cur.execute("""UPDATE customers SET
                status='ACTIVE', customer_name=?, company=?,
                source_group_id=?, route_plan=?, current_company=?,
                report_section=?, last_update=?, updated_at=?
                WHERE case_id=?""",
                (name, company, source_group_id, route_plan, current_company,
                 report_section, text, now, case_id))
            return case_id
        else:
            # 業務員透過 LINE 建立的新案件直接 ACTIVE（PENDING 是 web /new-customer 表單專用狀態）
            # 已測試：改 ACTIVE 後日報正確顯示，結案流程不受影響
            case_id = short_id()
            cur.execute("""INSERT INTO customers
                (case_id,customer_name,id_no,source_group_id,company,route_plan,current_company,report_section,last_update,status,created_at,updated_at)
                VALUES (?,?,?,?,?,?,?,?,?,'ACTIVE',?,?)""",
                (case_id, name, id_no, source_group_id, company, route_plan, current_company, report_section, text, now, now))
            return case_id


def update_customer(case_id, company=None, text=None, from_group_id="", status=None,
                    name=None, source_group_id=None, route_plan=None,
                    current_company=None, report_section=None,
                    approved_amount=None, disbursement_date=None,
                    signing_area=None, signing_salesperson=None,
                    signing_company=None, signing_time=None, signing_location=None,
                    notify_amount=None, notify_period=None, concurrent_companies=None,
                    id_no=None):
    """更新客戶（Bug 5/6 修復：context manager + transaction）

    UPDATE + INSERT case_logs 包在同一交易內，確保原子性。
    """
    now = now_iso()
    fields, values = [], []
    for col, val in [("company", company), ("last_update", text), ("status", status),
                     ("customer_name", name), ("source_group_id", source_group_id),
                     ("route_plan", route_plan), ("current_company", current_company),
                     ("report_section", report_section),
                     ("approved_amount", approved_amount),
                     ("disbursement_date", disbursement_date),
                     ("signing_area", signing_area),
                     ("signing_salesperson", signing_salesperson),
                     ("signing_company", signing_company),
                     ("signing_time", signing_time),
                     ("signing_location", signing_location),
                     ("notify_amount", notify_amount),
                     ("notify_period", notify_period),
                     ("concurrent_companies", concurrent_companies),
                     ("id_no", id_no)]:
        if val is not None:
            fields.append(f"{col} = ?"); values.append(val)
    fields.append("updated_at = ?"); values.append(now); values.append(case_id)
    with db_conn(commit=True) as conn:
        cur = conn.cursor()
        # 先讀 before 快照（供還原用）— 只存關鍵欄位避免龐大
        cur.execute("SELECT * FROM customers WHERE case_id=?", (case_id,))
        before_row = cur.fetchone()
        snapshot = None
        if before_row:
            snap_fields = ["company", "current_company", "concurrent_companies",
                           "route_plan", "report_section", "approved_amount",
                           "notify_amount", "notify_period", "disbursement_date",
                           "status", "id_no", "customer_name", "source_group_id",
                           "signing_area", "signing_salesperson", "signing_company",
                           "signing_time", "signing_location"]
            snapshot = json.dumps({k: before_row[k] for k in snap_fields if k in before_row.keys()},
                                  ensure_ascii=False)
        cur.execute(f"UPDATE customers SET {', '.join(fields)} WHERE case_id=?", values)
        cur.execute("SELECT * FROM customers WHERE case_id=?", (case_id,))
        row = cur.fetchone()
        if row and text is not None:
            cur.execute("INSERT INTO case_logs (case_id,customer_name,id_no,company,message_text,from_group_id,created_at,snapshot_json) VALUES (?,?,?,?,?,?,?,?)",
                (row["case_id"], row["customer_name"], row["id_no"], row["company"], text, from_group_id, now, snapshot))


def find_active_by_id_no(id_no):
    id_no = normalize_id_no(id_no)
    if not id_no: return None
    with db_conn() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM customers WHERE id_no=? AND status='ACTIVE' ORDER BY updated_at DESC LIMIT 1", (id_no,))
        return cur.fetchone()


def find_active_by_id_no_in_group(id_no, group_id):
    """在指定群組內查 ACTIVE；跨群組獨立案時用（避免抓到別群組案件）"""
    id_no = normalize_id_no(id_no)
    if not id_no or not group_id: return None
    with db_conn() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM customers WHERE id_no=? AND source_group_id=? AND status='ACTIVE' ORDER BY updated_at DESC LIMIT 1", (id_no, group_id))
        return cur.fetchone()


def find_all_active_by_id_no(id_no):
    """查所有同身分證的 ACTIVE 案件（可能跨群組多筆，用於 A 群回貼時跳按鈕選擇）"""
    id_no = normalize_id_no(id_no)
    if not id_no: return []
    with db_conn() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM customers WHERE id_no=? AND status='ACTIVE' ORDER BY updated_at DESC", (id_no,))
        return cur.fetchall()


def find_active_by_name(name):
    with db_conn() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM customers WHERE customer_name=? AND status='ACTIVE' ORDER BY updated_at DESC", (name,))
        return cur.fetchall()


def find_any_by_name(name):
    with db_conn() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM customers WHERE customer_name=? ORDER BY updated_at DESC", (name,))
        return cur.fetchall()


def find_any_by_id_no(id_no):
    id_no = normalize_id_no(id_no)
    if not id_no: return []
    with db_conn() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM customers WHERE id_no=? ORDER BY updated_at DESC", (id_no,))
        return cur.fetchall()


def save_pending_action(action_id, action_type, payload):
    """Bug 3: 用 db_conn(commit=True) 進 BEGIN IMMEDIATE 防並發競態 + 連線洩漏"""
    with db_conn(commit=True) as conn:
        cur = conn.cursor()
        cur.execute(
            "INSERT OR REPLACE INTO pending_actions (action_id,action_type,payload,created_at) VALUES (?,?,?,?)",
            (action_id, action_type, json.dumps(payload, ensure_ascii=False), now_iso()))


def get_pending_action_and_delete(action_id):
    """Bug 3: 原子化讀取+刪除，防快速重複點擊造成重複處理"""
    with db_conn(commit=True) as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM pending_actions WHERE action_id=?", (action_id,))
        row = cur.fetchone()
        if not row:
            return None
        cur.execute("DELETE FROM pending_actions WHERE action_id=?", (action_id,))
        data = dict(row)
        try:
            data["payload"] = json.loads(data["payload"])
        except Exception:
            data["payload"] = {}
        return data


def get_pending_action(action_id):
    """保留供唯讀查詢用（不刪除）。注意：消費按鈕請改用 get_pending_action_and_delete()"""
    with db_conn() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM pending_actions WHERE action_id=?", (action_id,))
        row = cur.fetchone()
        if not row:
            return None
        data = dict(row)
        try:
            data["payload"] = json.loads(data["payload"])
        except Exception:
            data["payload"] = {}
        return data


def delete_pending_action(action_id):
    with db_conn(commit=True) as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM pending_actions WHERE action_id=?", (action_id,))


# =========================
# 核准金額 / 撥款名單工具
# =========================
def extract_approved_amount(text: str) -> str:
    """
    從訊息抓取核准金額（改良版）。
    支援格式：
    - 20萬 / 6W / 6.5萬 / (5).(8)萬
    - 降撥4萬 / 本利攤50萬
    - 120,000 / 40000（大數字自動換算）
    - 核准金額20萬 / 最高核貸金額20萬
    注意：「核准30期」的30是期數不是金額，會自動跳過
    """
    # 1. 先找有萬/W字的（最可靠）
    wan_patterns = [
        r"\((\d+)\)\s*\.\s*\((\d+)\)\s*萬",   # (5).(8)萬
        r"降撥\s*(\d+(?:\.\d+)?)\s*萬",             # 降撥4萬
        r"本利攤\s*(\d+(?:\.\d+)?)\s*萬",           # 本利攤50萬
        r"金額[:：]?\s*(\d+(?:\.\d+)?)\s*萬",       # 金額20萬
        r"核[准貸]\s*金額?\s*[:：]?\s*(\d+(?:\.\d+)?)\s*萬",  # 核准金額20萬
        r"(\d+(?:\.\d+)?)\s*[Ww萬]",                 # 20萬 / 6W
    ]
    for pat in wan_patterns:
        m = re.search(pat, text)
        if m:
            if m.lastindex == 2:  # (5).(8)萬格式
                return f"{m.group(1)}.{m.group(2)}萬"
            raw = m.group(1)
            try:
                num = float(raw)
                if num > 0:
                    return f"{int(num)}萬" if num == int(num) else f"{num}萬"
            except Exception:
                pass

    # 2. 找大數字（超過10000視為元，換算成萬）
    for m in re.finditer(r"(\d{1,3}(?:,\d{3})+|\d{5,})", text):
        num_str = m.group(1).replace(",", "")
        try:
            num = int(num_str)
            if num >= 10000:
                wan = num // 10000
                return f"{wan}萬"
        except Exception:
            pass

    # 3. 最後才用「核准+數字」，但排除期數
    m = re.search(r"核准\s*(\d+)", text)
    if m:
        num = int(m.group(1))
        after = text[m.end():][:5]
        if num <= 60 and any(c in after for c in ["期", "，", ",", "/"]):
            pass  # 這是期數，跳過
        elif num > 0:
            return f"{num}萬"

    return ""



async def extract_approved_amount_with_ai(text: str) -> str:
    api_key = os.getenv("ANTHROPIC_API_KEY", "")
    if not api_key:
        return ""
    try:
        resp = req_lib.post(
            "https://api.anthropic.com/v1/messages",
            headers={"x-api-key": api_key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
            json={"model": "claude-haiku-4-5-20251001", "max_tokens": 100, "messages": [{"role": "user", "content": f"""貸款核准訊息如下，請判斷是否為真正核准並抓出金額。
訊息：{text}
只回傳JSON，不要其他文字：{{"is_approved": true或false, "amount": "金額字串或空字串"}}
金額格式如：20萬、6萬、50萬。注意：期數不是金額，月付不是金額，初估/待補/需補/照會=false。"""}]},
            timeout=15,
        )
        if resp.status_code == 200:
            import json as jl
            result = resp.json()["content"][0]["text"].strip().replace("```json","").replace("```","").strip()
            data = jl.loads(result)
            if data.get("is_approved") and data.get("amount"):
                return data["amount"]
    except Exception as e:
        print(f"AI解析失敗: {e}")
    return ""


def parse_disbursement_list(text: str) -> Dict:
    """
    解析撥款名單格式：
    04/01 21 撥款名單
    吳翰杰
    蔡文杰
    ...
    04/01 亞太排撥
    洪莉萍

    回傳 {"04/01": {"21": ["吳翰杰","蔡文杰"], "亞太": ["洪莉萍"]}}
    """
    result = {}
    current_date = ""
    current_company = ""

    # 撥款名單標頭正則（有日期）
    DISB_HEADER_RE = re.compile(
        r"(\d{1,2}/\d{1,2})\s*([\w一-鿿]+?)\s*(撥款名單|預計排撥|排撥|今日撥款|商品撥款|機車撥款|汽車撥款|撥款)",
        re.IGNORECASE
    )
    # 無日期標頭：「貸救補 今日撥款」
    DISB_HEADER_NODATE_RE = re.compile(
        r"^([\w一-鿿]+?)\s*(撥款名單|預計排撥|排撥|今日撥款|商品撥款|機車撥款|汽車撥款|撥款)\s*$",
        re.IGNORECASE
    )

    today_date = datetime.now().strftime("%#m/%#d") if os.name == "nt" else datetime.now().strftime("%-m/%-d")

    lines = text.splitlines()
    for line in lines:
        line = line.strip()
        if not line:
            continue

        m = DISB_HEADER_RE.search(line)
        if m:
            current_date = m.group(1)
            current_company = m.group(2).strip()
            if current_date not in result:
                result[current_date] = {}
            if current_company not in result[current_date]:
                result[current_date][current_company] = []
            continue

        # 無日期標頭（如「貸救補 今日撥款」），用今日日期
        m2 = DISB_HEADER_NODATE_RE.match(line)
        if m2:
            current_date = current_date or today_date
            current_company = m2.group(1).strip()
            if current_date not in result:
                result[current_date] = {}
            if current_company not in result[current_date]:
                result[current_date][current_company] = []
            continue

        # 名字行（2-4個中文字，或含英文的姓名）
        if current_date and current_company:
            name_m = re.match(r"^([一-鿿]{2,6})\s*$", line)
            if name_m:
                result[current_date][current_company].append(name_m.group(1))

    return result


def is_disbursement_list(text: str) -> bool:
    """判斷是否為撥款名單（多行格式，標頭+人名行）"""
    lines = [l.strip() for l in (text or "").splitlines() if l.strip()]
    if len(lines) < 2:
        return False
    # 至少一行要是標頭（含撥款關鍵詞）
    header_re = re.compile(
        r"(撥款名單|排撥|預計排撥|今日撥款|商品撥款|機車撥款|汽車撥款|撥款)"
    )
    has_header = False
    header_idx = -1
    for i, line in enumerate(lines):
        if header_re.search(line):
            has_header = True
            header_idx = i
            break
    if not has_header:
        return False
    # 標頭後面至少要有一行純人名（2-4個中文字）
    name_re = re.compile(r"^[一-鿿]{2,6}$")
    for line in lines[header_idx + 1:]:
        if name_re.match(line):
            return True
    return False


def get_admin_group_for_sales(sales_group_id: str) -> Optional[str]:
    """
    找對應的行政群ID。
    優先找有設定 linked_sales_group_id 對應的，
    找不到就找任何啟用的行政群（所有業務群共用一個行政群的情況）。
    """
    conn = get_conn(); cur = conn.cursor()
    # 先找有明確對應的
    cur.execute("""
        SELECT group_id FROM groups
        WHERE group_type='ADMIN_GROUP' AND linked_sales_group_id=? AND is_active=1
        LIMIT 1
    """, (sales_group_id,))
    row = cur.fetchone()
    if row:
        conn.close()
        return row["group_id"]
    # 找不到就用任何啟用的行政群
    cur.execute("""
        SELECT group_id FROM groups
        WHERE group_type='ADMIN_GROUP' AND is_active=1
        LIMIT 1
    """)
    row = cur.fetchone(); conn.close()
    return row["group_id"] if row else None


def get_all_admin_groups() -> List[str]:
    """取得所有啟用的行政群ID（用於廣播撥款名單）"""
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT group_id FROM groups WHERE group_type='ADMIN_GROUP' AND is_active=1")
    rows = cur.fetchall(); conn.close()
    return [r["group_id"] for r in rows]


# =========================
# 狀態摘要提取
# =========================
_INTERNAL_ACTION_KEYWORDS = [
    "設送件", "核准金額修改為", "轉送", "加送/備註", "加送",
    "改名為",            # 改名動作文字
    "改身分證",          # 改身分證動作文字
    "重啟案件",          # 重啟動作
    "還原到",            # 還原動作
    "不送：",            # A 方案移除同送
    "婉拒（從同送",      # reject_company 移除同送
]


def extract_status_summary(first_line: str, customer_name: str) -> str:
    """
    從訊息第一行提取狀態關鍵字，用於日報顯示。
    標準化標籤：
    - 有核准 → 「核准」
    - 補照會/照會 → 「已補照會」
    - 補申覆/申覆 → 「已補申覆」
    - 補件類 → 「已補資料」
    - NA/未接照會 → 「NA」
    - 其他 → 顯示前20字
    """
    if not first_line:
        return ""
    # 內部動作（@AI 指令產生的記錄文字）不顯示為業務狀態
    if any(kw in first_line for kw in _INTERNAL_ACTION_KEYWORDS):
        return ""
    # 「待核准/待核準」= 還有缺要補（資料/照會都有可能）— 不算核准，先查是否真正核准
    text_wo_pending = first_line.replace("待核准", "").replace("待核準", "")
    if "核准" in text_wo_pending or "核準" in text_wo_pending:
        # 核准後還需接照會 → 「核准待照會」（排除「已接照會」「照會完」= 真正完成）
        has_need_contact = any(w in first_line for w in ["接照會", "待照會", "需照會", "等照會", "再照會"])
        has_done_contact = any(w in first_line for w in ["已接照會", "照會完", "照會過", "接完照會"])
        if has_need_contact and not has_done_contact:
            return "核准待照會"
        return "核准"
    # 補時段（照會時段）：「補時段 11:00~12:00」或「補時段 11點~12點」
    # 業務告知客戶能接到照會電話的時段
    m_notif_time = re.search(r"補時段\s*(\d{1,2}(?:[:：]\d{1,2}|點)?\s*[~～\-至到]\s*\d{1,2}(?:[:：]\d{1,2}|點)?)", first_line)
    if m_notif_time:
        time_range = m_notif_time.group(1).replace("~", "~").replace("-", "~").replace("至", "~").replace("到", "~").replace("：", ":")
        return f"照會時段 {time_range}"

    # 「待核准」下方判斷會處理具體缺項（補照會/補申覆/補件類）；若都沒命中，最後 fallback「待核准」
    # 等保書細分：根據訊息內其他關鍵字決定實際子狀態
    if "等保書" in first_line or "等保人" in first_line:
        # 換保人類：更換/不合格/另找/換一位/換個
        if any(k in first_line for k in ["更換", "換一位", "換個", "換另", "不合格", "不符合", "重找", "另找", "另外找", "再找"]):
            return "換保人"
        # 補保人類：沒保人/還沒/需補/提供保人
        if any(k in first_line for k in ["補保人", "沒保人", "沒有保人", "找保人", "提供保人"]):
            return "補保人"
        # 簽名類
        if any(k in first_line for k in ["簽名", "等簽", "親簽", "待簽"]):
            return "等保人簽"
        # 補資料/文件類
        if any(k in first_line for k in ["保人補", "保書補", "補資料", "補件", "資料不齊"]):
            return "保人補件"
        # 什麼特定關鍵字都沒 → 顯示「等保書」
        return "等保書"
    # 已補照會 / 照會完 / 接完照會 → 已完成（先查，避免命中下面的「補照會」）
    if "已補照會" in first_line or "照會完" in first_line or "接完照會" in first_line:
        return "已補照會"
    if "補照會" in first_line:
        # 如果有時段（時間字或數字:數字）→ 顯示時段
        import re as _re
        tm = _re.search(r"(\d{1,2}[:：點]\d{0,2}(?:[-~至到]\d{1,2}[:：點]?\d{0,2})?)", first_line)
        if tm:
            return f"補照會{tm.group(1)}"
        return "待補照會"
    if "照會" in first_line:
        return "已送件"
    # 補件/申覆判斷：
    # - 明確「已補」/「待補」關鍵字優先
    # - 業務主動語氣（業務說/我補/幫補/補好/補完）→ 已補
    # - 只有「補XX」含糊 → 預設待補（需跳按鈕由業務確認）
    _business_done_markers = ["業務說", "我補", "我已補", "幫補", "主動補", "補好了", "補完了",
                              "補過了", "業務補", "補完成"]
    _explicit_pending_markers = ["待補", "請補", "要補", "缺", "未補", "還沒補", "尚缺",
                                  "請提供", "麻煩補", "需補"]

    # 申覆類
    if "申覆" in first_line:
        # 明確已補：「已補 XX 申覆」或「申覆完/通過/好」
        if "已補" in first_line or "申覆完" in first_line or "申覆通過" in first_line or "申覆好" in first_line or "補完申覆" in first_line or "補好申覆" in first_line:
            return "已補申覆"
        # 明確待補：「待補/請補/要補/缺」等
        if any(m in first_line for m in _explicit_pending_markers):
            return "待補申覆"
        # C 方案：業務主動語氣 → 已補
        if any(m in first_line for m in _business_done_markers):
            return "已補申覆"
        # 含糊「補 X 申覆」→ 預設待補
        return "待補申覆"

    # 補件類
    _bu_list = ["補件", "補資料", "補行照", "補聯徵", "補保人", "補薪轉", "補照片", "補時段",
                "補JCIC", "補jcic", "補在職", "補存摺", "補勞保", "補駕照",
                "缺聯徵", "缺資料", "缺薪轉", "缺JCIC", "缺jcic", "缺保人", "缺在職", "缺存摺"]
    if "已補" in first_line or "補好" in first_line or "補完" in first_line:
        return "已補資料"
    if any(w in first_line for w in _bu_list):
        # 業務主動語氣 → 已補
        if any(m in first_line for m in _business_done_markers):
            return "已補資料"
        return "待補資料"
    if "未接照會" in first_line or first_line.strip().endswith("NA") or " NA" in first_line:
        return "NA"
    # 「待核准」fallback：沒命中具體補件/照會 → 直接顯示「待核准」
    if "待核准" in first_line or "待核準" in first_line:
        return "待核准"

    # 去掉客戶姓名
    text = first_line.replace(customer_name, "").strip()

    # 去掉身分證/居留證（避免身分證末碼被當狀態文字）
    text = re.sub(r"[A-Z][A-Z0-9]\d{8}", "", text, flags=re.IGNORECASE).strip()

    # 去掉日期
    text = re.sub(r"^\d{1,4}/\d{1,2}(/\d{1,2})?[-－]?\s*", "", text).strip()

    # 去掉公司名稱（COMPANY_LIST + COMPANY_ALIAS 別名）
    all_companies = list(COMPANY_LIST) + list(COMPANY_ALIAS.keys())
    for c in sorted(all_companies, key=len, reverse=True):
        text = text.replace(c, "").strip()

    # 去掉【N】之後的內容
    text = re.sub(r"【\d+】.*", "", text).strip()

    # 去掉空括號（公司名被清掉後留下的括號）
    text = re.sub(r"[\(（][\s]*[\)）]", "", text).strip()

    # 去掉多餘標點
    text = re.sub(r"^[-－/\s()（）]+|[-－/\s()（）]+$", "", text).strip()

    # 只取前20字
    return text[:20] if text else ""


# =========================
# 日報產生
# =========================
# 公司子分類歸到日報主欄位
COMPANY_SECTION_MAP = {
    # 和裕所有方案 → 全部歸到「和裕」欄位
    "和裕商品": "和裕", "維力商品貸": "和裕",
    "和裕機車": "和裕", "和裕機": "和裕",
    "維力機車專": "和裕", "維力機車貸": "和裕",
    # 亞太所有方案 → 全部歸到「亞太」欄位
    "亞太商品": "亞太", "亞太工會": "亞太", "亞太工": "亞太",
    "亞太機車": "亞太", "亞太機": "亞太",
    # 21方案 → 全部歸到「21」欄位
    "21機車": "21", "21汽車": "21", "21汽": "21",
    "21機": "21", "21機25": "21", "21機25萬": "21",
    "21機車12萬": "21", "21機車25萬": "21",
    # 分貝方案 → 各自欄位
    "分貝汽": "分貝汽車",
    "分貝機": "分貝機車",
    # 創鉅方案 → 全部歸到「創鉅」欄位
    "創鉅手機": "創鉅", "創鉅手": "創鉅",
    "創鉅機車": "創鉅", "創鉅機": "創鉅",
    # 麻吉方案 → 全部歸到「麻吉」欄位
    "麻吉機車": "麻吉", "麻吉機": "麻吉",
    "麻吉手機": "麻吉", "麻吉手": "麻吉",
    # 帶金額版本 → 主公司
    "亞太工15": "亞太", "亞太25": "亞太", "亞太15": "亞太",
    "亞太汽車": "亞太",
    "亞太機25萬": "亞太", "亞太機車25萬": "亞太", "亞太機車15萬": "亞太",
    "21商品": "21",
    "合信機車": "合信", "合信手機": "合信", "合信二輪": "合信",
    "第一商品": "第一", "第一機車": "第一",
    "喬美40": "喬美", "喬美14": "喬美",
    "興達機車": "興達", "興達機": "興達",
    "維力": "和裕",
    "預付手機": "預付手機分期",
    # 新鑫/土地 → 房地
    "新鑫": "房地", "土地": "房地",
    # 向下相容舊字
    "貸就補": "貸救補",
    # 民間方案簡稱
    "鄉": "鄉民", "研": "商品貸", "當": "當舖專案", "商": "商品貸",
    "銀角": "零卡", "慢點付": "零卡", "分期趣": "零卡",
    # 鼎多 = 喬美
    "鼎多": "喬美",
}

def normalize_section(section: str) -> str:
    """日報區塊規範化。優先查 COMPANY_SECTION_MAP，fallback 用 COMPANY_ALIAS。
    例：「元大」→ COMPANY_SECTION_MAP 無 → COMPANY_ALIAS「元大→銀行」→ 歸「銀行」區塊
    """
    if section in COMPANY_SECTION_MAP:
        return COMPANY_SECTION_MAP[section]
    aliased = COMPANY_ALIAS.get(section)
    if aliased:
        # alias 後再查一次（例：「元大→銀行」，銀行在 REPORT_SECTION_2 直接顯示）
        return COMPANY_SECTION_MAP.get(aliased, aliased)
    return section


def build_section_map(all_rows) -> Dict[str, List[str]]:
    """把客戶列表轉成 section_map"""
    section_map: Dict[str, List[str]] = {}
    today_str = datetime.now().strftime("%Y-%m-%d")
    for row in all_rows:
        report_sec = row["report_section"] or ""
        current_co = row["current_company"] or row["company"] or ""
        section = report_sec or current_co or "送件"
        section = normalize_section(section)
        created = row["created_at"] or ""
        date_str = created[5:10].replace("-", "/") if created else ""
        company_str = current_co

        # 如果待撥款但還在送其他公司 → 兩個區塊都顯示
        extra_section = None
        if report_sec == "待撥款" and current_co:
            approved_companies = [h.get("company","") for h in get_all_approved(row["route_plan"] or "")]
            still_sending = current_co not in approved_companies and not any(current_co in ac or ac in current_co for ac in approved_companies)
            if still_sending:
                extra_section = normalize_section(current_co)

        last_update = row["last_update"] or ""
        first_line = last_update.splitlines()[0].strip() if last_update.strip() else ""
        status_short = extract_status_summary(first_line, row["customer_name"])

        # 讀取每家公司各自的狀態
        try:
            company_status = json.loads(row["company_status"] or "{}")
        except Exception:
            company_status = {}

        def get_section_status(sec_name):
            """取得該區塊對應公司的狀態（各家獨立，沒有就不顯示）"""
            if sec_name in company_status:
                cs_text = company_status[sec_name]
                # 從所有行找有意義的狀態（第一行通常是姓名+公司，狀態常在第二行）
                for ln in cs_text.splitlines():
                    ln = ln.strip()
                    if not ln:
                        continue
                    s = extract_status_summary(ln, row["customer_name"])
                    if s:
                        return s
                return ""
            # 如果是同時送件的公司但還沒收到A群回貼 → 不顯示狀態
            concurrent_list = [c.strip() for c in (row["concurrent_companies"] or "").split(",") if c.strip()]
            is_concurrent = any(normalize_section(c) == sec_name for c in concurrent_list)
            if is_concurrent:
                return ""
            # 如果客戶已有其他公司的狀態紀錄（例如前一家婉拒後推進），status_short 可能是
            # 前一家公司的狀態摘要（如「亞太婉拒」），不該顯示在當前送件公司的區塊。
            if company_status:
                return ""
            # status_short 來自 last_update 第一行：
            # - 訊息有提公司（如「喬美 婉拒」）→ 只套該家區塊（不污染其他）
            # - 訊息沒提公司（如「呂布 補申覆」）→ 套所有在送的區塊（泛狀態）
            if status_short:
                mentioned_co = extract_company(first_line) or ""
                if not mentioned_co:
                    return status_short  # 泛狀態、所有區塊都套
                if normalize_section(mentioned_co) == sec_name:
                    return status_short  # 訊息提到這家 → 套
                return ""  # 訊息提到別家 → 不套
            return status_short

        if section == "待撥款":
            created = row["created_at"] or ""
            created_date = created[5:10].replace("-", "/") if created else date_str
            amount = row["approved_amount"] or ""
            # 對保時間（已排定對保的待撥款案件）
            signing_time = (row["signing_time"] or "").strip()
            signing_date = signing_time.split()[0] if signing_time else ""
            pending_tag = f"(對保{signing_date})" if signing_date else "(待撥款)"
            # 從 route_plan 歷史找核准公司（每家各自顯示撥款狀態）
            approved_list = get_all_approved(row["route_plan"] or "")
            if approved_list:
                parts = []
                for ap in approved_list:
                    co = ap.get('company') or ''
                    amt = ap.get('amount') or ''
                    disb = ap.get('disbursed') or ''
                    if disb:
                        parts.append(f"{co}{amt}(撥款{disb})")
                    else:
                        parts.append(f"{co}{amt}{pending_tag}")
                amount_str = "-核准" + "/".join(parts)
            elif amount:
                disb_date = row["disbursement_date"] or ""
                disb_str = f"(撥款{disb_date})" if disb_date else pending_tag
                amount_str = f"-核准{amount}{disb_str}"
            else:
                amount_str = ""
            line = f"{created_date}-{row['customer_name']}{amount_str}"
        else:
            sec_status = get_section_status(section)
            line = f"{date_str}-{row['customer_name']}-{company_str}"
            if sec_status:
                line += f"-{sec_status}"
        # 今日新進件標記
        if created[:10] == today_str:
            line = "🆕" + line
        section_map.setdefault(section, []).append(line)
        # 還在送其他公司 → 也加到該公司區塊（不帶撥款資訊）
        if extra_section and extra_section != section:
            extra_line = f"{date_str}-{row['customer_name']}-{company_str}"
            if created[:10] == today_str:
                extra_line = "🆕" + extra_line
            section_map.setdefault(extra_section, []).append(extra_line)
        # 同時送件的公司也要顯示
        concurrent_str = row["concurrent_companies"] or ""
        if concurrent_str:
            for co in concurrent_str.split(","):
                co = co.strip()
                if not co:
                    continue
                co_section = normalize_section(co)
                if co_section != section:
                    co_status = get_section_status(co_section)
                    co_line = f"{date_str}-{row['customer_name']}-{co}"
                    if co_status:
                        co_line += f"-{co_status}"
                    if created[:10] == today_str:
                        co_line = "🆕" + co_line
                    section_map.setdefault(co_section, []).append(co_line)
    return section_map


def build_segment(sections: List[str], section_map: Dict, shown: set) -> str:
    """把指定 sections 的內容組成一段文字"""
    lines = []
    for sec in sections:
        if sec in section_map:
            lines.append(sec)
            lines.extend(section_map[sec])
            lines.append("——————————————")
            shown.add(sec)
    return "\n".join(lines)


def generate_report_lines(group_id: str) -> List[str]:
    """
    產生日報，分三段對話框：
    第1段：貸款方案（麻吉~手機分期）+ 送件 + 待撥款
    第2段：民間方案（銀行/零卡/商品貸/代書/當舖/核准）
    第3段：房地（房地/核准(房地)）
    每段都只在有客戶時才顯示，每段超過4500字再自動切割。
    """
    group_name = get_group_name(group_id)
    today = datetime.now().strftime("%m/%d")
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT * FROM customers WHERE source_group_id=? AND status='ACTIVE' ORDER BY updated_at DESC", (group_id,))
    all_rows = cur.fetchall(); conn.close()

    if not all_rows:
        return [f"📊 {group_name} 日報 {today}\n（目前無有效案件）"]

    section_map = build_section_map(all_rows)
    shown = set()
    segments = []

    # 第1段：貸款方案 + 送件
    seg1 = build_segment(REPORT_SECTION_1, section_map, shown)
    if seg1:
        segments.append(f"📊 {group_name} 日報 {today}\n{seg1}")

    # 第2段：民間方案
    seg2 = build_segment(REPORT_SECTION_2, section_map, shown)
    if seg2:
        segments.append(seg2)

    # 第3段：房地
    seg3 = build_segment(REPORT_SECTION_3, section_map, shown)
    if seg3:
        segments.append(seg3)

    # 未歸類的（不在任何section的，排除待撥款）補到第1段後面
    extra = []
    for sec, lines in section_map.items():
        if sec not in shown and sec != "待撥款":
            extra.append(sec)
            extra.extend(lines)
            extra.append("——————————————")
    if extra:
        if segments:
            segments[0] += "\n" + "\n".join(extra)
        else:
            segments.append(f"📊 {group_name} 日報 {today}\n" + "\n".join(extra))

    # 第4段：待撥款（排在最後）
    seg4 = build_segment(REPORT_SECTION_4, section_map, shown)
    if seg4:
        segments.append(seg4)

    # 待撥款超過7天提醒（獨立一段）
    overdue_lines = []
    for row in all_rows:
        if (row["report_section"] or "") == "待撥款" and not (row["disbursement_date"] or ""):
            created = row["created_at"] or ""
            if created:
                try:
                    days = (datetime.now() - datetime.fromisoformat(created.replace("Z",""))).days
                    if days >= 7:
                        signing_time = (row["signing_time"] or "").strip()
                        signing_date = signing_time.split()[0] if signing_time else ""
                        extra = f"，對保{signing_date}" if signing_date else ""
                        overdue_lines.append(f"  {row['customer_name']}（{days}天{extra}）")
                except Exception:
                    pass
    if overdue_lines:
        segments.append(f"⏰ 待撥款超過7天（{len(overdue_lines)}筆）\n" + "\n".join(overdue_lines))

    if not segments:
        return [f"📊 {group_name} 日報 {today}\n（目前無有效案件）"]

    # 每段超過4500字再切割
    final_segments = []
    for seg in segments:
        if len(seg) <= 4500:
            final_segments.append(seg)
        else:
            lines = seg.splitlines()
            current = ""
            for line in lines:
                if len(current) + len(line) + 1 > 4500:
                    final_segments.append(current.strip())
                    current = line + "\n"
                else:
                    current += line + "\n"
            if current.strip():
                final_segments.append(current.strip())

    return final_segments


def search_customer_info(name: str, group_id: str) -> str:
    rows = find_any_by_name(name)
    active = [r for r in rows if r["status"] == "ACTIVE" and r["source_group_id"] == group_id]
    if not active:
        active = [r for r in rows if r["status"] == "ACTIVE"]
    if not active:
        closed = [r for r in rows if r["status"] != "ACTIVE"]
        if closed:
            return f"❌ {name} 已結案（{closed[0]['updated_at'][:10]}）"
        return f"❌ 找不到客戶：{name}"
    r = active[0]
    route_data = parse_route_json(r["route_plan"] or "")
    order, idx = route_data.get("order", []), route_data.get("current_index", 0)
    history = route_data.get("history", [])
    lines = [f"👤 {name}"]
    if r["id_no"]:
        lines.append(f"身分證：{r['id_no']}")
    lines.append(f"所屬群組：{get_group_name(r['source_group_id'])}")
    if order:
        current = order[idx] if idx < len(order) else "（已完成）"
        lines.append(f"目前送件：{current}（第{idx+1}/{len(order)}家）")
        if idx + 1 < len(order):
            lines.append(f"下一家：{order[idx+1]}")
        if history:
            lines.append("送件歷程：")
            for h in history[-3:]:
                lines.append(f"  {h.get('date','')} {h.get('company','')} → {h.get('status','')}")
    last_update = (r["last_update"] or "").strip()
    if last_update:
        last_lines = [l.strip() for l in last_update.splitlines() if l.strip()]
        if last_lines:
            lines.append(f"最新進度：{last_lines[0][:80]}")
            for extra in last_lines[1:4]:
                lines.append(f"  {extra[:80]}")
    return "\n".join(lines)


# =========================
# Quick Reply 按鈕
# =========================
def send_reopen_case_buttons(reply_token, block_text, closed_rows, source_group_id, push_to_a_after_reopen=False):
    action_id = short_id()
    closed_rows = sorted(closed_rows, key=lambda x: x["updated_at"], reverse=True)[:1]
    save_pending_action(action_id, "reopen_or_new_case", {
        "block_text": block_text, "case_ids": [r["case_id"] for r in closed_rows],
        "source_group_id": source_group_id, "push_to_a_after_reopen": push_to_a_after_reopen,
    })
    items = [make_quick_reply_item(f'重啟-{c["customer_name"]}-{c["company"] or "未填"}',
                                    f"REOPEN_CASE|{action_id}|{c['case_id']}") for c in closed_rows]
    items += [make_quick_reply_item("建立新案件", f"CREATE_NEW_CASE|{action_id}"),
              make_quick_reply_item("取消", f"CANCEL_REOPEN|{action_id}")]
    reply_quick_reply(reply_token, "⚠️ 此客戶案件已結案，請選擇要重啟原案件或建立新案件", items)


def send_ambiguous_case_buttons(reply_token, block_text, matches):
    action_id = short_id()
    save_pending_action(action_id, "route_a_case", {"block_text": block_text, "case_ids": [m["case_id"] for m in matches]})
    items = [make_quick_reply_item(
        f"{c['customer_name']}-{c['company'] or '未填'}-{get_group_name(c['source_group_id'])}",
        f"SELECT_CASE|{action_id}|{c['case_id']}") for c in matches[:10]]
    reply_quick_reply(reply_token, "⚠️ 多筆同名客戶，請選擇要回貼的案件", items)


def send_transfer_case_buttons(reply_token, customer, source_group_id, block_text, allow_new=True):
    # Bug 6: 驗證 source_group_id 合法性，避免按鈕回調寫入不存在的群組
    new_g = get_group_name(source_group_id) if source_group_id else ""
    if not source_group_id or not new_g or new_g == "未知群組":
        reply_text(reply_token, "⚠️ 無法轉送：來源群組不存在或尚未註冊，請聯絡管理員")
        return
    action_id = short_id()
    save_pending_action(action_id, "transfer_customer", {
        "case_id": customer["case_id"], "target_group_id": source_group_id,
        "block_text": block_text, "name": extract_name(block_text) or customer["customer_name"],
    })
    old_g = get_group_name(customer["source_group_id"])
    # 三按鈕語意：沿用 = 兩邊各有獨立案件；轉移 = 原群組搬到新群組；取消
    items = [make_quick_reply_item(f"沿用(兩邊都有)", f"CREATE_NEW_FROM_TRANSFER|{action_id}"),
             make_quick_reply_item(f"轉移到{new_g}", f"CONFIRM_TRANSFER|{action_id}"),
             make_quick_reply_item("取消", f"CANCEL_TRANSFER|{action_id}")]
    reply_quick_reply(reply_token, f"⚠️ {old_g} 已有同名客戶，請選擇：沿用(兩邊各建獨立案)/轉移(搬到{new_g})/取消", items)


def send_same_name_supplement_buttons(reply_token, block_text, matches, source_group_id, want_push_a):
    """本群組多筆同名補件時 → 跳按鈕讓業務選要更新哪位"""
    action_id = short_id()
    save_pending_action(action_id, "select_case_for_supplement", {
        "block_text": block_text, "source_group_id": source_group_id,
        "want_push_a": want_push_a, "case_ids": [m["case_id"] for m in matches]
    })
    items = []
    for r in matches[:12]:
        id4 = (r["id_no"] or "")[-4:] or "無"
        co = r["current_company"] or r["company"] or "未填"
        items.append(make_quick_reply_item(
            f"{co}-末4:{id4}", f"SELECT_SUPPLEMENT|{action_id}|{r['case_id']}"))
    items.append(make_quick_reply_item("取消", f"CANCEL_SUPPLEMENT|{action_id}"))
    name = extract_name(block_text) or matches[0]["customer_name"]
    reply_quick_reply(reply_token, f"⚠️ 本群組有 {len(matches)} 位「{name}」，要更新哪位？", items)


def send_same_name_diff_id_buttons(reply_token, block_text, matches, source_group_id,
                                     new_id, new_name, new_company):
    """本群組已有同姓名但身分證不同 → 跳按鈕讓業務確認是同一人（身分證打錯）或新客戶"""
    action_id = short_id()
    save_pending_action(action_id, "same_name_diff_id", {
        "block_text": block_text, "source_group_id": source_group_id,
        "new_id": new_id, "new_name": new_name, "new_company": new_company,
        "case_ids": [m["case_id"] for m in matches]
    })
    items = []
    for r in matches[:10]:
        id4 = (r["id_no"] or "")[-4:] or "無"
        co = r["current_company"] or r["company"] or "未填"
        items.append(make_quick_reply_item(
            f"同一人-{co}末4:{id4}", f"SAME_PERSON|{action_id}|{r['case_id']}"))
    items.append(make_quick_reply_item("不同人(建新)", f"NEW_PERSON|{action_id}"))
    items.append(make_quick_reply_item("取消", f"CANCEL_SAMENAME|{action_id}"))
    reply_quick_reply(reply_token,
                      f"⚠️ 本群組已有「{new_name}」，新打的身分證({new_id})不同，是同一人還是新客戶？",
                      items)


def send_confirm_new_case_buttons(reply_token, block_text, existing_customer, source_group_id):
    action_id = short_id()
    save_pending_action(action_id, "confirm_new_case_with_existing_id", {
        "block_text": block_text, "source_group_id": source_group_id,
        "existing_case_id": existing_customer["case_id"],
    })
    old_g, new_g = get_group_name(existing_customer["source_group_id"]), get_group_name(source_group_id)
    name = extract_name(block_text) or existing_customer["customer_name"]
    # 三按鈕語意：沿用 = 兩邊各有獨立案件（新群組建新案、原群組保留）；轉移 = 搬到新群組；取消
    items = [make_quick_reply_item(f"沿用(兩邊都有)", f"FORCE_CREATE_NEW|{action_id}"),
             make_quick_reply_item(f"轉移到{new_g}", f"TRANSFER_FROM_CONFIRM|{action_id}"),
             make_quick_reply_item("取消", f"CANCEL_NEW_CASE|{action_id}")]
    reply_quick_reply(reply_token, f"⚠️ {name} 身分證已存在於{old_g}，請選擇：沿用(兩邊各建獨立案)/轉移(搬到{new_g})/取消", items)


# =========================
# 特殊指令
# =========================
def parse_special_command(text: str, group_id: str) -> Optional[Dict]:
    clean = strip_ai_trigger(text).strip()
    # 文字正規化（全形→半形、異體字統一、空白合併）
    clean = normalize_command_text(clean)

    # 批次結案：@AI 批次結案\n姓名1\n姓名2\n...
    if clean.startswith("批次結案"):
        rest = clean[len("批次結案"):].strip()
        names = [n.strip() for n in rest.splitlines() if n.strip()]
        if names:
            return {"type": "batch_close", "names": names}

    # 批次婉拒：@AI 批次婉拒\n姓名1\n姓名2\n...
    if clean.startswith("批次婉拒"):
        rest = clean[len("批次婉拒"):].strip()
        names = [n.strip() for n in rest.splitlines() if n.strip()]
        if names:
            return {"type": "batch_reject", "names": names}

    # 待撥款名單：@AI 待撥款
    if re.match(r"^待撥款$", clean):
        return {"type": "pending_disbursement"}

    # 統計：@AI 統計
    if re.match(r"^統計$", clean):
        return {"type": "stats"}

    # 群組ID查詢（不分大小寫）
    if re.match(r"^群組[Ii][Dd]$", clean):
        return {"type": "group_id"}

    # 日報
    if re.match(r"^日報$", clean):
        return {"type": "report"}

    # 指令速查卡（只留兩個：@AI 格式 / @AI 說明）
    if clean in ["格式", "說明"]:
        return {"type": "help"}

    # 查詢：支援所有格式
    # @AI 查 彭駿為 / 彭駿為 查@AI / 彭駿為@AI 查
    m = re.match(r"^查\s*([\u4e00-\u9fff]{2,6})$", clean)
    if m:
        return {"type": "search", "name": m.group(1)}
    m = re.match(r"^([\u4e00-\u9fff]{2,6})\s*查$", clean)
    if m:
        return {"type": "search", "name": m.group(1)}

    # 轉下一家
    m = re.match(r"^([\u4e00-\u9fff]{2,6})\s*轉下一家$", clean)
    if m:
        return {"type": "advance", "name": m.group(1), "target": None}

    # 加送：@AI 姓名 送公司[+公司] [金額/期數] → 加入同時送件清單，不換掉原本的
    m = re.match(r"^([\u4e00-\u9fff]{2,6})\s*送\s*([^\s]+)(?:\s+(\d+(?:\.\d+)?)\s*萬?\s*[/／]\s*(\d+)\s*期?)?\s*$", clean)
    if m and "下一家" not in m.group(2) and "轉" not in m.group(2):
        return {"type": "add_concurrent", "name": m.group(1), "company": m.group(2).strip(),
                "notify_amount": m.group(3) or "", "notify_period": m.group(4) or ""}

    # 轉指定公司（支援多家同送「轉A+B」、可帶送件金額「轉A+B 100萬/24期」）
    m = re.match(r"^([\u4e00-\u9fff]{2,6})\s*轉([^\s]+)(?:\s+(\d+(?:\.\d+)?)\s*萬?\s*[/／]\s*(\d+)\s*期?)?\s*$", clean)
    if m and m.group(2).strip() != "下一家":
        return {"type": "advance", "name": m.group(1), "target": m.group(2).strip(),
                "notify_amount": m.group(3) or "", "notify_period": m.group(4) or ""}

    # 取消核准：@AI 姓名 公司 取消核准
    m = re.match(r"^([\u4e00-\u9fff]{2,6})\s*(.+?)\s*取消核准$", clean)
    if m:
        return {"type": "cancel_approval", "name": m.group(1), "company": m.group(2).strip()}

    # 修改核准金額：@AI 姓名 公司 核准 金額（姓名和公司之間必須有空格）
    # 異體字（核準→核准）由 normalize_command_text 統一；用 lookbehind 排除「待核准」
    m = re.match(r"^([\u4e00-\u9fff]{2,6})\s+(.+?)\s*(?<!待)核准\s*(.+)$", clean)
    if m:
        return {"type": "update_amount", "name": m.group(1), "company": m.group(2).strip(), "amount": m.group(3).strip()}

    # 核准金額（未指定公司 → 用當前 current_company）：姓名核准 金額
    m = re.match(r"^([\u4e00-\u9fff]{2,6})\s*(?<!待)核准\s*(.+)$", clean)
    if m:
        return {"type": "update_amount", "name": m.group(1), "company": "", "amount": m.group(2).strip()}

    # 撥款（日期在前、有公司）：@AI 姓名 公司 M/D 撥款
    # 排除「結案/違約金/取消核准」等動詞字被當公司名
    m = re.match(r"^([\u4e00-\u9fff]{2,6})\s+(?!結案|違約金|取消核准)(\S+)\s+(\d{1,2}/\d{1,2})\s*撥款$", clean)
    if m:
        return {"type": "disbursed", "name": m.group(1),
                "company": m.group(2).strip(), "date": m.group(3)}

    # 撥款（日期在前、無公司）：@AI 姓名 M/D 撥款
    m = re.match(r"^([\u4e00-\u9fff]{2,6})\s+(\d{1,2}/\d{1,2})\s*撥款$", clean)
    if m:
        return {"type": "disbursed", "name": m.group(1),
                "company": "", "date": m.group(2)}

    # 撥款指定公司（日期在後）：@AI 姓名 公司 撥款 M/D
    # 排除「結案 已撥款」等會被誤判（「結案 已」當公司）
    m = re.match(r"^([\u4e00-\u9fff]{2,6})\s+(?!結案|違約金|取消核准)(.+?)\s*撥款\s*(\d{1,2}/\d{1,2})?$", clean)
    if m:
        return {"type": "disbursed", "name": m.group(1),
                "company": m.group(2).strip(), "date": m.group(3) or ""}

    # 撥款（無公司、日期在後或省略）：@AI 姓名 撥款 M/D
    # 要先於 missing_verb，否則 4/19 會被誤判為 4萬/19期
    m = re.match(r"^([\u4e00-\u9fff]{2,6})\s*撥款\s*(\d{1,2}/\d{1,2})?$", clean)
    if m:
        return {"type": "disbursed", "name": m.group(1),
                "company": "", "date": m.group(2) or ""}

    # 防錯：看起來像「姓名 公司 金額/期數」但缺動詞（轉/送/同送/核准）
    # handler 內會判斷：動詞字、亂公司、合法公司 三種情境分別提示
    m = re.match(r"^([\u4e00-\u9fff]{2,6})\s+([^\s@]+?)\s*(\d+(?:\.\d+)?)\s*萬?\s*[/／]\s*(\d+)\s*期?\s*$", clean)
    if m:
        return {"type": "missing_verb", "name": m.group(1), "companies_raw": m.group(2),
                "amount": m.group(3), "period": m.group(4)}

    # 防錯：金額期數沒用「/」分隔（空白或其他符號）
    # 例：「周馮鈺婷 轉喬美 100 24」→ 提示使用者用「N萬/N期」
    m = re.match(r"^([\u4e00-\u9fff]{2,6})\s*(轉|送)([^\s]+)\s+(\d+(?:\.\d+)?)\s*萬?\s+(\d+)\s*期?\s*$", clean)
    if m:
        return {"type": "bad_amount_format", "name": m.group(1), "verb": m.group(2),
                "target": m.group(3), "n1": m.group(4), "n2": m.group(5)}

    # 缺日期的送件順序：姓名-公司/公司/...（沒前綴日期）
    # 例：「郭富城-第一/亞太機/21機/銀行」
    m = re.match(r"^([\u4e00-\u9fff]{2,6})-([^\s]+)$", clean)
    if m and "/" in m.group(2):
        cos = [c.strip() for c in m.group(2).split("/") if c.strip()]
        valid = _get_valid_company_names()
        if any(COMPANY_ALIAS.get(c, c) in valid for c in cos):
            return {"type": "missing_date_hint", "name": m.group(1), "tail": m.group(2)}

    # 防錯：訊息沒任何空白但含中文 + 數字/ → 可能是「姓名公司黏在一起」
    # 例：「周馮鈺婷喬美+房地100萬/120期」
    if " " not in clean and "/" in clean and any(c in clean for c in COMPANY_LIST):
        return {"type": "no_space_hint", "raw": clean}

    # 改名：@AI 舊名 改名 新名
    m = re.match(r"^([\u4e00-\u9fff]{2,6})\s*改名\s*([\u4e00-\u9fff]{2,6})$", clean)
    if m:
        return {"type": "rename", "old_name": m.group(1), "new_name": m.group(2)}

    # 改身分證：@AI 姓名 改身分證 新ID（尾巴可加備註）
    m = re.match(r"^([\u4e00-\u9fff]{2,6})\s*改身分證\s*([A-Z][A-Z0-9]\d{8})(?:\s+.*)?$", clean, re.IGNORECASE)
    if m:
        return {"type": "change_id", "name": m.group(1), "new_id": m.group(2).upper()}
    # 改身分證 + ID 格式錯（例如少一位、打錯字）→ 提示格式，不落到新客戶建立
    m = re.match(r"^([\u4e00-\u9fff]{2,6})\s*改身分證\s*(.+)$", clean, re.IGNORECASE)
    if m:
        return {"type": "bad_id_format", "name": m.group(1), "raw_id": m.group(2).strip()}

    # 重啟：@AI 姓名 重啟
    m = re.match(r"^([\u4e00-\u9fff]{2,6})\s*重啟$", clean)
    if m:
        return {"type": "reopen", "name": m.group(1)}

    # 查歷史：@AI 姓名 歷史
    m = re.match(r"^([\u4e00-\u9fff]{2,6})\s*歷史$", clean)
    if m:
        return {"type": "history", "name": m.group(1)}

    # 還原：@AI 姓名 還原 N（N 預設 1 = 最近一筆）
    m = re.match(r"^([\u4e00-\u9fff]{2,6})\s*還原\s*(\d+)?$", clean)
    if m:
        try:
            idx = int(m.group(2) or "1")
        except Exception:
            idx = 1
        return {"type": "restore", "name": m.group(1), "index": idx}

    # 結案帶原因：@AI 姓名 結案 原因
    m = re.match(r"^([一-鿿]{2,6})\s*結案\s+(.+)$", clean)
    if m:
        return {"type": "close", "name": m.group(1), "reason": m.group(2).strip()}

    # 結案
    m = re.match(r"^([一-鿿]{2,6})\s*(已結案|結案)$", clean)
    if m:
        return {"type": "close", "name": m.group(1)}

    # 批次結案/婉拒格式錯：姓名出現在「批次結案/婉拒」前
    # 例：「黎明 王陽明 批次結案」→ 提示換行格式
    m = re.match(r"^(.+)\s+批次(結案|婉拒)$", clean)
    if m:
        return {"type": "bad_batch_format", "action": m.group(2), "raw": clean}

    # 【A 方案】姓名 公司 結案 → 客戶已核准時、只從同送清單移除該家（不結案客戶）
    # 客戶未核准時 → handler 會退回普通結案並提示
    m = re.match(r"^([一-鿿]{2,6})\s+(.+?)\s*結案$", clean)
    if m:
        return {"type": "remove_concurrent_or_close", "name": m.group(1),
                "company": m.group(2).strip()}

    # 婉拒 轉XXX（明確指定要婉拒的公司 + 跳到指定公司）
    # 例：「王陽明 裕融 婉拒 轉亞太」= 婉拒裕融（不管在 current 或 concurrent）+ current 改亞太
    m = re.match(r"^([一-鿿]{2,6})\s+(.+?)\s*婉拒\s*轉\s*(.+)$", clean)
    if m:
        return {"type": "reject_to", "name": m.group(1),
                "company": m.group(2).strip(), "target": m.group(3).strip()}

    # 婉拒 轉XXX（沒指定公司，婉拒當前 current_company）
    m = re.match(r"^([一-鿿]{2,6})\s*婉拒\s*轉\s*(.+)$", clean)
    if m:
        return {"type": "reject_to", "name": m.group(1),
                "company": "", "target": m.group(2).strip()}

    # 婉拒 + 加送 複合指令：@AI 姓名 公司A婉拒 送公司B
    # 例：吳瑞銘 喬美婉拒 送貸10（喬美婉拒、加送貸救補）
    m = re.match(r"^([一-鿿]{2,6})\s+(.+?)\s*婉拒\s*送\s*([^\s]+)\s*$", clean)
    if m:
        return {"type": "reject_and_add", "name": m.group(1),
                "reject_company": m.group(2).strip(),
                "add_company": m.group(3).strip()}

    # 婉拒（推到下一家）
    m = re.match(r"^([一-鿿]{2,6})\s*婉拒$", clean)
    if m:
        return {"type": "reject", "name": m.group(1)}

    # 確認婉拒（房地/銀行/C 類別的二次確認指令）：@AI 姓名 公司 確認婉拒
    m = re.match(r"^([一-鿿]{2,6})\s+(.+?)\s*確[定認]婉拒$", clean)
    if m:
        return {"type": "confirm_reject_company", "name": m.group(1), "company": m.group(2).strip()}

    # 指定某家婉拒（不跳轉）：@AI 姓名 公司 婉拒
    # 例：林俊杰同送房地+銀行，銀行評估不過 → 「林俊杰 銀行 婉拒」
    m = re.match(r"^([一-鿿]{2,6})\s+(.+?)\s*婉拒$", clean)
    if m:
        return {"type": "reject_company", "name": m.group(1), "company": m.group(2).strip()}

    # 違約金已支付 xxxx（有收到違約金）
    m = re.match(r"^([一-鿿]{2,6})\s*違約金已支付\s*([\d,，]+)", clean)
    if m:
        amt = m.group(2).replace(",","").replace("，","")
        return {"type": "penalty", "name": m.group(1), "penalty": amt}

    # 照會：多種格式都支援
    # @AI 照會 王小明 亞太+和裕
    m = re.match(r"^照會\s*([\u4e00-\u9fff]{2,6})(?:\s+(.+))?$", clean)
    if m:
        return {"type": "notification", "name": m.group(1), "company": (m.group(2) or "").strip()}
    # @AI 王小明 照會 亞太+和裕
    m = re.match(r"^([\u4e00-\u9fff]{2,6})\s*照會(?:\s+(.+))?$", clean)
    if m:
        return {"type": "notification", "name": m.group(1), "company": (m.group(2) or "").strip()}
    # @AI 王小明 亞太+和裕 照會
    m = re.match(r"^([\u4e00-\u9fff]{2,6})\s+(.+?)\s*照會$", clean)
    if m:
        return {"type": "notification", "name": m.group(1), "company": m.group(2).strip()}

    # 對保派件（A 訊息）：辦理方案/核准金額/客戶姓名/對保地區
    if "客戶姓名" in clean and "對保地區" in clean:
        fields = {}
        for line in clean.splitlines():
            mm = re.match(r"^\s*(辦理方案|核准金額|客戶姓名|對保地區)\s*[:：]\s*(.+?)\s*$", line)
            if mm:
                fields[mm.group(1)] = mm.group(2).strip()
        name = fields.get("客戶姓名", "")
        if name:
            return {
                "type": "signing_request",
                "name": name,
                "plan": fields.get("辦理方案", ""),
                "amount": fields.get("核准金額", ""),
                "area": fields.get("對保地區", ""),
            }

    # 對保時間地點（B 訊息）：對保 XXX XX對保 / 時間 X / 地點 X
    lines_b = [ln.strip() for ln in clean.splitlines() if ln.strip()]
    if lines_b:
        m = re.match(r"^對保\s+([\u4e00-\u9fff]{2,6})\s+(.+?)對保\s*$", lines_b[0])
        if m:
            salesperson = m.group(1)
            signing_co = m.group(2).strip()
            time_str = ""
            loc_str = ""
            for line in lines_b[1:]:
                mt = re.match(r"^時間\s*[:：]?\s*(.+)$", line)
                if mt:
                    time_str = mt.group(1).strip()
                    continue
                ml = re.match(r"^地點\s*[:：]?\s*(.+)$", line)
                if ml:
                    loc_str = ml.group(1).strip()
            return {
                "type": "signing_schedule",
                "salesperson": salesperson,
                "signing_company": signing_co,
                "time": time_str,
                "location": loc_str,
            }

    # ===== 常見錯誤格式防錯（給正確範例提示）=====
    # 轉但沒指定公司
    m = re.match(r"^([\u4e00-\u9fff]{2,6})\s*轉\s*$", clean)
    if m:
        return {"type": "format_hint", "hint": f"轉哪家？\n例：@AI {m.group(1)} 轉喬美\n帶金額：@AI {m.group(1)} 轉喬美 50萬/24期"}
    # 送但沒指定公司
    m = re.match(r"^([\u4e00-\u9fff]{2,6})\s*送\s*$", clean)
    if m:
        return {"type": "format_hint", "hint": f"送哪家？\n例：@AI {m.group(1)} 送第一\n帶金額：@AI {m.group(1)} 送第一 30萬/24期"}
    # 核准但沒公司沒金額
    m = re.match(r"^([\u4e00-\u9fff]{2,6})\s*核准\s*$", clean)
    if m:
        return {"type": "format_hint", "hint": f"哪家核准？金額多少？\n例：@AI {m.group(1)} 第一 核准 30萬"}
    # 違約金沒金額
    m = re.match(r"^([\u4e00-\u9fff]{2,6})\s*違約金[已支付]*\s*$", clean)
    if m:
        return {"type": "format_hint", "hint": f"違約金多少？\n例：@AI {m.group(1)} 違約金已支付30000"}
    # 改名沒新名
    m = re.match(r"^([\u4e00-\u9fff]{2,6})\s*改名\s*$", clean)
    if m:
        return {"type": "format_hint", "hint": f"改成什麼名字？\n例：@AI {m.group(1)} 改名 新名字"}
    # 改身分證沒 ID
    m = re.match(r"^([\u4e00-\u9fff]{2,6})\s*改身分證\s*$", clean)
    if m:
        return {"type": "format_hint", "hint": f"新身分證？\n例：@AI {m.group(1)} 改身分證 A123456789"}
    # 取消核准沒指定公司
    m = re.match(r"^([\u4e00-\u9fff]{2,6})\s*取消核准\s*$", clean)
    if m:
        return {"type": "format_hint", "hint": f"要取消哪家的核准？\n例：@AI {m.group(1)} 第一 取消核准"}
    # 撥款動詞在前（順序錯）
    m = re.match(r"^撥款\s+([\u4e00-\u9fff]{2,6})", clean)
    if m:
        return {"type": "format_hint", "hint": f"撥款指令格式：姓名在前\n例：@AI {m.group(1)} 撥款 4/18\n或：@AI {m.group(1)} 裕融 撥款 4/18"}
    # 結案動詞在前
    m = re.match(r"^結案\s+([\u4e00-\u9fff]{2,6})", clean)
    if m:
        return {"type": "format_hint", "hint": f"結案指令格式：姓名在前\n例：@AI {m.group(1)} 結案"}
    # 婉拒動詞在前
    m = re.match(r"^婉拒\s+([\u4e00-\u9fff]{2,6})", clean)
    if m:
        return {"type": "format_hint", "hint": f"婉拒指令格式：姓名在前\n例：@AI {m.group(1)} 婉拒"}

    return None


def generate_notification_text(r: dict, company: str = "") -> str:
    """根據客戶 DB 資料自動產生照會注意事項文字"""
    def v(k): return (r.get(k, "") or "").strip()

    name = v("customer_name")
    # 公司：優先用參數，其次 current_company，最後 company
    co = company or v("current_company") or v("company") or ""

    # 居住地：同戶籍→戶籍，否則→現居地
    live_type = "戶籍" if v("live_same_as_reg") == "1" else "現居地"
    live_years = v("live_years") or "0"
    live_status = v("live_status") or "自有"

    # 年資
    co_years = v("company_years") or "0"

    # 月薪：轉成「N萬」或「N萬N」格式
    salary_str = ""
    try:
        sal = float(v("company_salary") or "0")
        if sal >= 10000:
            wan = int(sal // 10000)
            remainder = int((sal % 10000) // 1000)
            salary_str = f"{wan}萬{remainder}" if remainder else f"{wan}萬"
        elif sal > 0:
            salary_str = f"{sal}"
        else:
            salary_str = "0"
    except Exception:
        salary_str = v("company_salary") or "0"

    # 送件金額/期數（notify_amount ≠ approved_amount）：
    # 優先用手動設定的送件金額 → 沒設則用 PLAN_INFO 方案預設 → 再沒有則用 eval_fund_need
    # 不 fallback 到 approved_amount（核准金額），避免客戶核准後照會訊息顯示核准金額造成誤導
    manual_amount = v("notify_amount")
    manual_period = v("notify_period")
    plan_info = PLAN_INFO.get(co)
    if manual_amount:
        amt_str = manual_amount if "萬" in manual_amount else f"{manual_amount}萬"
        amount_line = f"{amt_str}/{manual_period}期" if manual_period else amt_str
    elif plan_info and plan_info[1]:
        amount_line = plan_info[1]
    else:
        amount_line = v("eval_fund_need") or ""

    # 學歷：轉成口語
    edu = v("education")
    EDU_MAP = {"專科/大學": "大學", "專科、大學": "大學", "高中/職": "高中",
               "高中職": "高中", "研究所以上": "研究所", "其他": "高中"}
    edu_spoken = EDU_MAP.get(edu, edu) if edu else "高中"

    # 資金用途
    fund = v("adminb_fund_use") or "家用"

    # 名下車貸狀況：檢查 debt_list
    car_loan_status = "名下無貸款"
    try:
        debt_list = json.loads(v("debt_list")) if v("debt_list") else []
        for d in debt_list:
            co_name = (d.get("co", "") or "").lower()
            dy = d.get("dy", "") or ""
            if any(w in co_name for w in ["車", "機車", "汽車"]):
                if "公路" in dy or "動保" in dy:
                    car_loan_status = "名下車貸正常繳"
                break
    except Exception:
        pass

    # 組合照會文字
    lines = [
        name,
        "照會注意事項",
        f"✅現居住{live_type}地址 居住{live_years}年 {live_status} 工作年資{co_years}年 月薪{salary_str}",
    ]
    if amount_line:
        lines.append(f"✅{amount_line}")
    lines += [
        "✅帳單簡訊條碼繳款",
        "✅ 姓名親簽",
        f"✅學歷詢問到麻煩說{edu_spoken}畢",
        f"✅資金用途 ：{fund}",
        "✅ 詢問任何法學，刑事，一律都說不是自己的",
        "✅假如問到在外面有沒有欠款公司要說沒有！",
        f"✅{car_loan_status}",
        "",
        "⚠️不好的一概否認喔，比如卡循卡債就算清償了，沒拉聯徵，會問的比較多！任何法學 如果詢問麻煩也都否認💯",
        "",
        "💌確認進件照會時間",
        "     🎀白天通知進件 中午或是下午以前注意來電",
    ]
    return "\n".join(lines)


def _get_valid_company_names():
    """所有系統認識的公司名（COMPANY_LIST + COMPANY_ALIAS + PLAN_INFO + COMPANY_SECTION_MAP）"""
    valid = set(COMPANY_LIST) | set(COMPANY_ALIAS.keys()) | set(COMPANY_ALIAS.values())
    valid |= {v[0] for v in PLAN_INFO.values()} | set(PLAN_INFO.keys())
    valid |= set(COMPANY_SECTION_MAP.keys()) | set(COMPANY_SECTION_MAP.values())
    return valid


def _validate_companies_or_warn(companies, reply_token, name):
    """驗證公司名清單，若有未知回覆警告 + 合法名清單 並回 False；全部合法回 True"""
    valid = _get_valid_company_names()
    unknown = [c for c in companies if c not in valid]
    if unknown:
        reply_text(reply_token,
                   f"⚠️ {name}：找不到公司「{'、'.join(unknown)}」\n"
                   f"常見公司名：亞太、喬美、第一、房地、21、裕融、和裕、麻吉、貸救補、鄉民、銀行、零卡、商品貸、代書、當舖")
        return False
    return True


def _field_display_label(field_name: str) -> str:
    """DB 欄位名翻譯成業務看得懂的中文（用於執行摘要）"""
    return {
        "company": "公司",
        "current_company": "當前公司",
        "concurrent_companies": "同送",
        "report_section": "日報區塊",
        "approved_amount": "核准金額",
        "notify_amount": "送件金額(萬)",
        "notify_period": "送件期數",
        "disbursement_date": "撥款日",
        "status": "狀態",
        "id_no": "身分證",
        "customer_name": "姓名",
        "route_plan": "送件順序",
        "source_group_id": "所屬群組",
    }.get(field_name, field_name)


def update_with_verify(case_id: str, changes: Dict, from_group_id: str = "", text_log: str = ""):
    """執行 update_customer 並實測 DB 前後變化，回傳 (ok, diff_lines, customer_name)。
    changes 是要 update 的欄位 dict；diff_lines 是實際變動的欄位摘要（「欄位: 舊 → 新」）。
    若 case_id 不存在 → ok=False；所有欄位都沒變 → diff_lines 空（表面無變化）。
    """
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT * FROM customers WHERE case_id=?", (case_id,))
    before = cur.fetchone()
    if not before:
        conn.close()
        return False, ["⚠️ 案件不存在"], ""
    before_dict = dict(before)
    conn.close()
    # 過濾 + key 轉換（update_customer 簽名用 name 對應 customer_name）
    allowed_keys = {"company", "status", "source_group_id", "route_plan",
                    "current_company", "report_section", "approved_amount",
                    "disbursement_date", "signing_area", "signing_salesperson",
                    "signing_company", "signing_time", "signing_location",
                    "notify_amount", "notify_period", "concurrent_companies",
                    "id_no", "name"}
    filtered = {}
    for k, v in changes.items():
        if k == "customer_name":
            filtered["name"] = v
        elif k in allowed_keys:
            filtered[k] = v
    kwargs = {"text": text_log, "from_group_id": from_group_id}
    kwargs.update(filtered)
    update_customer(case_id, **kwargs)
    conn2 = get_conn(); cur2 = conn2.cursor()
    cur2.execute("SELECT * FROM customers WHERE case_id=?", (case_id,))
    after = cur2.fetchone()
    conn2.close()
    after_dict = dict(after) if after else {}
    diffs = []
    for k in changes.keys():
        if k in ("text", "from_group_id"):
            continue
        old_v = before_dict.get(k)
        new_v = after_dict.get(k)
        if old_v != new_v:
            old_disp = old_v if old_v not in (None, "") else "(空)"
            new_disp = new_v if new_v not in (None, "") else "(空)"
            diffs.append(f"• {_field_display_label(k)}: {old_disp} → {new_disp}")
    return True, diffs, before_dict.get("customer_name", "")


def normalize_command_text(text: str) -> str:
    """@AI 指令文字正規化：全形→半形、異體字統一、多空白合併。
    parse_special_command 進入前先跑，讓 regex 只需寫一套標準寫法。
    """
    if not text:
        return ""
    # 全形英數/標點 → 半形（U+FF01~U+FF5E）
    out = []
    for ch in text:
        code = ord(ch)
        if 0xFF01 <= code <= 0xFF5E:
            out.append(chr(code - 0xFEE0))
        elif code == 0x3000:  # 全形空白
            out.append(" ")
        else:
            out.append(ch)
    text = "".join(out)
    # 符號統一
    text = text.replace("／", "/").replace("＋", "+")
    # 異體字統一（台灣業務可能打「核准/核準/核凖、身份證/身分證」）
    text = text.replace("核準", "核准").replace("核凖", "核准")
    text = text.replace("身份證", "身分證")
    # 多個空白/Tab → 單一空白
    text = re.sub(r"[ \t]+", " ", text).strip()
    return text


def _validate_new_case_fields(date: str, name: str, id_no: str):
    """建客戶前檢查資料，回傳警告訊息列表（不 block 建立，只提示）。"""
    warnings = []
    if date:
        m = re.match(r"^(\d{1,4})/(\d{1,2})", date)
        if m:
            y = int(m.group(1))
            # 合理：民國 100-130（2011-2041）、西元 2020-2100；3 位數 or 2 位數民國年常見
            if 1 <= y <= 3:
                warnings.append(f"日期「{date}」民國年太小，可能打錯")
            elif 4 <= y < 80:  # 民國 4~79 (1915~1990)，年紀太大的客戶？
                warnings.append(f"日期「{date}」對應民國{y}年({1911+y}年)，確認無誤？")
            elif 131 <= y < 1000:
                warnings.append(f"日期「{date}」民國年太大，可能打錯")
    if id_no:
        if not re.match(r"^[A-Z][A-Z0-9]\d{8}$", id_no, re.IGNORECASE):
            warnings.append(f"身分證「{id_no}」格式不標準（應為 1 英文字母 + 1 英數 + 8 位數字）")
        elif not validate_tw_id_checksum(id_no):
            warnings.append(f"身分證「{id_no}」校驗位不符，可能打錯一碼")
    if name:
        if len(name) > 5:
            warnings.append(f"姓名「{name}」長度 {len(name)} 字，確認無誤？")
    return warnings


def _resolve_target(name: str, group_id: str, reply_token: str):
    """依姓名在本群組找 ACTIVE 客戶（輕量版：多筆回警告不跳按鈕）"""
    rows = find_active_by_name(name)
    same = [r for r in rows if r["source_group_id"] == group_id]
    if len(same) == 1:
        return same[0]
    if len(same) > 1:
        lines = [f"⚠️ 本群組有 {len(same)} 位「{name}」，請用身分證或更精確識別："]
        for r in same:
            id4 = (r["id_no"] or "")[-4:] or "無"
            co = r["current_company"] or r["company"] or "未填"
            lines.append(f"  - {co}（身分證末4：{id4}）")
        reply_text(reply_token, "\n".join(lines))
        return None
    if rows:
        return rows[0]
    reply_text(reply_token,
               f"❌ 找不到客戶：{name}\n"
               f"可能原因：\n"
               f"  • 姓名打錯（2-6 個中文字）\n"
               f"  • 客戶還沒建立\n"
               f"  • 客戶已結案 → 先打「@AI {name} 重啟」")
    return None


def _check_active_or_warn(target, reply_token, action_label: str, customer_name: str = ""):
    """破壞性指令前檢查 target 是否為 ACTIVE；非 ACTIVE 則警告並回 False"""
    if not target:
        return False
    status = target["status"] if "status" in target.keys() else None
    if status == "ACTIVE":
        return True
    status_label = {"CLOSED": "已結案", "PENALTY": "違約金結案",
                    "ABANDONED": "已放棄", "REJECTED": "全數婉拒",
                    "PENDING": "尚未啟用"}.get(status, f"狀態 {status}")
    name = customer_name or (target["customer_name"] if "customer_name" in target.keys() else "此客戶")
    reply_text(reply_token,
               f"⚠️ {name} {status_label}，無法執行「{action_label}」\n"
               f"如需操作請先打「@AI {name} 重啟」讓客戶回到進行中")
    return False


# 台灣身分證字母對應值（首碼轉換）
_ID_LETTER_MAP = {
    'A': 10, 'B': 11, 'C': 12, 'D': 13, 'E': 14, 'F': 15, 'G': 16, 'H': 17,
    'I': 34, 'J': 18, 'K': 19, 'L': 20, 'M': 21, 'N': 22, 'O': 35, 'P': 23,
    'Q': 24, 'R': 25, 'S': 26, 'T': 27, 'U': 28, 'V': 29, 'W': 32, 'X': 30,
    'Y': 31, 'Z': 33,
}


def validate_tw_id_checksum(id_no: str) -> bool:
    """台灣身分證校驗位。居留證第 2 碼為字母則僅驗格式、跳過 checksum。"""
    if not id_no or len(id_no) != 10:
        return False
    id_no = id_no.upper()
    first = id_no[0]
    if first not in _ID_LETTER_MAP:
        return False
    if not id_no[1].isdigit():
        return id_no[1] in _ID_LETTER_MAP and id_no[2:].isdigit()
    if not id_no[1:].isdigit():
        return False
    first_val = _ID_LETTER_MAP[first]
    digits = [first_val // 10, first_val % 10] + [int(d) for d in id_no[1:]]
    weights = [1, 9, 8, 7, 6, 5, 4, 3, 2, 1, 1]
    total = sum(d * w for d, w in zip(digits, weights))
    return total % 10 == 0


_PRIVATE_LOAN_SECTIONS = {"銀行", "零卡", "當舖", "當舖專案", "代書",
                           "商品貸", "鄉民", "房地", "新鑫", "慢點付", "分期趣", "銀角"}


def _is_private_loan_company(co: str) -> bool:
    """判斷是否民間方案（不需告知金額）：直接名稱、alias、section 都判"""
    if co in _PRIVATE_LOAN_SECTIONS:
        return True
    aliased = COMPANY_ALIAS.get(co)
    if aliased and aliased in _PRIVATE_LOAN_SECTIONS:
        return True
    return normalize_section(co) in _PRIVATE_LOAN_SECTIONS


_HELP_MENU_TITLE = "📖 指令速查 — 點下方按鈕看該類指令"

_HELP_SEND = """🔹 送件操作

【第一次送件】同時送多家 + 產生照會訊息
  打：@AI 王小明 裕融+第一 照會
  說明：系統會產生照會訊息給客戶，把兩家設成同送

━━━━━━━━━━
【轉】原本在送那家不送了、改送別家
  改一家：@AI 王小明 轉喬美
  帶金額：@AI 王小明 轉喬美 50萬/24期

【轉兩家】原本那家不送了、改成同送兩家
  打：@AI 王小明 轉喬美+房地
  帶金額：@AI 王小明 轉喬美+房地 100萬/120期

━━━━━━━━━━
【加送】原本那家保留、再多加一家
  加一家：@AI 王小明 送第一
  帶金額：@AI 王小明 送第一 30萬/24期

【加送兩家】原本保留、再多加兩家一起送
  打：@AI 王小明 送喬美+房地"""

_HELP_APPROVAL = """🔹 審核結果

【核准】某家核准了，系統自動移到日報「待撥款」區塊
  打：@AI 王小明 第一 核准 30萬

【結案】客戶整筆不送了（從日報拿掉）
  打：@AI 王小明 結案
  加原因：@AI 王小明 結案 已撥款

【只結案不送那幾家（留下核准）】
  情境：已經有一家核准，其他同送的不送了
  一家不送：@AI 王小明 第一 結案
  多家不送：@AI 王小明 第一/亞太 結案
  多家不送：@AI 王小明 第一+亞太 結案
  說明：只把那幾家從同送清單拿掉
       核准那家 + 客戶本身不動，日報還在待撥款

【婉拒】
  當前那家被婉拒、自動推下一家：
    @AI 王小明 婉拒
  當前那家被婉拒、跳到指定那家：
    @AI 王小明 婉拒轉亞太
  同送多家、指定某家被婉拒（不跳轉）：
    @AI 王小明 銀行 婉拒
    @AI 王小明 房地 婉拒
    情境：送房地+銀行，銀行評估不過 → 移除銀行、房地繼續送
  同送多家、指定某家被婉拒+跳指定家：
    @AI 王小明 裕融 婉拒 轉亞太
  婉拒 A + 馬上加送 B（複合指令）：
    @AI 王小明 喬美婉拒 送貸10
    情境：喬美不送了、改送貸救補
  補充：系列類用類別名即可
       銀行 ↔ 元大/渣打、房地 ↔ 房地一胎/二胎
       亞太 ↔ 亞太機車15萬/25萬/商品

  ⚠️ 重要：業務群用格式化訊息婉拒（如「3/4-王小明-房地婉拒」）
     只會被「記錄」、不會自動推下一家
     要推到下一家請打：@AI 王小明 婉拒

【撥款】
  有日期：@AI 王小明 裕融 撥款 4/18
  沒日期（= 今天）：@AI 王小明 裕融 撥款
  補充：只有一家核准時公司可以省略
       日期放前後都可以
       多家核准一定要指定哪家

【取消核准】核准打錯、要作廢那筆核准
  打：@AI 王小明 第一 取消核准
  說明：把那家從待撥款拿掉、核准金額清空

【違約金結案】核准後客戶放棄，收到違約金
  打：@AI 王小明 違約金已支付30000
  說明：狀態記成「違約金結案」，日報會把這筆移出待撥款"""

_HELP_UNDO = """🔹 做錯救急

【看歷史】先看這個確認編號
  打：@AI 王小明 歷史
  說明：列這客戶最近 10 筆動作、有編號

【還原】做錯了、回到上一步
  第一次先打：@AI 王小明 還原 1
  還不對再打：@AI 王小明 還原 2
  說明：一次只能退一步，不能直接跳還原 3 或 4
       還原 1 後如果不夠，再打還原 2 繼續退

【重啟】結案客戶要再送件
  打：@AI 王小明 重啟
  說明：客戶狀態從結案改回進行中，回到日報"""

_HELP_TOOLS = """🔹 其他小工具

【查詢類】
  查這客戶：@AI 查 王小明
  產日報：@AI 日報
  看統計：@AI 統計
    備註：今日/本月 進件、核准、結案數
  待撥款名單：@AI 待撥款
  查群組ID：@AI 群組ID

【修改類】
  改名：@AI 舊名 改名 新名
  改身分證：@AI 王小明 改身分證 A123456789

【送件小工具】推下一家
  打：@AI 王小明 轉下一家
  情境：送件順序已經設好（如 [裕融/第一/亞太]）
       當前那家不送了、自動跳到下一家

【批次操作】
  批次結案：
    @AI 批次結案
    王小明
    陳某某
  批次婉拒：
    @AI 批次婉拒
    王小明
    陳某某"""


_HELP_PAIRING = """🔹 對保流程（業務群內部操作，不用 @AI）

【派對保】行政/業務打給對保員
  範本：
    辦理方案：裕融
    核准金額：50萬
    客戶姓名：王小明
    對保地區：台北

  說明：
    系統會紀錄對保地區、客戶標「派對保」狀態

━━━━━━━━━━
【對保員接單 + 回時間地點】對保員打
  範本：
    對保 張三 新光對保
    時間：4/20 下午2點
    地點：台北 XX 路

  說明：
    第一行格式「對保 業務員 公司對保」
    時間、地點各一行
    系統自動紀錄

━━━━━━━━━━
【對保完成】對保員打
  訊息含以下任一：「對好」、「對保完成」、「不收不簽」
  例：王小明 對好
  例：王小明 對保完成

━━━━━━━━━━
【告知照會時段】業務告訴 A 群、客戶可接照會的時段
  範本：
    王小明 和裕補時段 11:00~12:00
  說明：
    系統會在日報顯示「照會時段 11:00~12:00」
    支援寫法：11:00~12:00、11點~12點、13:00到14:00
    這不是對保、是告訴 A 群什麼時候打電話給客戶"""

_HELP_NEWCASE = """🔹 開新客戶（這個不用加 @AI）

範本：
  115/4/17-董卓 A122999876
  房屋無貸款

說明：
  第一行（必填）：日期-姓名 身分證
  第二行（選填）：客戶備註

━━━━━━━━━━
🔹 一次設好送件順序（也不用 @AI）

範本：
  04/18-羅志祥-裕融/和潤/第一/亞太機/21機

說明：
  每家用「/」分開，系統依序送（失敗自動推下一家）"""

_HELP_SYMBOLS = """🔸 符號代表什麼

「/」有三種用途：
  1. 金額跟期數中間（50萬/24期）
  2. 日期月跟日中間（4/18）
  3. 送件順序公司分隔（裕融/和潤/第一）

「-」日期和姓名中間：
  開新客戶：115/4/17-董卓 A1...
  送件順序：04/18-羅志祥-裕融/...

「+」多家一起（同送/加送）：
  例：喬美+房地、裕融+第一

空格：姓名和動詞之間（例：王小明 轉 喬美）

━━━━━━━━━━
🔸 系統自動幫你（不用特別記）

・全形「１２３」→ 變半形「123」
・「核準／核准」、「身份證／身分證」都認
・「萬」「期」可省（100/24 = 100萬/24期）

━━━━━━━━━━
🔸 系統自動擋錯（看到警告是在幫你）

・結案客戶再操作 → 提醒重啟
・金額 0 / 超大 → 提醒確認
・身分證格式錯 → 警告
・5 秒內重複指令 → 忽略
・找不到客戶 → 不誤建怪客戶
・公司名打錯 → 列合法清單
・送件順序重複公司 → 自動去重
・本群組多位同名 → 跳按鈕選"""


_GUIDE_SHELL = """<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>指令速查 - 貸款系統</title>
<style>
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: -apple-system, "Segoe UI", "PingFang TC", "Microsoft JhengHei", sans-serif;
       background: #f5f5f0; color: #3a3229; line-height: 1.6; }
.wrap { max-width: 1100px; margin: 0 auto; padding: 20px 16px 60px; }
h1 { font-size: 22px; color: #4e7055; margin-bottom: 6px; }
.subtitle { color: #8a7a68; font-size: 13px; margin-bottom: 20px; }

.search-box { position: sticky; top: 0; background: #f5f5f0; padding: 10px 0; z-index: 10; border-bottom: 1px solid #e5ded4; }
.search-box input { width: 100%; padding: 10px 14px; border: 2px solid #d4c9b8; border-radius: 8px;
                    font-size: 15px; background: #fff; }
.search-box input:focus { outline: none; border-color: #4e7055; }

.nav-tabs { display: grid; grid-template-columns: repeat(4, 1fr); gap: 10px; margin: 14px 0; }
.nav-tab { padding: 14px 10px; background: #e5ded4; color: #4a3e30; border-radius: 10px;
           text-decoration: none; font-size: 14px; font-weight: 600; text-align: center;
           border: 1px solid transparent; transition: all 0.15s; }
.nav-tab:hover { background: #4e7055; color: #fff; transform: translateY(-1px); }
@media (max-width: 600px) {
  .nav-tabs { grid-template-columns: repeat(2, 1fr); }
  .nav-tab { padding: 16px 8px; font-size: 15px; }
}

.section { background: #fff; border-radius: 10px; padding: 20px; margin-bottom: 18px;
           border: 1px solid #e5ded4; scroll-margin-top: 80px; }
.section h2 { color: #4e7055; font-size: 18px; margin-bottom: 14px; padding-bottom: 8px;
              border-bottom: 2px solid #e5ded4; }

.help-raw { background: #fafaf5; border-left: 4px solid #4e7055; padding: 16px 20px;
            border-radius: 8px; font-family: "Consolas", "Menlo", "PingFang TC", monospace;
            font-size: 14px; line-height: 1.9; white-space: pre-wrap; color: #3a3229;
            overflow-x: auto; }

.copyable { background: #fff; padding: 2px 8px; border-radius: 4px;
            border: 1px solid #d4c9b8; cursor: pointer; color: #2d5016;
            display: inline-block; transition: all 0.15s; user-select: all; }
.copyable:hover { background: #4e7055; color: #fff; border-color: #4e7055; }
.copyable.copied { background: #166534; color: #fff; border-color: #166534; }
.copyable::after { content: " 📋"; opacity: 0.4; font-size: 11px; }
.copyable:hover::after, .copyable.copied::after { opacity: 1; }
.copyable.copied::after { content: " ✓"; }

.copy-block { display: inline-block; background: #fff; padding: 10px 14px;
              border-radius: 6px; border: 1px solid #d4c9b8; cursor: pointer;
              color: #2d5016; transition: all 0.15s; white-space: pre;
              font-family: inherit; margin: 2px 0; }
.copy-block:hover { border-color: #4e7055; background: #f0f5f0; }
.copy-block.copied { background: #dcfce7; border-color: #166534; }
.copy-block::before { content: "📋 點此整段複製 ↓"; display: block; font-size: 11px;
                      color: #8a7a68; margin-bottom: 6px; font-family: -apple-system, sans-serif; }
.copy-block.copied::before { content: "✓ 已複製"; color: #166534; font-weight: bold; }


.section.hidden { display: none; }
.back-link { color: #4e7055; text-decoration: none; font-size: 13px; }
.back-link:hover { text-decoration: underline; }
</style>
</head>
<body>
<div class="wrap">
  <a href="/report" class="back-link">← 回日報頁</a>
  <h1>📖 指令速查</h1>
  <p class="subtitle">業務 / 行政 操作 LINE Bot 的所有指令（和 LINE「@AI 說明」同一份內容）<br>
    <span style="color:#4e7055;">💡 小方框 = 單行指令、點一下複製；大方框 = 多行整段複製（派對保、批次結案...）</span></p>

  <div class="search-box">
    <input type="text" id="search" placeholder="🔍 搜尋指令或關鍵字（例：轉、核准、還原）">
  </div>

  <div class="nav-tabs">
    <a href="#send" class="nav-tab">📨 送件</a>
    <a href="#approval" class="nav-tab">✅ 核准/結案</a>
    <a href="#undo" class="nav-tab">🔙 做錯救急</a>
    <a href="#tools" class="nav-tab">🛠 其他工具</a>
    <a href="#newcase" class="nav-tab">🆕 開新客戶</a>
    <a href="#pairing" class="nav-tab">🤝 對保流程</a>
    <a href="#symbols" class="nav-tab">📝 符號說明</a>
  </div>

  __SECTIONS__
</div>

<script>
function _flash(el){
  el.classList.add("copied");
  setTimeout(function(){ el.classList.remove("copied"); }, 1500);
}
function copyInline(span){
  var text = (span.innerText || span.textContent).trim();
  navigator.clipboard.writeText(text).then(function(){ _flash(span); });
}
function copyBlock(el){
  var text = el.getAttribute("data-copy") || el.innerText;
  navigator.clipboard.writeText(text).then(function(){ _flash(el); });
}

var searchInput = document.getElementById("search");
if (searchInput) {
  searchInput.addEventListener("input", function(){
    var q = this.value.trim().toLowerCase();
    document.querySelectorAll(".section").forEach(function(sec){
      var text = sec.innerText.toLowerCase();
      if(!q || text.indexOf(q) >= 0){ sec.classList.remove("hidden"); }
      else { sec.classList.add("hidden"); }
    });
  });
}
</script>
</body>
</html>"""


def _process_help_to_html(text: str) -> str:
    """把 _HELP_X 原始字串轉成 HTML 片段（自動辨識 @AI 指令、範本區塊並加複製功能）。"""
    import html as _h
    import re as _re
    lines = text.split("\n")
    out = []
    i = 0
    n = len(lines)
    while i < n:
        line = lines[i]
        m_block = _re.match(r'^  (範本|批次結案|批次婉拒)：\s*$', line)
        if m_block:
            out.append(_h.escape(line))
            block_lines = []
            j = i + 1
            while j < n and lines[j].startswith("    "):
                block_lines.append(lines[j][4:])
                j += 1
            copy_text = "\n".join(block_lines).rstrip()
            attr_copy = _h.escape(copy_text, quote=True).replace("\n", "&#10;")
            out.append(
                '    <span class="copy-block" data-copy="'
                + attr_copy
                + '" onclick="copyBlock(this)">'
                + _h.escape(copy_text)
                + '</span>'
            )
            i = j
            continue
        if "@AI" in line:
            esc = _h.escape(line)
            esc = _re.sub(
                r'(@AI[^\n]+)',
                r'<span class="copyable" onclick="copyInline(this)">\1</span>',
                esc,
            )
            out.append(esc)
        else:
            m_ex = _re.match(r'^(  例：)(.+)$', line)
            if m_ex:
                out.append(
                    _h.escape(m_ex.group(1))
                    + '<span class="copyable" onclick="copyInline(this)">'
                    + _h.escape(m_ex.group(2))
                    + '</span>'
                )
            else:
                out.append(_h.escape(line))
        i += 1
    return "\n".join(out)


def _render_guide_html() -> str:
    """從 _HELP_* 動態渲染指令速查頁（和 LINE「@AI 說明」單一來源）。"""
    sections = [
        ("send", "📨 送件", _HELP_SEND),
        ("approval", "✅ 核准 / 結案", _HELP_APPROVAL),
        ("undo", "🔙 做錯救急", _HELP_UNDO),
        ("tools", "🛠 其他工具", _HELP_TOOLS),
        ("newcase", "🆕 開新客戶", _HELP_NEWCASE),
        ("pairing", "🤝 對保流程", _HELP_PAIRING),
        ("symbols", "📝 符號說明", _HELP_SYMBOLS),
    ]
    parts = []
    for sid, title, txt in sections:
        body = txt.split("\n", 1)[1] if "\n" in txt else txt
        body = body.lstrip("\n")
        parts.append(
            f'<section id="{sid}" class="section">\n'
            f'  <h2>{title}</h2>\n'
            f'  <pre class="help-raw">{_process_help_to_html(body)}</pre>\n'
            f'</section>'
        )
    return _GUIDE_SHELL.replace("__SECTIONS__", "\n".join(parts))


def _split_company_amount(item: str):
    """把「裕融100萬60期」或「裕融 100/60」等字串拆成 (公司, 金額, 期數)。
    用途：照會/送/轉指令切「+」後，每段可能有「公司+金額+期數」黏在一起。
    回傳：(company, amount, period) 全部為字串，無則 ""
    """
    if not item:
        return ("", "", "")
    item = item.strip()
    # 1) 整段就是合法公司名 → 直接回
    if item in COMPANY_ALIAS or item in COMPANY_LIST:
        return (item, "", "")
    # 2) 從合法公司名集合找最長前綴（避免「21機」被當「21」+「機」切）
    all_names = sorted(set(list(COMPANY_LIST) + list(COMPANY_ALIAS.keys())),
                       key=len, reverse=True)
    for name in all_names:
        if item.startswith(name):
            rest = item[len(name):].strip()
            if not rest:
                return (name, "", "")
            m = re.match(r"^(\d+(?:\.\d+)?)\s*萬?\s*[/／]?\s*(\d+)?\s*期?\s*$", rest)
            if m:
                return (name, m.group(1), m.group(2) or "")
            return (name, "", "")
    # 3) 沒命中合法公司 → 嘗試「第一個數字前」當公司
    m = re.search(r"\d", item)
    if m and m.start() > 0:
        co = item[:m.start()].strip()
        rest = item[m.start():].strip()
        m2 = re.match(r"^(\d+(?:\.\d+)?)\s*萬?\s*[/／]?\s*(\d+)?\s*期?\s*$", rest)
        if m2:
            return (co, m2.group(1), m2.group(2) or "")
        return (co, "", "")
    return (item, "", "")


def _build_plan_info_hint(companies, manual_amount="", manual_period=""):
    """產生「請告知客戶送件金額」提示字串。
    規則：
    - 民間方案（銀行/當舖/零卡/代書等）→ 不需告知金額
    - 有手動金額 → 所有公司都套用
    - 無手動金額 → 每家查 PLAN_INFO；無資料標「金額未定」
    - 全部都是民間方案 → 不顯示提醒
    """
    if not companies:
        return ""
    lines = []
    amt_str = per_str = ""
    if manual_amount:
        amt_str = manual_amount if "萬" in manual_amount else f"{manual_amount}萬"
        per_str = f"/{manual_period}期" if manual_period else ""
    for co in companies:
        if _is_private_loan_company(co):
            continue  # 民間方案不提金額
        if manual_amount:
            lines.append(f"  • {co} {amt_str}{per_str}（你指定）")
        else:
            plan = PLAN_INFO.get(co)
            if plan and plan[1]:
                lines.append(f"  • {co} {plan[1]}")
            else:
                lines.append(f"  • {co} 金額未定，請手動告知")
    if not lines:
        return ""  # 全部都是民間方案，不顯示
    return "💬 請告知客戶送件金額：\n" + "\n".join(lines)


def _validate_amount_or_warn(amount_str: str, reply_token, name: str, label: str = "金額"):
    """金額合理性：0 或負數 block；> 1000 萬警告 block 讓確認"""
    if not amount_str:
        return True
    m = re.search(r"(\d+(?:\.\d+)?)", amount_str)
    if not m:
        return True
    amt = float(m.group(1))
    if amt <= 0:
        reply_text(reply_token, f"❌ {name}：{label} 不能 = 0 或負數（你打的：{amount_str}），已阻擋")
        return False
    if amt > 1000:
        reply_text(reply_token,
                   f"⚠️ {name}：{label} 超過 1000 萬（你打的：{amount_str}）\n"
                   f"如果真的是這個數字，請打：@AI {name} {label} {amount_str} 確認")
        return False
    return True


# 重複訊息偵測（5 秒內相同訊息視為手滑）
_recent_msgs: Dict[str, float] = {}
_DUP_WINDOW_SEC = 5


def is_duplicate_message(group_id: str, content: str) -> bool:
    """5 秒內同群組收到一模一樣訊息 → True（caller 忽略）"""
    import time as _time
    now = _time.time()
    stale = [k for k, t in _recent_msgs.items() if t < now - _DUP_WINDOW_SEC]
    for k in stale:
        _recent_msgs.pop(k, None)
    key = f"{group_id}|{(content or '')[:200]}"
    if key in _recent_msgs:
        return True
    _recent_msgs[key] = now
    return False


def _resolve_target_strict(cmd: Dict, name: str, group_id: str, reply_token: str, action_label: str):
    """破壞性指令用：多筆同名跳按鈕讓使用者選。
    - cmd 若含 _forced_case_id（來自按鈕 callback），直接取該 case
    - 本群組 1 筆 → 回該筆
    - 本群組多筆 → 跳按鈕 EXEC_CMD，回傳 None
    - 本群組 0 筆、其他群組有 → 回最近更新那筆（非破壞性公司查詢情境）
    - 完全找不到 → 回覆錯誤
    """
    forced = cmd.get("_forced_case_id") if cmd else None
    if forced:
        conn = get_conn(); cur = conn.cursor()
        cur.execute("SELECT * FROM customers WHERE case_id=?", (forced,))
        c = cur.fetchone(); conn.close()
        return c
    rows = find_active_by_name(name)
    same = [r for r in rows if r["source_group_id"] == group_id]
    if len(same) == 1:
        return same[0]
    if len(same) > 1:
        action_id = short_id()
        # 存 cmd 副本（不含 _forced_case_id）到 pending_action
        cmd_copy = {k: v for k, v in cmd.items() if k != "_forced_case_id"}
        save_pending_action(action_id, "select_case_for_cmd", {
            "cmd": cmd_copy, "group_id": group_id,
            "case_ids": [r["case_id"] for r in same], "label": action_label
        })
        items = []
        for r in same[:12]:
            id4 = (r["id_no"] or "")[-4:] or "無"
            co = r["current_company"] or r["company"] or "未填"
            items.append(make_quick_reply_item(
                f"{co}-末4:{id4}", f"EXEC_CMD|{action_id}|{r['case_id']}"))
        items.append(make_quick_reply_item("取消", f"CANCEL_CMD|{action_id}"))
        reply_quick_reply(reply_token, f"⚠️ 本群組有 {len(same)} 位「{name}」，要{action_label}哪一位？", items)
        return None
    if rows:
        return rows[0]
    reply_text(reply_token,
               f"❌ 找不到客戶：{name}\n"
               f"可能原因：\n"
               f"  • 姓名打錯（2-6 個中文字）\n"
               f"  • 客戶還沒建立\n"
               f"  • 客戶已結案 → 先打「@AI {name} 重啟」")
    return None


def handle_special_command(cmd: Dict, reply_token: str, group_id: str):
    try:
        _handle_special_command_inner(cmd, reply_token, group_id)
    except Exception as e:
        import traceback
        traceback.print_exc()
        reply_text(reply_token, f"❌ 系統處理失敗，請截圖給管理員\n錯誤代碼：{type(e).__name__}")


def _handle_special_command_inner(cmd: Dict, reply_token: str, group_id: str):
    t = cmd["type"]

    if t == "group_id":
        gname = get_group_name(group_id)
        reply_text(reply_token, f"📋 此群組資訊\n名稱：{gname}\nID：{group_id}")
        return

    if t == "report":
        try:
            segs = generate_report_lines(group_id)
            reply_text(reply_token, segs[0])
            for seg in segs[1:]:
                push_text(group_id, seg)
        except Exception as e:
            import traceback
            traceback.print_exc()
            reply_text(reply_token, f"❌ 日報產生失敗，請截圖給管理員\n錯誤代碼：{type(e).__name__}")
        return

    if t == "search":
        reply_text(reply_token, search_customer_info(cmd["name"], group_id))
        return

    if t == "format_hint":
        # 通用格式提示 handler（給常見錯誤格式友善提示）
        hint = cmd.get("hint", "指令格式錯，請看 @AI 說明")
        reply_text(reply_token, f"⚠️ {hint}")
        return

    if t == "bad_batch_format":
        action = cmd.get("action", "結案")
        reply_text(reply_token,
                   f"⚠️ 批次{action}格式不對\n"
                   f"要這樣打（每個姓名換一行）：\n"
                   f"  @AI 批次{action}\n"
                   f"  黎明\n"
                   f"  王陽明")
        return

    if t == "batch_close":
        names = cmd["names"]
        results = []
        for name in names:
            rows = find_active_by_name(name)
            same = [r for r in rows if r["source_group_id"] == group_id]
            target = same[0] if same else (rows[0] if rows else None)
            if not target:
                results.append(f"  {name} ❌ 找不到客戶")
            else:
                update_customer(target["case_id"], status="CLOSED",
                                text=f"{name} 結案", from_group_id=group_id)
                push_text(target["source_group_id"], f"{name} 結案")
                results.append(f"  {name} ✅ 已結案")
        ok = sum(1 for r in results if "✅" in r)
        fail = len(results) - ok
        header = f"📋 批次結案 {len(names)} 筆（成功 {ok}"
        if fail:
            header += f"，失敗 {fail}"
        header += "）"
        reply_text(reply_token, header + "\n" + "\n".join(results))
        return

    if t == "batch_reject":
        names = cmd["names"]
        results = []
        for name in names:
            rows = find_active_by_name(name)
            same = [r for r in rows if r["source_group_id"] == group_id]
            target = same[0] if same else (rows[0] if rows else None)
            if not target:
                results.append(f"  {name} ❌ 找不到客戶")
            else:
                route = target["route_plan"] or ""
                current = get_current_company(route)
                next_co = get_next_company(route)
                new_route = advance_route(route, "婉拒")
                update_customer(target["case_id"], route_plan=new_route,
                                current_company=next_co or current,
                                text=f"{name} {current} 婉拒", from_group_id=group_id)
                push_text(target["source_group_id"], f"{name} {current} 婉拒")
                if next_co:
                    results.append(f"  {name} ✅ {current}婉拒 → {next_co}")
                else:
                    results.append(f"  {name} ✅ {current}婉拒（無下一家）")
        ok = sum(1 for r in results if "✅" in r)
        fail = len(results) - ok
        header = f"📋 批次婉拒 {len(names)} 筆（成功 {ok}"
        if fail:
            header += f"，失敗 {fail}"
        header += "）"
        reply_text(reply_token, header + "\n" + "\n".join(results))
        return

    if t == "pending_disbursement":
        conn = get_conn(); cur = conn.cursor()
        if group_id == A_GROUP_ID:
            cur.execute("SELECT customer_name, current_company, company, approved_amount, route_plan, created_at, source_group_id FROM customers WHERE status='ACTIVE' AND report_section='待撥款' ORDER BY created_at")
        else:
            cur.execute("SELECT customer_name, current_company, company, approved_amount, route_plan, created_at, source_group_id FROM customers WHERE status='ACTIVE' AND report_section='待撥款' AND source_group_id=? ORDER BY created_at", (group_id,))
        rows = cur.fetchall(); conn.close()
        if not rows:
            reply_text(reply_token, "📋 目前沒有待撥款客戶")
            return
        lines = [f"📋 待撥款名單（{len(rows)} 筆）\n"]
        for r in rows:
            name = r["customer_name"]
            co = r["current_company"] or r["company"] or ""
            amt = r["approved_amount"] or ""
            created = (r["created_at"] or "")[:10]
            approved_list = get_all_approved(r["route_plan"] or "")
            if approved_list:
                parts = [(h.get("company") or "") + (h.get("amount") or "") for h in approved_list]
                amt_str = "/".join(parts)
            elif amt:
                amt_str = amt
            else:
                amt_str = "未知金額"
            gname = get_group_name(r["source_group_id"]) if group_id == A_GROUP_ID else ""
            grp_tag = f"({gname})" if gname else ""
            lines.append(f"{created}-{name}{grp_tag}-核准{amt_str}")
        reply_text(reply_token, "\n".join(lines))
        return

    if t == "stats":
        conn = get_conn(); cur = conn.cursor()
        today = now_iso()[:10]
        month_start = today[:7] + "-01"
        is_a = group_id == A_GROUP_ID
        grp_filter = "" if is_a else " AND source_group_id=?"
        grp_param = [] if is_a else [group_id]
        # 今日
        cur.execute(f"SELECT COUNT(*) as c FROM customers WHERE date(created_at)=?{grp_filter}", [today] + grp_param)
        today_new = cur.fetchone()["c"]
        cur.execute(f"SELECT COUNT(*) as c FROM customers WHERE status IN ('CLOSED','PENALTY','ABANDONED','REJECTED') AND date(updated_at)=?{grp_filter}", [today] + grp_param)
        today_closed = cur.fetchone()["c"]
        # 本月
        cur.execute(f"SELECT COUNT(*) as c FROM customers WHERE created_at>=?{grp_filter}", [month_start] + grp_param)
        month_new = cur.fetchone()["c"]
        cur.execute(f"SELECT COUNT(*) as c FROM customers WHERE status IN ('CLOSED','PENALTY','ABANDONED','REJECTED') AND updated_at>=?{grp_filter}", [month_start] + grp_param)
        month_closed = cur.fetchone()["c"]
        # 核准
        cur.execute(f"SELECT route_plan FROM customers WHERE status='ACTIVE' AND route_plan IS NOT NULL AND route_plan!=''{grp_filter}", grp_param)
        all_routes = cur.fetchall()
        today_approved = 0
        month_approved = 0
        for r in all_routes:
            approved = get_all_approved(r["route_plan"])
            if approved:
                for a in approved:
                    d = a.get("date", "")
                    if d == today:
                        today_approved += 1
                    if d >= month_start:
                        month_approved += 1
        # 待撥款
        cur.execute(f"SELECT COUNT(*) as c FROM customers WHERE status='ACTIVE' AND report_section='待撥款'{grp_filter}", grp_param)
        total_pending = cur.fetchone()["c"]
        cur.execute(f"SELECT COUNT(*) as c FROM customers WHERE status='ACTIVE'{grp_filter}", grp_param)
        total_active = cur.fetchone()["c"]
        conn.close()
        msg = (f"📊 統計資訊\n\n"
               f"【今日 {today}】\n"
               f"  新進件：{today_new}\n"
               f"  核准：{today_approved}\n"
               f"  結案：{today_closed}\n\n"
               f"【本月】\n"
               f"  新進件：{month_new}\n"
               f"  核准：{month_approved}\n"
               f"  結案：{month_closed}\n\n"
               f"目前活躍：{total_active}　待撥款：{total_pending}")
        reply_text(reply_token, msg)
        return

    if t == "rename":
        old_name = cmd["old_name"]
        new_name = cmd["new_name"]
        rows = find_active_by_name(old_name)
        if not rows:
            # 也找結案客戶
            conn2 = get_conn(); cur2 = conn2.cursor()
            cur2.execute("SELECT * FROM customers WHERE customer_name=? ORDER BY updated_at DESC LIMIT 1", (old_name,))
            r = cur2.fetchone(); conn2.close()
            if r: rows = [r]
        same = [r for r in rows if r["source_group_id"] == group_id]
        target = same[0] if same else (rows[0] if rows else None)
        if not target:
            reply_text(reply_token, f"❌ 找不到客戶：{old_name}"); return
        conn = get_conn(); cur = conn.cursor()
        cur.execute("UPDATE customers SET customer_name=?, updated_at=? WHERE case_id=?",
                    (new_name, now_iso(), target["case_id"]))
        conn.commit(); conn.close()
        # 寫 case_log
        update_customer(target["case_id"],
                        text=f"{old_name} 改名為 {new_name}",
                        from_group_id=group_id)
        reply_text(reply_token, f"✅ 已更新姓名：「{old_name}」→「{new_name}」")
        return

    if t == "bad_id_format":
        name = cmd["name"]
        raw = cmd["raw_id"]
        reply_text(reply_token,
                   f"⚠️ 身分證「{raw}」格式不對\n"
                   f"要這樣打：A123456789（1 個英文字母 + 9 個數字 = 共 10 位）\n"
                   f"例：@AI {name} 改身分證 A123456789")
        return

    if t == "change_id":
        name = cmd["name"]
        new_id = cmd["new_id"]
        # 改身分證允許結案客戶（跟 rename 一致）
        rows = find_active_by_name(name)
        if not rows:
            conn2 = get_conn(); cur2 = conn2.cursor()
            cur2.execute("SELECT * FROM customers WHERE customer_name=? ORDER BY updated_at DESC LIMIT 1", (name,))
            r = cur2.fetchone(); conn2.close()
            if r: rows = [r]
        same = [r for r in rows if r["source_group_id"] == group_id]
        target = same[0] if same else (rows[0] if rows else None)
        if not target:
            reply_text(reply_token,
                       f"❌ 找不到客戶：{name}\n"
                       f"可能原因：姓名打錯 / 客戶還沒建立"); return
        old_id = target["id_no"] or "無"
        conn = get_conn(); cur = conn.cursor()
        cur.execute("UPDATE customers SET id_no=?, updated_at=? WHERE case_id=?",
                    (new_id, now_iso(), target["case_id"]))
        conn.commit(); conn.close()
        update_customer(target["case_id"],
                        text=f"{name} 改身分證 {old_id} → {new_id}",
                        from_group_id=group_id)
        reply_text(reply_token, f"✅ {name} 身分證已更新為 {new_id}\n（原 {old_id} → 新 {new_id}）")
        return

    if t == "disbursed":
        name = cmd["name"]
        co_raw = cmd.get("company", "")
        disb_date = cmd["date"] or datetime.now().strftime("%-m/%-d") if os.name != "nt" else cmd["date"] or datetime.now().strftime("%#m/%#d")
        target = _resolve_target_strict(cmd, name, group_id, reply_token, "撥款")
        if not target:
            return
        if not _check_active_or_warn(target, reply_token, "撥款", name):
            return
        # 查所有核准家
        all_approved = get_all_approved(target["route_plan"] or "")
        approved_cos = [a.get("company", "") for a in all_approved if a.get("company")]
        # 判斷撥哪家
        if co_raw:
            # 有指定公司
            co = COMPANY_ALIAS.get(co_raw, co_raw)
            if not _validate_companies_or_warn([co], reply_token, name):
                return
            if co not in approved_cos:
                lst = "、".join([f'{a.get("company","")} {a.get("amount","")}' for a in all_approved]) or "無"
                reply_text(reply_token,
                           f"⚠️ {name} 的 {co} 沒核准、不能撥款\n"
                           f"已核准：{lst}")
                return
            company = co
        else:
            # 沒指定公司
            if not approved_cos:
                reply_text(reply_token, f"⚠️ {name} 還沒核准、不能撥款")
                return
            if len(approved_cos) > 1:
                lst = "、".join([f'{a.get("company","")} {a.get("amount","")}' for a in all_approved])
                reply_text(reply_token,
                           f"⚠️ {name} 有多家核准（{lst}），要撥哪家？\n"
                           f"請打：@AI {name} [公司] 撥款 {disb_date}")
                return
            company = approved_cos[0]
        # 撥款日期合理性檢查（未來日期/早於建案日期）
        disb_warnings = []
        try:
            dm = re.match(r"(\d{1,2})/(\d{1,2})", disb_date)
            if dm:
                now = datetime.now()
                month, day = int(dm.group(1)), int(dm.group(2))
                try:
                    disb_dt = datetime(now.year, month, day)
                    if (disb_dt - now).days > 30:
                        disb_warnings.append(f"撥款日 {disb_date} 超過 30 天後，是否打錯月份？")
                    if target["created_at"]:
                        try:
                            created_dt = datetime.fromisoformat(str(target["created_at"])[:19])
                            if disb_dt.date() < created_dt.date():
                                disb_warnings.append(f"撥款日 {disb_date} 早於建案日 {created_dt.strftime('%m/%d')}")
                        except Exception:
                            pass
                except ValueError:
                    disb_warnings.append(f"撥款日 {disb_date} 日期無效（例如 2/30）")
        except Exception:
            pass
        # 更新撥款日期到 route_plan history 和 disbursement_date
        new_route = set_disbursed_in_history(target["route_plan"] or "", company, disb_date)
        update_customer(target["case_id"], disbursement_date=disb_date,
                        route_plan=new_route, report_section="待撥款",
                        signing_area="", signing_salesperson="",
                        signing_company="", signing_time="", signing_location="",
                        text=f"{name} {company} 撥款{disb_date}",
                        from_group_id=group_id)
        msg = f"✅ {name} {company} 已撥款：{disb_date}"
        if disb_warnings:
            msg += "\n⚠️ " + "；".join(disb_warnings)
        reply_text(reply_token, msg)
        return

    if t == "signing_request":
        name = cmd["name"]
        area = cmd["area"]
        rows = find_active_by_name(name)
        same = [r for r in rows if r["source_group_id"] == group_id]
        target = same[0] if same else (rows[0] if rows else None)
        if not target:
            reply_text(reply_token, f"❌ 找不到客戶：{name}"); return
        update_customer(target["case_id"],
                        signing_area=area,
                        text=f"{name} 派對保（地區：{area}）",
                        from_group_id=group_id)
        reply_text(reply_token, f"✅ {name} 已記錄派對保：{area}")
        return

    if t == "signing_schedule":
        conn = get_conn(); cur = conn.cursor()
        cur.execute(
            "SELECT * FROM customers WHERE status='ACTIVE' AND source_group_id=? "
            "AND signing_area IS NOT NULL AND signing_area!='' "
            "AND (signing_time IS NULL OR signing_time='') "
            "ORDER BY updated_at DESC LIMIT 1",
            (group_id,))
        target = cur.fetchone(); conn.close()
        if not target:
            reply_text(reply_token,
                       "❌ 找不到待對保客戶\n"
                       "請先派對保（範本）：\n"
                       "  辦理方案：裕融\n"
                       "  核准金額：50萬\n"
                       "  客戶姓名：王小明\n"
                       "  對保地區：台北"); return
        update_customer(target["case_id"],
                        signing_salesperson=cmd["salesperson"],
                        signing_company=cmd["signing_company"],
                        signing_time=cmd["time"],
                        signing_location=cmd["location"],
                        text=f"{target['customer_name']} 對保 {cmd['salesperson']} {cmd['signing_company']} {cmd['time']} {cmd['location']}",
                        from_group_id=group_id)
        reply_text(reply_token,
                   f"✅ {target['customer_name']} 對保已記錄\n"
                   f"  業務：{cmd['salesperson']}\n"
                   f"  公司：{cmd['signing_company']}\n"
                   f"  時間：{cmd['time']}\n"
                   f"  地點：{cmd['location']}")
        return

    if t == "help":
        items = [
            make_quick_reply_item("📨 送件", "HELP|send"),
            make_quick_reply_item("✅ 核准/結案", "HELP|approval"),
            make_quick_reply_item("🔙 做錯救急", "HELP|undo"),
            make_quick_reply_item("🛠 其他工具", "HELP|tools"),
            make_quick_reply_item("🆕 開新客戶", "HELP|newcase"),
            make_quick_reply_item("🤝 對保流程", "HELP|pairing"),
            make_quick_reply_item("📝 符號說明", "HELP|symbols"),
        ]
        reply_quick_reply(reply_token, _HELP_MENU_TITLE, items)
        return


    if t == "history":
        name = cmd["name"]
        target = _resolve_target_strict(cmd, name, group_id, reply_token, "查歷史")
        if not target:
            return
        conn = get_conn(); cur = conn.cursor()
        cur.execute("""SELECT id, created_at, message_text FROM case_logs
                       WHERE case_id=? ORDER BY id DESC LIMIT 10""",
                    (target["case_id"],))
        rows = cur.fetchall(); conn.close()
        if not rows:
            reply_text(reply_token, f"📋 {name} 尚無操作紀錄"); return
        lines = [f"📋 {name} 最近操作（最新在上）："]
        for i, r in enumerate(rows, 1):
            ts = (r["created_at"] or "")[5:16].replace("T", " ")  # MM-DD HH:MM
            first_line = ((r["message_text"] or "").splitlines() or [""])[0]
            msg = first_line[:30]
            lines.append(f"[{i}] {ts} - {msg}")
        lines.append("\n要還原：@AI 姓名 還原 N（N = 編號）")
        lines.append("例：@AI " + name + " 還原 1  → 回到第 1 筆之前的狀態")
        reply_text(reply_token, "\n".join(lines))
        return

    if t == "restore":
        name = cmd["name"]
        idx = int(cmd.get("index", 1) or 1)
        if idx < 1:
            reply_text(reply_token, "❌ 編號要 1 以上\n例：@AI 王小明 還原 1（回到最近一次動作之前）"); return
        target = _resolve_target_strict(cmd, name, group_id, reply_token, "還原")
        if not target:
            return
        case_id = target["case_id"]
        conn = get_conn(); cur = conn.cursor()
        cur.execute("""SELECT id, created_at, snapshot_json, message_text FROM case_logs
                       WHERE case_id=? ORDER BY id DESC LIMIT ?""",
                    (case_id, idx))
        rows = cur.fetchall(); conn.close()
        if len(rows) < idx:
            reply_text(reply_token,
                       f"❌ {name} 只有 {len(rows)} 筆紀錄，無法還原到第 {idx} 筆\n"
                       f"請先打「@AI {name} 歷史」確認編號")
            return
        target_log = rows[idx - 1]
        snapshot_json = target_log["snapshot_json"]
        if not snapshot_json:
            reply_text(reply_token,
                       f"❌ 第 {idx} 筆沒有備份資料\n"
                       f"（舊紀錄沒存備份，只有近期動作才能還原）")
            return
        try:
            snapshot = json.loads(snapshot_json)
        except Exception:
            reply_text(reply_token, "❌ 備份資料損毀，無法還原")
            return
        # 套回快照欄位
        ok_v, diffs, cust_name = update_with_verify(
            case_id, snapshot, from_group_id=group_id,
            text_log=f"{name} 還原到第 {idx} 筆之前")
        if not ok_v:
            reply_text(reply_token, "⚠️ 案件不存在，無法還原")
            return
        ts = (target_log["created_at"] or "")[5:16].replace("T", " ")
        if diffs:
            msg = f"✅ {cust_name} 已還原到 {ts} 之前的狀態\n" + "\n".join(diffs)
        else:
            msg = f"ℹ️ {cust_name}：還原完成，但狀態無實際變動（可能已經是這個版本）"
        reply_text(reply_token, msg)
        return

    if t == "reopen":
        name = cmd["name"]
        conn = get_conn(); cur = conn.cursor()
        cur.execute("SELECT * FROM customers WHERE customer_name=? AND status IN ('CLOSED','PENALTY','ABANDONED','REJECTED') ORDER BY updated_at DESC LIMIT 1", (name,))
        target = cur.fetchone(); conn.close()
        if not target:
            reply_text(reply_token, f"❌ 找不到已結案客戶：{name}"); return
        update_customer(target["case_id"], status="ACTIVE",
                        report_section="",
                        text=f"{name} 重啟案件",
                        from_group_id=group_id)
        reply_text(reply_token, f"✅ {name} 已重啟，恢復到日報")
        return

    if t == "cancel_approval":
        name = cmd["name"]
        company = cmd["company"].strip()
        if not company:
            reply_text(reply_token,
                       f"⚠️ 要取消哪家的核准？\n"
                       f"例：@AI {name} 第一 取消核准")
            return
        target = _resolve_target_strict(cmd, name, group_id, reply_token, "取消核准")
        if not target:
            return
        if not _check_active_or_warn(target, reply_token, "取消核准", name):
            return
        # 從 route_plan history 移除該公司的核准記錄
        route = target["route_plan"] or ""
        data = parse_route_json(route) if route else {"order":[], "current_index":0, "history":[]}
        history = data.get("history", [])
        new_history = []
        removed = False
        for h2 in history:
            hc = h2.get("company", "")
            if (company in hc or hc in company) and h2.get("status") in ("核准", "待撥款", "撥款"):
                removed = True
                continue
            new_history.append(h2)
        data["history"] = new_history
        new_route = json.dumps(data, ensure_ascii=False)
        # 判斷是否還有其他核准 → 沒有就從待撥款區塊移除
        still_approved = any(h2.get("status") in ("核准", "待撥款", "撥款") and h2.get("amount") for h2 in new_history)
        new_section = "待撥款" if still_approved else ""
        new_amount = "" if not still_approved else target["approved_amount"]
        update_customer(target["case_id"], route_plan=new_route,
                        approved_amount=new_amount or None,
                        report_section=new_section,
                        text=f"{name} {company} 取消核准", from_group_id=group_id)
        push_text(target["source_group_id"], f"{name} {company} 取消核准")
        base = f"✅ {name} 已取消 {company} 核准"
        if still_approved:
            base += "（仍有其他核准保留）"
        else:
            base += "（已從待撥款移除）"
        reply_text(reply_token,
                   f"{base}\n\n"
                   f"⚠️ 但 {company} 還在這客戶的送件清單裡、日報還會顯示\n\n"
                   f"接下來怎麼選：\n"
                   f"  【情況 A】{company} 不送了\n"
                   f"    打 → @AI {name} {company} 結案\n"
                   f"  【情況 B】{company} 還要繼續送（例如要談新金額）\n"
                   f"    不用動作、日報繼續顯示、等新核准")
        return

    if t == "update_amount":
        name = cmd["name"]
        company = cmd["company"]
        amount = cmd["amount"]
        target = _resolve_target_strict(cmd, name, group_id, reply_token, "核准金額")
        if not target:
            return
        if not _check_active_or_warn(target, reply_token, "更新核准金額", name):
            return
        if not _validate_amount_or_warn(amount, reply_token, name, "核准金額"):
            return
        # 未指定公司 → 用當前 current_company（防錯：使用者常忘了打公司）
        if not company:
            company = target["current_company"] or target["company"] or ""
        if not company:
            reply_text(reply_token, f"❌ {name} 找不到當前送件公司，請明確指定：{name} 公司 核准 {amount}")
            return
        # 套 COMPANY_ALIAS 規範化（例：「貸10」→「貸救補」）
        company = COMPANY_ALIAS.get(company, company)
        route = target["route_plan"] or ""
        new_route = update_company_amount_in_history(route, company, amount)
        # 核准金額更新時，一律移到「待撥款」區塊（房地/當鋪/C 等都一致）
        update_customer(target["case_id"], route_plan=new_route,
                        approved_amount=amount,
                        report_section="待撥款",
                        text=f"{name} {company} 核准金額修改為 {amount}",
                        from_group_id=group_id)
        reply_text(reply_token, f"✅ {name} {company} 核准金額已更新為 {amount}，已移到待撥款")
        return

    if t == "missing_date_hint":
        name = cmd["name"]
        tail = cmd["tail"]
        today = datetime.now().strftime("%#m/%#d") if os.name == "nt" else datetime.now().strftime("%-m/%-d")
        reply_text(reply_token,
                   f"⚠️ 你是不是忘了日期？\n"
                   f"送件順序不用 @AI，格式是：M/D-姓名-公司/公司/...\n"
                   f"直接貼這段就好：\n"
                   f"  {today}-{name}-{tail}")
        return

    if t == "no_space_hint":
        raw = cmd["raw"]
        reply_text(reply_token,
                   f"⚠️ 打太黏看不懂，請加空白分開：\n"
                   f"  你打的：{raw}\n"
                   f"  要這樣：王小明 轉喬美 100萬/24期")
        return

    if t == "bad_amount_format":
        name = cmd["name"]
        verb = cmd["verb"]
        target = cmd["target"]
        n1 = cmd["n1"]
        n2 = cmd["n2"]
        reply_text(reply_token,
                   f"⚠️ 金額和期數中間要加「/」\n"
                   f"  要這樣：{name} {verb}{target} {n1}萬/{n2}期\n"
                   f"  你打的：{name} {verb}{target} {n1} {n2}")
        return

    if t == "missing_verb":
        name = cmd["name"]
        companies_raw = cmd.get("companies_raw", "")
        amount = cmd.get("amount", "")
        period = cmd.get("period", "")
        cos_list = [c.strip() for c in re.split(r"[/+＋,，]", companies_raw) if c.strip()]
        # 情況 1：第一個字是動詞（例：王小明 撥款 100萬/24期）
        VERB_HINT = {
            "撥款": f"@AI {name} 撥款 4/19  ← M/D 日期",
            "結案": f"@AI {name} 結案",
            "婉拒": f"@AI {name} 婉拒",
            "核准": f"@AI {name} 第一 核准 30萬  ← 先打公司再打金額",
            "重啟": f"@AI {name} 重啟",
            "歷史": f"@AI {name} 歷史",
            "還原": f"@AI {name} 還原 1",
            "改名": f"@AI {name} 改名 新名字",
            "取消核准": f"@AI {name} 第一 取消核准",
        }
        if cos_list and cos_list[0] in VERB_HINT:
            verb = cos_list[0]
            reply_text(reply_token,
                       f"⚠️ 「{verb}」不是公司名、是動作\n"
                       f"你要的是：{VERB_HINT[verb]}")
            return
        # 情況 2：公司名不合法
        cos_canon = [COMPANY_ALIAS.get(c, c) for c in cos_list]
        valid = _get_valid_company_names()
        unknown = [c for c in cos_canon if c not in valid]
        if unknown:
            reply_text(reply_token,
                       f"⚠️ 找不到公司：{'、'.join(unknown)}\n"
                       f"是不是打錯 / 漏寫公司名？\n"
                       f"常見：亞太、喬美、第一、房地、21、裕融、和裕、麻吉、貸救補、鄉民、銀行、零卡、商品貸、代書、當舖")
            return
        # 情況 3：公司名都合法 → 缺動詞
        reply_text(reply_token,
                   f"⚠️ 你沒說要做什麼，請加動詞：\n"
                   f"  • 原本在送的不送了、改送這家 → {name} 轉{companies_raw} {amount}萬/{period}期\n"
                   f"  • 原本那家保留、再加一家 → {name} 送{companies_raw} {amount}萬/{period}期\n"
                   f"  • 這家已核准 → {name} {companies_raw} 核准 {amount}萬")
        return

    if t == "close":
        name = cmd["name"]
        reason = cmd.get("reason", "")
        target = _resolve_target_strict(cmd, name, group_id, reply_token, "結案")
        if not target:
            return
        if not _check_active_or_warn(target, reply_token, "結案", name):
            return
        close_text = f"{name} 結案（{reason}）" if reason else f"{name} 結案"
        update_customer(target["case_id"], status="CLOSED",
                        text=close_text, from_group_id=group_id)
        push_text(target["source_group_id"], close_text)
        reply_text(reply_token, f"✅ {name} 已結案，從日報移除" + (f"\n原因：{reason}" if reason else ""))
        return

    if t == "remove_concurrent_or_close":
        # 【A 方案】姓名 公司 結案（支援多家：第一/亞太、第一+亞太 等）
        # 客戶已核准 → 從同送清單移除這些公司（不動客戶狀態）
        # 客戶未核准 → 退回普通結案並提示
        name = cmd["name"]
        co_raw = cmd["company"]
        target = _resolve_target_strict(cmd, name, group_id, reply_token, "結案")
        if not target:
            return
        if not _check_active_or_warn(target, reply_token, "結案", name):
            return
        cos_list = [c.strip() for c in re.split(r"[/+＋,，]", co_raw) if c.strip()]
        cos = [COMPANY_ALIAS.get(c, c) for c in cos_list]
        if not _validate_companies_or_warn(cos, reply_token, name):
            return
        approved = (target["approved_amount"] or "").strip()
        current_co = target["current_company"] or ""
        concurrent_str = target["concurrent_companies"] or ""
        concurrent_list = [c.strip() for c in concurrent_str.split(",") if c.strip()]
        if not approved:
            close_text = f"{name} 結案"
            update_customer(target["case_id"], status="CLOSED",
                            text=close_text, from_group_id=group_id)
            push_text(target["source_group_id"], close_text)
            reply_text(reply_token,
                       f"⚠️ {name} 尚未核准，「{co_raw} 結案」已視為整筆結案\n"
                       f"若要只取消某家送件，請先該家核准後再打此指令")
            return
        removed = []
        skipped_current = []
        skipped_not_in = []
        new_concurrent = list(concurrent_list)
        # 用 normalize_section 做模糊比對：「亞太」對應「亞太機車15萬/25萬、亞太商品」等同系列
        current_norm = normalize_section(current_co) if current_co else ""
        for co in cos:
            co_norm = normalize_section(co)
            if co_norm and co_norm == current_norm:
                skipped_current.append(co)
                continue
            matches = [c for c in new_concurrent if normalize_section(c) == co_norm]
            if matches:
                for m in matches:
                    new_concurrent.remove(m)
                    removed.append(m)
            else:
                skipped_not_in.append(co)
        if removed:
            text_note = f"{name} 不送：{'、'.join(removed)}（留 {current_co} 核准 {approved}）"
            update_customer(target["case_id"],
                            concurrent_companies=",".join(new_concurrent),
                            text=text_note, from_group_id=group_id)
            push_text(target["source_group_id"], text_note)
        msgs = []
        if removed:
            msgs.append(f"✅ 已從同送清單移除：{'、'.join(removed)}")
        if skipped_current:
            msgs.append(f"⚠️ {'、'.join(skipped_current)} 是核准那家，沒處理\n　 （作廢核准用：@AI {name} {skipped_current[0]} 取消核准）")
        if skipped_not_in:
            msgs.append(f"⚠️ 同送清單裡沒有：{'、'.join(skipped_not_in)}")
        msgs.append(f"核准保留：{current_co} {approved}")
        msgs.append(f"還在送：{'、'.join(new_concurrent) if new_concurrent else '無'}")
        reply_text(reply_token, "\n".join(msgs))
        return

    if t == "reject":
        name = cmd["name"]
        target = _resolve_target_strict(cmd, name, group_id, reply_token, "婉拒")
        if not target:
            return
        if not _check_active_or_warn(target, reply_token, "婉拒", name):
            return
        route = target["route_plan"] or ""
        current, next_co = get_current_company(route), get_next_company(route)
        new_route = advance_route(route, "婉拒")
        concurrent_list = [c.strip() for c in (target["concurrent_companies"] or "").split(",") if c.strip()]
        # 決定新 current：
        # 1. route 有下一家 → 用 route 下一家
        # 2. route 沒下一家、concurrent 有家 → 從 concurrent 第一家升上來（避免婉拒那家殘留）
        # 3. 都沒有 → current 清空（日報不顯示原那家）
        if next_co:
            new_current = next_co
            new_concurrent = concurrent_list
            promoted_from_concurrent = ""
        elif concurrent_list:
            new_current = concurrent_list[0]
            new_concurrent = concurrent_list[1:]
            promoted_from_concurrent = new_current
        else:
            new_current = ""
            new_concurrent = concurrent_list
            promoted_from_concurrent = ""
        update_kw = {"route_plan": new_route, "current_company": new_current,
                     "concurrent_companies": ",".join(new_concurrent),
                     "text": f"{name} {current} 婉拒", "from_group_id": group_id}
        if not new_current:
            # 已全數婉拒：清 company 避免日報 fallback 顯示被婉拒那家
            update_kw["company"] = ""
        update_customer(target["case_id"], **update_kw)
        # 回貼業務群
        push_msg = f"{name} {current} 婉拒"
        if new_current:
            push_msg += f"\n➡️ 下一家：{new_current}"
        push_text(target["source_group_id"], push_msg)
        if new_current:
            extra = f"（從同送清單升上來）" if promoted_from_concurrent else ""
            reply_text(reply_token, f"✅ {name} {current} 婉拒\n➡️ 下一家：{new_current}{extra}")
        else:
            reply_text(reply_token,
                       f"✅ {name} {current} 婉拒\n"
                       f"⚠️ 已全數婉拒、請手動處理：\n"
                       f"  不送了 → @AI {name} 結案\n"
                       f"  再送別家 → @AI {name} 送XXX")
        return

    if t == "reject_and_add":
        # @AI 姓名 公司A婉拒 送公司B ── 一次婉拒 A + 加送 B
        name = cmd["name"]
        reject_co_raw = cmd["reject_company"]
        add_co_raw = cmd["add_company"]
        target = _resolve_target_strict(cmd, name, group_id, reply_token, "婉拒送")
        if not target:
            return
        if not _check_active_or_warn(target, reply_token, "婉拒送", name):
            return
        reject_co = COMPANY_ALIAS.get(reject_co_raw, reject_co_raw)
        add_co = COMPANY_ALIAS.get(add_co_raw, add_co_raw)
        if not _validate_companies_or_warn([reject_co, add_co], reply_token, name):
            return
        reject_norm = normalize_section(reject_co)
        add_norm = normalize_section(add_co)
        current_co = target["current_company"] or ""
        concurrent_list = [c.strip() for c in (target["concurrent_companies"] or "").split(",") if c.strip()]
        # 先處理婉拒部分
        rejected = []
        new_current = current_co
        route_update = {}
        if current_co and normalize_section(current_co) == reject_norm:
            # 婉拒 current
            route = target["route_plan"] or ""
            new_route = advance_route(route, "婉拒")
            rejected = [current_co]
            # current 交給下面的加送決定（add_co 取代）
            new_current = ""
            route_update = {"route_plan": new_route}
        else:
            matches = [c for c in concurrent_list if normalize_section(c) == reject_norm]
            if matches:
                concurrent_list = [c for c in concurrent_list if c not in matches]
                rejected = matches
            else:
                reply_text(reply_token,
                           f"⚠️ {name} 目前沒在送 {reject_co}，無法婉拒\n"
                           f"目前在送：{current_co or '無'}\n"
                           f"同送：{','.join(concurrent_list) or '無'}")
                return
        # 加送部分：檢查 add_co 是否已在送
        if normalize_section(new_current) == add_norm or any(normalize_section(c) == add_norm for c in concurrent_list):
            reply_text(reply_token, f"⚠️ {name} 已經在送 {add_co}，不用再加送")
            return
        # 若 current 剛被婉拒空掉、把 add_co 放 current；否則放 concurrent
        if not new_current:
            new_current = add_co
        else:
            concurrent_list.append(add_co)
        # 寫入 DB
        text_note = f"{name} {'、'.join(rejected)} 婉拒 + 加送 {add_co}"
        update_customer(target["case_id"],
                        current_company=new_current,
                        concurrent_companies=",".join(concurrent_list),
                        text=text_note, from_group_id=group_id,
                        **route_update)
        push_text(target["source_group_id"], text_note)
        all_active = ([new_current] if new_current else []) + concurrent_list
        reply_text(reply_token,
                   f"✅ {name}\n"
                   f"  {'、'.join(rejected)} 婉拒\n"
                   f"  加送：{add_co}\n"
                   f"  現在在送：{'、'.join(all_active) if all_active else '無'}")
        return

    if t in ("reject_company", "confirm_reject_company"):
        # @AI 姓名 公司 婉拒 ── 指定某家婉拒（不跳轉）
        # 情境：林俊杰同送房地+銀行，銀行評估不過 → 「林俊杰 銀行 婉拒」
        # confirm_reject_company = 房地/銀行/C 類別的二次確認指令
        name = cmd["name"]
        co_raw = cmd["company"]
        is_confirmed = (t == "confirm_reject_company")
        target = _resolve_target_strict(cmd, name, group_id, reply_token, "婉拒")
        if not target:
            return
        if not _check_active_or_warn(target, reply_token, "婉拒", name):
            return
        co = COMPANY_ALIAS.get(co_raw, co_raw)
        if not _validate_companies_or_warn([co], reply_token, name):
            return
        co_norm = normalize_section(co)
        # 房地/銀行/C 需二次確認（非 confirm 指令時攔截）
        if not is_confirmed and co_norm in ("房地", "銀行", "零卡"):
            type_label = {"房地": "房地", "銀行": "銀行", "零卡": "C（零卡）"}[co_norm]
            reply_text(reply_token,
                       f"⚠️ 是 {type_label} 所有方案都婉拒嗎？\n"
                       f"確認請打：@AI {name} {co} 確認婉拒\n"
                       f"不是的話：@AI {name} [具體方案] 婉拒（例：房地一胎）")
            return
        current_co = target["current_company"] or ""
        concurrent_list = [c.strip() for c in (target["concurrent_companies"] or "").split(",") if c.strip()]
        # 情況 1：是 current（含同系列匹配，例：亞太 ↔ 亞太機車15萬、銀行 ↔ 元大）→ 走一般婉拒邏輯
        if current_co and normalize_section(current_co) == co_norm:
            route = target["route_plan"] or ""
            current = get_current_company(route) or current_co
            next_co = get_next_company(route)
            new_route = advance_route(route, "婉拒")
            # route 沒下一家時從 concurrent 第一家升上來（避免婉拒那家殘留當 current）
            if next_co:
                new_current = next_co
                new_concurrent = concurrent_list
                promoted = ""
            elif concurrent_list:
                new_current = concurrent_list[0]
                new_concurrent = concurrent_list[1:]
                promoted = new_current
            else:
                new_current = ""
                new_concurrent = concurrent_list
                promoted = ""
            update_kw = {"route_plan": new_route, "current_company": new_current,
                         "concurrent_companies": ",".join(new_concurrent),
                         "text": f"{name} {current} 婉拒", "from_group_id": group_id}
            if not new_current:
                update_kw["company"] = ""  # 已全數婉拒：清 company 避免日報 fallback
            update_customer(target["case_id"], **update_kw)
            if new_current:
                push_text(target["source_group_id"], f"{name} {current} 婉拒\n➡️ 下一家：{new_current}")
                extra = f"（從同送清單升上來）" if promoted else ""
                reply_text(reply_token, f"✅ {name} {current} 婉拒\n➡️ 下一家：{new_current}{extra}")
            else:
                push_text(target["source_group_id"], f"{name} {current} 婉拒")
                reply_text(reply_token,
                           f"✅ {name} {current} 婉拒\n"
                           f"⚠️ 已全數婉拒、請手動處理：\n"
                           f"  不送了 → @AI {name} 結案\n"
                           f"  再送別家 → @AI {name} 送XXX")
            return
        # 情況 2：在 concurrent（含同系列匹配）→ 從同送清單移除
        matches = [c for c in concurrent_list if normalize_section(c) == co_norm]
        if matches:
            new_concurrent = [c for c in concurrent_list if c not in matches]
            text_note = f"{name} {'、'.join(matches)} 婉拒（從同送清單移除）"
            update_customer(target["case_id"],
                            concurrent_companies=",".join(new_concurrent),
                            text=text_note, from_group_id=group_id)
            push_text(target["source_group_id"], text_note)
            remain = ",".join(new_concurrent) if new_concurrent else "無"
            all_active = ([current_co] if current_co else []) + new_concurrent
            reply_text(reply_token,
                       f"✅ {name} {'、'.join(matches)} 婉拒（從同送清單移除）\n"
                       f"現在在送：{'、'.join(all_active) if all_active else '無'}")
            return
        # 情況 3：兩邊都沒有
        all_active = ([current_co] if current_co else []) + concurrent_list
        reply_text(reply_token,
                   f"⚠️ {name} 目前沒在送 {co}\n"
                   f"現在在送：{'、'.join(all_active) if all_active else '無'}")
        return

    if t == "reject_to":
        name = cmd["name"]
        target_co = cmd["target"]
        reject_company = cmd.get("company", "")  # 可空（舊格式「婉拒轉XX」沒指定）
        target = _resolve_target_strict(cmd, name, group_id, reply_token, "婉拒轉")
        if not target:
            return
        # 規範化公司名（套 alias）
        target_co = COMPANY_ALIAS.get(target_co, target_co)
        # 驗證 target 公司名
        if not _validate_companies_or_warn([target_co], reply_token, name):
            return
        route = target["route_plan"] or ""
        current = get_current_company(route)
        reject_co = COMPANY_ALIAS.get(reject_company, reject_company) if reject_company else current
        if reject_company and not _validate_companies_or_warn([reject_co], reply_token, name):
            return
        # 從 concurrent_companies 移除婉拒的公司（不留在同送清單）
        concurrent_str = target["concurrent_companies"] or ""
        concurrent_list = [c.strip() for c in concurrent_str.split(",") if c.strip()]
        concurrent_list = [c for c in concurrent_list
                           if not (reject_co and (reject_co in c or c in reject_co))]
        new_concurrent = ",".join(concurrent_list)
        # 更新 route_plan：加婉拒歷史 + 把 target_co 設為 current
        data = parse_route_json(route)
        order = data.get("order", []) or []
        history = data.get("history", []) or []
        # 避免重複寫歷史
        if reject_co and not any(h.get("company") == reject_co and h.get("status") == "婉拒"
                                  for h in history):
            history.append({"company": reject_co, "status": "婉拒", "date": now_iso()[:10]})
        if target_co not in order:
            order.append(target_co)
        data["order"] = order
        data["current_index"] = order.index(target_co)
        data["history"] = history
        new_route = json.dumps(data, ensure_ascii=False)
        update_customer(target["case_id"], route_plan=new_route,
                        current_company=target_co,
                        concurrent_companies=new_concurrent,
                        text=f"{name} {reject_co} 婉拒，轉送 {target_co}",
                        from_group_id=group_id)
        # 回貼業務群
        push_text(target["source_group_id"], f"{name} {reject_co} 婉拒\n➡️ 跳轉到：{target_co}")
        reply_text(reply_token, f"✅ {name} {reject_co} 婉拒\n➡️ 跳轉到：{target_co}")
        return

    if t == "penalty":
        name = cmd["name"]
        amt = cmd["penalty"]
        rows = find_active_by_name(name)
        same = [r for r in rows if r["source_group_id"] == group_id]
        target = same[0] if same else (rows[0] if rows else None)
        if not target:
            reply_text(reply_token, f"❌ 找不到客戶：{name}"); return
        has_approved = bool(target["approved_amount"])
        reason = "核准後放棄" if has_approved else "辦理中放棄"
        update_customer(target["case_id"], status="PENALTY",
                        text=f"{name} 違約金已支付 ${amt}（{reason}）", from_group_id=group_id)
        push_text(target["source_group_id"], f"{name} 違約金已支付 ${amt}（{reason}）")
        reply_text(reply_token, "✅ " + name + " 已結案\n原因：" + reason + "\n違約金：$" + amt)
        return

    if t == "notification":
        name = cmd["name"]
        company = cmd.get("company", "")
        rows = find_active_by_name(name)
        same = [r for r in rows if r["source_group_id"] == group_id]
        target = same[0] if same else (rows[0] if rows else None)
        if not target:
            reply_text(reply_token, f"❌ 找不到客戶：{name}"); return
        # 處理同時送件（+ 分隔）：每個 item 拆公司名和金額
        concurrent_names = []      # 純公司名（給 concurrent_companies 存）
        concurrent_with_amt = []   # (公司, 金額, 期數) tuple（給照會訊息顯示）
        first_amount = first_period = ""
        if "+" in company:
            items = [c.strip() for c in company.split("+") if c.strip()]
            for i, item in enumerate(items):
                co, amt, per = _split_company_amount(item)
                co = COMPANY_ALIAS.get(co, co)
                concurrent_names.append(co)
                concurrent_with_amt.append((co, amt, per))
                if i == 0 and amt:  # 第一家的金額 → 記作 notify_amount
                    first_amount, first_period = amt, per
            company = concurrent_names[0] if concurrent_names else company
        else:
            # 單家也拆（可能「裕融100萬60期」）
            co, amt, per = _split_company_amount(company)
            if co:
                company = COMPANY_ALIAS.get(co, co)
                if amt:
                    first_amount, first_period = amt, per
        # 照會時如果在送件區塊，移到公司區塊
        if (target["report_section"] or "") == "送件":
            update_customer(target["case_id"], report_section="",
                            text=f"{name} 照會", from_group_id=group_id)
        # 存同時送件公司（只存純公司名）+ 記錄第一家金額
        # 修：沒帶新金額時清掉舊 notify_amount（避免上次送件的金額污染這次照會）
        update_kw = {}
        if concurrent_names:
            update_kw["concurrent_companies"] = ",".join(concurrent_names)
        if first_amount:
            update_kw["notify_amount"] = first_amount
            update_kw["notify_period"] = first_period
        else:
            # 沒帶新金額 → 清舊的，照會 fallback 用 PLAN_INFO 預設
            update_kw["notify_amount"] = ""
            update_kw["notify_period"] = ""
        update_customer(target["case_id"],
                        text=f"{name} 照會 {company}", from_group_id=group_id, **update_kw)
        r = dict(target)
        # 重新讀更新後的資料（才有最新 notify_amount）
        conn_re = get_conn(); cur_re = conn_re.cursor()
        cur_re.execute("SELECT * FROM customers WHERE case_id=?", (target["case_id"],))
        fresh = cur_re.fetchone(); conn_re.close()
        if fresh:
            r = dict(fresh)
        txt = generate_notification_text(r, company)
        reply_text(reply_token, txt)
        # 第二則：同時送件資訊（列出各家金額，民間方案跳過）
        if concurrent_with_amt:
            lines = []
            for co, amt, per in concurrent_with_amt:
                if amt:
                    per_str = f"/{per}期" if per else ""
                    lines.append(f"{co} {amt}萬{per_str}（指定）")
                else:
                    info = PLAN_INFO.get(co)
                    if info and info[1]:
                        lines.append(f"{info[0]} {info[1]}")
                    else:
                        lines.append(co)
            push_text(group_id, "📌 同時送件：\n" + "\n".join(lines))
        return

    if t == "set_amount":
        name, amount = cmd["name"], cmd["amount"]
        target = _resolve_target_strict(cmd, name, group_id, reply_token, "設金額")
        if not target:
            return
        update_customer(target["case_id"], approved_amount=amount, from_group_id=group_id)
        reply_text(reply_token, f"✅ {name} 核准金額已更新：{amount}")
        return

    if t == "add_concurrent":
        name = cmd["name"]
        company_raw = cmd["company"]
        notify_amount = cmd.get("notify_amount", "")
        notify_period = cmd.get("notify_period", "")
        target = _resolve_target_strict(cmd, name, group_id, reply_token, "加送")
        if not target:
            return
        if not _check_active_or_warn(target, reply_token, "加送", name):
            return
        # 切「+」後每個 item 拆公司名和金額（例：「裕融100萬60期」→ 公司=裕融、金額=100、期=60）
        new_companies_raw = []
        for item in re.split(r"[+＋]", company_raw):
            item = item.strip()
            if not item:
                continue
            co, amt, per = _split_company_amount(item)
            new_companies_raw.append(co)
            if amt and not notify_amount:
                notify_amount = amt
                notify_period = per
        if not new_companies_raw:
            reply_text(reply_token, f"❌ 找不到公司名：{company_raw}\n常見：亞太、喬美、第一、房地、裕融、和裕、麻吉、21..."); return
        if not _validate_companies_or_warn(new_companies_raw, reply_token, name):
            return
        current_co = target["current_company"] or ""
        concurrent_list = [c.strip() for c in (target["concurrent_companies"] or "").split(",") if c.strip()]
        added = []         # 真加送的
        specified = []     # 具體化備註的 (co, aliased_big_cat)
        new_current_co = None
        for co in new_companies_raw:
            # 若 co 有 alias 對應到某大類（例：元大→銀行），且該大類已在 current/concurrent
            # → 具體化備註替換（不加新家）
            aliased = COMPANY_ALIAS.get(co)
            specified_here = False
            if aliased:
                if aliased == current_co:
                    # current 大類（銀行）→ 具體化為 co（元大）
                    new_current_co = co
                    specified.append((co, aliased))
                    specified_here = True
                else:
                    # 找 concurrent 有沒有該大類
                    for i, c in enumerate(concurrent_list):
                        if c == aliased:
                            concurrent_list[i] = co
                            specified.append((co, aliased))
                            specified_here = True
                            break
            if not specified_here:
                # 正常加送（但若 co 已存在或有相近 → 不重複）
                if co not in concurrent_list and not any(co in p or p in co for p in concurrent_list):
                    concurrent_list.append(co)
                    added.append(co)
        update_kw = {"concurrent_companies": ",".join(concurrent_list),
                     "text": f"{name} 加送/備註 {company_raw}", "from_group_id": group_id}
        if new_current_co:
            update_kw["current_company"] = new_current_co
        if notify_amount:
            update_kw["notify_amount"] = notify_amount
            update_kw["notify_period"] = notify_period
        update_customer(target["case_id"], **update_kw)
        # 組訊息
        msg_parts = [f"✅ {name}"]
        if added:
            msg_parts.append(f"已加送：{'、'.join(added)}")
        if specified:
            for co, big in specified:
                msg_parts.append(f"備註：{big} → {co}")
        msg = "\n".join(msg_parts)
        # 金額提示（民間方案 skip）
        hint = _build_plan_info_hint(added, notify_amount, notify_period)
        if hint:
            msg += "\n" + hint
        reply_text(reply_token, msg)
        return

    if t == "advance":
        name, target_co = cmd["name"], cmd.get("target")
        notify_amount = cmd.get("notify_amount", "")
        notify_period = cmd.get("notify_period", "")
        target = _resolve_target_strict(cmd, name, group_id, reply_token, "轉送")
        if not target:
            return
        if not _check_active_or_warn(target, reply_token, "轉送", name):
            return
        route = target["route_plan"] or ""
        current = get_current_company(route)
        if target_co:
            # 解析目標：可能含「+」多公司同送，每個 item 拆公司名和金額
            targets = []
            for item in re.split(r"[+＋]", target_co):
                item = item.strip()
                if not item:
                    continue
                co, amt, per = _split_company_amount(item)
                co = COMPANY_ALIAS.get(co, co)
                targets.append(co)
                # 若尾巴沒金額且這裡偵測到 → 記錄第一家金額（給 notify）
                if amt and not notify_amount:
                    notify_amount = amt
                    notify_period = per
            # 防錯：驗證公司名
            if not _validate_companies_or_warn(targets, reply_token, name):
                return
            if len(targets) == 1:
                # 單公司：維持原邏輯（試 advance_route_to，失敗改 current_company）
                single_co = targets[0]
                new_route, ok, err = advance_route_to(route, single_co, "轉送")
                update_kw = {"current_company": single_co,
                             "text": f"{name} 轉送{single_co}", "from_group_id": group_id}
                if ok:
                    update_kw["route_plan"] = new_route
                if notify_amount:
                    update_kw["notify_amount"] = notify_amount
                    update_kw["notify_period"] = notify_period
                update_customer(target["case_id"], **update_kw)
                push_text(target["source_group_id"], f"{name} 已轉送：{current} → {single_co}")
                hint = _build_plan_info_hint([single_co], notify_amount, notify_period)
                msg = f"✅ {name} 已轉送：{current} → {single_co}"
                if hint:
                    msg += "\n" + hint
                reply_text(reply_token, msg)
            else:
                # 多公司（轉 A+B = 同送）：清掉原 route，重建為 [A, B, ...]，第一家 current、其餘 concurrent
                data = parse_route_json(route)
                history = data.get("history", []) or []
                if current:
                    history.append({"company": current, "status": "轉送", "date": now_iso()[:10]})
                new_route = make_route_json(targets, current_index=0, history=history)
                first_co = targets[0]
                concurrent_str = ",".join(targets[1:])
                update_kw = {"route_plan": new_route, "current_company": first_co,
                             "concurrent_companies": concurrent_str,
                             "text": f"{name} 轉送{'+'.join(targets)}", "from_group_id": group_id}
                if notify_amount:
                    update_kw["notify_amount"] = notify_amount
                    update_kw["notify_period"] = notify_period
                update_customer(target["case_id"], **update_kw)
                targets_str = "+".join(targets)
                hint = _build_plan_info_hint(targets, notify_amount, notify_period)
                push_text(target["source_group_id"], f"{name} 已轉送（同送）：{current} → {targets_str}")
                msg = f"✅ {name} 已轉送（同送）：{current} → {targets_str}"
                if hint:
                    msg += "\n" + hint
                reply_text(reply_token, msg)
        else:
            next_co = get_next_company(route)
            if not next_co:
                reply_text(reply_token, f"⚠️ {name} 沒設下一家要送的公司\n若要設送件順序：04/18-{name}-裕融/第一/亞太（不用 @AI）"); return
            new_route = advance_route(route, "轉送")
            update_customer(target["case_id"], route_plan=new_route, current_company=next_co,
                            text=f"{name} 轉下一家→{next_co}", from_group_id=group_id)
            push_text(target["source_group_id"], f"{name} 已轉送：{current} → {next_co}")
            reply_text(reply_token, f"✅ {name} 已轉送：{current} → {next_co}")


# =========================
# 主要業務邏輯
# =========================

def parse_single_approval_line(line: str) -> Dict:
    """解析單公司核准/婉拒格式：03/04-黃娫柔-房地核准20萬"""
    m = SINGLE_APPROVAL_RE.match(line.strip())
    if not m:
        return {}
    company = m.group(3).strip()
    for alias, real in COMPANY_ALIAS.items():
        if alias in company:
            company = real
            break
    return {
        "date": m.group(1), "name": m.group(2),
        "company": company, "status": m.group(4), "amount": m.group(5) or "",
    }


def is_single_approval_line(line: str) -> bool:
    return bool(parse_single_approval_line(line))


def handle_new_case_block(block_text, source_group_id, reply_token) -> Optional[str]:
    f = parse_header_fields(block_text)
    first_line = extract_first_line(block_text)
    # 抓公司前先把身分證移除，避免 U121558670 的「21」被當成公司
    first_line_no_id = re.sub(r"[A-Z][A-Z0-9]\d{8}", "", first_line, flags=re.IGNORECASE)
    name, id_no, company = f.get("name", ""), f.get("id_no", ""), extract_company(first_line_no_id)
    if not name or not id_no:
        return None
    # 欄位合理性檢查（日期/身分證/姓名異常 → 警告但不擋，讓業務看到確認）
    date_val = f.get("date", "")
    warnings = _validate_new_case_fields(date_val, name, id_no)
    # 先查本群組：若本群組已有 ACTIVE（同身分證），直接更新
    same_group = find_active_by_id_no_in_group(id_no, source_group_id)
    if same_group:
        update_customer(same_group["case_id"], company=company or same_group["company"] or "",
                        text=block_text, from_group_id=source_group_id, name=name)
        return f"🔄 已更新客戶：{name}"
    # 本群組同身分證沒有，但可能有同姓名（身分證不同）→ 跳按鈕確認是同人還是新客戶
    name_matches = find_active_by_name(name)
    same_group_by_name = [r for r in name_matches if r["source_group_id"] == source_group_id]
    if same_group_by_name:
        send_same_name_diff_id_buttons(reply_token, block_text, same_group_by_name,
                                        source_group_id, id_no, name, company)
        return "QUICK_REPLY_SENT"
    # 本群組無（含姓名），檢查別群組是否已有同身分證 → 跳沿用/轉移按鈕
    existing = find_active_by_id_no(id_no)
    if existing:
        send_confirm_new_case_buttons(reply_token, block_text, existing, source_group_id)
        return "QUICK_REPLY_SENT"
    create_customer_record(name, id_no, company, source_group_id, block_text)
    msg = f"🆕 已建立客戶：{name}"
    if warnings:
        msg += "\n⚠️ 資料檢查：\n" + "\n".join(f"  • {w}" for w in warnings) + \
               "\n（客戶仍已建立，如需修正可用 @AI 姓名 改身分證/改名）"
    return msg


def handle_route_order_block(block_text, source_group_id, reply_token) -> Optional[str]:
    parsed = parse_route_order_line(extract_first_line(block_text))
    if not parsed:
        return None
    name, companies = parsed["name"], parsed["companies"]
    dupe_count = parsed.get("dupe_count", 0)
    route_json = make_route_json(companies)
    current_co = companies[0]
    # 判斷第一家是否為民間方案 → 直接放對應區塊，不放送件
    private_keywords = ["銀行", "零卡", "商品貸", "代書", "當舖", "鄉民", "房地", "新鑫", "慢點付", "分期趣", "銀角", "刷卡換現", "鄉"]
    is_private = any(k in current_co for k in private_keywords)
    init_section = "" if is_private else "送件"
    # @AI 尾巴的送件金額/期數（非核准金額）
    n_amt, n_per = extract_notify_amount_period(block_text)
    notify_kw = {}
    if n_amt:
        notify_kw["notify_amount"] = n_amt
        notify_kw["notify_period"] = n_per or ""
    rows = find_active_by_name(name)
    same = [r for r in rows if r["source_group_id"] == source_group_id]
    dupe_suffix = f"（已去除重複 {dupe_count} 家）" if dupe_count > 0 else ""
    if same:
        update_customer(same[0]["case_id"], route_plan=route_json, current_company=current_co,
                        report_section=init_section,
                        text=block_text, from_group_id=source_group_id, **notify_kw)
        return f"📋 已更新 {name} 送件順序：{'/'.join(companies)}{dupe_suffix}"
    other = [r for r in rows if r["source_group_id"] != source_group_id]
    if other:
        send_transfer_case_buttons(reply_token, other[0], source_group_id, block_text, allow_new=True)
        return "QUICK_REPLY_SENT"
    create_customer_record(name, "", current_co, source_group_id, block_text,
                           route_plan=route_json, current_company=current_co)
    # 新建客戶：貸款方案先放送件，民間方案直接放對應區塊
    conn = get_conn(); cur = conn.cursor()
    if not is_private:
        cur.execute("UPDATE customers SET report_section='送件' WHERE customer_name=? AND source_group_id=? AND status='ACTIVE' ORDER BY created_at DESC LIMIT 1", (name, source_group_id))
    if n_amt:
        cur.execute("UPDATE customers SET notify_amount=?, notify_period=? WHERE customer_name=? AND source_group_id=? AND status='ACTIVE' ORDER BY created_at DESC LIMIT 1",
                    (n_amt, n_per or "", name, source_group_id))
    conn.commit(); conn.close()
    return f"🆕 已建立客戶 {name}，送件順序：{'/'.join(companies)}{dupe_suffix}"


def parse_transfer_line(line: str) -> Dict:
    """解析「8/5-戴君哲-轉21」「8/11-林曉薇-轉麻吉 6/18」「4/17-王思婷-轉喬美+麻吉機」，回傳 dict 或 {}"""
    m = TRANSFER_RE.match(line.strip())
    if not m:
        return {}
    raw = m.group(3).strip()
    # 支援 A+B+C 多公司（也支援全形＋）
    parts = [p.strip() for p in re.split(r"[+＋]", raw) if p.strip()]
    targets = [COMPANY_ALIAS.get(p, p) for p in parts]
    return {"date": m.group(1), "name": m.group(2), "targets": targets}


def handle_transfer_block(block_text, source_group_id, reply_token) -> Optional[str]:
    """處理「8/5-戴君哲-轉21」或「4/17-王思婷-轉喬美+麻吉機」格式的轉送指令"""
    parsed = parse_transfer_line(extract_first_line(block_text))
    if not parsed:
        return None
    name = parsed["name"]
    targets = parsed["targets"]
    if not targets:
        return None
    # @AI 尾巴的送件金額/期數（非核准金額）
    n_amt, n_per = extract_notify_amount_period(block_text)
    notify_kw = {}
    if n_amt:
        notify_kw["notify_amount"] = n_amt
        notify_kw["notify_period"] = n_per or ""
    rows = find_active_by_name(name)
    same = [r for r in rows if r["source_group_id"] == source_group_id]
    target = same[0] if same else (rows[0] if rows else None)
    if not target:
        return f"❌ 找不到客戶：{name}"
    route = target["route_plan"] or ""
    current = get_current_company(route)
    first_co = targets[0]
    if len(targets) == 1:
        # 單公司：維持原行為（推進到該公司或直接改 current_company）
        new_route, ok, err = advance_route_to(route, first_co, "轉送")
        if not ok:
            update_customer(target["case_id"], current_company=first_co,
                            text=f"{name} 轉送{first_co}", from_group_id=source_group_id, **notify_kw)
            return f"✅ {name} 已轉送：{current or '無'} → {first_co}"
        update_customer(target["case_id"], route_plan=new_route, current_company=first_co,
                        text=f"{name} 轉送{first_co}", from_group_id=source_group_id, **notify_kw)
        return f"✅ {name} 已轉送：{current or '無'} → {first_co}"
    # 多公司（「轉 A+B」= 同送）：
    # 第一家為 current_company，其餘加入 concurrent_companies → 日報同時顯示 A、B 區塊
    data = parse_route_json(route)
    history = data.get("history", []) or []
    # 把目前公司標記為「轉送」，保留歷史
    if current:
        history.append({"company": current, "status": "轉送", "date": now_iso()[:10]})
    new_route = make_route_json(targets, current_index=0, history=history)
    targets_str = "+".join(targets)
    # 第二家以後加入同送清單
    cur_concurrent = target["concurrent_companies"] or ""
    existing = [c.strip() for c in cur_concurrent.split(",") if c.strip()]
    for co in targets[1:]:
        if co not in existing and not any(co in e or e in co for e in existing):
            existing.append(co)
    concurrent_str = ",".join(existing) if existing else ""
    update_customer(target["case_id"], route_plan=new_route, current_company=first_co,
                    concurrent_companies=concurrent_str,
                    text=f"{name} 轉送{targets_str}", from_group_id=source_group_id, **notify_kw)
    return f"✅ {name} 已轉送（同送）：{current or '無'} → {targets_str}"


def handle_a_case_block(block_text, reply_token) -> Optional[str]:
    if is_blocked(block_text):
        return "❌ 含禁止關鍵字，已略過"
    id_no = extract_id_no(block_text)
    name = extract_name(block_text)
    # Bug 16: 用姓名鎖防止同客戶並發更新
    with get_name_lock(name):
        return _handle_a_case_block_locked(block_text, reply_token, id_no, name)


def _handle_a_case_block_locked(block_text, reply_token, id_no, name, forced_case_id: str = "") -> Optional[str]:
    # 異體字統一：核準→核准、身份證→身分證、全形→半形
    block_text = normalize_command_text(block_text)
    customer = None
    # 從 SELECT_CASE 按鈕選擇而來：固定用選中的 case，跳過 find 邏輯
    if forced_case_id:
        conn = get_conn(); cur = conn.cursor()
        cur.execute("SELECT * FROM customers WHERE case_id=?", (forced_case_id,))
        customer = cur.fetchone(); conn.close()
        if not customer:
            return f"⚠️ 選中的案件不存在（case_id={forced_case_id}）"
    if not customer and id_no:
        # 同身分證可能有多筆獨立案（B 群、C 群各一）→ 跳按鈕讓 A 群選
        candidates = find_all_active_by_id_no(id_no)
        if len(candidates) == 1:
            customer = candidates[0]
        elif len(candidates) > 1:
            send_ambiguous_case_buttons(reply_token, block_text, candidates)
            return "QUICK_REPLY_SENT"
    if not customer and name:
        matches = find_active_by_name(name)
        if len(matches) == 1:
            customer = matches[0]
        elif len(matches) > 1:
            send_ambiguous_case_buttons(reply_token, block_text, matches)
            return "QUICK_REPLY_SENT"
    if not customer:
        return f"⚠️ 找不到對應客戶：{name or extract_first_line(block_text)}"

    company = extract_company(block_text) or customer["company"] or ""
    new_status = "CLOSED" if is_closed_text(block_text) else None
    # 「待核准」= 還有缺、還沒真正核准，不觸發待撥款；先把「待核准/待核準」字串移除再比對
    text_wo_pending = block_text.replace("待核准", "").replace("待核準", "")
    # C（零卡）特例：「有額度 / 有換 / 可換 / 能換」視同核准直接排撥款（零卡可馬上換現）
    is_c_card = (company == "零卡" or COMPANY_ALIAS.get(company) == "零卡"
                 or normalize_section(company) == "零卡")
    is_c_confirmed = is_c_card and any(w in text_wo_pending for w in ["有額度", "有換", "可換", "能換"])
    is_approved = (any(w in text_wo_pending for w in ["核准", "核準", "過件", "通過", "核貸"])
                   or is_c_confirmed) and new_status != "CLOSED"
    # 「撤件」「客戶撤件」「客戶自行撤件」不算婉拒（只記錄，不推進 route；後續若業務明確打「轉XXX」才動）
    is_reject = not is_approved and any(w in block_text for w in ["婉拒", "申覆失敗", "建議維持原審", "不予承作", "無法再進件", "無法承作", "30日內有進件", "已建檔", "不提供申覆", "退件", "無法核貸", "無法進件"])
    route = customer["route_plan"] or ""
    new_route, next_co = route, ""
    # 檢查公司是否在同時送件清單或 route 裡
    concurrent_list = [c.strip() for c in (customer["concurrent_companies"] or "").split(",") if c.strip()]
    is_in_concurrent = any(company in c or c in company for c in concurrent_list if company)
    route_companies = parse_route_json(route).get("order", []) if route else []
    current_co = customer["current_company"] or customer["company"] or ""
    is_in_route = any(company and (company in rc or rc in company) for rc in route_companies)
    is_current = company and (company in current_co or current_co in company)
    # 如果公司不在路線、不是 current、不在同送清單 → 跳按鈕詢問
    if company and not is_in_concurrent and not is_in_route and not is_current:
        # 偵測任何「補X」關鍵字（婉拒後恢復類：補保人/補JCIC/補聯徵/補信用/補件/補薪轉...）
        is_supplement_recovery = any(kw in block_text for kw in [
            "補保人", "補JCIC", "補jcic", "補申覆", "補聯徵", "補信用",
            "補件", "補資料", "補薪轉", "補照片", "補時段", "補照會",
            "補行照", "補在職", "補存摺", "補勞保", "補駕照", "補身分證"
        ])
        action_id = short_id()
        save_pending_action(action_id, "unknown_company", {
            "case_id": customer["case_id"], "block_text": block_text,
            "company": company, "name": customer["customer_name"],
            "is_supplement": is_supplement_recovery
        })
        items = [
            make_quick_reply_item(f"再送{company}", f"UNKNOWN_CO_REROUTE|{action_id}"),
            make_quick_reply_item(f"同送{company}", f"UNKNOWN_CO_CONCURRENT|{action_id}"),
        ]
        if is_supplement_recovery:
            items.insert(0, make_quick_reply_item("補件恢復", f"UNKNOWN_CO_SUPPLEMENT|{action_id}"))
        items.append(make_quick_reply_item("取消", f"UNKNOWN_CO_CANCEL|{action_id}"))
        prompt = f"⚠️ {company} 不在 {customer['customer_name']} 的路線裡，請選擇："
        if is_supplement_recovery:
            prompt = f"⚠️ {customer['customer_name']} 的 {company} 已婉拒/不在路線，收到補件訊息，請選擇："
        reply_quick_reply(reply_token, prompt, items)
        return "QUICK_REPLY_SENT"
    if is_reject and route and not is_in_concurrent:
        # 不在同時送件清單 → 正常推進 route
        next_co = get_next_company(route)
        new_route = advance_route(route, "婉拒")

    # A群回貼時，如果客戶在「送件」區塊，清掉讓它移到公司區塊
    cur_report_sec = customer["report_section"] or ""
    # 核准時抓金額，判斷是否移到待撥款區
    approved_amount = None
    new_report_section = "" if cur_report_sec == "送件" else None
    ai_amount_needed = False
    if is_approved:
        quick_amount = extract_approved_amount(block_text)
        if quick_amount:
            approved_amount = quick_amount
            new_route = update_company_amount_in_history(new_route, company, quick_amount)
        else:
            ai_amount_needed = True
        new_report_section = "待撥款"

    # 核准或婉拒時，從同時送件移除該公司
    concurrent = customer["concurrent_companies"] or ""
    if concurrent and (is_approved or is_reject):
        parts = [c.strip() for c in concurrent.split(",") if c.strip()]
        parts = [c for c in parts if company not in c and c not in company]
        concurrent = ",".join(parts)
        conn3 = get_conn(); cur3 = conn3.cursor()
        cur3.execute("UPDATE customers SET concurrent_companies=? WHERE case_id=?", (concurrent, customer["case_id"]))
        conn3.commit(); conn3.close()

    # 更新該公司的狀態（每家公司各存一份）
    if company:
        try:
            cs_str = customer["company_status"] or "{}"
            cs = json.loads(cs_str)
        except Exception:
            cs = {}
        # 用 normalize_section 統一 key（亞太商品→亞太）
        co_key = normalize_section(company)
        cs[co_key] = block_text
        conn4 = get_conn(); cur4 = conn4.cursor()
        cur4.execute("UPDATE customers SET company_status=? WHERE case_id=?",
                     (json.dumps(cs, ensure_ascii=False), customer["case_id"]))
        conn4.commit(); conn4.close()

    if is_approved:
        update_customer(customer["case_id"], text=block_text,
                        from_group_id=A_GROUP_ID, status=new_status,
                        route_plan=new_route,
                        approved_amount=approved_amount,
                        report_section=new_report_section)
    else:
        # 婉拒/其他：正常更新 company + current_company（推進送件順序）
        update_customer(customer["case_id"], company=company, text=block_text,
                        from_group_id=A_GROUP_ID, status=new_status,
                        route_plan=new_route,
                        current_company=next_co if next_co else None,
                        approved_amount=approved_amount,
                        report_section=new_report_section)

    ok, err = push_text(customer["source_group_id"], block_text)
    if not ok:
        return f"❌ 找到客戶：{customer['customer_name']}，但回貼{get_group_name(customer['source_group_id'])}失敗：{err}"

    gname = get_group_name(customer["source_group_id"])
    msg = f"✅ 已結案並回貼到{gname}：{customer['customer_name']}" if new_status == "CLOSED" else f"✅ 已回貼到{gname}：{customer['customer_name']}"
    if is_reject:
        if is_in_concurrent:
            # 同時送件婉拒 → 從清單移除，剩下的還在送
            remaining = [c for c in concurrent_list if company not in c and c not in company]
            if remaining:
                msg += f"\n➡️ 剩下同時送件：{'、'.join(remaining)}"
            else:
                msg += f"\n➡️ 下一家：{next_co}" if next_co else f"\n⚠️ {customer['customer_name']} 已無下一家送件方案"
        else:
            msg += f"\n➡️ 下一家：{next_co}" if next_co else f"\n⚠️ {customer['customer_name']} 已無下一家送件方案"
    # 核准時在A群顯示金額確認
    if is_approved and approved_amount:
        msg += f"\n💰 核准金額：{approved_amount}（已存入）"

    if ai_amount_needed:
        import threading
        # Bug 16: 只捕獲 id_no/name/case_id/company，不要捕獲整個 customer 物件
        cap_id_no = customer["id_no"]
        cap_name = customer["customer_name"]
        cap_case_id = customer["case_id"]
        cap_company = company

        def ai_parse_and_update():
            import asyncio, traceback
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            try:
                try:
                    ai_amount = loop.run_until_complete(extract_approved_amount_with_ai(block_text))
                except Exception as ex:
                    # Bug 4: Anthropic API 異常或網路錯誤，主動推訊息通知 A 群
                    print(f"[ai_parse_and_update] Anthropic API error: {ex}")
                    traceback.print_exc()
                    push_text(A_GROUP_ID,
                        f"⚠️ {cap_name} AI 金額辨識異常（{type(ex).__name__}），"
                        f"請手動補：{cap_name} 核准金額XX萬@AI")
                    return

                if not ai_amount:
                    push_text(A_GROUP_ID, f"⚠️ {cap_name} 核准金額無法辨識\n請手動補：{cap_name} 核准金額XX萬@AI")
                    return

                # Bug 16: 進姓名鎖，確保不與其他訊息打架
                with get_name_lock(cap_name):
                    # Bug 16: 重新從 DB 讀最新狀態（不用快照）
                    try:
                        fresh = find_active_by_id_no(cap_id_no) if cap_id_no else None
                    except Exception as ex:
                        print(f"[ai_parse_and_update] DB read error: {ex}")
                        push_text(A_GROUP_ID, f"⚠️ {cap_name} 寫入失敗，請手動處理")
                        return
                    if not fresh:
                        push_text(A_GROUP_ID,
                            f"⚠️ {cap_name} AI 辨識完成（金額：{ai_amount}），"
                            f"但客戶在背景處理期間已被結案，未更新。\n"
                            f"如需處理，請手動重啟案件。")
                        return

                    try:
                        cur_route = fresh["route_plan"] or ""
                        new_r = update_company_amount_in_history(cur_route, cap_company, ai_amount)
                        update_customer(fresh["case_id"], approved_amount=ai_amount,
                                        route_plan=new_r, from_group_id=A_GROUP_ID)
                        push_text(A_GROUP_ID, f"💰 {cap_name} {cap_company} 核准金額已辨識：{ai_amount}")
                    except Exception as ex:
                        print(f"[ai_parse_and_update] update error: {ex}")
                        traceback.print_exc()
                        push_text(A_GROUP_ID,
                            f"⚠️ {cap_name} 金額已辨識為 {ai_amount} 但寫入失敗，請手動補")
            except Exception as ex:
                # 最後的防線：任何未預期的例外都不能讓執行緒靜默死掉
                print(f"[ai_parse_and_update] unexpected: {ex}")
                import traceback as _tb; _tb.print_exc()
            finally:
                try:
                    loop.close()
                except Exception:
                    pass
        threading.Thread(target=ai_parse_and_update, daemon=True).start()

    return msg


# =========================
# 按鈕指令處理
# =========================
def handle_command_text(text: str, reply_token: str) -> bool:

    def get_action(action_id, expected_type):
        a = get_pending_action(action_id)
        if not a or a["action_type"] != expected_type:
            reply_text(reply_token, "⚠️ 找不到待確認資料")
            return None
        return a

    if text.startswith("FORCE_CREATE_NEW|"):
        _, action_id = text.split("|", 1)
        a = get_action(action_id, "confirm_new_case_with_existing_id")
        if not a: return True
        p = a["payload"]; block_text = p.get("block_text", ""); sg = p.get("source_group_id", "")
        existing_case_id = p.get("existing_case_id", "")
        # 查既存客戶的原群
        conn = get_conn(); cur = conn.cursor()
        cur.execute("SELECT source_group_id FROM customers WHERE case_id=?", (existing_case_id,))
        existing = cur.fetchone(); conn.close()
        original_sg = existing["source_group_id"] if existing else ""
        name = extract_name(block_text)
        new_case_id = create_customer_record(name, extract_id_no(block_text), extract_company(block_text), sg, block_text)
        reply_text(reply_token, f"🆕 已在{get_group_name(sg)}建立新案件：{name}")
        # 通知原群 + 取消新建按鈕（30 分內有效）
        if original_sg and original_sg != sg and new_case_id:
            undo_id = short_id()
            save_pending_action(undo_id, "dup_use_undo", {
                "new_case_id": new_case_id, "original_sg": original_sg, "new_sg": sg,
                "customer_name": name,
            })
            buttons = [
                make_quick_reply_item("❌ 取消新建", f"DUP_USE_UNDO|{undo_id}"),
                make_quick_reply_item("✅ 知道了", f"DISMISS_NOTIFY|{undo_id}"),
            ]
            push_text_with_buttons(original_sg,
                f"ℹ️ {name} 在 {get_group_name(sg)} 也建了獨立案件（你這邊不變）\n"
                f"如果是誤按，30 分內可取消那邊的新建：",
                buttons)
        delete_pending_action(action_id); return True

    if text.startswith("USE_EXISTING_CASE|"):
        _, action_id = text.split("|", 1)
        a = get_action(action_id, "confirm_new_case_with_existing_id")
        if not a: return True
        conn = get_conn(); cur = conn.cursor()
        cur.execute("SELECT * FROM customers WHERE case_id=?", (a["payload"].get("existing_case_id",""),))
        c = cur.fetchone(); conn.close()
        if not c: reply_text(reply_token, "⚠️ 原案件不存在"); delete_pending_action(action_id); return True
        reply_text(reply_token, f"✅ 已沿用{get_group_name(c['source_group_id'])}案件：{c['customer_name']}")
        delete_pending_action(action_id); return True

    if text.startswith("TRANSFER_FROM_CONFIRM|"):
        _, action_id = text.split("|", 1)
        a = get_action(action_id, "confirm_new_case_with_existing_id")
        if not a: return True
        p = a["payload"]; block_text = p.get("block_text", ""); sg = p.get("source_group_id", "")
        conn = get_conn(); cur = conn.cursor()
        cur.execute("SELECT * FROM customers WHERE case_id=?", (p.get("existing_case_id",""),))
        c = cur.fetchone(); conn.close()
        if not c: reply_text(reply_token, "⚠️ 原案件不存在"); delete_pending_action(action_id); return True
        original_sg = c["source_group_id"] or ""
        cust_name = c["customer_name"]
        update_customer(c["case_id"], company=extract_company(block_text) or c["company"] or "",
                        text=block_text, from_group_id=sg,
                        name=extract_name(block_text) or cust_name, source_group_id=sg)
        reply_text(reply_token, f"✅ 已轉移到{get_group_name(sg)}:{cust_name}")
        # 通知原群 + 復原轉移按鈕（30 分內有效）
        if original_sg and original_sg != sg:
            undo_id = short_id()
            save_pending_action(undo_id, "transfer_undo", {
                "case_id": c["case_id"], "original_sg": original_sg, "new_sg": sg,
                "customer_name": cust_name,
            })
            buttons = [
                make_quick_reply_item("🔄 復原轉移", f"TRANSFER_UNDO|{undo_id}"),
                make_quick_reply_item("✅ 知道了", f"DISMISS_NOTIFY|{undo_id}"),
            ]
            push_text_with_buttons(original_sg,
                f"⚠️ {cust_name} 已從本群轉移到 {get_group_name(sg)}\n"
                f"你這邊日報不會再顯示這客戶。如果是誤按，30 分內可復原：",
                buttons)
        delete_pending_action(action_id); return True

    if text.startswith("CANCEL_NEW_CASE|"):
        _, action_id = text.split("|", 1)
        delete_pending_action(action_id); reply_text(reply_token, "✅ 已取消"); return True

    # ===== 跨群組重複客戶通知的復原按鈕 =====
    def _check_undo_expired(created_at_iso, minutes=30):
        """判斷 pending_action 是否超過 N 分鐘"""
        try:
            from datetime import datetime as _dt
            created = _dt.fromisoformat(created_at_iso)
            return (_dt.now() - created).total_seconds() > minutes * 60
        except Exception:
            return False

    if text.startswith("TRANSFER_UNDO|"):
        _, undo_id = text.split("|", 1)
        a = get_pending_action(undo_id)
        if not a or a["action_type"] != "transfer_undo":
            reply_text(reply_token, "⚠️ 找不到復原資料（可能已超過 30 分）")
            return True
        if _check_undo_expired(a.get("created_at", "")):
            reply_text(reply_token, "⚠️ 超過 30 分、已無法復原\n若仍要轉回、請手動用：@AI 姓名 (相關指令)")
            delete_pending_action(undo_id); return True
        p = a["payload"]
        case_id = p["case_id"]; original_sg = p["original_sg"]; new_sg = p["new_sg"]; name = p["customer_name"]
        update_customer(case_id, source_group_id=original_sg,
                        text=f"{name} 轉移復原（從 {get_group_name(new_sg)} 回到本群）",
                        from_group_id=original_sg)
        reply_text(reply_token, f"✅ 已復原：{name} 回到本群")
        push_text(new_sg, f"ℹ️ {name} 的轉移已被原群業務復原，客戶不再屬於 {get_group_name(new_sg)}")
        delete_pending_action(undo_id); return True

    if text.startswith("DUP_USE_UNDO|"):
        _, undo_id = text.split("|", 1)
        a = get_pending_action(undo_id)
        if not a or a["action_type"] != "dup_use_undo":
            reply_text(reply_token, "⚠️ 找不到取消資料（可能已超過 30 分）")
            return True
        if _check_undo_expired(a.get("created_at", "")):
            reply_text(reply_token, "⚠️ 超過 30 分、已無法取消\n若仍要刪除該案、請在該群手動結案")
            delete_pending_action(undo_id); return True
        p = a["payload"]
        new_case_id = p["new_case_id"]; new_sg = p["new_sg"]; name = p["customer_name"]
        update_customer(new_case_id, status="CLOSED",
                        text=f"{name} 沿用撤銷（原群業務取消新建）",
                        from_group_id=p["original_sg"])
        reply_text(reply_token, f"✅ 已取消：{name} 在 {get_group_name(new_sg)} 的新建案已結案")
        push_text(new_sg, f"ℹ️ {name} 的新建案已被原群業務撤銷（已結案）")
        delete_pending_action(undo_id); return True

    if text.startswith("DISMISS_NOTIFY|"):
        _, undo_id = text.split("|", 1)
        delete_pending_action(undo_id)
        reply_text(reply_token, "✅ 已關閉提示")
        return True

    # 多筆同名時使用者選了某個 case → 重新執行原指令
    if text.startswith("EXEC_CMD|"):
        parts = text.split("|", 2)
        if len(parts) < 3: return False
        _, action_id, case_id = parts
        a = get_action(action_id, "select_case_for_cmd")
        if not a: return True
        payload = a["payload"]
        cmd = dict(payload.get("cmd") or {})
        cmd["_forced_case_id"] = case_id
        try:
            handle_special_command(cmd, reply_token, payload.get("group_id", ""))
        finally:
            delete_pending_action(action_id)
        return True

    if text.startswith("CANCEL_CMD|"):
        _, action_id = text.split("|", 1)
        delete_pending_action(action_id); reply_text(reply_token, "✅ 已取消"); return True

    # 同姓名不同身分證：使用者選了「是同一人，更新既有案件的身分證 + 套用新訊息」
    if text.startswith("SAME_PERSON|"):
        parts = text.split("|", 2)
        if len(parts) < 3: return False
        _, action_id, case_id = parts
        a = get_action(action_id, "same_name_diff_id")
        if not a: return True
        p = a["payload"]
        new_id = p.get("new_id", "")
        new_company = p.get("new_company", "")
        block_text = p.get("block_text", "")
        source_group_id = p.get("source_group_id", "")
        conn = get_conn(); cur = conn.cursor()
        cur.execute("SELECT * FROM customers WHERE case_id=?", (case_id,))
        c = cur.fetchone()
        if not c:
            reply_text(reply_token, "⚠️ 原案件不存在"); conn.close()
            delete_pending_action(action_id); return True
        old_id = c["id_no"] or "無"
        cur.execute("UPDATE customers SET id_no=?, updated_at=? WHERE case_id=?",
                    (normalize_id_no(new_id), now_iso(), case_id))
        conn.commit(); conn.close()
        # 更新客戶資料：公司（若新訊息有指定）+ 原訊息寫入 case_logs 讓日報更新
        update_customer(case_id, company=new_company or c["company"] or "",
                        text=block_text + f"\n[身分證 {old_id} → {new_id}]",
                        from_group_id=source_group_id,
                        name=p.get("new_name", c["customer_name"]))
        reply_text(reply_token, f"✅ 已更新 {c['customer_name']} 身分證：{old_id} → {new_id}\n日報會顯示既有案件")
        delete_pending_action(action_id); return True

    if text.startswith("NEW_PERSON|"):
        _, action_id = text.split("|", 1)
        a = get_action(action_id, "same_name_diff_id")
        if not a: return True
        p = a["payload"]
        new_name = p.get("new_name", "")
        new_id = p.get("new_id", "")
        new_company = p.get("new_company", "")
        new_sg = p.get("source_group_id", "")
        block_text = p.get("block_text", "")
        if not new_name or not new_sg:
            reply_text(reply_token, "⚠️ 資料不完整，無法建立（請聯絡管理員）")
            delete_pending_action(action_id); return True
        # 建客戶 + 設 report_section 讓日報看得到（民間方案直接對應區塊，其他進「送件」）
        private_keywords = ["銀行", "零卡", "商品貸", "代書", "當舖", "鄉民", "房地", "新鑫"]
        init_section = "" if any(k in new_company for k in private_keywords) else "送件"
        case_id = create_customer_record(new_name, new_id, new_company, new_sg, block_text,
                                          current_company=new_company, report_section=init_section)
        reply_text(reply_token, f"🆕 已建立新客戶：{new_name}（身分證 {new_id}）\n已顯示在日報")
        delete_pending_action(action_id); return True

    if text.startswith("CANCEL_SAMENAME|"):
        _, action_id = text.split("|", 1)
        delete_pending_action(action_id); reply_text(reply_token, "✅ 已取消，未建立案件"); return True

    # 補件時多筆同名，使用者選了要更新哪筆
    if text.startswith("SELECT_SUPPLEMENT|"):
        parts = text.split("|", 2)
        if len(parts) < 3: return False
        _, action_id, case_id = parts
        a = get_action(action_id, "select_case_for_supplement")
        if not a: return True
        p = a["payload"]
        block_text = p.get("block_text", "")
        source_group_id = p.get("source_group_id", "")
        want_push_a = p.get("want_push_a", False)
        conn = get_conn(); cur = conn.cursor()
        cur.execute("SELECT * FROM customers WHERE case_id=?", (case_id,))
        c = cur.fetchone(); conn.close()
        if not c:
            reply_text(reply_token, "⚠️ 案件不存在"); delete_pending_action(action_id); return True
        company = extract_company(block_text) or c["company"] or ""
        new_status = "CLOSED" if is_closed_text(block_text) else None
        # 偵測核准/金額（normalize 後「核準」已統一為「核准」）
        text_wo_pending = block_text.replace("待核准", "")
        is_approved = any(w in text_wo_pending for w in ["核准", "過件", "通過", "核貸"]) and new_status != "CLOSED"
        changes = {"company": company, "status": new_status}
        if is_approved:
            amt = extract_approved_amount(block_text) or ""
            if amt:
                changes["approved_amount"] = amt
            changes["report_section"] = "待撥款"
        ok, diffs, cust_name = update_with_verify(case_id, changes,
                                                    from_group_id=source_group_id,
                                                    text_log=block_text)
        if not ok:
            reply_text(reply_token, "⚠️ 案件不存在")
            delete_pending_action(action_id); return True
        pushed = False
        if want_push_a and new_status != "CLOSED":
            ok_push, _ = push_text(A_GROUP_ID, block_text); pushed = ok_push
        if diffs:
            msg = f"✅ 已更新客戶：{cust_name}\n" + "\n".join(diffs)
        else:
            msg = f"ℹ️ {cust_name}：訊息已記錄（無實際欄位變動）"
        if pushed: msg += f"\n✅ 已回貼A群"
        reply_text(reply_token, msg); delete_pending_action(action_id); return True

    if text.startswith("CANCEL_SUPPLEMENT|"):
        _, action_id = text.split("|", 1)
        delete_pending_action(action_id); reply_text(reply_token, "✅ 已取消"); return True

    # 指令速查卡分類按鈕：HELP|send / approval / undo / tools / newcase / symbols
    if text.startswith("HELP|"):
        cat = text.split("|", 1)[1] if "|" in text else ""
        mapping = {
            "send": _HELP_SEND,
            "approval": _HELP_APPROVAL,
            "undo": _HELP_UNDO,
            "tools": _HELP_TOOLS,
            "newcase": _HELP_NEWCASE,
            "pairing": _HELP_PAIRING,
            "symbols": _HELP_SYMBOLS,
        }
        content = mapping.get(cat, "⚠️ 未知分類，請重打「@AI 格式」查主選單")
        reply_text(reply_token, content)
        return True

    if text.startswith("CONFIRM_TRANSFER|"):
        _, action_id = text.split("|", 1)
        a = get_action(action_id, "transfer_customer")
        if not a: return True
        p = a["payload"]; block_text = p.get("block_text", ""); tg = p.get("target_group_id", "")
        conn = get_conn(); cur = conn.cursor()
        cur.execute("SELECT * FROM customers WHERE case_id=?", (p.get("case_id",""),))
        c = cur.fetchone(); conn.close()
        if not c: reply_text(reply_token, "⚠️ 原案件不存在"); delete_pending_action(action_id); return True
        update_customer(c["case_id"], company=extract_company(block_text) or c["company"] or "",
                        text=block_text, from_group_id=tg, name=p.get("name", c["customer_name"]), source_group_id=tg)
        reply_text(reply_token, f"✅ 已改到{get_group_name(tg)}")
        delete_pending_action(action_id); return True

    if text.startswith("KEEP_OLD_CASE|"):
        _, action_id = text.split("|", 1)
        a = get_action(action_id, "transfer_customer")
        if not a: return True
        p = a["payload"]; block_text = p.get("block_text", "")
        conn = get_conn(); cur = conn.cursor()
        cur.execute("SELECT * FROM customers WHERE case_id=?", (p.get("case_id",""),))
        c = cur.fetchone(); conn.close()
        if not c: reply_text(reply_token, "⚠️ 原案件不存在"); delete_pending_action(action_id); return True
        update_customer(c["case_id"], company=extract_company(block_text) or c["company"] or "",
                        text=block_text, from_group_id=c["source_group_id"])
        reply_text(reply_token, f"✅ 已沿用{get_group_name(c['source_group_id'])}原案件")
        delete_pending_action(action_id); return True

    if text.startswith("CREATE_NEW_FROM_TRANSFER|"):
        _, action_id = text.split("|", 1)
        a = get_action(action_id, "transfer_customer")
        if not a: return True
        p = a["payload"]; block_text = p.get("block_text", ""); tg = p.get("target_group_id", "")
        name = extract_name(block_text)
        create_customer_record(name, extract_id_no(block_text), extract_company(block_text), tg, block_text)
        reply_text(reply_token, f"🆕 已建立新案件：{name}")
        delete_pending_action(action_id); return True

    if text.startswith("CANCEL_TRANSFER|"):
        _, action_id = text.split("|", 1)
        delete_pending_action(action_id); reply_text(reply_token, "✅ 已取消"); return True

    if text.startswith("SELECT_CASE|"):
        parts = text.split("|", 2)
        if len(parts) < 3: return False
        _, action_id, case_id = parts
        a = get_action(action_id, "route_a_case")
        if not a: return True
        block_text = a["payload"].get("block_text", "")
        # delegate 給完整 A 群處理流程（會寫 company_status / 核准移待撥款 / 婉拒推 route 等）
        id_no = extract_id_no(block_text)
        name = extract_name(block_text)
        with get_name_lock(name):
            result = _handle_a_case_block_locked(block_text, reply_token, id_no, name, forced_case_id=case_id)
        if result and result != "QUICK_REPLY_SENT":
            reply_text(reply_token, result)
        delete_pending_action(action_id); return True

    # 先處理二次確認（CONFIRM_UNKNOWN_CO_XXX|）
    is_confirm = text.startswith("CONFIRM_UNKNOWN_CO_")
    is_first = (text.startswith("UNKNOWN_CO_REROUTE|") or text.startswith("UNKNOWN_CO_CONCURRENT|") or text.startswith("UNKNOWN_CO_SUPPLEMENT|") or text.startswith("UNKNOWN_CO_CANCEL|"))
    if is_first or is_confirm:
        if is_first:
            action_type = text.split("|")[0].replace("UNKNOWN_CO_", "")  # REROUTE/CONCURRENT/SUPPLEMENT/CANCEL
            action_id = text.split("|", 1)[1]
        else:
            action_type = text.split("|")[0].replace("CONFIRM_UNKNOWN_CO_", "")
            action_id = text.split("|", 1)[1]
        a = get_action(action_id, "unknown_company")
        if not a: return True
        p = a["payload"]
        case_id = p.get("case_id", "")
        company = p.get("company", "")
        block_text = p.get("block_text", "")
        name = p.get("name", "")
        # 第一次點 → 顯示二次確認
        if is_first and action_type != "CANCEL":
            act_label = {"REROUTE": "再送", "CONCURRENT": "同送", "SUPPLEMENT": "補件恢復"}.get(action_type, action_type)
            items = [
                make_quick_reply_item(f"✅ 確定{act_label}{company}", f"CONFIRM_UNKNOWN_CO_{action_type}|{action_id}"),
                make_quick_reply_item("↩️ 返回", f"UNKNOWN_CO_CANCEL|{action_id}"),
            ]
            reply_quick_reply(reply_token, f"⚠️ 確認要將 {company} {act_label}給 {name} 嗎？", items)
            return True
        # 取消或二次確認 → 執行
        if action_type == "CANCEL":
            reply_text(reply_token, f"已取消：{name} 的 {company} 訊息不處理")
            delete_pending_action(action_id)
            return True
        conn = get_conn(); cur = conn.cursor()
        cur.execute("SELECT * FROM customers WHERE case_id=?", (case_id,))
        cust = cur.fetchone(); conn.close()
        if not cust:
            reply_text(reply_token, "⚠️ 客戶已不存在"); delete_pending_action(action_id); return True
        # 先計算新的 company_status（把原訊息寫入，日報才能抓到狀態摘要）
        try:
            cs_raw = cust["company_status"] or ""
            cs_dict = json.loads(cs_raw) if cs_raw else {}
        except Exception:
            cs_dict = {}
        co_key = normalize_section(company)
        cs_dict[co_key] = block_text  # 各路徑都把訊息當 company 的最新狀態
        new_company_status = json.dumps(cs_dict, ensure_ascii=False)
        if action_type in ("REROUTE", "SUPPLEMENT"):
            route = cust["route_plan"] or ""
            data = parse_route_json(route) if route else {"order":[], "current_index":0, "history":[]}
            order = data.get("order", [])
            if company not in order:
                order.append(company)
            data["order"] = order
            data["current_index"] = order.index(company)
            new_route = json.dumps(data, ensure_ascii=False)
            act_verb = "再送" if action_type == "REROUTE" else "補件恢復"
            update_customer(case_id, route_plan=new_route, current_company=company,
                            text=f"{name} {act_verb} {company}\n{block_text}", from_group_id=A_GROUP_ID)
            # 另外直接 SQL 寫 company_status（update_customer 簽名未含此欄）
            with db_conn(commit=True) as _cn:
                _cn.cursor().execute("UPDATE customers SET company_status=? WHERE case_id=?",
                                     (new_company_status, case_id))
            reply_text(reply_token, f"✅ 已將 {company} 加入 {name} 的送件順序並設為當前公司")
        else:  # CONCURRENT
            concurrent = cust["concurrent_companies"] or ""
            parts2 = [c.strip() for c in concurrent.split(",") if c.strip()]
            if company not in parts2:
                parts2.append(company)
            new_concurrent = ",".join(parts2)
            update_customer(case_id, concurrent_companies=new_concurrent,
                            text=f"{name} 同送 {company}\n{block_text}", from_group_id=A_GROUP_ID)
            with db_conn(commit=True) as _cn:
                _cn.cursor().execute("UPDATE customers SET company_status=? WHERE case_id=?",
                                     (new_company_status, case_id))
            reply_text(reply_token, f"✅ 已將 {company} 加入 {name} 的同時送件清單")
        delete_pending_action(action_id)
        return True

    if text.startswith("REOPEN_CASE|"):
        parts = text.split("|", 2)
        if len(parts) < 3: return False
        _, action_id, case_id = parts
        a = get_action(action_id, "reopen_or_new_case")
        if not a: return True
        p = a["payload"]; block_text = p.get("block_text", ""); push_to_a = p.get("push_to_a_after_reopen", False)
        conn = get_conn(); cur = conn.cursor()
        cur.execute("SELECT * FROM customers WHERE case_id=?", (case_id,))
        c = cur.fetchone(); conn.close()
        if not c: reply_text(reply_token, "⚠️ 原案件不存在"); delete_pending_action(action_id); return True
        update_customer(case_id, company=extract_company(block_text) or c["company"] or "",
                        text=block_text, from_group_id=c["source_group_id"], status="ACTIVE")
        pushed = False
        if push_to_a:
            ok, _ = push_text(A_GROUP_ID, block_text); pushed = ok
        msg = f"✅ 已重啟案件：{c['customer_name']}"
        if pushed: msg += f"\n✅ 已回貼A群：{c['customer_name']}"
        reply_text(reply_token, msg); delete_pending_action(action_id); return True

    if text.startswith("CREATE_NEW_CASE|"):
        _, action_id = text.split("|", 1)
        a = get_action(action_id, "reopen_or_new_case")
        if not a: return True
        p = a["payload"]; block_text = p.get("block_text", ""); sg = p.get("source_group_id", "")
        push_to_a = p.get("push_to_a_after_reopen", False); name = extract_name(block_text)
        create_customer_record(name, extract_id_no(block_text), extract_company(block_text), sg, block_text)
        pushed = False
        if push_to_a:
            ok, _ = push_text(A_GROUP_ID, block_text); pushed = ok
        msg = f"🆕 已建立新案件：{name}"
        if pushed: msg += f"\n✅ 已回貼A群：{name}"
        reply_text(reply_token, msg); delete_pending_action(action_id); return True

    if text.startswith("CANCEL_REOPEN|"):
        _, action_id = text.split("|", 1)
        delete_pending_action(action_id); reply_text(reply_token, "✅ 已取消"); return True

    return False


# =========================
# 撥款名單處理
# =========================
def handle_disbursement_list(text: str, reply_token: str):
    """
    解析A群撥款名單：
    1. 去DB查每個客戶的核准金額和所屬業務群
    2. 更新撥款日期
    3. 組合排撥名單推給所有行政群（格式：日期 排撥名單\n姓名\t公司+金額\t業務群）
    4. A群回報結果（成功幾筆/找不到幾筆）
    """
    parsed = parse_disbursement_list(text)
    if not parsed:
        reply_text(reply_token,
                   "⚠️ 撥款名單格式看不懂\n"
                   "範本：\n"
                   "  04/18 裕融 撥款名單\n"
                   "  王小明\n"
                   "  陳某某")
        return

    all_push_lines = []
    disb_date_str = ""
    result_lines = []  # (idx, status, name, msg)
    global_idx = 0

    for disb_date, companies in parsed.items():
        disb_date_str = disb_date
        for company, names in companies.items():
            for name in names:
                global_idx += 1

                # 找客戶（進行中或待撥款）
                rows = find_active_by_name(name)
                if not rows:
                    conn = get_conn(); cur = conn.cursor()
                    cur.execute(
                        "SELECT * FROM customers WHERE customer_name=? ORDER BY updated_at DESC LIMIT 1",
                        (name,))
                    row = cur.fetchone(); conn.close()
                    rows = [row] if row else []

                if not rows:
                    result_lines.append((global_idx, False, name, f"⚠️ 找不到對應客戶：{name}"))
                    continue

                target = rows[0]
                # Bug 14: 若業務群不存在，顯示 "-" 而非 "未知群組" 以免日報統計出錯
                sales_group_name = get_group_name(target["source_group_id"])
                if sales_group_name == "未知群組":
                    sales_group_name = "-"
                # 優先從 route_plan history 找這家公司的金額，找不到再用 approved_amount
                approved_amount = get_amount_from_history(target["route_plan"] or "", company)
                if not approved_amount:
                    approved_amount = target["approved_amount"] or ""

                # 更新撥款日期（寫到 route_plan history + disbursement_date）
                new_route = set_disbursed_in_history(target["route_plan"] or "", company, disb_date)
                update_customer(
                    target["case_id"],
                    disbursement_date=disb_date,
                    route_plan=new_route,
                    report_section="待撥款",
                    signing_area="",
                    signing_salesperson="",
                    signing_company="",
                    signing_time="",
                    signing_location="",
                    text=name + " " + company + " 撥款" + disb_date,
                    from_group_id=A_GROUP_ID,
                )
                result_lines.append((global_idx, True, name, f"✅ {name}"))

                # 組合排撥名單行：姓名\t公司+金額\t業務群名稱
                company_amount = company + (" " + approved_amount if approved_amount else "")
                push_line = name + "\t" + company_amount + "\t" + sales_group_name
                all_push_lines.append(push_line)

    # 推送排撥名單給所有行政群
    all_admin_gids = get_all_admin_groups()
    pushed_ok = False
    if all_admin_gids and all_push_lines:
        push_msg = disb_date_str + " 排撥名單" + chr(10) + chr(10).join(all_push_lines)
        for admin_gid in all_admin_gids:
            ok, _ = push_text(admin_gid, push_msg)
            if ok:
                pushed_ok = True

    # 組合A群回覆
    ok_count = sum(1 for _, s, _, _ in result_lines if s)
    fail_count = len(result_lines) - ok_count

    msg_lines = []
    if pushed_ok:
        msg_lines.append(f"✅ 撥款名單已推送行政群，共{ok_count}筆")
    else:
        msg_lines.append(f"✅ 撥款名單處理完成，共{ok_count}筆")

    for idx, success, name, detail in result_lines:
        if success:
            msg_lines.append(f"第{idx}筆：✅ {name}")
        else:
            msg_lines.append(f"第{idx}筆：{detail}")

    reply_text(reply_token, chr(10).join(msg_lines))


def _check_ambiguous_supplement(block_text, target_row, name):
    """含糊補件偵測：訊息是補件類 + 沒指公司 + 客戶同送多家 → 回提示文字。
    否則回 None（放行給原流程處理）。
    """
    bu_markers = ["補申覆", "補照會", "補薪轉", "補聯徵", "補照片", "補保人",
                  "補在職", "補存摺", "補勞保", "補駕照", "補行照",
                  "補JCIC", "補jcic", "補件", "補資料"]
    first_line = block_text.splitlines()[0] if block_text else ""
    bu_type = next((m for m in bu_markers if m in first_line), "")
    if not bu_type:
        return None
    # 訊息有指定公司 → 不觸發
    mentioned = extract_company(first_line) or ""
    if mentioned:
        return None
    # 計算在送家數
    current_co = target_row["current_company"] or ""
    concurrent_str = target_row["concurrent_companies"] or ""
    concurrent_list = [c.strip() for c in concurrent_str.split(",") if c.strip()]
    all_sending = ([current_co] if current_co else []) + concurrent_list
    if len(all_sending) < 2:
        return None
    return (f"⚠️ {name} 同時送 {len(all_sending)} 家（{'、'.join(all_sending)}）\n"
            f"「{bu_type}」請指明是哪一家：\n"
            f"例：{name} {all_sending[0]} {bu_type}")


def handle_bc_case_block(block_text, source_group_id, reply_token, source_text="") -> Optional[str]:
    if is_blocked(block_text):
        return "❌ 含禁止關鍵字，已略過"
    # Bug 16: 用姓名鎖防止同客戶並發
    name_for_lock = extract_name(block_text)
    with get_name_lock(name_for_lock):
        return _handle_bc_case_block_locked(block_text, source_group_id, reply_token, source_text)


def _handle_bc_case_block_locked(block_text, source_group_id, reply_token, source_text="") -> Optional[str]:
    # 異體字統一：核準→核准、身份證→身分證、全形→半形
    block_text = normalize_command_text(block_text)
    # 照會注意事項（等同已送件，提取欄位資料）
    if is_notification_briefing(block_text):
        result = handle_notification_briefing(block_text, source_group_id, reply_token)
        if result:
            return result
    if is_route_order_line(extract_first_line(block_text)):
        return handle_route_order_block(block_text, source_group_id, reply_token)
    # 轉送格式：8/5-戴君哲-轉21
    if parse_transfer_line(extract_first_line(block_text)):
        return handle_transfer_block(block_text, source_group_id, reply_token)
    # 單公司核准/婉拒格式：03/04-黃娫柔-房地核准20萬
    if is_single_approval_line(extract_first_line(block_text)):
        parsed = parse_single_approval_line(extract_first_line(block_text))
        name, company, status, amount = parsed["name"], parsed["company"], parsed["status"], parsed["amount"]
        rows = find_active_by_name(name)
        same = [r for r in rows if r["source_group_id"] == source_group_id and r["status"] == "ACTIVE"]
        target = same[0] if same else (rows[0] if rows else None)
        if not target:
            create_customer_record(name, "", company, source_group_id, block_text)
            return f"🆕 已建立客戶：{name}（{company} {status}）"
        route = target["route_plan"] or ""
        # 婉拒：
        # - 房地/銀行/C（零卡）是「類別底下多家」→ 不推、加提示要業務確認是否所有方案都無法送
        # - 其他公司（裕融/亞太/喬美等）→ 自動推下一家
        if status == "婉拒":
            company_section = normalize_section(company)
            if company_section in ("房地", "銀行", "零卡"):
                type_label = {"房地": "房地", "銀行": "銀行", "零卡": "C（零卡）"}[company_section]
                update_customer(target["case_id"], company=company, text=block_text,
                                from_group_id=source_group_id)
                return (f"⚠️ 已記錄 {name} {company}婉拒\n"
                        f"是 {type_label} 所有方案都無法送嗎？婉拒是否確認？\n"
                        f"要推到下一家請打：@AI {name} 婉拒")
            # 其他公司 → 自動推下一家
            current_co = get_current_company(route) or target["current_company"] or company
            next_co = get_next_company(route)
            new_route = advance_route(route, "婉拒")
            update_kw = {"route_plan": new_route, "current_company": next_co or "",
                         "company": company, "text": block_text, "from_group_id": source_group_id}
            if not next_co:
                update_kw["company"] = ""  # 已無下一家：清 company 避免日報殘留
            update_customer(target["case_id"], **update_kw)
            msg = f"✅ {name} {company} 婉拒"
            if next_co:
                msg += f"\n➡️ 下一家：{next_co}"
            else:
                msg += f"\n⚠️ 已無下一家可推、請手動結案或送新家"
            return msg
        # 核准：正常處理（移到待撥款）
        if amount:
            route = update_company_amount_in_history(route, company, amount)
        section = "待撥款" if status in ("核准", "核準") else None
        update_customer(target["case_id"], company=company, text=block_text,
                        from_group_id=source_group_id, route_plan=route,
                        report_section=section, approved_amount=amount or None)
        msg = f"已更新：{name} {company} {status}"
        if amount:
            msg += f" {amount}"
        return msg
    if looks_like_new_case_block(block_text):
        return handle_new_case_block(block_text, source_group_id, reply_token)

    name = extract_name(block_text)
    id_no = extract_id_no(block_text)
    company = extract_company(block_text)
    if not name or name in IGNORE_NAME_SET:
        return None

    # want_push_a：原始訊息有@AI觸發 且 訊息含補件相關動作
    # 用 source_text（未去掉@AI的原始文字）來判斷@AI觸發
    # 這樣「彭駿為 補案件@AI」也能正確判斷
    raw_for_trigger = source_text or block_text
    has_bu_keyword = any(w in block_text for w in [
        "補", "照會", "缺資料", "補件", "補資料", "補照片",
        "補時段", "補聯徵", "補保人", "補行照", "補照會",
    ])
    # 民間方案（銀行/零卡/商品貸/代書/當舖/鄉民/房地）不推 A 群，只記錄
    is_private_loan = any(w in block_text for w in [
        "銀行", "零卡", "商品貸", "代書", "當舖", "鄉民", "房地", "新鑫",
        "慢點付", "分期趣", "銀角", "刷卡換現"
    ])
    want_push_a = has_ai_trigger(raw_for_trigger) and has_bu_keyword and not is_private_loan
    has_action = has_business_action_word(block_text)

    if id_no:
        # 先查本群組有無 ACTIVE（支援「兩邊獨立案」：跨群組可各有一筆 ACTIVE）
        same_group = find_active_by_id_no_in_group(id_no, source_group_id)
        if same_group:
            new_status = "CLOSED" if is_closed_text(block_text) else None
            update_customer(same_group["case_id"], company=company or same_group["company"] or "",
                            text=block_text, from_group_id=source_group_id, status=new_status, name=name)
            pushed = False
            if want_push_a and new_status != "CLOSED":
                ok, _ = push_text(A_GROUP_ID, block_text); pushed = ok
            msg = f"已更新客戶：{name}"
            if pushed: msg += f"\n✅ 已回貼A群：{name}"
            return msg
        # 本群組沒有，檢查別群組是否已有此 id_no
        existing = find_active_by_id_no(id_no)
        if existing:
            if is_closed_text(block_text):
                return f"⚠️ 同身分證案件存在於{get_group_name(existing['source_group_id'])}：{name}"
            send_transfer_case_buttons(reply_token, existing, source_group_id, block_text, allow_new=True)
            return "QUICK_REPLY_SENT"
        create_customer_record(name, id_no, company, source_group_id, block_text)
        pushed = False
        if want_push_a:
            ok, _ = push_text(A_GROUP_ID, block_text); pushed = ok
        msg = f"🆕 已建立客戶：{name}"
        if pushed: msg += f"\n✅ 已回貼A群：{name}"
        return msg

    any_rows = find_any_by_name(name)
    same_rows = [r for r in any_rows if r["source_group_id"] == source_group_id]
    same_active = [r for r in same_rows if r["status"] == "ACTIVE"]
    same_closed = [r for r in same_rows if r["status"] != "ACTIVE"]
    other_active = [r for r in any_rows if r["source_group_id"] != source_group_id and r["status"] == "ACTIVE"]
    other_closed = [r for r in any_rows if r["source_group_id"] != source_group_id and r["status"] != "ACTIVE"]

    if has_action and same_closed and not same_active and not is_closed_text(block_text):
        send_reopen_case_buttons(reply_token, block_text, same_closed, source_group_id, push_to_a_after_reopen=want_push_a)
        return "QUICK_REPLY_SENT"
    if has_action and same_active:
        # 多筆同名 → 跳按鈕讓使用者選
        if len(same_active) > 1:
            send_same_name_supplement_buttons(reply_token, block_text, same_active,
                                               source_group_id, want_push_a)
            return "QUICK_REPLY_SENT"
        c = same_active[0]
        # 含糊補件（補件訊息 + 沒指公司 + 多家在送）→ 提示業務指定公司
        amb_hint = _check_ambiguous_supplement(block_text, c, name)
        if amb_hint:
            return amb_hint
        new_status = "CLOSED" if is_closed_text(block_text) else None
        update_customer(c["case_id"], company=company or c["company"] or "",
                        text=block_text, from_group_id=source_group_id, status=new_status)
        pushed = False
        if want_push_a and new_status != "CLOSED":
            ok, _ = push_text(A_GROUP_ID, block_text); pushed = ok
        msg = f"已更新客戶：{name}"
        if pushed: msg += f"\n✅ 已回貼A群：{name}"
        return msg
    if not has_action and not is_format_trigger(block_text):
        return None
    if same_active:
        # 多筆同名 → 跳按鈕
        if len(same_active) > 1:
            send_same_name_supplement_buttons(reply_token, block_text, same_active,
                                               source_group_id, want_push_a)
            return "QUICK_REPLY_SENT"
        r = same_active[0]
        # 含糊補件提示
        amb_hint = _check_ambiguous_supplement(block_text, r, name)
        if amb_hint:
            return amb_hint
        update_customer(r["case_id"], company=company or r["company"] or "",
                        text=block_text, from_group_id=source_group_id)
        pushed = False
        if want_push_a:
            ok, _ = push_text(A_GROUP_ID, block_text); pushed = ok
        msg = f"已更新客戶：{name}"
        if pushed: msg += f"\n✅ 已回貼A群：{name}"
        return msg
    if other_active:
        send_transfer_case_buttons(reply_token, other_active[0], source_group_id, block_text, allow_new=True)
        return "QUICK_REPLY_SENT"
    if same_closed and is_closed_text(block_text):
        r = same_closed[0]
        update_customer(r["case_id"], company=company or r["company"] or "",
                        text=block_text, from_group_id=source_group_id, status="CLOSED")
        return f"已更新客戶：{name}"
    if same_closed:
        send_reopen_case_buttons(reply_token, block_text, same_closed, source_group_id, push_to_a_after_reopen=want_push_a)
        return "QUICK_REPLY_SENT"
    if other_closed:
        send_transfer_case_buttons(reply_token, other_closed[0], source_group_id, block_text, allow_new=True)
        return "QUICK_REPLY_SENT"
    # 【B 方案】無 id_no 一律不建新客戶：找不到就警告，要求業務用「日期-姓名-身分證」格式
    # 避免打錯指令、姓名拼錯、格式亂 → 誤建怪客戶污染日報
    reply_text(reply_token,
               f"⚠️ 找不到客戶「{name}」\n"
               f"可能原因：\n"
               f"  • 姓名打錯（2-6 個中文字）\n"
               f"  • 客戶還沒建立（建客戶格式：04/18-姓名 A123456789）\n"
               f"  • 想操作結案客戶 → 先打「@AI {name} 重啟」\n"
               f"不知道怎麼打：@AI 說明")
    return None


# =========================
# 按鈕指令處理
# =========================
# =========================
# 撥款名單處理
# =========================
# =========================
# Webhook
# =========================
def process_event(event: dict):
    try:
        _process_event_inner(event)
    except Exception as e:
        import traceback
        traceback.print_exc()
        reply_token = (event or {}).get("replyToken", "")
        if reply_token:
            reply_text(reply_token, f"❌ 系統處理失敗，請截圖給管理員\n錯誤代碼：{type(e).__name__}")


def _process_event_inner(event: dict):
    # Bug 12: 防呆檢查 event 型別與必要欄位
    if not isinstance(event, dict):
        return
    if event.get("type") != "message":
        return
    message = event.get("message", {})
    if not isinstance(message, dict) or message.get("type") != "text":
        return
    text = (message.get("text") or "").strip()
    reply_token = event.get("replyToken") or ""
    source = event.get("source") or {}
    if not isinstance(source, dict):
        return
    group_id = source.get("groupId")
    if not text:
        return
    # Bug 2: reply_token 空值時 LINE API 會錯誤，直接忽略該事件
    if not reply_token:
        return
    # 5 秒內同群組同內容 → 視為重複按，直接忽略（防手滑重複送）
    if group_id and is_duplicate_message(group_id, text):
        return
    if handle_command_text(text, reply_token):
        return

    # 任何群組都可以查群組ID（方便設定用）
    if has_ai_trigger(text):
        clean = strip_ai_trigger(text).strip()
        if clean.upper() == "群組ID":
            gname = get_group_name(group_id)
            reply_text(reply_token, f"📋 此群組資訊\n名稱：{gname}\nID：{group_id}")
            return

    # A 群優先（避免 A 群同時被註冊為 SALES_GROUP 時走錯邏輯）
    if group_id == A_GROUP_ID:
        # 撥款名單（不需要@AI觸發，去掉@AI再判斷）
        disb_text = strip_ai_trigger(text).strip() if has_ai_trigger(text) else text
        if is_disbursement_list(disb_text):
            handle_disbursement_list(disb_text, reply_token)
            return

        if not has_ai_trigger(text):
            return
        # A群特殊指令（排除 update_amount，讓核准訊息走正常 A 群流程回貼業務群）
        cmd = parse_special_command(text, group_id)
        if cmd and cmd["type"] != "update_amount":
            handle_special_command(cmd, reply_token, group_id)
            return
        clean = strip_ai_trigger(text)
        if clean in {"", "助理", "AI助理"}:
            return
        blocks = split_multi_cases(clean) or [clean]
        results = []
        for idx, block in enumerate(blocks, 1):
            result = handle_a_case_block(block, reply_token)
            if result and result != "QUICK_REPLY_SENT":
                results.append(f"第{idx}筆：{result}" if len(blocks) > 1 else result)
        if results:
            reply_text(reply_token, "\n".join(results))
        return

    sales_ids = get_sales_group_ids()
    admin_ids = get_admin_group_ids()

    # 業務群 / 行政群
    if group_id in sales_ids or group_id in admin_ids:
        raw = text
        # 特殊 @AI 指令
        if has_ai_trigger(raw):
            cmd = parse_special_command(raw, group_id)
            if cmd:
                handle_special_command(cmd, reply_token, group_id)
                return

        is_creation = looks_like_new_case_block(raw) or any(looks_like_new_case_block(b) for b in split_multi_cases(raw))
        is_route = is_route_order_line(extract_first_line(raw)) or any(is_route_order_line(extract_first_line(b)) for b in split_multi_cases(raw))
        should_process = is_creation or is_route or has_ai_trigger(raw)
        if not should_process:
            return

        process_text = raw if (is_creation or is_route) else strip_ai_trigger(raw)
        if process_text in {"", "助理", "AI助理"}:
            return

        blocks = split_multi_cases(process_text) or [process_text]
        results, conflicts, qr_sent = [], [], False

        for idx, block in enumerate(blocks, 1):
            result = handle_bc_case_block(block, group_id, reply_token, source_text=raw)
            if result == "QUICK_REPLY_SENT":
                if len(blocks) > 1:
                    name = extract_name(block) or f"第{idx}筆"
                    conflicts.append(f"⚠️ {name} 有衝突，請手動確認")
                    continue
                else:
                    qr_sent = True; break
            if result:
                results.append(f"第{idx}筆：{result}" if len(blocks) > 1 else result)

        all_msgs = results + conflicts
        if not qr_sent and all_msgs:
            reply_text(reply_token, "\n".join(all_msgs))
        return

    # Bug 2: 未知群組 — 自動註冊為 UNASSIGNED 待審 + 主動回覆
    if group_id:
        try:
            with db_conn(commit=True) as conn:
                cur = conn.cursor()
                cur.execute("SELECT 1 FROM groups WHERE group_id=?", (group_id,))
                if not cur.fetchone():
                    # 自動建立待審群組
                    cur.execute("""INSERT INTO groups
                        (group_id, group_name, group_type, is_active, created_at)
                        VALUES (?, ?, 'UNASSIGNED', 0, ?)""",
                        (group_id, f"待審群組 {group_id[:8]}", now_iso()))
                    # 不主動回覆，等管理員打 @AI 群組ID 才會顯示
                    pass
        except Exception as e:
            print(f"[unknown_group] failed: {e}")


@app.post("/callback")
async def callback(request: Request, background_tasks: BackgroundTasks):
    body_bytes = await request.body()
    if LINE_CHANNEL_SECRET:
        signature = request.headers.get("X-Line-Signature", "")
        expected = base64.b64encode(
            hmac.new(LINE_CHANNEL_SECRET.encode("utf-8"), body_bytes, hashlib.sha256).digest()
        ).decode("utf-8")
        if not signature or not hmac.compare_digest(signature, expected):
            return JSONResponse({"status": "forbidden"}, status_code=403)
    try:
        body = json.loads(body_bytes) if body_bytes else {}
    except Exception:
        return {"status": "ok"}
    if not isinstance(body, dict):
        return {"status": "ok"}
    events = body.get("events", [])
    if not isinstance(events, list):
        return {"status": "ok"}
    for event in events:
        if not isinstance(event, dict):
            continue
        background_tasks.add_task(process_event, event)
    return {"status": "ok"}


# =========================
# 管理 API
# =========================
@app.post("/admin/reset_data")
async def reset_data(request: Request):
    """
    清除所有測試資料（customers + case_logs + pending_actions）
    群組設定保留不動。
    需要帶 {"confirm": "yes"} 才會執行，避免誤觸。
    """
    role = check_auth(request)
    if role != "admin":
        return JSONResponse({"status": "error", "message": "無權限"}, status_code=403)
    body = await request.json()
    if body.get("confirm") != "yes":
        return {"status": "error", "message": '請帶 {"confirm": "yes"} 才會執行清除'}
    try:
        with db_conn(commit=True) as conn:
            cur = conn.cursor()
            cur.execute("DELETE FROM customers")
            cur.execute("DELETE FROM case_logs")
            cur.execute("DELETE FROM pending_actions")
            cur.execute("DELETE FROM sqlite_sequence WHERE name IN ('customers','case_logs','pending_actions')")
        return {"status": "ok", "message": "✅ 已清除所有案件資料，群組設定保留"}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.get("/admin/reset_data")
def reset_data_page(request: Request):
    """清除測試資料的網頁介面"""
    from fastapi.responses import RedirectResponse
    role = check_auth(request)
    if role != "admin":
        return RedirectResponse("/login", status_code=303)
    return HTMLResponse("""<!DOCTYPE html><html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>清除測試資料</title></head><body>
    <h2>⚠️ 清除測試資料</h2>
    <p>此操作將清除所有客戶案件紀錄，群組設定不受影響。</p>
    <p><b>清除後無法復原（除非從 Render Snapshot 還原）</b></p>
    <button onclick="doReset()" style="background:red;color:white;padding:10px 20px;font-size:16px;border:none;border-radius:6px;cursor:pointer">
        確認清除所有測試資料
    </button>
    <div id="result" style="margin-top:20px;font-size:18px;"></div>
    <br><a href="/">回首頁</a>
    <script>
    async function doReset() {
        if (!confirm('確定要清除所有案件資料嗎？')) return;
        const r = await fetch('/admin/reset_data', {
            method: 'POST',
            headers: {'Content-Type': 'application/json'},
            body: JSON.stringify({confirm: 'yes'})
        });
        const data = await r.json();
        document.getElementById('result').innerText = data.message;
    }
    </script>
    </body></html>""")


VALID_GROUP_TYPES = {"SALES_GROUP", "ADMIN_GROUP", "A_GROUP", "UNASSIGNED"}

@app.post("/admin/add_group")
async def add_group(request: Request):
    """新增群組 API（Bug 3 修復：類型 whitelist + 防覆蓋）

    Body: {group_id, group_name, group_type: SALES_GROUP/ADMIN_GROUP/A_GROUP}

    行為：
    - 若 group_id 已存在 → 改用 UPDATE 只更新 name/type/linked，**保留 password_hash**
    - 若不存在 → INSERT 新群組
    - 類型必須在 whitelist 內，否則拒絕
    """
    role = check_auth(request)
    if role != "admin":
        return JSONResponse({"status": "error", "message": "無權限"}, status_code=403)
    body = await request.json()
    gid = body.get("group_id", "").strip()
    gname = body.get("group_name", "").strip()
    gtype = body.get("group_type", "SALES_GROUP").strip()
    if not gid or not gname:
        return {"status": "error", "message": "group_id 和 group_name 必填"}
    # Bug 3-1: 類型 whitelist 驗證
    if gtype not in VALID_GROUP_TYPES:
        return {"status": "error",
                "message": f"無效的群組類型：{gtype}（需為 SALES_GROUP/ADMIN_GROUP/A_GROUP）"}
    linked = body.get("linked_sales_group_id", "").strip()
    try:
        with db_conn(commit=True) as conn:
            cur = conn.cursor()
            # Bug 3-2: 先檢查是否已存在
            cur.execute("SELECT 1 FROM groups WHERE group_id=?", (gid,))
            existing = cur.fetchone()
            if existing:
                # 已存在 → UPDATE 不動 password_hash 和 created_at
                cur.execute("""UPDATE groups
                    SET group_name=?, group_type=?, is_active=1, linked_sales_group_id=?
                    WHERE group_id=?""",
                    (gname, gtype, linked or None, gid))
                msg = f"已更新群組：{gname}（{gtype}）— 密碼保留不動"
            else:
                # 新增
                cur.execute("""INSERT INTO groups
                    (group_id,group_name,group_type,is_active,linked_sales_group_id,created_at)
                    VALUES (?,?,?,1,?,?)""",
                    (gid, gname, gtype, linked or None, now_iso()))
                msg = f"已新增群組：{gname}（{gtype}）— 請到「密碼管理」設定密碼"
            if linked:
                msg += f"，對應業務群：{get_group_name(linked)}"
            return {"status": "ok", "message": msg}
    except Exception as e:
        return {"status": "error", "message": str(e)}



# =========================
# 密碼工具
# =========================
import hashlib as _hl

def hash_pw(password: str) -> str:
    """密碼雜湊 — 使用 PBKDF2-SHA256，20 萬次 iteration（Bug 1 修復）"""
    import os as _os
    salt = _os.urandom(16).hex()
    dk = _hl.pbkdf2_hmac("sha256", password.encode(), bytes.fromhex(salt), 200_000)
    return f"pbkdf2$200000${salt}${dk.hex()}"

def verify_pw(password: str, stored: str) -> bool:
    """驗證密碼 — 支援新格式 PBKDF2 + 舊格式 SHA-256（向後相容）

    舊密碼能繼續登入，但 Web 端建議下次登入成功後重新 hash 升級。
    使用 hmac.compare_digest 防時序攻擊。
    """
    if not stored:
        return False
    try:
        import hmac as _hmac
        # 新格式：pbkdf2$iterations$salt$hash
        if stored.startswith("pbkdf2$"):
            _, iters, salt, hashed = stored.split("$", 3)
            dk = _hl.pbkdf2_hmac("sha256", password.encode(), bytes.fromhex(salt), int(iters))
            return _hmac.compare_digest(dk.hex(), hashed)
        # 舊格式：salt:hash（向後相容）
        if ":" in stored:
            salt, hashed = stored.split(":", 1)
            calc = _hl.sha256((salt + password).encode()).hexdigest()
            return _hmac.compare_digest(calc, hashed)
    except Exception:
        return False
    return False

def needs_pw_upgrade(stored: str) -> bool:
    """檢查密碼是否為舊格式，需要升級到 PBKDF2"""
    return bool(stored) and not stored.startswith("pbkdf2$")

def get_setting(key: str) -> str:
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT value FROM settings WHERE key=?", (key,))
    row = cur.fetchone(); conn.close()
    return row["value"] if row else ""

def set_setting(key: str, value: str):
    conn = get_conn(); cur = conn.cursor()
    cur.execute("INSERT OR REPLACE INTO settings (key,value,updated_at) VALUES (?,?,?)",
        (key, value, now_iso()))
    conn.commit(); conn.close()

def get_group_password(group_id: str) -> str:
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT password_hash FROM groups WHERE group_id=?", (group_id,))
    row = cur.fetchone(); conn.close()
    return row["password_hash"] if row and row["password_hash"] else ""

def is_login_locked(identifier: str) -> bool:
    # locked_until 用 datetime 比較而非字串比較，避免 NULL/格式錯誤誤判
    with db_conn() as conn:
        cur = conn.cursor()
        cur.execute("SELECT attempts, locked_until FROM login_attempts WHERE identifier=?", (identifier,))
        row = cur.fetchone()
    if not row:
        return False
    lu = row["locked_until"]
    if not lu:
        return False
    try:
        lu_dt = datetime.strptime(lu, "%Y-%m-%d %H:%M:%S")
        return lu_dt > datetime.now()
    except (ValueError, TypeError):
        return False

def record_login_fail(identifier: str):
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT attempts FROM login_attempts WHERE identifier=?", (identifier,))
    row = cur.fetchone()
    if row:
        attempts = (row["attempts"] or 0) + 1  # 防 NULL
        locked_until = ""
        if attempts >= 5:
            from datetime import timedelta
            locked_until = (datetime.now() + timedelta(minutes=15)).strftime("%Y-%m-%d %H:%M:%S")
        cur.execute("UPDATE login_attempts SET attempts=?, locked_until=?, updated_at=? WHERE identifier=?",
            (attempts, locked_until, now_iso(), identifier))
    else:
        cur.execute("INSERT INTO login_attempts (identifier,attempts,locked_until,updated_at) VALUES (?,1,'',?)",
            (identifier, now_iso()))
    conn.commit(); conn.close()

def clear_login_fail(identifier: str):
    conn = get_conn(); cur = conn.cursor()
    cur.execute("DELETE FROM login_attempts WHERE identifier=?", (identifier,))
    conn.commit(); conn.close()

# =========================
# 網頁
# =========================
REPORT_PASSWORD = os.getenv("REPORT_PASSWORD", "admin123")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "admin_secret")


def _ensure_sessions_table():
    """確保 sessions 表存在（相容舊版 DB）"""
    try:
        conn = get_conn(); cur = conn.cursor()
        cur.execute("""CREATE TABLE IF NOT EXISTS sessions (
            token TEXT PRIMARY KEY NOT NULL,
            role TEXT NOT NULL,
            group_id TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            expires_at TEXT NOT NULL
        )""")
        conn.commit(); conn.close()
    except Exception:
        pass


def _create_session(role: str, group_id: str = "") -> str:
    """產生隨機 session token 並存入 DB"""
    from datetime import timedelta
    _ensure_sessions_table()
    token = secrets.token_urlsafe(32)
    now = now_iso()
    expires = (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d %H:%M:%S")
    try:
        conn = get_conn(); cur = conn.cursor()
        cur.execute("INSERT INTO sessions (token,role,group_id,created_at,expires_at) VALUES (?,?,?,?,?)",
            (token, role, group_id, now, expires))
        conn.commit(); conn.close()
    except Exception as e:
        print(f"Session create error: {e}")
    return token


def _is_session_expired(expires_at_str) -> bool:
    """Bug 10: 用 datetime 比較時間，不用字串比較

    NULL/格式錯誤 → 視為過期（安全預設）
    """
    if not expires_at_str:
        return True
    try:
        expires = datetime.strptime(expires_at_str, "%Y-%m-%d %H:%M:%S")
        return expires < datetime.now()
    except (ValueError, TypeError):
        return True


def _lookup_session(token: str) -> Optional[Dict]:
    """查找 session，回傳 {role, group_id} 或 None

    Bug 10/18 修復：用 datetime 比較 + db_conn context manager 防連線洩漏
    """
    if not token:
        return None
    try:
        _ensure_sessions_table()
        with db_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT role, group_id, expires_at FROM sessions WHERE token=?", (token,))
            row = cur.fetchone()
        if not row:
            return None
        # Bug 10: 用 datetime 比較
        if _is_session_expired(row["expires_at"]):
            # Bug 18: 用 db_conn 確保例外時連線會關
            with db_conn(commit=True) as conn:
                cur = conn.cursor()
                cur.execute("DELETE FROM sessions WHERE token=?", (token,))
            return None
        return {"role": row["role"], "group_id": row["group_id"]}
    except Exception as e:
        print(f"Session lookup error: {e}")
        return None


def _delete_session(token: str):
    """登出時刪除 session（Bug 18: 用 db_conn 防連線洩漏）"""
    if not token:
        return
    try:
        with db_conn(commit=True) as conn:
            cur = conn.cursor()
            cur.execute("DELETE FROM sessions WHERE token=?", (token,))
    except Exception as e:
        print(f"Session delete error: {e}")


# =========================
# form_tokens helpers（客戶自助填單 Magic Link）
# =========================
def create_form_token(group_id: str, note: str = "", days: int = 7) -> str:
    import secrets as _secrets
    from datetime import datetime as _dt, timedelta as _td
    token = _secrets.token_urlsafe(24)
    created = now_iso()
    expires = (_dt.now() + _td(days=days)).strftime("%Y-%m-%d %H:%M:%S")
    with db_conn(commit=True) as conn:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO form_tokens (token, group_id, note, created_at, expires_at) VALUES (?,?,?,?,?)",
            (token, group_id, note, created, expires),
        )
    return token


def get_form_token(token: str):
    """回傳 row（可用）或 None（不存在/過期/已用/已撤回）。"""
    if not token:
        return None
    with db_conn() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM form_tokens WHERE token=?", (token,))
        row = cur.fetchone()
    if not row:
        return None
    if row["used_at"] or row["revoked_at"]:
        return None
    if _is_session_expired(row["expires_at"]):
        return None
    return row


def consume_form_token(token: str, case_id: str):
    with db_conn(commit=True) as conn:
        cur = conn.cursor()
        cur.execute(
            "UPDATE form_tokens SET used_at=?, case_id=? WHERE token=?",
            (now_iso(), case_id, token),
        )


def revoke_form_token(token: str):
    with db_conn(commit=True) as conn:
        cur = conn.cursor()
        cur.execute("UPDATE form_tokens SET revoked_at=? WHERE token=?", (now_iso(), token))


def list_form_tokens(limit: int = 100):
    with db_conn() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM form_tokens ORDER BY created_at DESC LIMIT ?", (limit,))
        return cur.fetchall()


PAGE_CSS = """
<meta name="viewport" content="width=device-width,initial-scale=1">
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f4f6f9;color:#1a1a2e;font-size:14px}
a{color:#3b82f6;text-decoration:none}
.btn{display:inline-flex;align-items:center;gap:6px;padding:7px 16px;border-radius:7px;border:1px solid #d1d5db;background:#fff;cursor:pointer;font-size:13px;color:#374151;transition:background .15s}
.btn:hover{background:#f3f4f6}
.btn-primary{background:#1a1a2e;color:#fff;border-color:#1a1a2e}
.btn-primary:hover{background:#2d2d4e}
.btn-danger{color:#dc2626;border-color:#fca5a5}
.btn-danger:hover{background:#fef2f2}
.input{width:100%;padding:8px 12px;border:1px solid #d1d5db;border-radius:7px;font-size:13px;background:#fff}
.input:focus{outline:none;border-color:#6366f1;box-shadow:0 0 0 3px rgba(99,102,241,.1)}
.badge{display:inline-flex;align-items:center;font-size:11px;padding:2px 7px;border-radius:4px;font-weight:500}
.b-ok{background:#dcfce7;color:#15803d}
.b-wait{background:#f3e8ff;color:#7c3aed}
.b-doc{background:#fff7ed;color:#c2410c}
.b-act{background:#eff6ff;color:#1d4ed8}
.b-rej{background:#fef2f2;color:#b91c1c}
.b-close{background:#f1f5f9;color:#64748b}
.topnav{background:#1a1a2e;color:#fff;padding:0 20px;display:flex;align-items:center;justify-content:space-between;height:52px;position:sticky;top:0;z-index:100}
.topnav-title{font-size:15px;font-weight:600}
.topnav-links{display:flex;gap:4px;align-items:center}
.nl{padding:6px 12px;border-radius:6px;color:rgba(255,255,255,.65);font-size:13px;transition:all .15s;white-space:nowrap}
.nl:hover{background:rgba(255,255,255,.1);color:#fff}
.nl.active{background:rgba(255,255,255,.15);color:#fff;font-weight:500}
.nl-drop{position:relative;display:inline-block}
.nl-drop-btn{padding:6px 12px;border-radius:6px;color:rgba(255,255,255,.65);font-size:13px;cursor:pointer;background:none;border:none;font-family:inherit;transition:all .15s;white-space:nowrap}
.nl-drop-btn:hover{background:rgba(255,255,255,.1);color:#fff}
.nl-drop.open .nl-drop-btn,.nl-drop.active-parent .nl-drop-btn{background:rgba(255,255,255,.15);color:#fff;font-weight:500}
.nl-drop-menu{display:none;position:absolute;top:calc(100% + 4px);right:0;background:#1a1a2e;border:1px solid rgba(255,255,255,.1);border-radius:8px;padding:4px;min-width:170px;box-shadow:0 4px 12px rgba(0,0,0,.3);z-index:110}
.nl-drop.open .nl-drop-menu{display:block}
.nl-drop-menu .nl{display:block;padding:8px 12px;border-radius:5px;white-space:nowrap}
.nl-drop-caret{font-size:10px;margin-left:4px;opacity:.7}
.menu-btn{display:none;background:none;border:none;color:#fff;font-size:22px;cursor:pointer;padding:4px 8px}
.mobile-menu{display:none;position:fixed;top:52px;left:0;right:0;background:#1a1a2e;z-index:99;padding:8px 16px 16px;border-bottom:2px solid #6366f1}
.mobile-menu.show{display:flex;flex-direction:column;gap:4px}
.mobile-menu .nl{padding:10px 14px;font-size:14px;border-radius:8px}
.page{max-width:1100px;margin:0 auto;padding:20px 16px}
.stats-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:20px}
.stat-card{background:#fff;border:1px solid #e5e7eb;border-radius:9px;padding:14px 16px}
.stat-num{font-size:24px;font-weight:600}
.stat-lbl{font-size:11px;color:#9ca3af;margin-top:3px}
.search-bar{display:flex;gap:10px;margin-bottom:16px}
.group-card{background:#fff;border:1px solid #e5e7eb;border-radius:10px;margin-bottom:12px;overflow:hidden}
.group-hd{padding:11px 16px;display:flex;align-items:center;justify-content:space-between;cursor:pointer;user-select:none;transition:background .15s}
.group-hd:hover{background:rgba(0,0,0,.02)}
.group-hd-left{display:flex;align-items:center;gap:10px}
.group-dot{width:10px;height:10px;border-radius:50%;flex-shrink:0}
.group-title{font-size:14px;font-weight:600}
.group-meta{display:flex;gap:6px;flex-wrap:wrap}
.gm-pill{background:#f3f4f6;color:#6b7280;padding:2px 8px;border-radius:4px;font-size:11px}
.chevron{color:#9ca3af;font-size:11px;transition:transform .2s}
.group-body{display:none;border-top:1px solid #f3f4f6}
.group-body.open{display:block}
.seg-tabs{display:flex;padding:0 16px;gap:0;border-bottom:1px solid #f3f4f6}
.seg-tab{padding:9px 14px;font-size:12px;font-weight:500;color:#9ca3af;cursor:pointer;border-bottom:2px solid transparent;margin-bottom:-1px;transition:all .15s}
.seg-tab:hover{color:#6366f1}
.seg-tab.active{color:#1a1a2e;border-bottom-color:#6366f1}
.seg-panel{display:none}
.seg-panel.active{display:block}
.sec-hd{padding:8px 16px 4px;font-size:11px;font-weight:600;color:#9ca3af;text-transform:uppercase;letter-spacing:.5px;background:#fafafa;border-bottom:1px solid #f3f4f6}
.cust-row{padding:8px 20px;display:flex;align-items:center;justify-content:space-between;border-bottom:1px solid #f9fafb;cursor:pointer;transition:background .1s}
.cust-row:last-child{border-bottom:none}
.cust-row:hover{background:#f8faff}
.cust-name{font-size:13px;font-weight:500;display:flex;align-items:center;gap:6px}
.cust-detail{font-size:11px;color:#9ca3af;margin-top:2px}
.cust-date{font-size:11px;color:#d1d5db;flex-shrink:0;margin-left:10px}
.empty-sec{padding:14px 20px;font-size:12px;color:#d1d5db;font-style:italic}
.modal-bg{display:none;position:fixed;inset:0;background:rgba(0,0,0,.45);z-index:200;align-items:center;justify-content:center}
.modal-bg.show{display:flex}
.modal{background:#fff;border-radius:12px;padding:24px;width:90%;max-width:480px;max-height:80vh;overflow-y:auto}
.modal-title{font-size:16px;font-weight:600;margin-bottom:16px}
.modal-close{float:right;cursor:pointer;color:#9ca3af;font-size:18px;line-height:1}
.form-row{margin-bottom:14px}
.form-row label{display:block;font-size:12px;color:#6b7280;margin-bottom:5px;font-weight:500}
.result-ok{background:#dcfce7;color:#15803d;padding:8px 12px;border-radius:6px;font-size:12px;margin-top:10px}
.result-err{background:#fef2f2;color:#dc2626;padding:8px 12px;border-radius:6px;font-size:12px;margin-top:10px}
.history-row{padding:10px 16px;border-bottom:1px solid #f3f4f6;display:flex;align-items:center;justify-content:space-between}
.history-row:last-child{border-bottom:none}
.search-result{background:#fff;border:1px solid #e5e7eb;border-radius:10px;padding:16px;margin-bottom:12px}
.route-line{padding:4px 0;font-size:12px;color:#6b7280;display:flex;align-items:center;gap:6px}
@media(max-width:768px){
  .topnav{height:52px;padding:0 12px}
  .topnav-links{display:none}
  .menu-btn{display:block}
  .stats-grid{grid-template-columns:repeat(2,1fr);gap:8px}
  .stat-card{padding:12px}
  .page{padding:14px 10px}
  .search-result{padding:12px}
  .modal{width:95%;padding:16px}
  .input{padding:10px 12px;font-size:14px}
  .btn{padding:9px 16px;font-size:14px}
  .btn-primary{padding:10px 18px}
  table{font-size:11px}
  th,td{padding:5px 6px}
  .group-hd{padding:12px 14px}
}
@media(max-width:480px){
  .stats-grid{grid-template-columns:1fr 1fr;gap:6px}
  .stat-num{font-size:20px}
  .stat-lbl{font-size:10px}
  .stat-card{padding:10px}
  .search-bar{flex-direction:column}
  .group-title{font-size:13px}
}
</style>
"""

GROUP_COLORS = ["#1a1a2e","#2d5016","#7c2d12","#1e3a5f","#4a1d6e","#1e4d4d",
                "#78350f","#164e63","#4c1d95","#052e16"]

def check_auth(request: Request) -> str:
    """回傳 'admin'/'adminB'/'normal'/'group_xxx'/'' """
    t = request.cookies.get("auth_token", "")
    session = _lookup_session(t)
    if not session:
        return ""
    role = session["role"]
    if role in ("admin", "adminB", "normal"):
        return role
    if role == "group" and session["group_id"]:
        return f"group_{session['group_id']}"
    return ""

def get_auth_group_id(request: Request) -> str:
    """業務角色回傳其群組ID"""
    role = check_auth(request)
    if role.startswith("group_"):
        return role.replace("group_","")
    return ""

def make_topnav(role: str, active: str) -> str:
    links = [("📊 日報","/report","report"),("🔍 查詢","/search","search"),
             ("📁 歷史","/history","history"),("📖 指令速查","/guide","guide")]
    if role in ("admin","adminB","normal") or role.startswith("group_"):
        links.append(("📋 客戶資料庫","/pending-customers","pending"))
        links.append(("➕ 新增客戶","/new-customer","new"))
    if role in ("admin","adminB"):
        links.append(("📋 行政B作業","/adminb","adminb"))
    admin_items = []
    if role == "admin":
        admin_items = [("⚙️ 群組管理","/admin/groups","admin"),
                       ("🔑 密碼管理","/admin/passwords","passwords"),
                       ("📝 操作紀錄","/admin/logs","logs"),
                       ("📄 申請書範本","/admin/templates","templates"),
                       ("💾 下載備份","/admin/download-db","download"),
                       ("☁️ Drive 備份","/admin/gdrive-backup","gdrive-backup"),
                       ("🗑️ 清除資料","/admin/reset_data","reset")]
    nav = "".join(f'<a class="nl {"active" if a==active else ""}" href="{u}">{n}</a>'
                  for n,u,a in links)
    if admin_items:
        is_active_parent = active in {a for _,_,a in admin_items}
        drop_items = "".join(f'<a class="nl {"active" if a==active else ""}" href="{u}">{n}</a>'
                             for n,u,a in admin_items)
        nav += (
            f'<div class="nl-drop{" active-parent" if is_active_parent else ""}" id="adminDrop">'
            f'<button class="nl-drop-btn" onclick="event.stopPropagation();document.getElementById(\'adminDrop\').classList.toggle(\'open\')">⚙️ 管理<span class="nl-drop-caret">▾</span></button>'
            f'<div class="nl-drop-menu">{drop_items}</div>'
            f'</div>'
        )
    nav += '<a class="nl" href="/logout">登出</a>'
    mobile_links = links + admin_items
    mobile_nav = "".join(f'<a class="nl {"active" if a==active else ""}" href="{u}" onclick="document.getElementById(\'mobileMenu\').classList.remove(\'show\')">{n}</a>'
                  for n,u,a in mobile_links)
    mobile_nav += '<a class="nl" href="/logout">登出</a>'
    script = '<script>document.addEventListener("click",function(e){var d=document.getElementById("adminDrop");if(d&&!d.contains(e.target))d.classList.remove("open");});</script>'
    return (f'<nav class="topnav"><div class="topnav-title">貸款案件管理</div>'
            f'<div class="topnav-links">{nav}</div>'
            f'<button class="menu-btn" onclick="document.getElementById(\'mobileMenu\').classList.toggle(\'show\')">☰</button></nav>'
            f'<div id="mobileMenu" class="mobile-menu">{mobile_nav}</div>{script}')

def get_badge(row) -> str:
    sec = row["report_section"] or ""
    last = row["last_update"] or ""
    first = last.splitlines()[0] if last else ""
    if sec == "待撥款": return '<span class="badge b-wait">待撥款</span>'
    if "核准" in first: return '<span class="badge b-ok">核准</span>'
    if "婉拒" in first: return '<span class="badge b-rej">婉拒</span>'
    if "補" in first or "缺" in first: return '<span class="badge b-doc">補件</span>'
    if "等保書" in first: return '<span class="badge b-doc">等保書</span>'
    if "照會" in first: return '<span class="badge b-doc">照會</span>'
    return '<span class="badge b-act">進行中</span>'

def get_status_type(row) -> str:
    """判斷客戶狀態類型"""
    row = dict(row)
    sec = row["report_section"] or ""
    last = row["last_update"] or ""
    first = last.splitlines()[0] if last else ""
    co = row["current_company"] or row["company"] or ""
    route_data = parse_route_json(row["route_plan"] or "")
    order = route_data.get("order", [])
    # 新客戶：有資料但沒有送件順序和公司
    if not co and not order and not sec:
        return "new"
    # 待撥款
    if sec == "待撥款":
        disbursed = row["disbursement_date"] if row["disbursement_date"] else ""
        if disbursed:
            return "paid_verified"
        return "paid_unverified"
    # 補件
    if any(w in first for w in ["補","缺資料","補件","補資料","補行照","補照會","補照片","補時段","補聯徵","補保人"]):
        return "supplement"
    # 結案
    if row["status"] != "ACTIVE":
        return "closed"
    return "active"

def classify_rows(rows):
    """把客戶分類到各狀態"""
    cats = {"new":[], "supplement":[], "active":[], "paid_unverified":[], "paid_verified":[], "closed":[]}
    for r in rows:
        t = get_status_type(r)
        cats[t].append(r)
    return cats

_report_role = ""  # 暫存日報頁面的角色

def render_customer_row(row, role="") -> str:
    role = role or _report_role
    """產生單一客戶列"""
    row = dict(row)
    created = row["created_at"] or ""
    date_str = created[5:10].replace("-","/") if created else ""
    last = row["last_update"] or ""
    first_line = last.splitlines()[0].strip() if last.strip() else ""
    status_summary = extract_status_summary(first_line, row["customer_name"])
    co = row["current_company"] or row["company"] or ""
    amt = row["approved_amount"] or ""
    disb = row["disbursement_date"] or ""
    route_data = parse_route_json(row["route_plan"] or "")
    order = route_data.get("order",[])
    idx = route_data.get("current_index",0)

    # 詳細資料 - 支援多家核准
    route_history = route_data.get("history", [])
    approved_list = [rh for rh in route_history if rh.get("status") in ("核准","待撥款","撥款") and rh.get("amount")]
    if (row["report_section"] or "") == "待撥款":
        if len(approved_list) > 1:
            parts = [(rh.get("company") or "") + (rh.get("amount") or "") for rh in approved_list if rh.get("amount")]
            sub = "多家核准：" + " + ".join(parts) + ("（撥款" + disb + "）" if disb else "（待撥款）")
        else:
            sub = co + (f" 核准{amt}" if amt else "") + (f"（撥款{disb}）" if disb else "（待撥款）")
    elif order and idx < len(order):
        next_co = order[idx+1] if idx+1 < len(order) else ""
        # 顯示各家進度摘要
        progress_parts = []
        for rh in route_history[-3:]:
            hco = rh.get("company","")
            hst = rh.get("status","")
            hamt = rh.get("amount","")
            if hst == "核准":
                progress_parts.append(hco + "核准" + hamt)
            elif hst == "婉拒":
                progress_parts.append(hco + "婉拒")
        sub = co + (f" → {next_co}" if next_co else f"（第{idx+1}/{len(order)}家）")
        if progress_parts:
            sub += "　" + " / ".join(progress_parts[-2:])
    else:
        sub = co + (f" · {status_summary}" if status_summary else "")

    # 詳細資料展開內容
    phone = row.get("phone","") or ""
    id_no = row.get("id_no","") or ""
    company = row.get("company_name_detail","") or row.get("company","") or ""
    salary = row.get("company_salary","") or ""
    labor = row.get("eval_labor_ins","") or ""
    fund = row.get("eval_fund_need","") or ""

    progress_html = ('<div style="margin-top:8px;font-size:12px;color:#4a3e30">最新進度：<b style="color:#2c2820">' + h(first_line[:80]) + '</b></div>') if first_line else ""
    cid = row["case_id"]
    if role == "admin":
        edit_progress_html = (
            '<div style="margin-top:8px;border-top:1px solid #ddd5ca;padding-top:8px">'
            '<div style="font-size:11px;color:#6a5e4e;font-weight:600;margin-bottom:4px">修改進度</div>'
            '<div style="display:flex;gap:6px">'
            '<input id="prog-' + h(cid) + '" value="' + h(first_line) + '" style="flex:1;padding:5px 8px;border:1px solid #c8bfb5;border-radius:5px;font-size:12px">'
            '<button onclick="saveProgress(\'' + h(cid) + '\')" style="background:#6a5e4e;color:#fff;border:none;padding:5px 12px;border-radius:5px;font-size:11px;cursor:pointer;white-space:nowrap">儲存</button>'
            '</div></div>'
        )
    else:
        edit_progress_html = ""
    detail_html = (
        '<div style="background:#f0ebe4;padding:12px 16px;border-top:1px solid #ddd5ca;">'
        '<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-bottom:6px">'
        '<div><div style="font-size:11px;color:#6a5e4e;font-weight:600">身分證</div><div style="font-size:13px;color:#2c2820;font-weight:500">' + h(id_no or "-") + '</div></div>'
        '<div><div style="font-size:11px;color:#6a5e4e;font-weight:600">電話</div><div style="font-size:13px;color:#2c2820;font-weight:500">' + h(phone or "-") + '</div></div>'
        '<div><div style="font-size:11px;color:#6a5e4e;font-weight:600">公司</div><div style="font-size:13px;color:#2c2820;font-weight:500">' + h(company or "-") + '</div></div>'
        '<div><div style="font-size:11px;color:#6a5e4e;font-weight:600">月薪</div><div style="font-size:13px;color:#2c2820;font-weight:500">' + (h(salary) + "萬" if salary else "-") + '</div></div>'
        '<div><div style="font-size:11px;color:#6a5e4e;font-weight:600">勞保</div><div style="font-size:13px;color:#2c2820;font-weight:500">' + h(labor or "-") + '</div></div>'
        '<div><div style="font-size:11px;color:#6a5e4e;font-weight:600">資金需求</div><div style="font-size:13px;color:#2c2820;font-weight:500">' + h(fund or "-") + '</div></div>'
        '</div>'
        + progress_html
        + edit_progress_html
        + '</div>'
    )

    cid = row["case_id"]
    cname = row["customer_name"]
    q = "'"
    row_onclick = 'onclick="togD(' + q + h(cid) + q + ')"'
    return (
        '<div style="border-bottom:1px solid #ddd5ca">'
        + '<div style="display:grid;grid-template-columns:24px 1fr auto;gap:6px;align-items:center;padding:10px 16px">'
        + '<input type="checkbox" class="batch-cb" value="' + h(cid) + '" onclick="event.stopPropagation()" style="width:16px;height:16px;cursor:pointer">'
        + '<div ' + row_onclick + ' style="cursor:pointer">'
        + '<div style="font-size:15px;font-weight:600;color:#1a1208">' + h(cname) + '</div>'
        + '<div style="font-size:13px;color:#4a3e30;margin-top:2px">' + h(sub) + '</div>'
        + '</div>'
        + '<div style="font-size:12px;color:#6a5e4e;white-space:nowrap">' + h(date_str) + '</div>'
        + '</div>'
        + '<div id="dd-' + h(cid) + '" style="display:none">' + detail_html + '</div>'
        + '</div>'
    )

def render_section_block(label, rows, color_bg, color_text, icon, gid_prefix="") -> str:
    """產生可收合的狀態區塊"""
    if not rows:
        return ""
    count = len(rows)
    sec_id = (gid_prefix + "_" if gid_prefix else "") + label.replace(" ","_").replace("/","_")
    rows_html = "".join(render_customer_row(r) for r in rows)
    onclick_js = "togS('" + h(sec_id) + "')"
    return (
        '<div style="border-top:1px solid #ddd5ca">'
        '<div onclick="' + onclick_js + '" style="display:flex;align-items:center;gap:8px;padding:8px 16px;cursor:pointer;background:#ece8e2">'
        '<span style="font-size:13px">' + icon + '</span>'
        '<span style="font-size:13px;font-weight:700;color:' + color_text + '">' + h(label) + '</span>'
        '<span style="font-size:12px;color:#4a3e30;font-weight:500">' + str(count) + '筆</span>'
        '<span id="sa-' + h(sec_id) + '" style="margin-left:auto;font-size:11px;color:#a09080">▶</span>'
        '</div>'
        '<div id="sc-' + h(sec_id) + '" style="display:none">' + rows_html + '</div>'
        '</div>'
    )

def build_row_map(rows) -> dict:
    m = {}
    for r in rows:
        sec = r["report_section"] or r["current_company"] or r["company"] or "送件"
        m.setdefault(sec, []).append(r)
    return m

def render_cust_rows(rows) -> str:
    if not rows: return '<div class="empty-sec">（無資料）</div>'
    return "".join(render_customer_row(r) for r in rows)

def render_seg(section_map: dict, sections: list, shown: set) -> str:
    html = ""
    for sec in sections:
        if sec in section_map:
            html += f'<div class="sec-hd">{h(sec)}</div>'
            html += render_cust_rows(section_map[sec])
            shown.add(sec)
    return html


@app.get("/guide", response_class=HTMLResponse)
def guide_page(request: Request):
    """指令速查頁（業務、行政都能看，需要登入）。內容動態從 _HELP_* 渲染，和 LINE 同步。"""
    from fastapi.responses import RedirectResponse
    role = check_auth(request)
    if not role:
        return RedirectResponse(url="/login?next=/guide", status_code=303)
    html = _render_guide_html()
    topnav = make_topnav(role, "guide")
    # 套共用 topnav CSS + 導航列到 /guide
    html = html.replace("</style>", f"\n{PAGE_CSS}\n</style>", 1)
    html = html.replace("<body>", f"<body>\n{topnav}\n", 1)
    # 移除原本的「回日報頁」連結（topnav 已含）
    html = html.replace('<a href="/report" class="back-link">← 回日報頁</a>', '')
    return html


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request, error: str = ""):
    if error == "locked":
        err_html = '<div style="background:#fef2f2;color:#dc2626;padding:8px 12px;border-radius:6px;font-size:12px;margin-bottom:14px">帳號已鎖定，請 15 分鐘後再試或聯繫管理員解鎖</div>'
    elif error:
        err_html = '<div style="background:#fef2f2;color:#dc2626;padding:8px 12px;border-radius:6px;font-size:12px;margin-bottom:14px">密碼錯誤，請重新輸入</div>'
    else:
        err_html = ""
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT group_id, group_name FROM groups WHERE group_type='SALES_GROUP' AND is_active=1 ORDER BY group_name")
    sales_groups = cur.fetchall(); conn.close()
    grp_opts = "".join(f'<option value="{h(g["group_id"])}">{h(g["group_name"])}</option>' for g in sales_groups)
    return f"""<!DOCTYPE html><html><head>{PAGE_CSS}<title>登入</title></head><body>
    <div style="min-height:100vh;display:flex;align-items:center;justify-content:center;background:#f5f0eb">
    <div style="background:#faf7f4;border:1px solid #ddd5ca;border-radius:12px;padding:36px 32px;width:360px;box-shadow:0 4px 24px rgba(0,0,0,.06)">
      <div style="text-align:center;margin-bottom:24px">
        <div style="font-size:22px;font-weight:700;color:#3a3530">貸款案件管理</div>
        <div style="font-size:12px;color:#9ca3af;margin-top:6px">請選擇身份並輸入密碼</div>
      </div>
      {err_html}
      <form method="post" action="/login">
        <div style="margin-bottom:14px">
          <label style="font-size:12px;color:#5a4e40;font-weight:600;display:block;margin-bottom:6px">身份</label>
          <select name="role" class="input" style="padding:9px 12px" onchange="document.getElementById('grp_sec').style.display=this.value==='group'?'block':'none'">
            <option value="admin">管理員</option>
            <option value="normal">行政A</option>
            <option value="adminB">行政B</option>
            <option value="group">業務（選群組）</option>
          </select>
        </div>
        <div id="grp_sec" style="display:none;margin-bottom:14px">
          <label style="font-size:12px;color:#5a4e40;font-weight:600;display:block;margin-bottom:6px">群組</label>
          <select name="group_id" class="input" style="padding:9px 12px">{grp_opts}</select>
        </div>
        <div style="margin-bottom:16px">
          <label style="font-size:12px;color:#5a4e40;font-weight:600;display:block;margin-bottom:6px">密碼</label>
          <input class="input" type="password" name="password" placeholder="輸入密碼" autofocus style="padding:10px 14px;font-size:15px">
        </div>
        <button type="submit" class="btn btn-primary" style="width:100%;justify-content:center;padding:10px;font-size:14px;background:#6a5e4e;border-color:#6a5e4e">登入</button>
      </form>
    </div></div></body></html>"""


@app.post("/login")
async def login_post(request: Request):
    from fastapi.responses import RedirectResponse
    form = await request.form()
    pw = form.get("password","")
    role = form.get("role","normal")
    group_id = form.get("group_id","")
    # 基本防護：group_id 僅允許字母數字（LINE 群組 ID 格式），防止 XSS payload 寫入 login_attempts.identifier
    if group_id and not re.fullmatch(r"[A-Za-z0-9]{1,64}", group_id):
        group_id = ""
    identifier = role if role != "group" else f"group_{group_id}"
    if is_login_locked(identifier):
        return RedirectResponse("/login?error=locked", status_code=303)
    ok = False
    session_role = ""
    session_group = ""
    stored = ""
    if role == "admin":
        stored = get_setting("admin_pw")
        # Bug 3: 移除明文 fallback，沒密碼就拒絕
        ok = verify_pw(pw, stored)
        session_role = "admin"
    elif role == "adminB":
        stored = get_setting("adminB_pw")
        ok = verify_pw(pw, stored)
        session_role = "adminB"
    elif role == "normal":
        stored = get_setting("report_pw")
        ok = verify_pw(pw, stored)
        session_role = "normal"
    elif role == "group" and group_id:
        stored = get_group_password(group_id)
        ok = verify_pw(pw, stored) if stored else False
        session_role = "group"
        session_group = group_id
    if ok:
        clear_login_fail(identifier)
        # Bug 1：密碼舊格式自動升級到 PBKDF2
        if stored and needs_pw_upgrade(stored):
            try:
                if role == "admin":
                    set_setting("admin_pw", hash_pw(pw))
                elif role == "adminB":
                    set_setting("adminB_pw", hash_pw(pw))
                elif role == "normal":
                    set_setting("report_pw", hash_pw(pw))
            except Exception as e:
                print(f"[pw_upgrade] failed: {e}")
        token = _create_session(session_role, session_group)
        resp = RedirectResponse("/report", status_code=303)
        # Bug 2：加 secure 旗標（本地開發可用 ENV=local 關閉）
        _secure = os.getenv("ENV", "production").lower() != "local"
        resp.set_cookie("auth_token", token, max_age=86400*7,
                        httponly=True, samesite="Lax", secure=_secure)
        return resp
    record_login_fail(identifier)
    return RedirectResponse("/login?error=1", status_code=303)


@app.get("/logout")
def logout(request: Request):
    from fastapi.responses import RedirectResponse
    token = request.cookies.get("auth_token", "")
    _delete_session(token)
    resp = RedirectResponse("/login", status_code=303)
    resp.delete_cookie("auth_token")
    return resp


@app.get("/", response_class=HTMLResponse)
def home_redirect(request: Request):
    from fastapi.responses import RedirectResponse
    return RedirectResponse("/login" if not check_auth(request) else "/report")


@app.post("/report/update-progress")
async def report_update_progress(request: Request):
    role = check_auth(request)
    if not role:
        return JSONResponse({"ok": False, "message": "未登入"}, status_code=401)
    data = await request.json()
    case_id = data.get("case_id", "")
    progress = data.get("progress", "").strip()
    if not case_id:
        return JSONResponse({"ok": False, "message": "資料不完整"})
    conn = get_conn(); cur = conn.cursor()
    cur.execute("UPDATE customers SET last_update=?, updated_at=? WHERE case_id=?", (progress, now_iso(), case_id))
    conn.commit(); conn.close()
    return JSONResponse({"ok": True, "message": "進度已更新" if progress else "進度已清空"})


@app.post("/report/batch-close")
async def report_batch_close(request: Request):
    role = check_auth(request)
    if not role:
        return JSONResponse({"ok": False, "message": "未登入"}, status_code=401)
    data = await request.json()
    case_ids = data.get("case_ids", [])
    if not case_ids:
        return JSONResponse({"ok": False, "message": "未選擇客戶"})
    conn = get_conn(); cur = conn.cursor()
    closed = 0
    for cid in case_ids:
        cur.execute("SELECT * FROM customers WHERE case_id=? AND status='ACTIVE'", (cid,))
        c = cur.fetchone()
        if c:
            update_customer(cid, status="CLOSED", text=f"{c['customer_name']} 網頁批次結案", from_group_id="WEB")
            closed += 1
    conn.close()
    return JSONResponse({"ok": True, "message": f"已結案 {closed} 筆"})


@app.get("/report/export")
def report_export(request: Request):
    from fastapi.responses import StreamingResponse
    role = check_auth(request)
    if not role: return RedirectResponse("/login")
    auth_group = get_auth_group_id(request)
    conn = get_conn(); cur = conn.cursor()
    if auth_group:
        cur.execute("SELECT * FROM customers WHERE source_group_id=? AND status='ACTIVE' ORDER BY created_at DESC", (auth_group,))
    else:
        cur.execute("SELECT * FROM customers WHERE status='ACTIVE' ORDER BY source_group_id, created_at DESC")
    rows = cur.fetchall(); conn.close()
    # CSV 格式
    lines = ["群組,建立日期,姓名,身分證,目前公司,核准金額,撥款日期,狀態"]
    for r in rows:
        gname = get_group_name(r["source_group_id"])
        created = (r["created_at"] or "")[:10]
        co = r["current_company"] or r["company"] or ""
        amt = r["approved_amount"] or ""
        disb = r["disbursement_date"] or ""
        sec = r["report_section"] or ""
        lines.append(f'{gname},{created},{r["customer_name"]},{r["id_no"] or ""},{co},{amt},{disb},{sec}')
    content = "\ufeff" + "\n".join(lines)  # BOM for Excel
    today = datetime.now().strftime("%Y%m%d")
    return StreamingResponse(
        iter([content.encode("utf-8-sig")]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="report_{today}.csv"'}
    )


@app.get("/report", response_class=HTMLResponse)
def report_web(request: Request):
    from fastapi.responses import RedirectResponse
    role = check_auth(request)
    if not role: return RedirectResponse("/login")
    auth_group = get_auth_group_id(request)
    global _report_role
    _report_role = role

    today = datetime.now().strftime("%m/%d")
    conn = get_conn(); cur = conn.cursor()

    # 業務只看自己群組
    if auth_group:
        cur.execute("SELECT * FROM groups WHERE group_id=? AND is_active=1", (auth_group,))
    else:
        cur.execute("SELECT * FROM groups WHERE group_type='SALES_GROUP' AND is_active=1 ORDER BY group_name")
    groups = cur.fetchall()

    # 統計
    total_new = 0; total_supp = 0; total_active = 0; total_unverified = 0; total_closed = 0
    month_start = datetime.now().strftime("%Y-%m-01")
    today_date = datetime.now().strftime("%Y-%m-%d")
    cur.execute("SELECT COUNT(*) as c FROM customers WHERE status IN ('CLOSED','PENALTY','ABANDONED','REJECTED') AND updated_at>=?", (month_start,))
    total_closed = cur.fetchone()["c"]
    cur.execute("SELECT COUNT(*) as c FROM customers WHERE date(created_at)=?", (today_date,))
    today_new = cur.fetchone()["c"]
    cur.execute("SELECT COUNT(*) as c FROM customers WHERE status='ACTIVE' AND report_section='待撥款'")
    total_pending_disb = cur.fetchone()["c"]

    groups_html = ""
    for i, grp in enumerate(groups):
        gid, gname = grp["group_id"], grp["group_name"]
        cur.execute("SELECT * FROM customers WHERE source_group_id=? AND status='ACTIVE' ORDER BY updated_at DESC", (gid,))
        active_rows = cur.fetchall()
        cur.execute("SELECT * FROM customers WHERE source_group_id=? AND status IN ('CLOSED','PENALTY','ABANDONED','REJECTED') AND updated_at>=? ORDER BY updated_at DESC", (gid, month_start))
        closed_rows = cur.fetchall()

        cats = classify_rows(active_rows)
        total_new += len(cats["new"])
        total_supp += len(cats["supplement"])
        total_active += len(cats["active"])
        total_unverified += len(cats["paid_unverified"])

        count = len(active_rows)
        color = GROUP_COLORS[i % len(GROUP_COLORS)]

        # 統計 pills
        pills = f'<span class="gm-pill">{count} 筆</span>'
        if cats["new"]: pills += f'<span class="gm-pill" style="background:#e0f2fe;color:#0369a1">新客戶 {len(cats["new"])}</span>'
        if cats["supplement"]: pills += f'<span class="gm-pill" style="background:#fef9c3;color:#854d0e">補件 {len(cats["supplement"])}</span>'
        if cats["active"]: pills += f'<span class="gm-pill" style="background:#f0fdf4;color:#166534">送件 {len(cats["active"])}</span>'
        if cats["paid_unverified"]: pills += f'<span class="gm-pill" style="background:#fef2f2;color:#991b1b">未對保 {len(cats["paid_unverified"])}</span>'
        if cats["paid_verified"]: pills += f'<span class="gm-pill" style="background:#dcfce7;color:#166534">已對保 {len(cats["paid_verified"])}</span>'

        # 各狀態區塊（用 gid 當 prefix，避免多群組 ID 衝突）
        secs_html = ""
        secs_html += render_section_block("新客戶－需排送件順序", cats["new"], "#e0f2fe", "#0369a1", "🆕", gid)
        secs_html += render_section_block("補件中", cats["supplement"], "#fef9c3", "#854d0e", "📋", gid)
        secs_html += render_section_block("送件中", cats["active"], "#f0fdf4", "#166534", "📤", gid)
        secs_html += render_section_block("待撥款－未對保", cats["paid_unverified"], "#fef2f2", "#991b1b", "💰", gid)
        secs_html += render_section_block("待撥款－已對保", cats["paid_verified"], "#dcfce7", "#166534", "✅", gid)
        if closed_rows:
            secs_html += render_section_block("本月結案", closed_rows, "#f8fafc", "#64748b", "📁", gid)
        if not secs_html:
            secs_html = '<div class="empty-sec">（目前無有效案件）</div>'

        toggle_js = "toggleGroup('" + h(gid) + "')"
        groups_html += (
            '<div class="group-card" style="margin-bottom:10px">'
            '<div class="group-hd" onclick="' + toggle_js + '">'
            '<div class="group-hd-left">'
            '<div class="group-dot" style="background:' + color + '"></div>'
            '<div class="group-title">' + h(gname) + '</div>'
            '<div class="group-meta">' + pills + '</div>'
            '</div>'
            '<div class="chevron" id="chev-' + h(gid) + '">▶</div>'
            '</div>'
            '<div class="group-body" id="body-' + h(gid) + '" style="display:none;border-top:1px solid #ddd5ca">' + secs_html + '</div>'
            '</div>'
        )

    conn.close()

    REPORT_CSS = """
    <style>
    :root{--bg1:#faf7f4;--bg2:#f0ebe4;--bg3:#ece8e2;--t1:#2c2820;--t2:#6a5e4e;--t3:#a09080;--bd:#ddd5ca}
    body{background:var(--bg3);color:var(--t1);font-family:'Microsoft JhengHei','PingFang TC',sans-serif;font-size:14px}
    .group-card{background:var(--bg1);border:1px solid var(--bd);border-radius:10px;overflow:hidden;margin-bottom:10px}
    .group-hd{padding:12px 16px;display:flex;align-items:center;justify-content:space-between;cursor:pointer;user-select:none}
    .group-hd:hover{background:var(--bg2)}
    .group-hd-left{display:flex;align-items:center;gap:10px;flex-wrap:wrap}
    .group-dot{width:10px;height:10px;border-radius:50%;flex-shrink:0}
    .group-title{font-size:15px;font-weight:600;color:var(--t1)}
    .group-meta{display:flex;gap:5px;flex-wrap:wrap}
    .gm-pill{padding:2px 8px;border-radius:20px;font-size:11px;font-weight:500}
    .chevron{color:var(--t3);font-size:11px}
    .group-body{display:none;border-top:1px solid var(--bd)}
    .group-body.open{display:block}
    .empty-sec{padding:14px 20px;font-size:12px;color:var(--t3);font-style:italic}
    .stat-card{background:var(--bg1);border:1px solid var(--bd);border-radius:9px;padding:14px 16px}
    .stat-num{font-size:24px;font-weight:600}
    .stat-lbl{font-size:11px;color:var(--t3);margin-top:3px}
    .group-body{overflow-x:auto}
    </style>"""

    return f"""<!DOCTYPE html><html><head>{PAGE_CSS}{REPORT_CSS}<title>日報 {today}</title></head><body>
    {make_topnav(role, "report")}
    <div class="page">
      <div class="stats-grid" style="grid-template-columns:repeat(5,1fr)">
        <div class="stat-card"><div class="stat-num" style="color:#0369a1">{total_new}</div><div class="stat-lbl">🆕 新客戶</div></div>
        <div class="stat-card"><div class="stat-num" style="color:#854d0e">{total_supp}</div><div class="stat-lbl">📋 補件中</div></div>
        <div class="stat-card"><div class="stat-num" style="color:#166534">{total_active}</div><div class="stat-lbl">📤 送件中</div></div>
        <div class="stat-card"><div class="stat-num" style="color:#991b1b">{total_unverified}</div><div class="stat-lbl">💰 未對保</div></div>
        <div class="stat-card"><div class="stat-num" style="color:#64748b">{total_closed}</div><div class="stat-lbl">📁 本月結案</div></div>
        <div class="stat-card"><div class="stat-num" style="color:#7c3aed">{today_new}</div><div class="stat-lbl">📅 今日進件</div></div>
        <div class="stat-card"><div class="stat-num" style="color:#ea580c">{total_pending_disb}</div><div class="stat-lbl">⏳ 待撥款</div></div>
      </div>
      <div style="margin-bottom:14px;text-align:right">
        <a href="/report/export" class="btn" style="background:#e8e2da;color:#4a3e30;font-size:12px;text-decoration:none">📥 匯出 CSV</a>
      </div>
      {groups_html}
    </div>
    <script>
    function toggleGroup(gid){{
      const body=document.getElementById("body-"+gid);
      const chev=document.getElementById("chev-"+gid);
      const open=body.style.display==="none"||body.style.display==="";
      body.style.display=open?"block":"none";
      chev.textContent=open?"▼":"▶";
    }}
    function togS(sid){{
      const el=document.getElementById("sc-"+sid);
      const ar=document.getElementById("sa-"+sid);
      const open=el.style.display==="none"||el.style.display==="";
      el.style.display=open?"block":"none";
      ar.textContent=open?"▼":"▶";
    }}
    function togD(cid){{
      const el=document.getElementById("dd-"+cid);
      el.style.display=el.style.display==="block"?"none":"block";
    }}
    function saveProgress(cid){{
      const el=document.getElementById('prog-'+cid);
      if(!el)return;
      const val=el.value.trim();
      if(!val && !confirm('確定要清空進度嗎？'))return;
      fetch('/report/update-progress',{{method:'POST',headers:{{'Content-Type':'application/json'}},body:JSON.stringify({{case_id:cid,progress:val}})}})
      .then(r=>r.json()).then(d=>{{alert(d.message);if(d.ok)location.reload();}});
    }}
    function batchClose(){{
      const cbs=document.querySelectorAll('.batch-cb:checked');
      if(!cbs.length){{alert('請先勾選客戶');return;}}
      if(!confirm('確定要批次結案 '+cbs.length+' 位客戶嗎？'))return;
      const ids=[...cbs].map(c=>c.value);
      fetch('/report/batch-close',{{method:'POST',headers:{{'Content-Type':'application/json'}},body:JSON.stringify({{case_ids:ids}})}})
      .then(r=>r.json()).then(d=>{{alert(d.message);location.reload();}});
    }}
    </script>
    <div style="position:fixed;bottom:20px;right:20px;z-index:999">
      <button onclick="batchClose()" class="btn btn-primary" style="box-shadow:0 4px 12px rgba(0,0,0,.15);font-size:13px;padding:10px 18px">批次結案</button>
    </div>
    </body></html>"""


def _build_timeline(case_id: str) -> str:
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT message_text, from_group_id, created_at FROM case_logs WHERE case_id=? ORDER BY created_at DESC LIMIT 20", (case_id,))
    logs = cur.fetchall(); conn.close()
    if not logs:
        return '<div style="font-size:12px;color:#999">無操作紀錄</div>'
    lines = []
    for log in logs:
        dt = (log["created_at"] or "")[:16].replace("T", " ")
        gname = get_group_name(log["from_group_id"])
        txt = (log["message_text"] or "")[:60]
        lines.append(f'<div style="padding:4px 0;border-bottom:1px solid #f0ebe4;font-size:11px"><span style="color:#999">{h(dt)}</span> <span style="color:#8b7355">{h(gname)}</span> {h(txt)}</div>')
    return "".join(lines)


@app.get("/search", response_class=HTMLResponse)
def search_page(request: Request, q: str = "", grp: str = "", date_from: str = "", date_to: str = ""):
    from fastapi.responses import RedirectResponse
    role = check_auth(request)
    if not role: return RedirectResponse("/login")

    conn2 = get_conn(); cur2 = conn2.cursor()
    cur2.execute("SELECT group_id, group_name FROM groups WHERE group_type='SALES_GROUP' AND is_active=1 ORDER BY group_name")
    all_groups2 = cur2.fetchall(); conn2.close()
    grp_opts2 = "<option value=''>全部群組</option>" + "".join(f'<option value="{h(g["group_id"])}" {"selected" if grp==g["group_id"] else ""}>{h(g["group_name"])}</option>' for g in all_groups2)
    results_html = ""
    if q or grp or date_from or date_to:
        conn = get_conn(); cur = conn.cursor()
        sql2 = "SELECT * FROM customers WHERE 1=1"; params2 = []
        if q: sql2 += " AND (customer_name LIKE ? OR id_no LIKE ?)"; params2 += [f"%{q}%", f"%{q}%"]
        if grp: sql2 += " AND source_group_id=?"; params2.append(grp)
        if date_from: sql2 += " AND DATE(created_at) >= ?"; params2.append(date_from)
        if date_to: sql2 += " AND DATE(created_at) <= ?"; params2.append(date_to)
        sql2 += " ORDER BY status ASC, updated_at DESC"
        cur.execute(sql2, params2)
        rows = cur.fetchall()
        conn.close()
        if not rows:
            results_html = '<div style="color:#9ca3af;padding:20px 0;text-align:center">找不到符合條件的客戶</div>'
        for row in rows:
            route_data = parse_route_json(row["route_plan"] or "")
            order, idx = route_data.get("order",[]), route_data.get("current_index",0)
            history = route_data.get("history",[])
            if row["status"] == "PENDING":
                badge = '<span class="badge b-doc">待處理</span>'
            elif row["status"] != "ACTIVE":
                badge = '<span class="badge b-close">已結案</span>'
            else:
                badge = get_badge(row)
            co = row["current_company"] or row["company"] or ""
            amt = row["approved_amount"] or ""
            disb = row["disbursement_date"] or ""

            route_html = ""
            if order:
                current_co = order[idx] if idx < len(order) else "（已完成）"
                next_co = order[idx+1] if idx+1 < len(order) else ""
                route_html += f'<div class="route-line">目前送件：<b>{h(current_co)}</b> （第{idx+1}/{len(order)}家）</div>'
                if next_co: route_html += f'<div class="route-line">下一家：{h(next_co)}</div>'
                if history:
                    route_html += '<div class="route-line" style="flex-wrap:wrap;gap:4px">歷程：'
                    def _fmt_history(hi):
                        co = h(hi.get("company",""))
                        st = h(hi.get("status",""))
                        amt = hi.get("amount") or ""
                        return f'{co}({st}{amt})' if amt else f'{co}({st})'
                    route_html += " → ".join(_fmt_history(hi) for hi in history[-5:])
                    route_html += '</div>'

            last = (row["last_update"] or "").splitlines()
            last_short = last[-1].strip()[:80] if last else ""
            amount_line = f'核准金額：{amt}' if amt else ""
            disb_line = f'撥款日期：{disb}' if disb else ""
            if role == "admin":
                tl = _build_timeline(row["case_id"])
                timeline_html = '<div onclick="var el=this.nextElementSibling;el.style.display=el.style.display===\'none\'?\'block\':\'none\'" style="cursor:pointer;font-size:12px;color:#8b7355;font-weight:600">▶ 操作歷程</div><div style="display:none;margin-top:6px;max-height:200px;overflow-y:auto">' + tl + '</div>'
            else:
                timeline_html = ""

            results_html += f'''<div class="search-result">
              <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:10px">
                <div style="display:flex;align-items:center;gap:8px">
                  <div style="font-size:16px;font-weight:600">{h(row["customer_name"])}</div>
                  {badge}
                </div>
                <div style="display:flex;align-items:center;gap:8px">
                  <div style="font-size:11px;color:#9ca3af">{h(get_group_name(row["source_group_id"]))}</div>
                  <a href="/edit-pending?case_id={h(row['case_id'])}" style="font-size:11px;color:#0369a1;text-decoration:none">✏️ 編輯</a>
                </div>
              </div>
              <div style="display:grid;grid-template-columns:1fr 1fr;gap:6px;font-size:12px;color:#6b7280;margin-bottom:10px">
                {"<div>身分證：" + h(row["id_no"]) + "</div>" if row["id_no"] else ""}
                {"<div>公司：" + h(co) + "</div>" if co else ""}
                {"<div>" + h(amount_line) + "</div>" if amount_line else ""}
                {"<div>" + h(disb_line) + "</div>" if disb_line else ""}
                <div>建立：{h(row["created_at"][:10]) if row["created_at"] else ""}</div>
                <div>更新：{h(row["updated_at"][:10]) if row["updated_at"] else ""}</div>
              </div>
              {route_html}
              {"<div style=\'margin-top:8px;padding:8px;background:#f9fafb;border-radius:6px;font-size:12px;color:#374151\'>" + h(last_short) + "</div>" if last_short else ""}
              <div style="margin-top:8px">
                {timeline_html}
              </div>
            </div>'''

    return f"""<!DOCTYPE html><html><head>{PAGE_CSS}<title>查詢客戶</title></head><body>
    {make_topnav(role, "search")}
    <div class="page">
      <form method="get" action="/search">
        <div style="background:#faf7f4;border:1px solid #ddd5ca;border-radius:10px;padding:14px 16px;margin-bottom:14px;display:flex;gap:10px;flex-wrap:wrap;align-items:flex-end;">
          <div><div style="font-size:12px;font-weight:600;color:#5a4e40;margin-bottom:3px">姓名／身分證</div><input class="input" name="q" value="{h(q)}" placeholder="搜尋..." autofocus style="width:150px"></div>
          <div><div style="font-size:12px;font-weight:600;color:#5a4e40;margin-bottom:3px">群組</div><select name="grp" class="input" style="width:110px">{grp_opts2}</select></div>
          <div><div style="font-size:12px;font-weight:600;color:#5a4e40;margin-bottom:3px">日期從</div><input type="date" name="date_from" value="{h(date_from)}" class="input" style="width:140px"></div>
          <div><div style="font-size:12px;font-weight:600;color:#5a4e40;margin-bottom:3px">日期至</div><input type="date" name="date_to" value="{h(date_to)}" class="input" style="width:140px"></div>
          <div style="display:flex;gap:6px;align-items:flex-end">
            <button type="submit" class="btn btn-primary">🔍 搜尋</button>
            <a href="/search" class="btn" style="background:#e8e2da;color:#4a3e30;">清除</a>
          </div>
        </div>
      </form>
      {results_html}
    </div></body></html>"""



def apply_adminb_rules(row: dict) -> dict:
    result = {}
    # 年資
    yr = str(row.get("company_years","") or "0").replace("年","").strip()
    try:
        y = float(yr.split()[0]) if yr else 0
        if y < 1:
            result["company_years_display"] = f"客戶填：{yr}年 → 填入：1年"
            result["company_years_val"] = "1"
            result["company_years_adj"] = True
        else:
            result["company_years_display"] = f"填入：{yr}年"
            result["company_years_val"] = yr
            result["company_years_adj"] = False
    except Exception:
        result["company_years_display"] = f"填入：1年（原值：{yr}）"
        result["company_years_val"] = "1"
        result["company_years_adj"] = True
    # 月薪
    sal = str(row.get("company_salary","") or "0")
    try:
        sv = float(sal.replace("萬","").replace("元","").replace(",","").strip())
        if "萬" in sal: sv *= 10000
        if sv < 35000:
            result["salary_display"] = f"客戶填：{sal} → 填入：3.5萬"
            result["salary_val"] = "3.5"
            result["salary_adj"] = True
        else:
            result["salary_display"] = f"填入：{sal}"
            result["salary_val"] = sal
            result["salary_adj"] = False
    except Exception:
        result["salary_display"] = f"填入：3.5萬（原值：{sal}）"
        result["salary_val"] = "3.5"
        result["salary_adj"] = True
    # 居住時間
    ly = str(row.get("live_years","") or "0").replace("年","").strip()
    try:
        lyv = float(ly) if ly else 0
        if lyv < 5:
            result["live_years_display"] = f"客戶填：{ly}年 → 填入：5年"
            result["live_years_val"] = "5"
            result["live_years_adj"] = True
        else:
            result["live_years_display"] = f"填入：{ly}年"
            result["live_years_val"] = ly
            result["live_years_adj"] = False
    except Exception:
        result["live_years_display"] = f"填入：5年（原值：{ly}）"
        result["live_years_val"] = "5"
        result["live_years_adj"] = True
    # 居住狀況
    ls = str(row.get("live_status","") or "")
    rent_kw = ["租","宿舍"]
    own_kw = ["自有","本人名下","配偶名下"]
    warnings = []
    if any(k in ls for k in rent_kw):
        result["live_status_display"] = f"客戶填：{ls} → 填入：親屬（父母或親友）"
        result["live_status_val"] = "親屬"
        result["live_status_adj"] = True
    elif any(k in ls for k in own_kw):
        result["live_status_display"] = f"填入：{ls}"
        result["live_status_val"] = ls
        result["live_status_adj"] = False
        warnings.append("⚠️ 居住狀況「自有」請確認無私設！")
    else:
        result["live_status_display"] = f"填入：{ls}"
        result["live_status_val"] = ls
        result["live_status_adj"] = False
    # 學歷
    edu = str(row.get("education","") or "")
    low_edu = ["國中","國小","小學","其他"]
    if any(k in edu for k in low_edu):
        result["edu_display"] = f"客戶填：{edu} → 填入：高中/職"
        result["edu_val"] = "高中/職"
        result["edu_adj"] = True
    else:
        result["edu_display"] = f"填入：{edu}"
        result["edu_val"] = edu
        result["edu_adj"] = False
    result["warnings"] = warnings
    return result


@app.get("/adminb", response_class=HTMLResponse)
def adminb_page(request: Request, case_id: str = "", saved: str = ""):
    from fastapi.responses import RedirectResponse
    role = check_auth(request)
    if not role: return RedirectResponse("/login")
    if role not in ("admin","adminB"): return RedirectResponse("/report")

    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT case_id, customer_name, id_no, source_group_id, created_at, status FROM customers WHERE status IN ('ACTIVE','PENDING') ORDER BY updated_at DESC LIMIT 200")
    all_customers = cur.fetchall()
    customer = None
    if case_id:
        cur.execute("SELECT * FROM customers WHERE case_id=?", (case_id,))
        row = cur.fetchone()
        if row: customer = dict(row)
    conn.close()

    cust_opts = "<option value=''>— 選擇客戶 —</option>"
    for cu in all_customers:
        sel = "selected" if case_id == cu["case_id"] else ""
        gname = get_group_name(cu["source_group_id"])
        tag = "📋表單" if cu["status"]=="PENDING" else "📊日報"
        cust_opts += f'<option value="{h(cu["case_id"])}" data-gid="{h(cu["source_group_id"])}" {sel}>[{tag}] {h(cu["customer_name"])} {h(cu["id_no"] or "")}（{h(gname)}）</option>'

    ADMINB_CSS = """<style>
body{background:#ece8e2;font-family:'Microsoft JhengHei','PingFang TC',sans-serif;font-size:14px;color:#2c2820;}
.ab-card{background:#faf7f4;border:1px solid #ddd5ca;border-radius:10px;padding:16px;margin-bottom:12px;}
.ab-sec{font-size:13px;font-weight:700;color:#4a3e30;margin-bottom:12px;padding-bottom:8px;border-bottom:1px solid #ece8e2;}
.ab-block{border-radius:7px;padding:12px 14px;margin-bottom:10px;}
.ab-lbl{font-size:11px;font-weight:600;color:#5a4e40;margin-bottom:4px;}
.ab-inp{width:100%;padding:7px 10px;border:1px solid #ddd5ca;border-radius:6px;font-size:13px;background:#fff;box-sizing:border-box;color:#1a1208;font-family:inherit;}
.ab-sel{width:100%;padding:7px 10px;border:1px solid #ddd5ca;border-radius:6px;font-size:13px;background:#fff;color:#1a1208;font-family:inherit;}
.ab-g2{display:grid;grid-template-columns:1fr 1fr;gap:10px;}
.ab-g3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px;}
.ab-plan-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-bottom:10px;}
.ab-plan{display:flex;align-items:center;gap:8px;padding:10px 12px;background:#fff;border:1px solid #ddd5ca;border-radius:8px;cursor:pointer;}
.ab-plan:has(input:checked){border:2px solid #6a5e4e;background:#f5f0eb;}
.ab-plan input{width:16px;height:16px;accent-color:#6a5e4e;flex-shrink:0;}
.ab-plan-name{font-size:13px;font-weight:600;color:#2c2820;}
.ab-plan-sub{font-size:11px;color:#6a5e4e;}
.rule-item{border-radius:6px;padding:8px 12px;font-size:13px;margin-bottom:6px;}
.rule-adj{background:#fef9c3;color:#854d0e;}
.rule-ok{background:#f0fdf4;color:#166534;}
.btn-save{background:#0369a1;color:#fff;border:none;padding:9px 20px;border-radius:8px;font-size:13px;cursor:pointer;font-weight:600;font-family:inherit;}
.btn-dl{background:#6a5e4e;color:#fff;border:none;padding:9px 16px;border-radius:8px;font-size:13px;cursor:pointer;font-weight:600;font-family:inherit;text-decoration:none;display:inline-block;}
.btn-qm{background:#4e7055;color:#fff;border:none;padding:9px 16px;border-radius:8px;font-size:13px;cursor:pointer;font-weight:600;font-family:inherit;text-decoration:none;display:inline-block;}
.ab-card{overflow-x:auto}
</style>"""

    # 群組選項
    conn2 = get_conn(); cur2 = conn2.cursor()
    cur2.execute("SELECT group_id, group_name FROM groups WHERE group_type='SALES_GROUP' AND is_active=1 ORDER BY group_name")
    ab_groups = cur2.fetchall(); conn2.close()
    grp_filter_opts = '<option value="">全部群組</option>' + "".join(
        f'<option value="{h(g["group_id"])}">{h(g["group_name"])}</option>' for g in ab_groups)

    if not customer:
        # 預先產生 JSON，避免 f-string 大括號衝突
        cust_json = json.dumps([{"id":cu["case_id"],"name":cu["customer_name"],"idno":cu["id_no"] or "","gid":cu["source_group_id"],"gname":get_group_name(cu["source_group_id"])} for cu in all_customers], ensure_ascii=False).replace("</", "<\\/").replace("\u2028", "\\u2028").replace("\u2029", "\\u2029")
        return """<!DOCTYPE html><html><head>""" + PAGE_CSS + ADMINB_CSS + """<title>行政B作業</title></head><body>
    """ + make_topnav(role,"adminb") + """
    <div class="page">
      <div class="ab-card">
        <div class="ab-sec">選擇客戶</div>
        <div style="display:flex;gap:8px;margin-bottom:10px;flex-wrap:wrap">
          <input type="text" id="ab-search" class="ab-inp" placeholder="輸入姓名或身分證搜尋..." autofocus style="flex:1;min-width:150px">
          <select id="ab-grp" class="ab-sel" style="width:130px">""" + grp_filter_opts + """</select>
          <button onclick="doSearch()" class="btn-save" style="padding:7px 16px">🔍 搜尋</button>
        </div>
        <div id="ab-results" style="max-height:400px;overflow-y:auto"></div>
      </div>
    </div>
    <script>
    const allCusts=""" + cust_json + """;
    function _escHtml(s){return String(s==null?'':s).replace(/[&<>"']/g,function(c){return({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'})[c];});}
    document.getElementById('ab-search').addEventListener('keydown',function(e){if(e.key==='Enter')doSearch()});
    function doSearch(){
      const q=document.getElementById('ab-search').value.trim().toLowerCase();
      const g=document.getElementById('ab-grp').value;
      const box=document.getElementById('ab-results');
      if(!q&&!g){box.innerHTML='<div style="color:#999;font-size:13px;padding:12px">請輸入姓名或身分證搜尋</div>';return;}
      const matches=allCusts.filter(c=>{
        const matchQ=!q||c.name.toLowerCase().includes(q)||(c.idno&&c.idno.toLowerCase().includes(q));
        const matchG=!g||c.gid===g;
        return matchQ&&matchG;
      }).slice(0,30);
      if(!matches.length){box.innerHTML='<div style="color:#999;font-size:13px;padding:12px">找不到符合的客戶</div>';return;}
      box.innerHTML=matches.map(c=>'<div onclick="location.href=\\'/adminb?case_id='+encodeURIComponent(c.id)+'\\'" style="padding:10px 14px;border-bottom:1px solid #ece8e2;cursor:pointer;display:flex;justify-content:space-between;align-items:center" onmouseover="this.style.background=\\'#f5f0eb\\'" onmouseout="this.style.background=\\'\\'"><div><div style="font-size:14px;font-weight:600">'+_escHtml(c.name)+'</div><div style="font-size:12px;color:#6a5e4e">'+_escHtml(c.idno||'-')+'</div></div><div style="font-size:11px;color:#999">'+_escHtml(c.gname)+'</div></div>').join('');
    }
    </script>
    </body></html>"""

    rules = apply_adminb_rules(customer)
    name = customer.get("customer_name","")
    id_no = customer.get("id_no","")
    phone = customer.get("phone","") or "-"
    gname = get_group_name(customer.get("source_group_id",""))
    created = (customer.get("created_at","") or "")[:10]
    co = customer.get("company_name_detail","") or customer.get("company","") or "-"
    salary = customer.get("company_salary","") or "-"
    labor = customer.get("eval_labor_ins","") or "-"
    sel_plans = customer.get("adminb_selected_plans","") or ""

    def rule_item(label, display, adj):
        cls = "rule-adj" if adj else "rule-ok"
        icon = "→" if adj else "✓"
        return f'<div class="rule-item {cls}"><span style="font-size:11px;font-weight:700;">{h(label)}</span>　{icon} {h(display)}</div>'

    saved_html = '<div style="background:#dcfce7;border:1px solid #86efac;border-radius:8px;padding:12px 16px;margin-bottom:12px;font-size:14px;font-weight:600;color:#166534;">✅ 資料已儲存完成</div>' if saved else ""

    rules_html = ""
    rules_html += rule_item("年資", rules["company_years_display"], rules["company_years_adj"])
    rules_html += rule_item("月薪", rules["salary_display"], rules["salary_adj"])
    rules_html += rule_item("居住時間", rules["live_years_display"], rules["live_years_adj"])
    rules_html += rule_item("居住狀況", rules["live_status_display"], rules["live_status_adj"])
    rules_html += rule_item("學歷", rules["edu_display"], rules["edu_adj"])

    warns_html = "".join(
        f'<div style="background:#fef2f2;border:1px solid #fca5a5;border-radius:7px;padding:10px 14px;font-size:13px;font-weight:700;color:#991b1b;margin-top:8px;">{h(w)}</div>'
        for w in rules["warnings"]
    )

    return f"""<!DOCTYPE html><html><head>{PAGE_CSS}{ADMINB_CSS}<title>行政B - {h(name)}</title></head><body>
    {make_topnav(role,"adminb")}
    <div class="page">
    <div style="margin-bottom:12px;">
      <input id="custSearch" class="ab-inp" placeholder="🔍 輸入姓名/群組快速篩選客戶..." style="margin-bottom:6px;">
      <select id="custSelect" class="ab-inp" onchange="if(this.value)location.href='/adminb?case_id='+this.value">{cust_opts}</select>
    </div>
    <div class="ab-card" id="custCard" style="position:sticky;top:0;z-index:20;">
      <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:12px;">
        <div><div style="font-size:20px;font-weight:700;color:#1a1208;">{h(name)}</div>
          <div style="font-size:13px;color:#6a5e4e;">{h(id_no)}　{h(gname)}　{h(created)}</div></div>
        <div style="display:flex;gap:8px;align-items:center">
          <a href="/edit-pending?case_id={h(case_id)}" style="font-size:12px;color:#0369a1;text-decoration:none;padding:4px 12px;border:1px solid #93c5fd;border-radius:20px">✏️ 編輯</a>
          <span style="background:#e0f2fe;color:#0369a1;font-size:12px;padding:4px 12px;border-radius:20px;font-weight:600;">ACTIVE</span>
        </div>
      </div>
      <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:8px;">
        <div><div style="font-size:11px;font-weight:600;color:#6a5e4e;">電話</div><div style="font-size:13px;color:#1a1208;font-weight:500;">{h(phone)}</div></div>
        <div><div style="font-size:11px;font-weight:600;color:#6a5e4e;">公司</div><div style="font-size:13px;color:#1a1208;font-weight:500;">{h(co)}</div></div>
        <div><div style="font-size:11px;font-weight:600;color:#6a5e4e;">月薪</div><div style="font-size:13px;color:#1a1208;font-weight:500;">{h(salary)}</div></div>
        <div><div style="font-size:11px;font-weight:600;color:#6a5e4e;">勞保</div><div style="font-size:13px;color:#1a1208;font-weight:500;">{h(labor)}</div></div>
      </div>
    </div>
    {saved_html}
    <div class="ab-card">
      <div class="ab-sec">確認資料（系統自動調整）</div>
      {rules_html}{warns_html}
    </div>
    <form method="post" action="/adminb/save">
    <input type="hidden" name="case_id" value="{h(case_id)}">
    <div class="ab-card">
      <div class="ab-sec">判別方案（勾選要送的方案）</div>
      <div style="font-size:11px;font-weight:700;color:#6a5e4e;margin-bottom:8px;">Excel 方案</div>
      <div class="ab-plan-grid">
        {"".join(f'<label class="ab-plan"><input type="checkbox" name="plans" value="{v}"><div><div class="ab-plan-name">{n}</div><div class="ab-plan-sub">{s}</div></div></label>' for v,n,s in [
          ("亞太商品","亞太商品","熊速貸 12萬 30期"),("亞太機車15萬","亞太機車15萬","機車動擔 15萬 36期"),
          ("亞太工會機車","亞太工會機車","工會機車動擔 15萬"),("亞太機車25萬","亞太機車25萬","機車動擔 25萬 48期"),
          ("和裕機車","和裕機車","維力機車專 15萬 24期"),("和裕商品","和裕商品","維力商品專 12萬 24期"),
          ("21機車12萬","21機車12萬","12萬 24期"),("21機車25萬","21機車25萬","25萬 48期"),
          ("21商品","21商品","12萬 24期"),("第一","第一","30萬 24期"),("貸救補","貸救補","10萬 24期"),
        ])}
        <label class="ab-plan" style="background:#4e7055;border-color:#3a5040;"><input type="checkbox" name="plans" value="喬美" style="accent-color:#fff;"><div><div class="ab-plan-name" style="color:#fff;">喬美</div><div class="ab-plan-sub" style="color:#c8e6ca;">PDF電子簽名</div></div></label>
      </div>
      <div style="font-size:11px;font-weight:700;color:#6a5e4e;margin-bottom:8px;">填寫表方案</div>
      <div class="ab-plan-grid">
        {"".join(f'<label class="ab-plan"><input type="checkbox" name="plans" value="{v}"><div><div class="ab-plan-name">{n}</div><div class="ab-plan-sub">{s}</div></div></label>' for v,n,s in [
          ("麻吉機車","麻吉機車","10萬 24期"),("麻吉手機","麻吉手機","10萬 24期"),
          ("手機分期","手機分期","填寫表"),("分貝機車","分貝機車","改裝填寫表"),
          ("分貝汽車","分貝汽車","改裝填寫表"),("21汽車","21汽車","貸款100% 利率16%"),
        ])}
      </div>
    </div>
    <div class="ab-card">
      <div class="ab-sec">補充資料</div>
      <div class="ab-block" data-plans="亞太商品,亞太機車15萬,亞太工會機車,亞太機車25萬" style="background:#e0f2fe;">
        <div style="font-size:12px;font-weight:700;color:#0369a1;margin-bottom:10px;">亞太（商品／機車／工會）</div>
        <div class="ab-g3" style="margin-bottom:10px;">
          <div><div class="ab-lbl">資金用途</div><select name="at_fund" class="ab-sel"><option value="">請選擇</option>{"".join(f'<option value="{o}" {"selected" if customer.get("adminb_fund_use","")==o else ""}>{o}</option>' for o in ["I-1教育費","I-2醫藥費","I-3出國旅遊","I-4創業","II-1購買交通工具","II-2購買手機","II-3購買3C產品","III-1交友","III-2健身&醫美","III-3美容課程","IV-1個人理財投資(含不動產、裝修、理財商品)","V-1生活周轉金","V-2整合負債(償還銀行/融資等)"])}</select></div>
          <div><div class="ab-lbl">行業類別</div><select name="at_industry" class="ab-sel"><option value="">請選擇</option>{"".join(f"<option>{o}</option>" for o in ["餐飲與服務業","製造業","建築與營造","軍警與公教","科技與資訊","運輸與物流","金融與保險業","批發與零售業","醫療與教育","農林漁牧業","自由職業","其他"])}</select></div>
          <div><div class="ab-lbl">職務</div><select name="at_role" class="ab-sel"><option value="">請選擇</option>{"".join(f"<option>{o}</option>" for o in ["行政與內勤","勞力與現場","銷售與業務","財務與專業","技術與工程","教學與醫護","管理與經營","自營與自由"])}</select></div>
        </div>
        <div style="font-size:11px;font-weight:700;color:#0369a1;margin-bottom:8px;">機車額外資料</div>
        <div class="ab-g3">
          <div><div class="ab-lbl">廠牌</div><input name="at_brand" class="ab-inp" placeholder="光陽機車" value="{h(customer.get('adminb_brand','') or '')}"></div>
          <div><div class="ab-lbl">牌照號碼</div><input name="at_plate" class="ab-inp" placeholder="MMX-0286" value="{h(customer.get('vehicle_plate','') or '')}"></div>
          <div><div class="ab-lbl">車輛型式</div><input name="at_vtype" class="ab-inp" placeholder="SE22BJ" value="{h(customer.get('adminb_vehicle_type','') or '')}"></div>
          <div><div class="ab-lbl">引擎號碼</div><input name="at_engine" class="ab-inp" placeholder="SE22BJ-105579" value="{h(customer.get('adminb_engine_no','') or '')}"></div>
          <div><div class="ab-lbl">排氣量</div><input name="at_disp" class="ab-inp" placeholder="111" value="{h(customer.get('adminb_displacement','') or '')}"></div>
          <div><div class="ab-lbl">顏色</div><input name="at_color" class="ab-inp" placeholder="綠" value="{h(customer.get('adminb_color','') or '')}"></div>
          <div><div class="ab-lbl">車身號碼</div><input name="at_body" class="ab-inp" placeholder="RFGBK..." value="{h(customer.get('adminb_body_no','') or '')}"></div>
          <div><div class="ab-lbl">出廠年月</div><input name="at_mfg" class="ab-inp" placeholder="2023/08" value="{h(customer.get('adminb_mfg_date','') or '')}"></div>
        </div>
      </div>
      <div class="ab-block" data-plans="和裕機車,和裕商品" style="background:#f0fdf4;">
        <div style="font-size:12px;font-weight:700;color:#166534;margin-bottom:10px;">和裕（機車／商品）</div>
        <div class="ab-g2" style="margin-bottom:10px;">
          <div><div class="ab-lbl">行業類別</div><select name="hr_industry" class="ab-sel"><option value="">請選擇</option>{"".join(f'<option {"selected" if customer.get("adminb_hr_industry","")==o else ""}>{o}</option>' for o in ["服務業","餐飲業","科技業","軍人","運輸業","倉儲業","金融業","製造業","營造業","電商網拍業","農狩林牧業","礦業","漁業","證券期貨業","保險業","不動產業","公教人員","水電燃氣業","通信業","社團個人服務","其它"])}</select></div>
          <div><div class="ab-lbl">照會時間</div><input name="hr_contact" class="ab-inp" placeholder="平日下午2-5點" value="{h(customer.get('adminb_contact_time','') or '')}"></div>
        </div>
        <div style="font-size:11px;font-weight:700;color:#166534;margin-bottom:8px;">撥款資料</div>
        <div class="ab-g2" style="margin-bottom:10px;">
          <div><div class="ab-lbl">銀行名稱</div><input name="hr_bank" class="ab-inp" placeholder="台灣銀行" value="{h(customer.get('adminb_bank','') or '')}"></div>
          <div><div class="ab-lbl">分行</div><input name="hr_branch" class="ab-inp" placeholder="苗栗分行" value="{h(customer.get('adminb_branch','') or '')}"></div>
        </div>
        <div style="font-size:11px;font-weight:700;color:#166534;margin-bottom:8px;">商品資料</div>
        <div class="ab-g2">
          <div><div class="ab-lbl">廠牌+商品</div><input name="hr_product" class="ab-inp" placeholder="三陽/安卓手機" value="{h(customer.get('adminb_product','') or '')}"></div>
          <div><div class="ab-lbl">型號或車號</div><input name="hr_model" class="ab-inp" placeholder="677-NSY/OPPO A77" value="{h(customer.get('adminb_model','') or '')}"></div>
        </div>
      </div>
      <div class="ab-block" data-plans="貸救補" style="background:#fef9c3;">
        <div style="font-size:12px;font-weight:700;color:#854d0e;margin-bottom:10px;">貸救補</div>
        <div class="ab-g2">
          <div><div class="ab-lbl">商品名稱</div><input name="lj_pname" class="ab-inp" placeholder="vivo" value="{h(customer.get('adminb_product_name','') or '')}"></div>
          <div><div class="ab-lbl">型號</div><input name="lj_pmodel" class="ab-inp" placeholder="vivo V60" value="{h(customer.get('adminb_product_model','') or '')}"></div>
        </div>
      </div>
      <div class="ab-block" data-plans="麻吉機車,麻吉手機" style="background:#fdf2f8;">
        <div style="font-size:12px;font-weight:700;color:#9d174d;margin-bottom:10px;">麻吉（機車／手機）</div>
        <div class="ab-g2">
          <div><div class="ab-lbl">商品廠牌</div><input name="mj_brand" class="ab-inp" placeholder="山葉 / iPhone" value="{h(customer.get('adminb_mj_brand','') or '')}"></div>
          <div><div class="ab-lbl">型號</div><input name="mj_model" class="ab-inp" placeholder="JQ5-063 / 16 Pro" value="{h(customer.get('adminb_mj_model','') or '')}"></div>
        </div>
      </div>
      <div class="ab-block" data-plans="21汽車" style="background:#fef2f2;">
        <div style="font-size:12px;font-weight:700;color:#991b1b;margin-bottom:10px;">21汽車（利率固定 16%）</div>
        <div class="ab-g2">
          <div><div class="ab-lbl">專案名稱</div><input name="car_proj" class="ab-inp" value="{h(customer.get('adminb_21car_project','') or '')}"></div>
          <div><div class="ab-lbl">車價（實際售價＝參考車價金額共用）</div><input name="car_price" class="ab-inp" value="{h(customer.get('adminb_21car_price','') or '')}"></div>
          <div><div class="ab-lbl">天書</div><input name="car_src" class="ab-inp" value="{h(customer.get('adminb_21car_ref_src','') or '')}"></div>
          <div><div class="ab-lbl">貸款金額</div><input name="car_amt" class="ab-inp" placeholder="50萬" value="{h(customer.get('adminb_21car_amount','') or '')}"></div>
          <div><div class="ab-lbl">期數</div><input name="car_period" class="ab-inp" placeholder="60" value="{h(customer.get('adminb_21car_period','') or '')}"></div>
          <div><div class="ab-lbl">月付金</div><input name="car_monthly" class="ab-inp" placeholder="9500" value="{h(customer.get('adminb_21car_monthly','') or '')}"></div>
          <div><div class="ab-lbl">資金用途</div><input name="car_fund" class="ab-inp" placeholder="購車" value="{h(customer.get('adminb_21car_fund','') or '')}"></div>
          <div><div class="ab-lbl">是否有信用卡</div><select name="car_hascc" class="ab-sel"><option value="">請選</option><option value="有" {"selected" if (customer.get('adminb_21car_hascc','') or '')=='有' else ''}>有</option><option value="無" {"selected" if (customer.get('adminb_21car_hascc','') or '')=='無' else ''}>無</option></select></div>
        </div>
      </div>
      <div class="ab-block" data-plans="喬美" style="background:#ede9fe;">
        <div style="font-size:12px;font-weight:700;color:#5b21b6;margin-bottom:10px;">喬美（PDF電子簽名）</div>
        <div class="ab-g2" style="margin-bottom:10px;">
          <div><div class="ab-lbl">手機型號</div><input name="qm_model" class="ab-inp" placeholder="iPhone 16 Pro Max" value="{h(customer.get('product_model','') or '')}"></div>
          <div><div class="ab-lbl">IMEI</div><input name="qm_imei" class="ab-inp" placeholder="356194482654922" value="{h(customer.get('product_imei','') or '')}"></div>
        </div>
        <div style="background:#fff;border:1px dashed #5b21b6;border-radius:8px;padding:12px;margin:10px 0;">
          <div style="font-size:12px;font-weight:700;color:#5b21b6;margin-bottom:8px;">📝 電子簽名（一鍵簽名+下載 PDF）</div>
          <div style="font-size:11px;color:#6b5b8e;margin-bottom:8px;">說明：下方畫板簽一次，即可同時用於 第1頁「申請人正楷簽名」+ 第2頁「立約定書人」</div>
          <canvas id="qmSignPad" width="500" height="150" style="border:2px solid #5b21b6;background:#fff;cursor:crosshair;display:block;border-radius:6px;width:100%;max-width:500px;"></canvas>
          <div style="margin-top:8px;display:flex;gap:8px;flex-wrap:wrap;">
            <button type="button" onclick="qmClear()" style="padding:6px 14px;background:#888;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:13px;">🗑 清除</button>
            <button type="button" onclick="qmSignAndDownload()" style="padding:6px 14px;background:#e91e63;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:13px;font-weight:700;">✍️ 簽名並下載喬美 PDF</button>
          </div>
        </div>
        <div style="font-size:11px;font-weight:700;color:#5b21b6;margin-bottom:8px;">信用資料</div>
        <div class="ab-g3">
          <div><div class="ab-lbl">發卡銀行</div><input name="qm_cbank" class="ab-inp" placeholder="星展銀行" value="{h(customer.get('adminb_credit_bank','') or '')}"></div>
          <div><div class="ab-lbl">卡號</div><input name="qm_cno" class="ab-inp" placeholder="6480-6295-1099-819" value="{h(customer.get('adminb_credit_no','') or '')}"></div>
          <div><div class="ab-lbl">有效日期（年/月）</div><div style="display:flex;gap:6px;"><input name="qm_exp_y" class="ab-inp" placeholder="年" style="width:50%;" value="{h((customer.get('adminb_credit_exp','') or '').split('/')[0] if '/' in (customer.get('adminb_credit_exp','') or '') else '')}"><input name="qm_exp_m" class="ab-inp" placeholder="月" style="width:50%;" value="{h((customer.get('adminb_credit_exp','') or '').split('/')[1] if '/' in (customer.get('adminb_credit_exp','') or '') else '')}"></div></div>
          <div><div class="ab-lbl">額度（萬）</div><input name="qm_climit" class="ab-inp" placeholder="10" value="{h(customer.get('adminb_credit_limit','') or '')}"></div>
          <div><div class="ab-lbl">近二月有無遲繳</div><select name="qm_clate" class="ab-sel"><option value="無">無</option><option value="有" {"selected" if (customer.get('adminb_credit_late','') or '')=='有' else ''}>有</option></select></div>
          <div><div class="ab-lbl">月付金</div><input name="qm_cpay" class="ab-inp" placeholder="3000" value="{h(customer.get('adminb_credit_pay','') or '')}"></div>
        </div>
      </div>
      <div class="ab-block" data-plans="分貝汽車,分貝機車,21汽車" style="background:#ece8e2;">
        <div style="font-size:12px;font-weight:700;color:#4a3e30;margin-bottom:8px;">照會時間（分貝汽車／分貝機車／21汽車）</div>
        <input name="contact_time" class="ab-inp" placeholder="平日下午2-5點" value="{h(customer.get('adminb_contact_time','') or '')}">
      </div>
    </div>
    <div class="ab-card">
      <div class="ab-sec">儲存與下載</div>
      <div style="display:flex;gap:10px;flex-wrap:wrap;align-items:center;">
        <button type="submit" class="btn-save">💾 儲存資料</button>
        <a href="/adminb/download-excel?case_id={h(case_id)}" class="btn-dl" onclick="return confirm('將根據勾選的方案下載 Excel/TXT，確定嗎？')">📥 下載EXCEL/TXT</a>
      </div>
      <div style="font-size:12px;color:#8a7a68;margin-top:8px;">請先儲存資料再下載；喬美 PDF 請使用「補充資料」區的「簽名並下載」按鈕</div>
    </div>

    <script>
    // 喬美簽名 canvas（補充資料區）
    var qmCanvas, qmCtx, qmDrawing = false;
    function qmInit() {{
      qmCanvas = document.getElementById('qmSignPad');
      if (!qmCanvas) return;
      qmCtx = qmCanvas.getContext('2d');
      qmCtx.lineWidth = 2.5;
      qmCtx.lineCap = 'round';
      qmCtx.strokeStyle = '#000';
      function getXY(e) {{
        var r = qmCanvas.getBoundingClientRect();
        var sx = qmCanvas.width / r.width;
        var sy = qmCanvas.height / r.height;
        if (e.touches) {{
          return [(e.touches[0].clientX - r.left) * sx, (e.touches[0].clientY - r.top) * sy];
        }}
        return [(e.clientX - r.left) * sx, (e.clientY - r.top) * sy];
      }}
      qmCanvas.onmousedown = function(e) {{ var p = getXY(e); qmDrawing = true; qmCtx.beginPath(); qmCtx.moveTo(p[0], p[1]); }};
      qmCanvas.onmousemove = function(e) {{ if(!qmDrawing) return; var p = getXY(e); qmCtx.lineTo(p[0], p[1]); qmCtx.stroke(); }};
      qmCanvas.onmouseup = function() {{ qmDrawing = false; }};
      qmCanvas.onmouseleave = function() {{ qmDrawing = false; }};
      qmCanvas.ontouchstart = function(e) {{ e.preventDefault(); var p = getXY(e); qmDrawing = true; qmCtx.beginPath(); qmCtx.moveTo(p[0], p[1]); }};
      qmCanvas.ontouchmove = function(e) {{ e.preventDefault(); if(!qmDrawing) return; var p = getXY(e); qmCtx.lineTo(p[0], p[1]); qmCtx.stroke(); }};
      qmCanvas.ontouchend = function() {{ qmDrawing = false; }};
    }}
    function qmClear() {{
      if (qmCtx) qmCtx.clearRect(0, 0, qmCanvas.width, qmCanvas.height);
    }}
    function qmSignAndDownload() {{
      if (!qmCanvas) {{ alert('簽名版未載入'); return; }}
      var imgData = qmCtx.getImageData(0, 0, qmCanvas.width, qmCanvas.height).data;
      var hasInk = false;
      for (var i = 3; i < imgData.length; i += 40) {{ if (imgData[i] > 0) {{ hasInk = true; break; }} }}
      if (!hasInk) {{ alert('請先簽名'); return; }}
      var dataUrl = qmCanvas.toDataURL('image/png');
      // 1. 先送出整個 adminB 表單（含型號/IMEI/信用卡等補充資料）
      var form = document.querySelector('form');
      var fd = new FormData(form);
      fetch('/adminb/save', {{ method: 'POST', body: fd }})
        .then(function() {{
          // 2. 存簽名
          return fetch('/adminb/save-signature', {{
            method: 'POST',
            headers: {{'Content-Type': 'application/json'}},
            body: JSON.stringify({{case_id: '{h(case_id)}', type: 'both', data: dataUrl}})
          }});
        }})
        .then(r => r.json()).then(d => {{
          if (d.ok) {{
            // 3. 下載 PDF
            window.location.href = '/adminb/download-qiaomei?case_id={h(case_id)}';
          }} else {{
            alert('簽名儲存失敗：' + (d.error || '未知錯誤'));
          }}
        }}).catch(function(e) {{ alert('錯誤：' + e); }});
    }}
    window.addEventListener('DOMContentLoaded', qmInit);
    </script>
    </form>
    </div>
    <script>
    var saved = "{sel_plans}".split(",").filter(Boolean);
    document.querySelectorAll('input[name="plans"]').forEach(function(cb){{
      if(saved.indexOf(cb.value)>=0) cb.checked = true;
    }});
    // === 補充資料區塊依勾選方案顯示/隱藏 ===
    function refreshBlocks() {{
      var checked = [];
      document.querySelectorAll('input[name="plans"]:checked').forEach(function(cb){{
        checked.push(cb.value);
      }});
      document.querySelectorAll('.ab-block[data-plans]').forEach(function(blk){{
        var mine = (blk.getAttribute('data-plans')||'').split(',').filter(Boolean);
        var hit = mine.some(function(p){{ return checked.indexOf(p) >= 0; }});
        blk.style.display = hit ? '' : 'none';
      }});
    }}
    document.querySelectorAll('input[name="plans"]').forEach(function(cb){{
      cb.addEventListener('change', refreshBlocks);
    }});
    refreshBlocks();
    // === 客戶下拉搜尋過濾 ===
    (function(){{
      var si = document.getElementById('custSearch');
      var se = document.getElementById('custSelect');
      if (!si || !se) return;
      var origOpts = Array.prototype.slice.call(se.options);
      si.addEventListener('input', function(){{
        var kw = si.value.trim().toLowerCase();
        se.innerHTML = '';
        origOpts.forEach(function(o){{
          if (!kw || o.text.toLowerCase().indexOf(kw) >= 0 || o.value === '') {{
            se.appendChild(o.cloneNode(true));
          }}
        }});
      }});
    }})();
    </script>
    </body></html>"""


@app.post("/adminb/save")
async def adminb_save(request: Request):
    from fastapi.responses import RedirectResponse
    role = check_auth(request)
    if not role or role not in ("admin","adminB"):
        return RedirectResponse("/login")
    form = await request.form()
    case_id = form.get("case_id","")
    plans = ",".join(form.getlist("plans"))
    ey = form.get("qm_exp_y","")
    em = form.get("qm_exp_m","")
    credit_exp = f"{ey}/{em}" if ey or em else ""
    conn = get_conn(); cur = conn.cursor()
    fields = {
        "adminb_selected_plans": plans,
        "adminb_industry": form.get("at_industry",""),
        "adminb_hr_industry": form.get("hr_industry",""),
        "adminb_brand": form.get("at_brand",""),
        "adminb_role": form.get("at_role",""),
        "adminb_contact_time": form.get("contact_time","") or form.get("hr_contact",""),
        "adminb_bank": form.get("hr_bank",""),
        "adminb_branch": form.get("hr_branch",""),
        "adminb_product": form.get("hr_product",""),
        "adminb_model": form.get("hr_model",""),
        "adminb_product_name": form.get("lj_pname",""),
        "adminb_product_model": form.get("lj_pmodel",""),
        "adminb_21car_project": form.get("car_proj",""),
        "adminb_21car_price": form.get("car_price",""),
        "adminb_21car_ref_src": form.get("car_src",""),
        "adminb_21car_ref_price": form.get("car_refp",""),
        "adminb_21car_rate": "16%",
        "adminb_21car_ref_price": form.get("car_price",""),
        "adminb_21car_amount": form.get("car_amt",""),
        "adminb_21car_period": form.get("car_period",""),
        "adminb_21car_monthly": form.get("car_monthly",""),
        "adminb_21car_fund": form.get("car_fund",""),
        "adminb_21car_hascc": form.get("car_hascc",""),
        "adminb_mj_brand": form.get("mj_brand",""),
        "adminb_mj_model": form.get("mj_model",""),
        "product_model": form.get("qm_model",""),
        "product_imei": form.get("qm_imei",""),
        "adminb_credit_bank": form.get("qm_cbank",""),
        "adminb_credit_no": form.get("qm_cno",""),
        "adminb_credit_exp": credit_exp,
        "adminb_credit_limit": form.get("qm_climit",""),
        "adminb_credit_late": form.get("qm_clate","無"),
        "adminb_credit_pay": form.get("qm_cpay",""),
        "adminb_vehicle_type": form.get("at_vtype",""),
        "adminb_engine_no": form.get("at_engine",""),
        "adminb_displacement": form.get("at_disp",""),
        "adminb_mfg_date": form.get("at_mfg",""),
        "adminb_color": form.get("at_color",""),
        "adminb_body_no": form.get("at_body",""),
        "adminb_fund_use": form.get("at_fund",""),
        "vehicle_plate": form.get("at_plate",""),
    }
    # 套用「確認資料」自動調整規則，回寫到主要欄位
    cur.execute("SELECT * FROM customers WHERE case_id=?", (case_id,))
    row = cur.fetchone()
    if row:
        rules = apply_adminb_rules(dict(row))
        fields["company_years"] = rules.get("company_years_val", "")
        fields["company_salary"] = rules.get("salary_val", "")
        fields["live_years"] = rules.get("live_years_val", "")
        fields["live_status"] = rules.get("live_status_val", "")
        fields["education"] = rules.get("edu_val", "")
    set_clause = ", ".join(f"{k}=?" for k in fields)
    vals = list(fields.values()) + [now_iso(), case_id]
    cur.execute(f"UPDATE customers SET {set_clause}, updated_at=? WHERE case_id=?", vals)
    conn.commit(); conn.close()
    return RedirectResponse(f"/adminb?case_id={case_id}&saved=1", status_code=303)




@app.get("/edit-pending", response_class=HTMLResponse)
def edit_pending_get(request: Request, case_id: str = ""):
    from fastapi.responses import RedirectResponse
    role = check_auth(request)
    if not role: return RedirectResponse("/login")
    if not case_id: return RedirectResponse("/pending-customers")
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT * FROM customers WHERE case_id=?", (case_id,))
    row = cur.fetchone()
    if not row: conn.close(); return RedirectResponse("/pending-customers")
    r = dict(row)
    cur.execute("SELECT group_id, group_name FROM groups WHERE group_type='SALES_GROUP' AND is_active=1 ORDER BY group_name")
    all_groups = cur.fetchall()
    conn.close()
    def v(k): return r.get(k,"") or ""
    grp_opts = '<option value="">請選擇</option>' + "".join(f'<option value="{h(g["group_id"])}" {"selected" if v("source_group_id")==g["group_id"] else ""}>{h(g["group_name"])}</option>' for g in all_groups)
    cities = ["台北市","新北市","桃園市","台中市","台南市","高雄市","基隆市","新竹市","新竹縣","苗栗縣","彰化縣","南投縣","雲林縣","嘉義市","嘉義縣","屏東縣","宜蘭縣","花蓮縣","台東縣","澎湖縣","金門縣","連江縣"]
    def csel(nm,val): return f'<select name="{nm}" class="ep">' + "".join(f'<option {"selected" if o==val else ""}>{o}</option>' for o in cities) + "</select>"
    lsame = v("live_same_as_reg")=="1"
    import json as _json
    try:
        debt_data = _json.loads(v("debt_list")) if v("debt_list") else []
    except Exception:
        debt_data = []
    # 負債明細轉成可編輯的 JSON 給前端
    import json as _json2
    debt_json = _json2.dumps(debt_data, ensure_ascii=False)
    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>編輯 {h(v("customer_name"))}</title>
{PAGE_CSS}
<style>
.page{{max-width:860px;margin:24px auto;padding:0 16px 40px;}}
.card{{background:#faf6f2;border-radius:10px;padding:18px;margin-bottom:14px;border:1px solid #ddd5ca;}}
.sec{{font-size:13px;font-weight:700;color:#5a4e40;margin-bottom:12px;padding-bottom:7px;border-bottom:1px solid #ddd5ca;}}
.g2{{display:grid;grid-template-columns:1fr 1fr;gap:12px;}}
.g3{{display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px;}}
label{{display:block;font-size:12px;font-weight:600;color:#5a4e40;margin-bottom:4px;}}
.ep{{width:100%;padding:8px 10px;border:1px solid #c8bfb5;border-radius:6px;font-size:13px;font-family:inherit;background:#fff;box-sizing:border-box;}}
.btn-s{{background:#6a5e4e;color:#fff;border:none;padding:10px 28px;border-radius:8px;font-size:15px;font-weight:700;cursor:pointer;font-family:inherit;}}
.btn-b{{background:#e8e2da;color:#4a3e30;border:1px solid #c8bfb5;padding:10px 20px;border-radius:8px;font-size:14px;text-decoration:none;display:inline-block;}}
</style></head><body>
{make_topnav(role, "edit")}
<div class="page">
<div style="font-size:20px;font-weight:700;margin-bottom:18px;">{h(v("customer_name"))} 資料編輯</div>
<form method="post" action="/edit-pending">
<input type="hidden" name="case_id" value="{h(case_id)}">
<div class="card"><div class="sec">所屬群組</div>
  <div><label>群組</label><select name="grp" class="ep">{grp_opts}</select></div>
</div>
<div class="card"><div class="sec">基本資料</div><div class="g2">
  <div><label>姓名</label><input name="cname" class="ep" value="{h(v("customer_name"))}"></div>
  <div><label>身分證</label><input name="idno" class="ep" value="{h(v("id_no"))}" style="text-transform:uppercase"></div>
  <div><label>出生年月日</label><input name="birth" class="ep" value="{h(v("birth_date"))}"></div>
  <div><label>行動電話</label><input name="phone" class="ep" value="{h(v("phone"))}"></div>
  <div><label>Email</label><input name="email" class="ep" value="{h(v("email"))}"></div>
  <div><label>LINE ID</label><input name="line" class="ep" value="{h(v("line_id"))}"></div>
  <div><label>電信業者</label><select name="carrier" class="ep"><option value="">請選擇</option><option {"selected" if v("carrier")=="中華電信" else ""}>中華電信</option><option {"selected" if v("carrier")=="遠傳電信" else ""}>遠傳電信</option><option {"selected" if v("carrier")=="台灣大哥大" else ""}>台灣大哥大</option><option {"selected" if v("carrier")=="台灣之星" else ""}>台灣之星</option><option {"selected" if v("carrier")=="亞太電信" else ""}>亞太電信</option><option {"selected" if v("carrier")=="其他" else ""}>其他</option></select></div>
  <div><label>客戶FB</label><input name="fb" class="ep" value="{h(v("fb"))}"></div>
  <div><label>婚姻狀態</label><select name="marry" class="ep"><option value="">請選擇</option><option {"selected" if v("marriage")=="未婚" else ""}>未婚</option><option {"selected" if v("marriage")=="已婚" else ""}>已婚</option></select></div>
  <div><label>最高學歷</label><select name="edu" class="ep"><option value="">請選擇</option><option {"selected" if v("education")=="高中/職" else ""}>高中/職</option><option {"selected" if v("education")=="專科/大學" else ""}>專科/大學</option><option {"selected" if v("education")=="研究所以上" else ""}>研究所以上</option><option {"selected" if v("education")=="其他" else ""}>其他</option></select></div>
</div></div>
<div class="card"><div class="sec">身分證發證資料</div><div class="g3">
  <div><label>發證日期</label><input name="iddate" class="ep" value="{h(v("id_issue_date"))}"></div>
  <div><label>發證地</label><select name="idplace" class="ep"><option value="">請選擇</option>{"".join(f'<option {"selected" if v("id_issue_place")==p else ""}>{p}</option>' for p in ["北市","新北市","桃市","中市","南市","高市","基市","竹市","竹縣","苗縣","彰縣","投縣","雲縣","嘉市","嘉縣","屏縣","宜縣","花縣","東縣","澎縣","金門","連江"])}</select></div>
  <div><label>換補發類別</label><select name="idtype" class="ep"><option value="">請選擇</option><option {"selected" if v("id_issue_type")=="初發" else ""}>初發</option><option {"selected" if v("id_issue_type")=="補發" else ""}>補發</option><option {"selected" if v("id_issue_type")=="換發" else ""}>換發</option></select></div>
</div></div>
<div class="card"><div class="sec">地址資料</div>
  <div style="font-size:13px;font-weight:700;color:#4a3e30;margin-bottom:8px">戶籍地址</div>
  <div class="g3" style="margin-bottom:10px"><div><label>縣市</label>{csel("rcity",v("reg_city"))}</div><div><label>區/鄉鎮</label><input name="rdist" class="ep" value="{h(v("reg_district"))}"></div><div><label>詳細地址</label><input name="raddr" class="ep" value="{h(v("reg_address"))}"></div></div>
  <div style="margin-bottom:10px"><label>戶籍電話</label><input name="rphone" class="ep" value="{h(v("reg_phone"))}" style="max-width:200px"></div>
  <label style="display:flex;align-items:center;gap:8px;font-size:14px;color:#3a3020;margin-bottom:10px;cursor:pointer;font-weight:600">
    <input type="checkbox" id="sameck" name="sameck" {"checked" if lsame else ""} onchange="document.getElementById('lsec').style.display=this.checked?'none':'block'">住家地址與戶籍相同</label>
  <div id="lsec" style="{"display:none" if lsame else ""};margin-bottom:10px">
    <div class="g3"><div><label>縣市</label>{csel("lcity",v("live_city"))}</div><div><label>區/鄉鎮</label><input name="ldist" class="ep" value="{h(v("live_district"))}"></div><div><label>詳細地址</label><input name="laddr" class="ep" value="{h(v("live_address"))}"></div></div>
  </div>
  <div class="g3">
    <div><label>現住電話</label><input name="lphone" class="ep" value="{h(v("live_phone"))}"></div>
    <div><label>居住狀況</label><select name="lstatus" class="ep"><option {"selected" if v("live_status")=="自有" else ""}>自有</option><option {"selected" if v("live_status")=="配偶" else ""}>配偶</option><option {"selected" if v("live_status")=="父母" else ""}>父母</option><option {"selected" if v("live_status")=="親屬" else ""}>親屬</option><option {"selected" if v("live_status")=="租屋" else ""}>租屋</option><option {"selected" if v("live_status")=="宿舍" else ""}>宿舍</option></select></div>
    <div><label>居住時間</label><div style="display:flex;gap:6px;align-items:center"><input name="lyear" class="ep" value="{h(v("live_years"))}" style="width:60px"><span>年</span><input name="lmon" class="ep" value="{h(v("live_months"))}" style="width:60px"><span>月</span></div></div>
  </div>
</div>
<div class="card"><div class="sec">職業資料</div><div class="g2">
  <div style="grid-column:1/-1"><label>公司名稱</label><input name="cmpname" class="ep" value="{h(v("company_name_detail"))}"></div>
  <div><label>公司電話</label><div style="display:flex;gap:6px;align-items:center"><select name="carea" class="ep" style="width:82px"><option value="">區碼</option>{"".join(f'<option {"selected" if v("company_phone_area")==a else ""}>{a}</option>' for a in ["02","03","037","04","049","05","06","07","08","089"])}<option {"selected" if v("company_phone_area")=="mobile" else ""} value="mobile">手機</option></select><input name="cnum" class="ep" value="{h(v("company_phone_num"))}"><input name="cext" class="ep" value="{h(v("company_phone_ext"))}" placeholder="分機" style="width:68px"></div></div>
  <div><label>職稱</label><input name="crole" class="ep" value="{h(v("company_role"))}"></div>
  <div><label>年資</label><div style="display:flex;gap:6px;align-items:center"><input name="cyear" class="ep" value="{h(v("company_years"))}" style="width:60px"><span>年</span><input name="cmon" class="ep" value="{h(v("company_months"))}" style="width:60px"><span>月</span></div></div>
  <div><label>月薪</label><input name="csal" class="ep" value="{h(v("company_salary"))}"></div>
  <div style="grid-column:1/-1"><label>公司地址</label><div class="g3">{csel("ccity",v("company_city"))}<input name="cdist" class="ep" value="{h(v("company_district"))}"><input name="caddr" class="ep" value="{h(v("company_address"))}"></div></div>
</div></div>
<div class="card"><div class="sec">聯絡人</div><div class="g2" style="margin-bottom:12px">
  <div><label>聯絡人1</label><input name="c1name" class="ep" value="{h(v("contact1_name"))}"></div>
  <div><label>關係</label><input name="c1rel" class="ep" value="{h(v("contact1_relation"))}"></div>
  <div><label>電話</label><input name="c1tel" class="ep" value="{h(v("contact1_phone"))}"></div>
  <div><label>知情</label><select name="c1know" class="ep"><option {"selected" if v("contact1_known")=="可知情" else ""}>可知情</option><option {"selected" if v("contact1_known")=="保密" else ""}>保密</option><option {"selected" if v("contact1_known")=="無可知情" else ""}>無可知情</option></select></div>
</div><div class="g2">
  <div><label>聯絡人2</label><input name="c2name" class="ep" value="{h(v("contact2_name"))}"></div>
  <div><label>關係</label><input name="c2rel" class="ep" value="{h(v("contact2_relation"))}"></div>
  <div><label>電話</label><input name="c2tel" class="ep" value="{h(v("contact2_phone"))}"></div>
  <div><label>知情</label><select name="c2know" class="ep"><option {"selected" if v("contact2_known")=="可知情" else ""}>可知情</option><option {"selected" if v("contact2_known")=="保密" else ""}>保密</option><option {"selected" if v("contact2_known")=="無可知情" else ""}>無可知情</option></select></div>
</div></div>
<div class="card"><div class="sec">貸款諮詢事項</div><div class="g2">
  <div><label>資金需求</label><input name="efund" class="ep" value="{h(v("eval_fund_need"))}"></div>
  <div><label>近三月送件</label><select name="esent" class="ep"><option {"selected" if v("eval_sent_3m")=="否" else ""}>否</option><option {"selected" if v("eval_sent_3m")=="是" else ""}>是</option></select></div>
  <div><label>當鋪私設</label><select name="eprivate" class="ep"><option {"selected" if v("eval_alert")=="無" else ""}>無</option><option {"selected" if v("eval_alert")=="有" else ""}>有</option></select></div>
  <div><label>勞保狀態</label><select name="elabor" class="ep"><option {"selected" if v("eval_labor_ins")=="公司保" else ""}>公司保</option><option {"selected" if v("eval_labor_ins")=="工會保" else ""}>工會保</option><option {"selected" if v("eval_labor_ins")=="自行投保" else ""}>自行投保</option><option {"selected" if v("eval_labor_ins")=="無勞保" else ""}>無勞保</option></select></div>
  <div><label>有無薪轉</label><select name="esal" class="ep"><option {"selected" if v("eval_salary_transfer")=="有薪轉" else ""}>有薪轉</option><option {"selected" if v("eval_salary_transfer")=="無薪轉" else ""}>無薪轉</option></select></div>
  <div><label>有無證照</label><select name="elicense" class="ep"><option {"selected" if v("eval_license")=="無" else ""}>無</option><option {"selected" if v("eval_license")=="有" else ""}>有</option></select></div>
  <div><label>貸款遲繳</label><select name="elate" class="ep"><option {"selected" if v("eval_late")=="無" else ""}>無</option><option {"selected" if v("eval_late")=="有" else ""}>有</option></select></div>
  <div><label>遲繳天數</label><input name="elateday" class="ep" value="{h(v("eval_late_days"))}"></div>
  <div><label>罰單欠費 $</label><input name="efine" class="ep" value="{h(v("eval_fine"))}"></div>
  <div><label>燃料稅 $</label><input name="efuel" class="ep" value="{h(v("eval_fuel_tax"))}"></div>
  <div><label>名下信用卡</label><input name="ecard" class="ep" value="{h(v("eval_credit_card"))}"></div>
  <div><label>動產/不動產</label><input name="eprop" class="ep" value="{h(v("eval_property"))}"></div>
  <div><label>法學</label><input name="elaw" class="ep" value="{h(v("eval_law"))}"></div>
  <div style="grid-column:1/-1"><label>備註</label><textarea name="enote" class="ep" style="min-height:60px">{h(v("eval_note"))}</textarea></div>
</div></div>
<div class="card"><div class="sec">負債明細</div>
  <div id="ep-debt-list"></div>
  <input type="hidden" name="debt_json" id="debt_json_input">
  <div style="display:flex;gap:8px;margin-top:8px;flex-wrap:wrap">
    <button type="button" onclick="addD('車貸')" style="background:#e8e2da;color:#4a3e30;border:1px dashed #a09080;border-radius:6px;padding:8px 16px;font-size:13px;cursor:pointer;font-weight:600">+ 新增車貸</button>
    <button type="button" onclick="addD('信貸')" style="background:#e8e2da;color:#4a3e30;border:1px dashed #a09080;border-radius:6px;padding:8px 16px;font-size:13px;cursor:pointer;font-weight:600">+ 新增信貸/其他</button>
  </div>
</div>
<div style="display:flex;gap:10px;margin-top:8px;align-items:center;flex-wrap:wrap">
  <button type="submit" class="btn-s">💾 儲存變更</button>
  <a href="/pending-customers" class="btn-b">取消</a>
  <div style="flex:1"></div>
  <button type="button" onclick="confirmDelete()" style="background:#b91c1c;color:#fff;border:none;padding:10px 22px;border-radius:8px;font-size:14px;font-weight:700;cursor:pointer;font-family:inherit;">🗑 刪除此客戶</button>
</div>
</form>
<form id="delete-form" method="post" action="/delete-customer" style="display:none;">
  <input type="hidden" name="case_id" value="{h(case_id)}">
</form>
<script>
function confirmDelete() {{
  if (confirm('⚠️ 確定要刪除這位客戶嗎？\\n刪除後無法復原，所有編輯紀錄都會一併清除。\\n\\n只能刪除 PENDING 狀態的客戶（已送件的案件不能刪）。')) {{
    document.getElementById('delete-form').submit();
  }}
}}
</script>
""" + """<script>
var existingDebts=""" + debt_json + """;
var dc=0;
function addD(type, preset){
  dc++;var n=dc;
  var isCar=(type==='車貸');
  var d=preset||{};
  var r=document.createElement('div');
  r.id='d'+n;r.className='d-row';
  var bg=isCar?'#f2f8f4':'#f8f5f1';
  var bc=isCar?'#4e7055':'#8a7a68';
  r.style.cssText='border-radius:8px;margin-bottom:8px;padding:10px 14px;background:'+bg+';border-left:4px solid '+bc+';';
  var lbl=isCar?'🚗 車貸':'💳 信貸/其他';
  var lc=isCar?'#4e7055':'#6a5e4e';
  var LS='font-size:12px;font-weight:500;color:#2c2820;height:17px;';
  var IS='width:100%;height:34px;padding:0 8px;border:0.5px solid #ddd5ca;border-radius:6px;font-size:13px;font-family:inherit;box-sizing:border-box;background:#fff;color:#2c2820;';
  var RS='width:100%;height:34px;padding:0 8px;border:0.5px solid #e8c0b0;border-radius:6px;font-size:13px;font-weight:500;color:#b84a35;background:#fff8f5;display:flex;align-items:center;justify-content:flex-end;box-sizing:border-box;';
  var SS='width:100%;height:34px;padding:0 6px;border:0.5px solid #ddd5ca;border-radius:6px;font-size:13px;font-family:inherit;box-sizing:border-box;background:#fff;color:#2c2820;';
  var FS='display:flex;flex-direction:column;gap:4px;';
  var GS='display:grid;grid-template-columns:repeat(5,1fr);gap:8px;margin-bottom:8px;';
  function esc(s){return String(s==null?'':s).replace(/"/g,'&quot;');}
  var pe=esc(d.pe||''),pa=esc(d.pa||'');
  var peDisplay=(pe||pa)?(pe+' / '+pa):'';
  var h='';
  h+='<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;">';
  h+='<span style="font-size:13px;font-weight:500;color:'+lc+';">'+lbl+'</span>';
  h+='<button type="button" onclick="rmD('+n+')" style="background:#f5ddd8;color:#b84a35;border:none;border-radius:5px;width:28px;height:28px;cursor:pointer;font-size:13px;">✕</button>';
  h+='</div>';
  h+='<div style="'+GS+'">';
  h+='<div style="'+FS+'"><div style="'+LS+'">貸款商家</div><input id="dc'+n+'" value="'+esc(d.co||'')+'" placeholder="裕融" style="'+IS+'"></div>';
  h+='<div style="'+FS+'"><div style="'+LS+'">貸款金額</div><input id="dl'+n+'" value="'+esc(d.lo||'')+'" placeholder="150000" type="number" style="'+IS+'"></div>';
  h+='<div style="'+FS+'"><div style="'+LS+'">期數／已繳</div><input id="dp'+n+'" value="'+esc(peDisplay)+'" placeholder="36 / 0" oninput="calcD('+n+')" style="'+IS+'"></div>';
  h+='<div style="'+FS+'"><div style="'+LS+'">月繳金額</div><input id="dm'+n+'" value="'+esc(d.mo||'')+'" placeholder="5265" type="number" oninput="calcD('+n+')" style="'+IS+'"></div>';
  h+='<div style="'+FS+'"><div style="'+LS+'">剩餘金額</div><div id="dr'+n+'" style="'+RS+'">-</div></div>';
  h+='</div>';
  if(isCar){
    h+='<div style="'+GS+'">';
    h+='<div style="'+FS+'"><div style="'+LS+'">設定日期（民國）</div><input id="dd'+n+'" value="'+esc(d.da||'')+'" placeholder="112/01" style="'+IS+'"></div>';
    var dyVal=d.dy||'無';
    var spVal=d.sp||'有';
    var dyOpts=['無','公路','動保','公路+動保'].map(function(o){return '<option'+(o===dyVal?' selected':'')+'>'+o+'</option>';}).join('');
    var spOpts=['有','無'].map(function(o){return '<option'+(o===spVal?' selected':'')+'>'+o+'</option>';}).join('');
    h+='<div style="'+FS+'"><div style="'+LS+'">動保／公路</div><select id="dg'+n+'" style="'+SS+'">'+dyOpts+'</select></div>';
    h+='<div style="'+FS+'"><div style="'+LS+'">空間</div><select id="ds'+n+'" style="'+SS+'">'+spOpts+'</select></div>';
    h+='<div style="height:49px;"></div><div style="height:49px;"></div>';
    h+='</div>';
  }else{
    h+='<input id="dd'+n+'" value="" style="display:none;"><input id="dg'+n+'" value="-" style="display:none;"><input id="ds'+n+'" value="-" style="display:none;">';
  }
  r.innerHTML=h;
  document.getElementById('ep-debt-list').appendChild(r);
  calcD(n);
}
function rmD(n){var el=document.getElementById('d'+n);if(el)el.remove();}
function calcD(n){
  var m=parseFloat((document.getElementById('dm'+n)||{}).value)||0;
  var dpv=((document.getElementById('dp'+n)||{}).value||'').trim();
  var parts=dpv.split('/');
  var p=parseFloat(parts[0])||0;
  var a=parseFloat(parts[1])||0;
  var el=document.getElementById('dr'+n);if(!el)return;
  if(m>0&&p>0){
    var rem=(m*p)-(m*a);
    el.textContent='$'+Math.round(rem).toLocaleString();
    el.style.color=rem>0?'#b84a35':'#4e7055';
  }else el.textContent='-';
}
function collectDebts(){
  var rows=[];
  for(var i=1;i<=dc;i++){
    var el=document.getElementById('d'+i);if(!el)continue;
    var co=(document.getElementById('dc'+i)||{}).value||'';
    if(!co)continue;
    var dpv=((document.getElementById('dp'+i)||{}).value||'').split('/');
    rows.push({
      co:co,
      lo:(document.getElementById('dl'+i)||{}).value||'',
      pe:(dpv[0]||'').trim(),
      mo:(document.getElementById('dm'+i)||{}).value||'',
      pa:(dpv[1]||'').trim(),
      re:(document.getElementById('dr'+i)||{}).textContent||'',
      da:(document.getElementById('dd'+i)||{}).value||'',
      dy:(document.getElementById('dg'+i)||{}).value||'',
      sp:(document.getElementById('ds'+i)||{}).value||''
    });
  }
  return rows;
}
// 載入既有負債：有 da 或 dy 非空非「-」→ 車貸，否則信貸
(function(){
  if(!existingDebts||!existingDebts.length){return;}
  existingDebts.forEach(function(d){
    var isCar=(d.da&&d.da!=='')||(d.dy&&d.dy!=='-'&&d.dy!=='')||(d.sp&&d.sp!=='-'&&d.sp!=='');
    addD(isCar?'車貸':'信貸',d);
  });
})();
document.querySelector('form').addEventListener('submit',function(){
  document.getElementById('debt_json_input').value=JSON.stringify(collectDebts());
});
</script>
</div></body></html>"""


@app.post("/edit-pending")
async def edit_pending_post(request: Request):
    from fastapi.responses import RedirectResponse
    role = check_auth(request)
    if not role: return RedirectResponse("/login")
    form = await request.form()
    f = dict(form)
    case_id = f.get("case_id","")
    if not case_id: return RedirectResponse("/pending-customers")
    live_same = "1" if f.get("sameck") else "0"
    now = now_iso()
    conn = get_conn(); cur = conn.cursor()
    cur.execute("""UPDATE customers SET
        customer_name=?,id_no=?,birth_date=?,phone=?,email=?,line_id=?,
        carrier=?,fb=?,
        source_group_id=?,marriage=?,education=?,
        id_issue_date=?,id_issue_place=?,id_issue_type=?,
        reg_city=?,reg_district=?,reg_address=?,reg_phone=?,live_same_as_reg=?,
        live_city=?,live_district=?,live_address=?,live_phone=?,live_status=?,live_years=?,live_months=?,
        company_name_detail=?,company_phone_area=?,company_phone_num=?,company_phone_ext=?,
        company_role=?,company_years=?,company_months=?,company_salary=?,
        company_city=?,company_district=?,company_address=?,
        contact1_name=?,contact1_relation=?,contact1_phone=?,contact1_known=?,
        contact2_name=?,contact2_relation=?,contact2_phone=?,contact2_known=?,
        updated_at=? WHERE case_id=?""",
        (f.get("cname",""),f.get("idno","").upper(),f.get("birth",""),f.get("phone",""),f.get("email",""),f.get("line",""),
         f.get("carrier",""),f.get("fb",""),
         f.get("grp",""),f.get("marry",""),f.get("edu",""),
         f.get("iddate",""),f.get("idplace",""),f.get("idtype",""),
         f.get("rcity",""),f.get("rdist",""),f.get("raddr",""),f.get("rphone",""),live_same,
         f.get("lcity","") if live_same=="0" else f.get("rcity",""),
         f.get("ldist","") if live_same=="0" else f.get("rdist",""),
         f.get("laddr","") if live_same=="0" else f.get("raddr",""),
         f.get("lphone",""),f.get("lstatus",""),f.get("lyear",""),f.get("lmon",""),
         f.get("cmpname",""),f.get("carea",""),f.get("cnum",""),f.get("cext",""),
         f.get("crole",""),f.get("cyear",""),f.get("cmon",""),f.get("csal",""),
         f.get("ccity",""),f.get("cdist",""),f.get("caddr",""),
         f.get("c1name",""),f.get("c1rel",""),f.get("c1tel",""),f.get("c1know",""),
         f.get("c2name",""),f.get("c2rel",""),f.get("c2tel",""),f.get("c2know",""),
         now, case_id))
    # 更新諮詢事項
    cur.execute("""UPDATE customers SET
        eval_fund_need=?,eval_sent_3m=?,eval_alert=?,eval_labor_ins=?,eval_salary_transfer=?,
        eval_license=?,eval_late=?,eval_late_days=?,eval_fine=?,eval_fuel_tax=?,eval_credit_card=?,eval_property=?,eval_law=?,eval_note=?,
        updated_at=? WHERE case_id=?""",
        (f.get("efund",""),f.get("esent",""),f.get("eprivate",""),f.get("elabor",""),f.get("esal",""),
         f.get("elicense",""),f.get("elate",""),f.get("elateday",""),f.get("efine",""),f.get("efuel",""),f.get("ecard",""),f.get("eprop",""),f.get("elaw",""),f.get("enote",""),
         now, case_id))
    # 更新負債明細
    debt_json_str = f.get("debt_json", "")
    if debt_json_str:
        cur.execute("UPDATE customers SET debt_list=?, updated_at=? WHERE case_id=?", (debt_json_str, now, case_id))
    conn.commit(); conn.close()
    return RedirectResponse("/edit-pending?case_id=" + case_id + "&saved=1", status_code=303)


@app.post("/delete-customer")
async def delete_customer(request: Request):
    """刪除單筆客戶（只允許 PENDING 狀態，防誤刪已送件案）"""
    from fastapi.responses import RedirectResponse
    role = check_auth(request)
    if not role or role not in ("admin", "adminB"):
        return HTMLResponse("<h3>❌ 無權限（需 admin 或 adminB）</h3>", status_code=403)
    form = await request.form()
    case_id = form.get("case_id", "").strip()
    if not case_id:
        return RedirectResponse("/pending-customers", status_code=303)
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT case_id, customer_name, status FROM customers WHERE case_id=?", (case_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return RedirectResponse("/pending-customers?err=notfound", status_code=303)
    if row["status"] != "PENDING":
        conn.close()
        return HTMLResponse(f"<h3>❌ 只能刪除 PENDING 狀態的客戶（此客戶狀態：{row['status']}）</h3>", status_code=400)
    cur.execute("DELETE FROM case_logs WHERE case_id=?", (case_id,))
    cur.execute("DELETE FROM customers WHERE case_id=?", (case_id,))
    conn.commit(); conn.close()
    return RedirectResponse("/pending-customers?deleted=1", status_code=303)


# 計算客戶資料完整度（填寫欄位比例 0-100）
_COMPLETENESS_EXCLUDE = {
    "id", "case_id", "status", "created_at", "updated_at", "last_update",
    "source_group_id", "route_plan", "current_company", "report_section",
    "customer_name", "created_by_role",
}


def calc_completeness(row: dict) -> int:
    """計算客戶資料填寫率：非空欄位 / 總欄位 × 100"""
    fields = [k for k in row.keys() if k not in _COMPLETENESS_EXCLUDE]
    if not fields:
        return 0
    filled = sum(1 for k in fields if (row.get(k) not in (None, "", "0")))
    return int(round(filled * 100 / len(fields)))


@app.get("/pending-customers", response_class=HTMLResponse)
def pending_customers_page(request: Request, q: str = "", grp: str = "", date_from: str = "", date_to: str = "", page: int = 1):
    from fastapi.responses import RedirectResponse
    role = check_auth(request)
    if not role: return RedirectResponse("/login")
    PAGE_SIZE = 50
    if page < 1: page = 1
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT group_id, group_name FROM groups WHERE group_type='SALES_GROUP' AND is_active=1 ORDER BY group_name")
    all_groups = cur.fetchall()

    where = "WHERE status='PENDING'"
    params: list = []
    if q: where += " AND (customer_name LIKE ? OR id_no LIKE ?)"; params += [f"%{q}%", f"%{q}%"]
    if grp: where += " AND source_group_id=?"; params.append(grp)
    if date_from: where += " AND DATE(created_at) >= ?"; params.append(date_from)
    if date_to: where += " AND DATE(created_at) <= ?"; params.append(date_to)

    cur.execute(f"SELECT COUNT(*) AS c FROM customers {where}", params)
    total = cur.fetchone()["c"]
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    if page > total_pages: page = total_pages
    offset = (page - 1) * PAGE_SIZE

    cur.execute(f"SELECT * FROM customers {where} ORDER BY created_at DESC LIMIT ? OFFSET ?", params + [PAGE_SIZE, offset])
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()

    # 一次取所有群組名稱（修掉迴圈內 N+1）
    group_name_map = get_all_group_names()

    grp_opts = "<option value=''>全部群組</option>" + "".join(
        f'<option value="{h(g["group_id"])}" {"selected" if grp==g["group_id"] else ""}>{h(g["group_name"])}</option>'
        for g in all_groups
    )

    rows_html = ""
    for r in rows:
        name = r.get("customer_name", "") or ""
        id_no = r.get("id_no", "") or ""
        created = (r.get("created_at", "") or "")[:10]
        gname = group_name_map.get(r.get("source_group_id", ""), "未知群組")
        case_id = r.get("case_id", "")
        pct = calc_completeness(r)
        dot_cls = "dot-green" if pct >= 70 else ("dot-yellow" if pct >= 30 else "dot-red")
        rows_html += f'''<tr>
            <td style="padding:11px 12px;width:36px;"><input type="checkbox" class="case-chk" value="{h(case_id)}"></td>
            <td style="padding:11px 12px;color:#3a2e1c;font-size:13px;font-weight:600;white-space:nowrap;">{h(created)}</td>
            <td style="padding:11px 12px;font-weight:700;font-size:14px;white-space:nowrap;color:#0f0a04;"><span class="dot {dot_cls}" title="完整度 {pct}%"></span>{h(name)}</td>
            <td style="padding:11px 12px;color:#1a1208;font-family:monospace;font-weight:600;">{h(id_no)}</td>
            <td style="padding:11px 12px;color:#1a1208;font-weight:600;">{h(gname)}</td>
            <td style="padding:11px 12px;text-align:center;">
                <a href="/edit-pending?case_id={h(case_id)}" style="background:#3a2e1c;color:#fff;padding:5px 14px;border-radius:6px;font-size:12px;text-decoration:none;font-weight:700;white-space:nowrap;">編輯</a>
            </td>
        </tr>'''
    empty = "" if rows else '<tr><td colspan="6" style="text-align:center;padding:30px;color:#4a3e30;font-weight:600;">目前沒有待確認客戶</td></tr>'

    # 建分頁連結（保留篩選參數）
    from urllib.parse import urlencode
    def page_url(p: int) -> str:
        qs = {"q": q, "grp": grp, "date_from": date_from, "date_to": date_to, "page": p}
        qs = {k: v for k, v in qs.items() if v not in ("", None)}
        return "/pending-customers?" + urlencode(qs)

    pager_html = ""
    if total_pages > 1:
        parts = []
        if page > 1:
            parts.append(f'<a href="{page_url(page-1)}">‹ 上一頁</a>')
        start = max(1, page - 3)
        end = min(total_pages, start + 6)
        start = max(1, end - 6)
        if start > 1:
            parts.append(f'<a href="{page_url(1)}">1</a>')
            if start > 2:
                parts.append('<span style="padding:6px 4px;color:#8a7a68;">…</span>')
        for p in range(start, end + 1):
            cls = ' class="current"' if p == page else ''
            parts.append(f'<a{cls} href="{page_url(p)}">{p}</a>')
        if end < total_pages:
            if end < total_pages - 1:
                parts.append('<span style="padding:6px 4px;color:#8a7a68;">…</span>')
            parts.append(f'<a href="{page_url(total_pages)}">{total_pages}</a>')
        if page < total_pages:
            parts.append(f'<a href="{page_url(page+1)}">下一頁 ›</a>')
        pager_html = '<div class="pager">' + "".join(parts) + '</div>'

    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>客戶資料庫</title>
{PAGE_CSS}
<style>
body{{background:#e4ddcd;}}
.page{{max-width:1040px;margin:24px auto;padding:0 16px 40px;color:#1a1208;}}
.card{{background:#f2ecdd;border:1px solid #b8ad9c;border-radius:10px;overflow:hidden;box-shadow:0 1px 3px rgba(60,45,25,0.08);}}
table{{width:100%;border-collapse:collapse;}}
thead tr{{background:#d8ccb0;}}
thead th{{position:sticky;top:0;background:#d8ccb0;z-index:2;border-bottom:2px solid #a89c82;}}
th{{padding:11px 12px;text-align:left;font-size:13px;font-weight:800;color:#1a1208;}}
tbody td{{color:#1a1208;}}
tbody tr{{border-bottom:1px solid #d8ccb0;}}
tbody tr:hover{{background:#ddd5c4;}}
h2{{font-size:19px;font-weight:800;color:#0f0a04;margin-bottom:14px;}}
input,select{{padding:7px 10px;border:1px solid #8a7e68;border-radius:6px;font-size:13px;font-family:inherit;color:#1a1208;background:#f2ede0;}}
input[type=checkbox]{{padding:0;width:16px;height:16px;cursor:pointer;}}
.dot{{display:inline-block;width:10px;height:10px;border-radius:50%;margin-right:7px;vertical-align:middle;}}
.dot-red{{background:#b91c1c;}}
.dot-yellow{{background:#b45309;}}
.dot-green{{background:#15803d;}}
.pager{{text-align:center;padding:16px;}}
.pager a{{display:inline-block;padding:6px 12px;margin:0 2px;border:1px solid #8a7e68;border-radius:6px;color:#1a1208;text-decoration:none;font-size:13px;background:#f2ecdd;font-weight:600;}}
.pager a:hover{{background:#d8ccb0;}}
.pager a.current{{background:#3a2e1c;color:#fff;border-color:#3a2e1c;font-weight:800;}}
.btn-export{{background:#2f5339;color:#fff;border:none;padding:8px 16px;border-radius:6px;font-size:13px;cursor:pointer;font-weight:700;font-family:inherit;}}
.btn-export:disabled{{background:#8a8275;cursor:not-allowed;}}
.filter-box{{background:#f2ecdd !important;border:1px solid #b8ad9c !important;}}
.filter-box label-text{{color:#1a1208 !important;}}
</style></head><body>
{make_topnav(role, "pending")}
<div class="page">
  <h2>客戶資料庫 共 {total} 筆</h2>
  <form method="get" action="/pending-customers">
    <div style="background:#f2ecdd;border:1px solid #b8ad9c;border-radius:10px;padding:14px 16px;margin-bottom:14px;display:flex;gap:10px;flex-wrap:wrap;align-items:flex-end;box-shadow:0 1px 3px rgba(60,45,25,0.08);">
      <div><div style="font-size:13px;font-weight:700;color:#1a1208;margin-bottom:4px">姓名／身分證</div><input name="q" value="{h(q)}" placeholder="搜尋..." style="width:150px"></div>
      <div><div style="font-size:13px;font-weight:700;color:#1a1208;margin-bottom:4px">群組</div><select name="grp" style="width:110px">{grp_opts}</select></div>
      <div><div style="font-size:13px;font-weight:700;color:#1a1208;margin-bottom:4px">日期從</div><input type="date" name="date_from" value="{h(date_from)}" style="width:140px"></div>
      <div><div style="font-size:13px;font-weight:700;color:#1a1208;margin-bottom:4px">日期至</div><input type="date" name="date_to" value="{h(date_to)}" style="width:140px"></div>
      <div style="display:flex;gap:6px;align-items:flex-end">
        <button type="submit" style="background:#3a2e1c;color:#fff;border:none;padding:8px 18px;border-radius:6px;font-size:13px;cursor:pointer;font-weight:700;font-family:inherit">🔍 搜尋</button>
        <a href="/pending-customers" style="background:#d8ccb0;color:#1a1208;border:1px solid #8a7e68;padding:8px 14px;border-radius:6px;font-size:13px;text-decoration:none;font-weight:600;">清除</a>
      </div>
      <div style="margin-left:auto;display:flex;gap:8px;align-items:flex-end;">
        <button type="button" id="btnExport" class="btn-export" disabled onclick="exportPdf()">📄 匯出 PDF（<span id="selCount">0</span>）</button>
      </div>
    </div>
  </form>
  <div class="card"><table>
    <thead><tr>
      <th style="width:36px;"><input type="checkbox" id="chkAll" onclick="toggleAll(this)"></th>
      <th>日期</th><th>姓名</th><th>身分證</th><th>群組</th><th>操作</th>
    </tr></thead>
    <tbody>{rows_html}{empty}</tbody>
  </table></div>
  {pager_html}
</div>
<script>
function updateSelCount() {{
  var n = document.querySelectorAll('.case-chk:checked').length;
  document.getElementById('selCount').textContent = n;
  document.getElementById('btnExport').disabled = (n === 0);
}}
function toggleAll(src) {{
  document.querySelectorAll('.case-chk').forEach(function(c) {{ c.checked = src.checked; }});
  updateSelCount();
}}
document.querySelectorAll('.case-chk').forEach(function(c) {{
  c.addEventListener('change', updateSelCount);
}});
function exportPdf() {{
  var ids = [];
  document.querySelectorAll('.case-chk:checked').forEach(function(c) {{ ids.push(c.value); }});
  if (ids.length === 0) return;
  window.open('/customer-pdf-batch?ids=' + encodeURIComponent(ids.join(',')));
}}
</script>
</body></html>"""


@app.get("/history", response_class=HTMLResponse)
def history_page(request: Request, group: str = "", month: str = "", q: str = ""):
    from fastapi.responses import RedirectResponse
    role = check_auth(request)
    if not role: return RedirectResponse("/login")
    auth_group = get_auth_group_id(request)

    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT * FROM groups WHERE group_type='SALES_GROUP' AND is_active=1 ORDER BY group_name")
    groups = cur.fetchall()

    # 業務只看自己群組
    if auth_group and not group:
        group = auth_group

    query = "SELECT * FROM customers WHERE status IN ('CLOSED','PENALTY','ABANDONED','REJECTED')"
    params = []
    if group:
        query += " AND source_group_id=?"
        params.append(group)
    if month:
        query += " AND updated_at LIKE ?"
        params.append(month + "%")
    if q:
        query += " AND (customer_name LIKE ? OR id_no LIKE ?)"
        params.extend([f"%{q}%", f"%{q}%"])
    query += " ORDER BY updated_at DESC LIMIT 300"
    cur.execute(query, params)
    rows = cur.fetchall()
    conn.close()

    # 月份選項（最近6個月）
    from datetime import datetime, timedelta
    month_opts = "<option value=\'\'>全部月份</option>"
    for i in range(6):
        dt = datetime.now().replace(day=1) - timedelta(days=i*28)
        ym = dt.strftime("%Y-%m")
        label = dt.strftime("%Y年%m月")
        sel = "selected" if month == ym else ""
        month_opts += f'<option value="{ym}" {sel}>{label}</option>'

    group_opts = "<option value=\'\'>全部群組</option>"
    for g in groups:
        sel = "selected" if g["group_id"]==group else ""
        group_opts += f'<option value="{h(g["group_id"])}" {sel}>{h(g["group_name"])}</option>'

    def get_close_badge(row):
        st = row.get("status","") or ""
        reason = row.get("close_reason","") or ""
        amt = row.get("approved_amount","") or ""
        penalty = row.get("penalty_amount","") or ""
        if st == "PENALTY":
            if "核准" in reason:
                label = "核准不撥款"
                bg, color = "#fef9c3", "#854d0e"
            else:
                label = "辦理中放棄"
                bg, color = "#fef9c3", "#854d0e"
            return f'<span style="background:{bg};color:{color};font-size:12px;padding:3px 10px;border-radius:20px;font-weight:600">{label}・違約金${h(penalty)}</span>'
        if st == "ABANDONED":
            if "核准" in reason:
                label = "核准不撥款"
            else:
                label = "辦理中放棄"
            return f'<span style="background:#ece8e2;color:#4a3e30;font-size:12px;padding:3px 10px;border-radius:20px;font-weight:600">{label}・未收違約金</span>'
        if st == "CLOSED" and amt:
            return '<span style="background:#dcfce7;color:#166534;font-size:12px;padding:3px 10px;border-radius:20px;font-weight:600">已撥款結案</span>'
        if st == "CLOSED":
            return '<span style="background:#f0fdf4;color:#166534;font-size:12px;padding:3px 10px;border-radius:20px;font-weight:600">結案</span>'
        if st == "REJECTED":
            return '<span style="background:#fee2e2;color:#991b1b;font-size:12px;padding:3px 10px;border-radius:20px;font-weight:600">全數婉拒</span>'
        return '<span style="background:#ece8e2;color:#4a3e30;font-size:12px;padding:3px 10px;border-radius:20px;font-weight:600">結案</span>'

    rows_html = ""
    if not rows:
        rows_html = '<div style="color:#6a5e4e;padding:24px;text-align:center;font-size:14px">沒有結案紀錄</div>'
    for row in rows:
        row = dict(row)
        badge = get_close_badge(row)
        co = row.get("current_company","") or row.get("company","") or ""
        amt = row.get("approved_amount","") or ""
        gname = get_group_name(row["source_group_id"])
        updated = row["updated_at"][:10].replace("-","/") if row["updated_at"] else ""
        # 多家核准
        route_data2 = parse_route_json(row.get("route_plan","") or "")
        all_approved = [rh for rh in route_data2.get("history",[]) if rh.get("status") in ("核准","待撥款","撥款") and rh.get("amount")]
        if len(all_approved) > 1:
            detail = "多家核准：" + " + ".join((rh.get("company") or "") + (rh.get("amount") or "") for rh in all_approved)
        elif all_approved:
            detail = all_approved[0].get("company","") + " 核准" + all_approved[0].get("amount","")
        else:
            detail = co + (" 核准" + amt if amt else "")
        rows_html += (
            '<div style="display:grid;grid-template-columns:1.4fr 1.2fr 1fr 0.7fr 0.8fr;gap:12px;align-items:center;padding:12px 16px;border-bottom:1px solid #ece8e2">'
            + '<div style="font-size:14px;font-weight:600;color:#1a1208">' + h(row["customer_name"]) + '</div>'
            + '<div style="font-size:13px;color:#3a3530">' + h(detail) + '</div>'
            + '<div>' + badge + '</div>'
            + '<div style="font-size:13px;color:#4a3e30">' + h(gname) + '</div>'
            + '<div style="font-size:12px;color:#6a5e4e">' + h(updated) + '</div>'
            + '</div>'
        )

    HIST_CSS = """
    <style>
    body{background:#ece8e2;color:#2c2820;font-family:'Microsoft JhengHei','PingFang TC',sans-serif}
    .input{background:#faf7f4;border:1px solid #ddd5ca;color:#2c2820;border-radius:7px;padding:8px 12px;font-size:14px;font-family:inherit}
    </style>"""

    return f"""<!DOCTYPE html><html><head>{PAGE_CSS}{HIST_CSS}<title>歷史紀錄</title></head><body>
    {make_topnav(role, "history")}
    <div class="page">
      <form method="get" action="/history" style="margin-bottom:16px">
        <div style="display:flex;gap:10px;flex-wrap:wrap;align-items:center">
          <select class="input" name="group" onchange="this.form.submit()" style="min-width:120px">{group_opts}</select>
          <select class="input" name="month" onchange="this.form.submit()" style="min-width:140px">{month_opts}</select>
          <input class="input" name="q" value="{h(q)}" placeholder="搜尋客戶姓名..." style="min-width:160px">
          <button type="submit" class="btn btn-primary" style="white-space:nowrap">搜尋</button>
          <span style="font-size:13px;color:#6a5e4e;margin-left:auto">共 {len(rows)} 筆</span>
        </div>
      </form>
      <div style="background:#faf7f4;border:1px solid #ddd5ca;border-radius:10px;overflow:hidden">
        <div style="display:grid;grid-template-columns:1.4fr 1.2fr 1fr 0.7fr 0.8fr;gap:12px;padding:9px 16px;background:#ece8e2;border-bottom:1px solid #ddd5ca">
          <div style="font-size:12px;font-weight:700;color:#4a3e30">客戶姓名</div>
          <div style="font-size:12px;font-weight:700;color:#4a3e30">方案/金額</div>
          <div style="font-size:12px;font-weight:700;color:#4a3e30">結案原因</div>
          <div style="font-size:12px;font-weight:700;color:#4a3e30">群組</div>
          <div style="font-size:12px;font-weight:700;color:#4a3e30">日期</div>
        </div>
        {rows_html if rows_html else '<div style="color:#6a5e4e;padding:24px;text-align:center;font-size:14px">沒有結案紀錄</div>'}
      </div>
    </div></body></html>"""


@app.get("/admin/groups", response_class=HTMLResponse)
def list_groups(request: Request):
    from fastapi.responses import RedirectResponse
    role = check_auth(request)
    if role != "admin": return RedirectResponse("/login")

    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT * FROM groups ORDER BY group_type, group_name")
    rows = cur.fetchall()
    cur.execute("SELECT group_id, group_name FROM groups WHERE group_type='SALES_GROUP' AND is_active=1")
    sales_groups = cur.fetchall()
    conn.close()

    sales_opts = "<option value=\'\'>（無）</option>"
    for sg in sales_groups:
        sales_opts += f'<option value="{h(sg["group_id"])}">{h(sg["group_name"])}</option>'

    type_labels = {"SALES_GROUP":"業務群","ADMIN_GROUP":"行政群","A_GROUP":"A群"}
    rows_html = ""
    for r in rows:
        linked_name = ""
        if r["linked_sales_group_id"]:
            conn2 = get_conn(); cur2 = conn2.cursor()
            cur2.execute("SELECT group_name FROM groups WHERE group_id=?", (r["linked_sales_group_id"],))
            ln = cur2.fetchone(); conn2.close()
            linked_name = ln["group_name"] if ln else r["linked_sales_group_id"]
        edit_btn = f'''<button onclick="openEdit(\'{h(r["group_id"])}\',\'{h(r["group_name"])}\',{1 if r["is_active"] else 0})" class="btn" style="font-size:11px;padding:3px 10px">✏️ 編輯</button>'''
        rows_html += f'''<tr style="border-bottom:1px solid #f3f4f6">
          <td style="padding:8px 12px;font-weight:500">{h(r["group_name"])}</td>
          <td style="padding:8px 12px;color:#6b7280">{h(type_labels.get(r["group_type"],r["group_type"]))}</td>
          <td style="padding:8px 12px;color:#6b7280">{h(linked_name) or "-"}</td>
          <td style="padding:8px 12px">{"✅" if r["is_active"] else "❌"}</td>
          <td style="padding:8px 12px;font-size:11px;color:#9ca3af;font-family:monospace;max-width:180px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="{h(r["group_id"])}" onclick="navigator.clipboard.writeText('{h(r["group_id"])}')">{h(r["group_id"])}</td>
          <td style="padding:8px 12px">{edit_btn}</td>
        </tr>'''

    return f"""<!DOCTYPE html><html><head>{PAGE_CSS}<title>群組管理</title></head><body>
    {make_topnav(role, "admin")}
    <div class="page">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px">
        <h2 style="font-size:18px;font-weight:600">群組管理</h2>
        <button class="btn btn-primary" onclick="document.getElementById('add-modal').classList.add('show')">+ 新增群組</button>
      </div>
      <div style="background:#fff;border:1px solid #e5e7eb;border-radius:10px;overflow-x:auto;margin-bottom:20px">
        <table style="width:100%;border-collapse:collapse;min-width:600px">
          <thead style="background:#fafafa;border-bottom:1px solid #f3f4f6">
            <tr>
              <th style="padding:8px 12px;text-align:left;font-size:12px;color:#6b7280">名稱</th>
              <th style="padding:8px 12px;text-align:left;font-size:12px;color:#6b7280">類型</th>
              <th style="padding:8px 12px;text-align:left;font-size:12px;color:#6b7280">對應業務群</th>
              <th style="padding:8px 12px;text-align:left;font-size:12px;color:#6b7280">啟用</th>
              <th style="padding:8px 12px;text-align:left;font-size:12px;color:#6b7280">Group ID</th>
              <th style="padding:8px 12px;text-align:left;font-size:12px;color:#6b7280"></th>
            </tr>
          </thead>
          <tbody>{rows_html}</tbody>
        </table>
      </div>

      <div style="background:#fff;border:1px solid #fca5a5;border-radius:10px;padding:16px">
        <div style="font-size:13px;font-weight:600;margin-bottom:8px;color:#dc2626">危險操作</div>
        <div style="font-size:12px;color:#6b7280;margin-bottom:12px">清除所有案件資料，群組設定不受影響。清除後無法復原。</div>
        <button class="btn btn-danger" onclick="doReset()">🗑️ 清除所有測試資料</button>
        <div id="reset-result" style="margin-top:10px;font-size:12px"></div>
      </div>
    </div>

    <div class="modal-bg" id="add-modal" onclick="if(event.target===this)this.classList.remove('show')">
      <div class="modal">
        <div class="modal-title">新增 / 更新群組 <span class="modal-close" onclick="document.getElementById('add-modal').classList.remove('show')">×</span></div>
        <div class="form-row"><label>群組 ID（在群組打 @AI 群組ID 取得）</label><input class="input" id="f_gid" placeholder="Cxxx..."></div>
        <div class="form-row"><label>群組名稱</label><input class="input" id="f_gname" placeholder="例：鉅烽行政群"></div>
        <div class="form-row"><label>群組類型</label>
          <select class="input" id="f_gtype" onchange="document.getElementById('linked-row').style.display=this.value==='ADMIN_GROUP'?'block':'none'">
            <option value="SALES_GROUP">業務群</option>
            <option value="ADMIN_GROUP">行政群</option>
            <option value="A_GROUP">A群/進度群</option>
          </select>
        </div>
        <div class="form-row" id="linked-row" style="display:none"><label>對應業務群（行政群才需要）</label><select class="input" id="f_linked">{sales_opts}</select></div>
        <button class="btn btn-primary" onclick="addGroup()" style="width:100%;justify-content:center">確認新增</button>
        <div id="add-result" style="margin-top:10px;font-size:12px"></div>
      </div>
    </div>

    <div class="modal-bg" id="edit-modal" onclick="if(event.target===this)this.classList.remove('show')">
      <div class="modal">
        <div class="modal-title">編輯群組 <span class="modal-close" onclick="document.getElementById('edit-modal').classList.remove('show')">×</span></div>
        <input type="hidden" id="e_gid">
        <div class="form-row"><label>群組 ID（不可修改）</label><input class="input" id="e_gid_show" disabled style="background:#f9fafb;color:#9ca3af"></div>
        <div class="form-row"><label>群組名稱</label><input class="input" id="e_gname" placeholder="輸入新名稱"></div>
        <div class="form-row"><label>狀態</label>
          <select class="input" id="e_active">
            <option value="1">啟用</option>
            <option value="0">停用</option>
          </select>
        </div>
        <button class="btn btn-primary" onclick="saveEdit()" style="width:100%;justify-content:center">儲存</button>
        <div id="edit-result" style="margin-top:10px;font-size:12px"></div>
      </div>
    </div>

    <script>
    function openEdit(gid, gname, isActive){{
      document.getElementById('e_gid').value = gid;
      document.getElementById('e_gid_show').value = gid;
      document.getElementById('e_gname').value = gname;
      document.getElementById('e_active').value = isActive;
      document.getElementById('edit-result').innerText = '';
      document.getElementById('edit-modal').classList.add('show');
    }}
    async function saveEdit(){{
      const gid = document.getElementById('e_gid').value;
      const gname = document.getElementById('e_gname').value.trim();
      const isActive = document.getElementById('e_active').value;
      const res = document.getElementById('edit-result');
      if(!gname){{res.className='result-err';res.innerText='名稱不可空白';return}}
      const r = await fetch('/admin/update_group', {{
        method:'POST',
        headers:{{'Content-Type':'application/json'}},
        body:JSON.stringify({{group_id:gid, group_name:gname, is_active:parseInt(isActive)}})
      }});
      const data = await r.json();
      res.className = data.status==='ok'?'result-ok':'result-err';
      res.innerText = data.message;
      if(data.status==='ok') setTimeout(()=>location.reload(),1200);
    }}
    async function addGroup(){{
      const gid=document.getElementById('f_gid').value.trim();
      const gname=document.getElementById('f_gname').value.trim();
      const gtype=document.getElementById('f_gtype').value;
      const linked=document.getElementById('f_linked').value;
      const res=document.getElementById('add-result');
      if(!gid||!gname){{res.className='result-err';res.innerText='群組ID和名稱必填';return}}
      const body={{group_id:gid,group_name:gname,group_type:gtype}};
      if(gtype==='ADMIN_GROUP'&&linked) body.linked_sales_group_id=linked;
      const r=await fetch('/admin/add_group',{{method:'POST',headers:{{\'Content-Type\':\'application/json\'}},body:JSON.stringify(body)}});
      const data=await r.json();
      res.className=data.status==='ok'?'result-ok':'result-err';
      res.innerText=data.message;
      if(data.status==='ok') setTimeout(()=>location.reload(),1500);
    }}
    async function doReset(){{
      if(!confirm('確定要清除所有案件資料嗎？此操作無法復原！')) return;
      const res=document.getElementById('reset-result');
      const r=await fetch('/admin/reset_data',{{method:'POST',headers:{{\'Content-Type\':\'application/json\'}},body:JSON.stringify({{confirm:'yes'}})}});
      const data=await r.json();
      res.style.color=data.status==='ok'?'#15803d':'#dc2626';
      res.innerText=data.message;
    }}
    </script>
    </body></html>"""

# =========================
# 更新群組 API
# =========================
@app.post("/admin/update_group")
async def update_group(request: Request):
    role = check_auth(request)
    if role != "admin":
        return JSONResponse({"status":"error","message":"無權限"}, status_code=403)
    data = await request.json()
    gid = data.get("group_id","").strip()
    gname = data.get("group_name","").strip()
    is_active = data.get("is_active", 1)
    if not gid or not gname:
        return JSONResponse({"status":"error","message":"群組ID和名稱必填"})
    try:
        conn = get_conn(); cur = conn.cursor()
        cur.execute("UPDATE groups SET group_name=?, is_active=? WHERE group_id=?",
            (gname, is_active, gid))
        conn.commit(); conn.close()
        return JSONResponse({"status":"ok","message":f"已更新：{gname}"})
    except Exception as e:
        return JSONResponse({"status":"error","message":str(e)})


@app.get("/admin/locked-accounts")
async def get_locked_accounts(request: Request):
    role = check_auth(request)
    if role != "admin":
        return JSONResponse({"ok": False, "locked": []})
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT identifier, attempts, locked_until FROM login_attempts WHERE locked_until > ? ORDER BY locked_until DESC",
        (now_iso(),))
    rows = cur.fetchall(); conn.close()
    locked = [{"identifier": r["identifier"], "attempts": r["attempts"],
               "locked_until": r["locked_until"][11:16] if r["locked_until"] else ""} for r in rows]
    return JSONResponse({"ok": True, "locked": locked})


@app.post("/admin/unlock-account")
async def unlock_account(request: Request):
    role = check_auth(request)
    if role != "admin":
        return JSONResponse({"ok": False, "message": "無權限"})
    data = await request.json()
    identifier = data.get("identifier","").strip()
    if not identifier:
        return JSONResponse({"ok": False, "message": "identifier 必填"})
    conn = get_conn(); cur = conn.cursor()
    cur.execute("DELETE FROM login_attempts WHERE identifier=?", (identifier,))
    conn.commit(); conn.close()
    return JSONResponse({"ok": True, "message": f"已解鎖：{identifier}"})


# =========================
# VBA 查詢 API
# =========================
@app.get("/api/customer-lookup")
async def customer_lookup(
    request: Request,
    date: str = "",
    name: str = "",
    id_no: str = "",
    secret: str = ""
):
    stored_secret = get_setting("vba_secret")
    if stored_secret:
        vba_ok = verify_pw(secret, stored_secret)
    else:
        vba_ok = (secret == os.getenv("VBA_SECRET", "vba_secret_2026"))
    if not vba_ok:
        return JSONResponse({"ok": False, "error": "無權限"}, status_code=403)
    if not date or not name or not id_no:
        return JSONResponse({"ok": False, "error": "請輸入日期、姓名、身分證"})
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT * FROM customers WHERE customer_name=? AND id_no=? ORDER BY created_at DESC", (name, id_no.upper()))
    rows = cur.fetchall(); conn.close()
    if not rows:
        return JSONResponse({"ok": False, "error": f"找不到客戶：{name} / {id_no}"})
    matched = None
    old_cases = []
    for row in rows:
        created = (row["created_at"] or "")[:10]
        created_fmt = created.replace("-", "/")
        # Bug 15: 先檢查長度，避免空字串/格式異常時 ValueError 被靜默吞掉
        created_roc = ""
        if len(created) >= 10:
            try:
                y = int(created[:4]) - 1911
                created_roc = f"{y}/{created[5:7]}/{created[8:10]}"
            except (ValueError, IndexError) as e:
                print(f"[date_parse] failed for {created!r}: {e}")
        if date in [created_fmt, created_roc]:
            if row["status"] == "ACTIVE":
                matched = row
            else:
                old_cases.append(row)
    if not matched:
        return JSONResponse({"ok": False, "error": f"找到客戶 {name} 但日期不符，請確認建立日期"})
    d = dict(matched)
    old_warning = ""
    if old_cases:
        oc = old_cases[0]
        old_warning = f"⚠️ 此客戶有舊案記錄\n舊案日期：{(oc['created_at'] or '')[:10]}\n舊案狀態：{oc['status']}\n請確認您填的是新案資料"
    return JSONResponse({
        "ok": True,
        "old_warning": old_warning,
        "data": {
            "customer_name": d.get("customer_name",""),
            "id_no": d.get("id_no",""),
            "birth_date": d.get("birth_date",""),
            "phone": d.get("phone",""),
            "email": d.get("email",""),
            "line_id": d.get("line_id",""),
            "marriage": d.get("marriage",""),
            "education": d.get("education",""),
            "id_issue_date": d.get("id_issue_date",""),
            "id_issue_place": d.get("id_issue_place",""),
            "id_issue_type": d.get("id_issue_type",""),
            "reg_city": d.get("reg_city",""),
            "reg_district": d.get("reg_district",""),
            "reg_address": d.get("reg_address",""),
            "reg_phone": d.get("reg_phone",""),
            "live_same_as_reg": d.get("live_same_as_reg",""),
            "live_city": d.get("live_city",""),
            "live_district": d.get("live_district",""),
            "live_address": d.get("live_address",""),
            "live_phone": d.get("live_phone",""),
            "live_status": d.get("live_status",""),
            "live_years": d.get("live_years",""),
            "live_months": d.get("live_months",""),
            "company_name_detail": d.get("company_name_detail",""),
            "company_industry": d.get("company_industry",""),
            "company_role": d.get("company_role",""),
            "company_phone_area": d.get("company_phone_area",""),
            "company_phone_num": d.get("company_phone_num",""),
            "company_phone_ext": d.get("company_phone_ext",""),
            "company_years": d.get("company_years",""),
            "company_salary": d.get("company_salary",""),
            "company_city": d.get("company_city",""),
            "company_district": d.get("company_district",""),
            "company_address": d.get("company_address",""),
            "bank_name": d.get("bank_name",""),
            "bank_branch": d.get("bank_branch",""),
            "bank_account": d.get("bank_account",""),
            "bank_holder": d.get("bank_holder",""),
            "contact1_name": d.get("contact1_name",""),
            "contact1_relation": d.get("contact1_relation",""),
            "contact1_phone": d.get("contact1_phone",""),
            "contact1_known": d.get("contact1_known",""),
            "contact2_name": d.get("contact2_name",""),
            "contact2_relation": d.get("contact2_relation",""),
            "contact2_phone": d.get("contact2_phone",""),
            "contact2_known": d.get("contact2_known",""),
            "product_model": d.get("product_model",""),
            "product_imei": d.get("product_imei",""),
            "vehicle_plate": d.get("vehicle_plate",""),
            "vehicle_type": d.get("vehicle_type",""),
            "vehicle_owner": d.get("vehicle_owner",""),
        }
    })
    # =========================
# =========================
# 新增客戶網頁（行政A）
# =========================
@app.get("/new-customer", response_class=HTMLResponse)
def new_customer_page(request: Request):
    from fastapi.responses import RedirectResponse
    role = check_auth(request)
    if not role: return RedirectResponse("/login")
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT * FROM groups WHERE is_active=1 AND group_type='SALES_GROUP' ORDER BY group_name")
    groups = cur.fetchall(); conn.close()
    group_opts = '<option value="">請選擇</option>' + "".join(f'<option value="{h(g["group_id"])}">{h(g["group_name"])}</option>' for g in groups)
    grp_map_js = json.dumps({g["group_id"]: g["group_name"] for g in groups}, ensure_ascii=False)
    pdf_js = _PDF_EXPORT_JS.replace("__GRP_MAP__", grp_map_js)
    HTML_PAGE = '<!DOCTYPE html>\n<html lang="zh-TW">\n<head>\n<meta charset="UTF-8">\n<meta name="viewport" content="width=device-width, initial-scale=1.0">\n<title>新增客戶資料</title>\n<style>\n* { box-sizing: border-box; margin: 0; padding: 0; }\nbody { font-family: \'Microsoft JhengHei\', \'PingFang TC\', sans-serif; background: #ece8e2; color: #2c2820; font-size: 14px; }\n.topnav { background: #3a3530; padding: 0 20px; display: flex; align-items: center; height: 50px; gap: 4px; }\n.topnav a { color: #c8bfb5; text-decoration: none; padding: 7px 14px; border-radius: 6px; font-size: 14px; }\n.topnav a.active { background: #7c6f5e; color: #fff; }\n.topnav a:hover { background: #4a4540; color: #fff; }\n.page { max-width: 820px; margin: 24px auto; padding: 0 16px 40px; }\n.page-header { display: flex; justify-content: space-between; align-items: center; margin-bottom: 20px; padding-bottom: 12px; border-bottom: 2px solid #8a7a68; }\nh2 { font-size: 20px; font-weight: 700; color: #2c2820; }\n.card { background: #faf6f2; border-radius: 10px; padding: 18px; margin-bottom: 14px; box-shadow: 0 1px 4px rgba(0,0,0,0.08); border: 1px solid #ddd5ca; }\n.section-title { font-size: 13px; font-weight: 700; color: #5a4e40; margin-bottom: 14px; padding-bottom: 7px; border-bottom: 1px solid #ddd5ca; letter-spacing: 0.5px; }\n.grid2 { display: grid; grid-template-columns: 1fr 1fr; gap: 12px; }\n.grid3 { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 12px; }\n.grid4 { display: grid; grid-template-columns: 1fr 1fr 1fr 1fr; gap: 12px; }\n.full { grid-column: 1 / -1; }\nlabel { display: block; font-size: 12px; color: #5a4e40; margin-bottom: 5px; font-weight: 600; }\n.req { color: #b84a35; }\ninput, select, textarea {\n  width: 100%; padding: 8px 11px; border: 1px solid #c8bfb5;\n  border-radius: 6px; font-size: 14px; font-family: inherit;\n  background: #fff; color: #2c2820; transition: border 0.15s;\n}\ninput:focus, select:focus, textarea:focus { outline: none; border-color: #7c6f5e; background: #fffcf9; }\ninput::placeholder, textarea::placeholder { color: #c8bfb5; }\ntextarea { resize: vertical; min-height: 65px; }\n.hint { font-size: 12px; color: #8a7a68; margin-top: 4px; }\n.auto-val { font-size: 15px; color: #b84a35; font-weight: 700; margin-top: 5px; }\n.vehicle-card { background: #f0ebe4; border: 1px solid #ccc5ba; border-radius: 8px; padding: 14px; margin-bottom: 10px; position: relative; }\n.vehicle-card .card-label { font-size: 13px; font-weight: 700; color: #5a4e40; margin-bottom: 10px; }\n.remove-btn { position: absolute; top: 12px; right: 12px; background: #f5ddd8; color: #b84a35; border: none; border-radius: 4px; padding: 4px 12px; font-size: 12px; cursor: pointer; font-weight: 600; }\n.debt-header { display: grid; grid-template-columns: 1.2fr 0.8fr 0.6fr 0.8fr 0.6fr 1fr 0.9fr 0.8fr 30px; gap: 5px; font-size: 12px; color: #5a4e40; font-weight: 600; margin-bottom: 6px; }\n.debt-row { display: grid; grid-template-columns: 1.2fr 0.8fr 0.6fr 0.8fr 0.6fr 1fr 0.9fr 0.8fr 30px; gap: 5px; margin-bottom: 7px; align-items: center; }\n.debt-row input { font-size: 13px; padding: 6px 8px; }\n.debt-remain { font-size: 13px; font-weight: 700; color: #b84a35; }\n.debt-del { background: #f5ddd8; color: #b84a35; border: none; border-radius: 4px; width: 30px; height: 32px; cursor: pointer; font-size: 15px; }\n.add-btn { background: #e8e2da; color: #4a3e30; border: 1px dashed #a09080; border-radius: 6px; padding: 8px 16px; font-size: 13px; cursor: pointer; margin-top: 8px; font-weight: 600; }\n.add-btn:hover { background: #ddd5ca; }\n.btn-row { display: flex; gap: 10px; margin-top: 22px; flex-wrap: wrap; }\n.btn { padding: 11px 26px; border-radius: 8px; font-size: 15px; font-weight: 700; cursor: pointer; border: none; font-family: inherit; }\n.btn-primary { background: #6a5e4e; color: #fff; }\n.btn-primary:hover { background: #5a4e40; }\n.btn-export { background: #4e7055; color: #fff; }\n.btn-export:hover { background: #3e5e45; }\n.btn-cancel { background: #ddd5ca; color: #4a3e30; }\n.btn-cancel:hover { background: #ccc5ba; }\n.ig { display: flex; gap: 6px; align-items: center; }\n.ig span { font-size: 14px; white-space: nowrap; line-height: 36px; color: #4a3e30; font-weight: 600; }\n</style>\n</head>\n<body>\n<div class="topnav">\n  <a href="/">&#128202; 日報</a>\n  <a href="/pending-customers">&#128203; 客戶資料庫</a>\n  <a href="/new-customer" class="active">&#10133; 新增客戶</a>\n</div>\n<div class="page">\n  <div class="page-header">\n    <h2>新增客戶資料</h2>\n\n  </div>\n  <form id="cf" method="post" action="/new-customer">\n    <div class="card">\n      <div class="section-title">所屬群組</div>\n      <div>\n        <div><label>群組</label><select name="grp"><option>B群</option><option>C群</option></select></div>\n      </div>\n    </div>\n    <div class="card">\n      <div class="section-title">基本資料</div>\n      <div class="grid2">\n        <div><label>客戶姓名 <span class="req">*</span></label><input name="cname" placeholder="王小明" required></div>\n        <div><label>身分證字號 <span class="req">*</span></label><input name="idno" placeholder="A123456789" style="text-transform:uppercase" required pattern="[A-Za-z][12]\\d{8}" title="格式：英文字母+1或2+8位數字"></div>\n        <div><label>出生年月日 <span class="req">*</span></label><input name="birth" placeholder="086/12/15"><div class="hint">民國年：086/12/15</div></div>\n        <div><label>行動電話 <span class="req">*</span></label><input name="phone" placeholder="0912-345678" pattern="09\\d{8}" title="格式：09開頭共10位數字"></div>\n        <div><label>電信業者</label><select name="carrier"><option value="">--- 請選擇 ---</option><option>中華電信</option><option>遠傳電信</option><option>台灣大哥大</option><option>台灣之星</option><option>亞太電信</option><option>其他</option></select></div>\n        <div><label>Email</label><input name="email" placeholder="example@gmail.com"></div>\n        <div><label>LINE ID</label><input name="line"></div>\n        <div><label>客戶FB</label><input name="fb" placeholder="Facebook名稱"></div>\n        <div><label>婚姻狀態</label><select name="marry"><option value="">--- 請選擇 ---</option><option>未婚</option><option>已婚</option></select></div>\n        <div><label>最高學歷</label><select name="edu"><option value="">--- 請選擇 ---</option><option>高中/職</option><option>專科/大學</option><option>研究所以上</option><option>其他</option></select></div>\n      </div>\n    </div>\n    <div class="card">\n      <div class="section-title">身分證發證資料</div>\n      <div class="grid3">\n        <div><label>發證日期 <span class="req">*</span></label><input name="iddate" placeholder="114/03/05"><div class="hint">民國年：114/03/05</div></div>\n        <div><label>發證地 <span class="req">*</span></label><select name="idplace"><option value="">--- 請選擇 ---</option><option>北市</option><option>新北市</option><option>桃市</option><option>中市</option><option>南市</option><option>高市</option><option>基市</option><option>竹市</option><option>竹縣</option><option>苗縣</option><option>彰縣</option><option>投縣</option><option>雲縣</option><option>嘉市</option><option>嘉縣</option><option>屏縣</option><option>宜縣</option><option>花縣</option><option>東縣</option><option>澎縣</option><option>金門</option><option>連江</option></select></div>\n        <div><label>換補發類別 <span class="req">*</span></label><select name="idtype"><option value="">--- 請選擇 ---</option><option>初發</option><option>補發</option><option>換發</option></select></div>\n      </div>\n    </div>\n    <div class="card">\n      <div class="section-title">地址資料</div>\n      <div style="font-size:13px;font-weight:700;color:#4a3e30;margin-bottom:10px">戶籍地址</div>\n      <div class="grid3" style="margin-bottom:10px">\n        <div><label>縣市 <span class="req">*</span></label><select name="rcity"><option value="">請選擇</option><option>台北市</option><option>新北市</option><option>桃園市</option><option>台中市</option><option>台南市</option><option>高雄市</option><option>基隆市</option><option>新竹市</option><option>新竹縣</option><option>苗栗縣</option><option>彰化縣</option><option>南投縣</option><option>雲林縣</option><option>嘉義市</option><option>嘉義縣</option><option>屏東縣</option><option>宜蘭縣</option><option>花蓮縣</option><option>台東縣</option><option>澎湖縣</option><option>金門縣</option><option>連江縣</option></select></div>\n        <div><label>區/鄉鎮</label><input name="rdist" placeholder="苗栗市"></div>\n        <div><label>詳細地址 <span class="req">*</span></label><input name="raddr" placeholder="新東街257號"></div>\n      </div>\n      <div style="margin-bottom:12px"><label>戶籍電話</label><input name="rphone" placeholder="037-123456" style="max-width:200px"></div>\n      <label style="display:flex;align-items:center;gap:8px;font-size:14px;color:#3a3020;margin-bottom:12px;cursor:pointer;font-weight:600">\n        <input type="checkbox" id="sameck" name="sameck" checked style="width:18px;height:18px;flex-shrink:0;accent-color:#6a5e4e" onchange="document.getElementById(\'lsec\').style.display=this.checked?\'none\':\'block\'">\n        住家地址與戶籍相同\n      </label>\n      <div id="lsec" style="display:none;margin-bottom:12px">\n        <div style="font-size:13px;font-weight:700;color:#4a3e30;margin-bottom:10px">住家地址</div>\n        <div class="grid3">\n          <div><label>縣市</label><select name="lcity"><option value="">請選擇</option><option>台北市</option><option>新北市</option><option>桃園市</option><option>台中市</option><option>台南市</option><option>高雄市</option><option>基隆市</option><option>新竹市</option><option>新竹縣</option><option>苗栗縣</option><option>彰化縣</option><option>南投縣</option><option>雲林縣</option><option>嘉義市</option><option>嘉義縣</option><option>屏東縣</option><option>宜蘭縣</option><option>花蓮縣</option><option>台東縣</option><option>澎湖縣</option><option>金門縣</option><option>連江縣</option></select></div>\n          <div><label>區/鄉鎮</label><input name="ldist"></div>\n          <div><label>詳細地址</label><input name="laddr"></div>\n        </div>\n      </div>\n      <div class="grid3">\n        <div><label>現住電話</label><input name="lphone" placeholder="037-123456"></div>\n        <div><label>居住狀況</label><select name="lstatus"><option value="">--- 請選擇 ---</option><option>自有</option><option>配偶</option><option>父母</option><option>親屬</option><option>租屋</option><option>宿舍</option></select></div>\n        <div><label>居住時間</label><div class="ig"><input name="lyear" placeholder="5" style="width:58px"><span>年</span><input name="lmon" placeholder="0" style="width:58px"><span>月</span></div></div>\n      </div>\n    </div>\n    <div class="card">\n      <div class="section-title">職業資料</div>\n      <div class="grid2">\n        <div class="full"><label>公司名稱 <span class="req">*</span></label><input name="cmpname" placeholder="嘉合企業社"></div>\n        <div><label>公司電話 <span class="req">*</span></label><div class="ig"><select name="carea" style="width:82px"><option value="">區碼</option><option>02</option><option>03</option><option>037</option><option>04</option><option>049</option><option>05</option><option>06</option><option>07</option><option>08</option><option>089</option><option value="mobile">手機</option></select><input name="cnum" placeholder="1234567"><input name="cext" placeholder="分機" style="width:68px"></div></div>\n        <div><label>職稱</label><input name="crole" placeholder="技工"></div>\n        <div><label>年資</label><div class="ig"><input name="cyear" placeholder="3" style="width:62px"><span>年</span><input name="cmon" placeholder="0" style="width:62px"><span>月</span></div></div>\n        <div><label>月薪（萬）</label><input name="csal" placeholder="3.5"></div>\n        <div class="full"><label>公司地址</label><div style="display:grid;grid-template-columns:1fr 1fr 2fr;gap:8px"><select name="ccity"><option value="">請選擇</option><option>台北市</option><option>新北市</option><option>桃園市</option><option>台中市</option><option>台南市</option><option>高雄市</option><option>基隆市</option><option>新竹市</option><option>新竹縣</option><option>苗栗縣</option><option>彰化縣</option><option>南投縣</option><option>雲林縣</option><option>嘉義市</option><option>嘉義縣</option><option>屏東縣</option><option>宜蘭縣</option><option>花蓮縣</option><option>台東縣</option><option>澎湖縣</option><option>金門縣</option><option>連江縣</option></select><input name="cdist" placeholder="區/鄉鎮"><input name="caddr" placeholder="詳細地址"></div></div>\n      </div>\n    </div>\n    <div class="card">\n      <div class="section-title">聯絡人資料</div>\n      <div style="margin-bottom:14px;padding-bottom:14px;border-bottom:1px solid #ddd5ca">\n        <div style="font-size:13px;color:#5a4e40;font-weight:600;margin-bottom:10px">聯絡人1（需為二等親）</div>\n        <div class="grid4">\n          <div><label>姓名 <span class="req">*</span></label><input name="c1name" placeholder="吳玉英"></div>\n          <div><label>關係</label><input name="c1rel" placeholder="母"></div>\n          <div><label>電話 <span class="req">*</span></label><input name="c1tel" placeholder="0978-055530"></div>\n          <div><label>可知情</label><select name="c1know"><option value="">請選擇</option><option>可知情</option><option>保密</option><option>無可知情</option></select></div>\n        </div>\n      </div>\n      <div>\n        <div style="font-size:13px;color:#5a4e40;font-weight:600;margin-bottom:10px">聯絡人2</div>\n        <div class="grid4">\n          <div><label>姓名 <span class="req">*</span></label><input name="c2name" placeholder="賴俊明"></div>\n          <div><label>關係</label><input name="c2rel" placeholder="友"></div>\n          <div><label>電話 <span class="req">*</span></label><input name="c2tel" placeholder="0919-616821"></div>\n          <div><label>可知情</label><select name="c2know"><option value="">請選擇</option><option>可知情</option><option>保密</option><option>無可知情</option></select></div>\n        </div>\n      </div>\n    </div>\n    <div class="card">\n      <div class="section-title">貸款諮詢事項</div>\n      <div class="grid2">\n        <div><label>資金需求</label><input name="efund" placeholder="$100,000"></div>\n        <div><label>近三月是否送件</label><select name="esent"><option value="">請選擇</option><option>否</option><option>是</option></select></div>\n        <div><label>當鋪私設</label><select name="eprivate"><option value="">請選擇</option><option>無</option><option>有</option></select></div>\n        <div><label>勞保狀態</label><select name="elabor"><option value="">請選擇</option><option>公司保</option><option>工會保</option><option>自行投保</option><option>無勞保</option></select></div>\n        <div><label>有無薪轉</label><select name="esal"><option value="">請選擇</option><option>有薪轉</option><option>無薪轉</option></select></div>\n        <div><label>有無證照</label><select name="elicense"><option value="">請選擇</option><option>無</option><option>有</option></select></div>\n        <div><label>貸款遲繳</label><select name="elate"><option value="">請選擇</option><option>無</option><option>有</option></select></div>\n        <div><label>遲繳天數</label><input name="elateday" placeholder="0"></div>\n        <div><label>罰單欠費金額 $</label><input name="efine" placeholder="0" type="number" min="0" oninput="calcF()"></div>\n        <div><label>燃料稅金額 $</label><input name="efuel" placeholder="0" type="number" min="0" oninput="calcF()"></div>\n        <div class="full"><label>欠費總額（自動計算）</label><div class="auto-val" id="tfees">$0</div></div>\n        <div><label>名下信用卡</label><input name="ecard" placeholder="銀行協商"></div>\n        <div><label>有無動產/不動產</label><input name="eprop" placeholder="有機車"></div>\n        <div><label>法學（幾條）</label><input name="elaw" placeholder="共1條"></div>\n        <div class="full"><label>備註</label><textarea name="enote" placeholder="其他說明..."></textarea></div>\n      </div>\n    </div>\n    <div class="card">\n      <div class="section-title">負債明細</div>\n      <div id="dlist"></div>\n      <div style="display:flex;gap:8px;margin-top:8px;flex-wrap:wrap;">\n        <button type="button" class="add-btn" onclick="addD(\'車貸\')">&#10133; 新增車貸</button>\n        <button type="button" class="add-btn" onclick="addD(\'信貸\')">&#10133; 新增信貸/其他</button>\n      </div>\n    </div>\n    <div class="btn-row">\n      <div id="err-box" style="display:none;background:#fef2f2;border:1px solid #fca5a5;border-radius:8px;padding:12px 16px;margin-bottom:12px;"></div><button type="button" class="btn btn-primary" onclick="doSubmit()">&#9989; 建立客戶</button>\n      <button type="button" class="btn btn-cancel" onclick="history.back()">取消</button>\n    </div>\n  </form>\n</div>\n<script>\nlet vc=0,dc=0;\nfunction gv(n){const e=document.querySelector(\'[name="\'+n+\'"]\');return e?e.value||\'\':\'\';}\n\nfunction addD(type){\n  dc++;const n=dc;\n  var isCar=(type===\'車貸\');\n  var r=document.createElement(\'div\');\n  r.id=\'d\'+n;r.className=\'d-row\';\n  var bg=isCar?\'#f2f8f4\':\'#f8f5f1\';\n  var bc=isCar?\'#4e7055\':\'#8a7a68\';\n  r.style.cssText=\'border-radius:8px;margin-bottom:8px;padding:10px 14px;background:\'+bg+\';border-left:4px solid \'+bc+\';\';\n  var lbl=isCar?\'🚗 車貸\':\'💳 信貸/其他\';\n  var lc=isCar?\'#4e7055\':\'#6a5e4e\';\n  var LS=\'font-size:12px;font-weight:500;color:#2c2820;height:17px;\';\n  var IS=\'width:100%;height:34px;padding:0 8px;border:0.5px solid #ddd5ca;border-radius:6px;font-size:13px;font-family:inherit;box-sizing:border-box;background:#fff;color:#2c2820;\';\n  var RS=\'width:100%;height:34px;padding:0 8px;border:0.5px solid #e8c0b0;border-radius:6px;font-size:13px;font-weight:500;color:#b84a35;background:#fff8f5;display:flex;align-items:center;justify-content:flex-end;box-sizing:border-box;\';\n  var SS=\'width:100%;height:34px;padding:0 6px;border:0.5px solid #ddd5ca;border-radius:6px;font-size:13px;font-family:inherit;box-sizing:border-box;background:#fff;color:#2c2820;\';\n  var FS=\'display:flex;flex-direction:column;gap:4px;\';\n  var GS=\'display:grid;grid-template-columns:repeat(5,1fr);gap:8px;margin-bottom:8px;\';\n  var h=\'\';\n  h+=\'<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;">\';\n  h+=\'<span style="font-size:13px;font-weight:500;color:\'+lc+\';">\'+lbl+\'</span>\';\n  h+=\'<button type="button" onclick="rmD(\'+n+\')" style="background:#f5ddd8;color:#b84a35;border:none;border-radius:5px;width:28px;height:28px;cursor:pointer;font-size:13px;">✕</button>\';\n  h+=\'</div>\';\n  h+=\'<div style="\'+GS+\'">\';\n  h+=\'<div style="\'+FS+\'"><div style="\'+LS+\'">貸款商家</div><input id="dc\'+n+\'" placeholder="裕融" style="\'+IS+\'"></div>\';\n  h+=\'<div style="\'+FS+\'"><div style="\'+LS+\'">貸款金額</div><input id="dl\'+n+\'" placeholder="150000" type="number" style="\'+IS+\'"></div>\';\n  h+=\'<div style="\'+FS+\'"><div style="\'+LS+\'">期數／已繳</div><input id="dp\'+n+\'" placeholder="36 / 0" oninput="calcD(\'+n+\')" style="\'+IS+\'"></div>\';\n  h+=\'<div style="\'+FS+\'"><div style="\'+LS+\'">月繳金額</div><input id="dm\'+n+\'" placeholder="5265" type="number" oninput="calcD(\'+n+\')" style="\'+IS+\'"></div>\';\n  h+=\'<div style="\'+FS+\'"><div style="\'+LS+\'">剩餘金額</div><div id="dr\'+n+\'" style="\'+RS+\'">-</div></div>\';\n  h+=\'</div>\';\n  if(isCar){\n    h+=\'<div style="\'+GS+\'">\';\n    h+=\'<div style="\'+FS+\'"><div style="\'+LS+\'">設定日期（民國）</div><input id="dd\'+n+\'" placeholder="112/01" style="\'+IS+\'"></div>\';\n    h+=\'<div style="\'+FS+\'"><div style="\'+LS+\'">動保／公路</div><select id="dg\'+n+\'" style="\'+SS+\'"><option>無</option><option>公路</option><option>動保</option><option>公路+動保</option></select></div>\';\n    h+=\'<div style="\'+FS+\'"><div style="\'+LS+\'">空間</div><select id="ds\'+n+\'" style="\'+SS+\'"><option>有</option><option>無</option></select></div>\';\n    h+=\'<div style="height:49px;"></div><div style="height:49px;"></div>\';\n    h+=\'</div>\';\n  }else{\n    h+=\'<input id="dd\'+n+\'" value="" style="display:none;"><input id="dg\'+n+\'" value="-" style="display:none;"><input id="ds\'+n+\'" value="-" style="display:none;">\';\n  }\n  r.innerHTML=h;\n  document.getElementById(\'dlist\').appendChild(r);\n}\nfunction rmD(n){\n  var el=document.getElementById(\'d\'+n);if(el)el.remove();\n}\nfunction calcD(n){\n  var m=parseFloat(document.getElementById(\'dm\'+n)?.value)||0;\n  var dpv=(document.getElementById(\'dp\'+n)?.value||\'\').trim();\n  var parts=dpv.split(\'/\');\n  var p=parseFloat(parts[0])||0;\n  var a=parseFloat(parts[1])||0;\n  var el=document.getElementById(\'dr\'+n);if(!el)return;\n  if(m>0&&p>0){\n    var rem=(m*p)-(m*a);\n    el.textContent=\'$\'+Math.round(rem).toLocaleString();\n    el.style.color=rem>0?\'#b84a35\':\'#4e7055\';\n  }else el.textContent=\'-\';\n}\n\nfunction calcF(){\n  const f=parseFloat(document.querySelector(\'[name="efine"]\')?.value)||0;\n  const u=parseFloat(document.querySelector(\'[name="efuel"]\')?.value)||0;\n  document.getElementById(\'tfees\').textContent=\'$\'+(f+u).toLocaleString();\n}\n\nfunction sec(t){\n  return \'<div style="background:#5a4e40;color:#fff;font-size:13px;font-weight:500;padding:7px 12px;border-radius:5px 5px 0 0;letter-spacing:0.5px;">\'+t+\'</div>\'\n    +\'<div style="border:1px solid #ccc5ba;border-top:none;border-radius:0 0 5px 5px;padding:8px 14px;margin-bottom:14px;background:#fdfaf7;">\';\n}\nfunction fl(l,v){\n  if(!v||v===\'-\'||v===\'0年0月\')return\'\';\n  return \'<div style="display:grid;grid-template-columns:110px 1fr;gap:6px;padding:7px 0;border-bottom:0.5px solid #ece8e2;">\'\n    +\'<span style="font-size:13px;color:#5a4e40;font-weight:500;line-height:1.5;">\'+l+\'</span>\'\n    +\'<span style="font-size:14px;font-weight:400;color:#2c2820;line-height:1.5;">\'+v+\'</span></div>\';\n}\nfunction fl2(items){\n  var h=\'<div style="display:grid;grid-template-columns:1fr 1fr;gap:0 28px;">\';\n  items.forEach(function(x){h+=fl(x[0],x[1]);});\n  return h+\'</div>\';\n}\n\nfunction exportPDF(){/* overridden by inject */}\n\n// page-init\n</script>\n</body>\n</html>\n'
    # 在表單前面插入智能填入區塊
    HTML_PAGE = HTML_PAGE.replace('<option>B群</option><option>C群</option>', group_opts)
    # 注入前端驗證JS
    # 把驗證JS注入到原本 HTML_PAGE 的 // page-init 位置
    inject_js = """
function showErr(msgs){
  var b=document.getElementById('err-box');
  b.innerHTML='<div style="color:#991b1b;font-weight:700;margin-bottom:8px;">請修正：</div>'+msgs.map(function(x){return '<div style="color:#b84a35;padding:2px 0;">• '+x+'</div>';}).join('');
  b.style.display='block';b.scrollIntoView({behavior:'smooth',block:'center'});
}
function doSubmit(){
  var e=[];
  var qq=function(s){return (document.querySelector(s)||{value:''}).value;};
  var n=qq('[name="cname"]').trim();
  var id=qq('[name="idno"]').trim().toUpperCase();
  var ph=qq('[name="phone"]').replace(/[- ()]/g,'');
  var em=qq('[name="email"]').trim();
  var li=qq('[name="line"]').trim();
  var ra=qq('[name="raddr"]').trim();
  var sa=document.getElementById('sameck').checked;
  var la=sa?ra:qq('[name="laddr"]').trim();
  var cn=qq('[name="cmpname"]').trim();
  var cp=qq('[name="cnum"]').trim();
  var cc=qq('[name="ccity"]');
  var ca=qq('[name="caddr"]').trim();
  var n1=qq('[name="c1name"]').trim();
  var t1=qq('[name="c1tel"]').replace(/[- ()]/g,'');
  var n2=qq('[name="c2name"]').trim();
  var t2=qq('[name="c2tel"]').replace(/[- ()]/g,'');
  if(!n)e.push('姓名不可空白');
  if(!id)e.push('身分證不可空白');
  else if(!/^[A-Z][A-Z0-9][0-9]{8}$/.test(id))e.push('身分證/居留證格式錯誤');
  if(!ph)e.push('行動電話不可空白');
  else if(ph.length!==10||!/^09/.test(ph))e.push('行動電話10碼');
  if(!em)e.push('Email必填');
  if(!li)e.push('LINE ID必填');
  if(!ra)e.push('戸籍地址不可空白');
  if(!sa&&!la)e.push('居住地址不可空白');
  var ly=qq('[name="lyear"]').trim();
  var ls2=qq('[name="lstatus"]');
  if(!ly)e.push('居住年數不可空白');
  if(!ls2)e.push('居住狀況不可空白');
  // 居住年數 vs 年齡
  var birthStr=qq('[name="birth"]').trim();
  if(ly&&birthStr){
    var bm=birthStr.match(/^(\d{2,4})\/(\d{1,2})\/(\d{1,2})$/);
    if(bm){
      var by=parseInt(bm[1]);
      if(by<200)by+=1911;
      var age=(new Date()).getFullYear()-by;
      if(age>0&&parseInt(ly)>age)e.push('居住年數('+ly+')不能大於年齡('+age+')');
    }
  }
  if(!cn)e.push('公司名稱不可空白');
  var isPackWork=cn&&(cn.indexOf('包工作')>=0||cn==='包工');
  if(!isPackWork){
    if(!cp)e.push('公司電話不可空白');
    if(!cc&&!ca)e.push('公司地址不可空白');
  }
  var selfkw=['自營','自己','自行','自營商','營業中','個人工作室'];
  if(cn&&selfkw.some(function(k){return cn===k||cn.indexOf(k)===0&&cn.length<=4;}))e.push('公司名稱請填明確行業（不可只寫「'+cn+'」）');
  if(!n1)e.push('聯絡人1姓名必填');
  if(!t1)e.push('聯絡人1電話必填');
  else if(t1.length!==10)e.push('聯絡人1電話10碼');
  if(!n2)e.push('聯絡人2姓名必填');
  if(!t2)e.push('聯絡人2電話必填');
  else if(t2.length!==10)e.push('聯絡人2電話10碼');
  if(ph&&t1&&ph===t1)e.push('申請人與聯絡人1電話相同');
  if(ph&&t2&&ph===t2)e.push('申請人與聯絡人2電話相同');
  if(t1&&t2&&t1===t2)e.push('聯絡人1耈2電話相同');
  if(e.length>0){showErr(e);return;}
  var w=[];
  var ls=qq('[name="lstatus"]');
  var rc=qq('[name="rcity"]');
  if(ls==='自有')w.push('居住「自有」請確認無私設');
  var lc=sa?rc:qq('[name="lcity"]');
  if(lc&&cc&&lc!==cc)w.push('居住('+lc+')與公司('+cc+')不同縣市，請確認距離合理');
  w.push('請逐字確認所有欄位填寫正確');
  var b2=document.getElementById('err-box');
  b2.innerHTML='<div style="color:#854d0e;font-weight:700;margin-bottom:8px;">人工確認事項：</div>'+w.map(function(x){return '<div style="color:#854d0e;padding:2px 0;">'+x+'</div>';}).join('')+'<div style="margin-top:10px;"><button onclick="doConfirmSubmit()" style="background:#6a5e4e;color:#fff;border:none;padding:8px 18px;border-radius:6px;cursor:pointer;">確認無誤，送出</button></div>';
  b2.style.display='block';b2.scrollIntoView({behavior:'smooth',block:'center'});
}
function collectDebt(){
  var rows=[];
  for(var i=1;i<=dc;i++){
    var el=document.getElementById('d'+i);
    if(!el)continue;
    var co=document.getElementById('dc'+i)?.value||'';
    if(!co)continue;
    rows.push({
      co:co,
      lo:document.getElementById('dl'+i)?.value||'',
      pe:(function(){var v=(document.getElementById('dp'+i)?.value||'').split('/');return (v[0]||'').trim();})(),
      mo:document.getElementById('dm'+i)?.value||'',
      pa:(function(){var v=(document.getElementById('dp'+i)?.value||'').split('/');return (v[1]||'').trim();})(),
      re:document.getElementById('dr'+i)?.textContent||'',
      da:document.getElementById('dd'+i)?.value||'',
      dy:document.getElementById('dg'+i)?.value||'',
      sp:document.getElementById('ds'+i)?.value||''
    });
  }
  var h=document.getElementById('debt_list_input');
  if(!h){h=document.createElement('input');h.type='hidden';h.name='debt_list';h.id='debt_list_input';document.getElementById('cf').appendChild(h);}
  h.value=JSON.stringify(rows);
}
function doConfirmSubmit(){collectDebt();localStorage.removeItem('nc_draft');document.forms[0].submit();}
function saveForm(){var data={};document.querySelectorAll('#cf input,#cf select,#cf textarea').forEach(function(el){if(!el.name)return;if(el.type==='checkbox')data[el.name]=el.checked;else data[el.name]=el.value;});localStorage.setItem('nc_draft',JSON.stringify(data));}
function restoreForm(){var raw=localStorage.getItem('nc_draft');if(!raw)return;var data=JSON.parse(raw);Object.keys(data).forEach(function(k){var el=document.querySelector('#cf [name="'+k+'"]');if(!el)return;if(el.type==='checkbox')el.checked=data[k];else el.value=data[k];});var ck=document.getElementById('sameck');if(ck)document.getElementById('lsec').style.display=ck.checked?'none':'block';}
window.addEventListener('DOMContentLoaded',function(){restoreForm();document.getElementById('cf').addEventListener('input',saveForm);document.getElementById('cf').addEventListener('change',saveForm);});
// page-init
"""
    HTML_PAGE = HTML_PAGE.replace('// page-init', inject_js + '\n' + (pdf_js or ''))
    return HTML_PAGE


@app.post("/new-customer")
async def new_customer_post(request: Request):
    from fastapi.responses import RedirectResponse
    role = check_auth(request)
    if not role: return RedirectResponse("/login")
    form = await request.form()
    f = dict(form)
    name = f.get("cname","").strip()
    id_no = f.get("idno","").strip().upper()
    phone = f.get("phone","").strip().replace("-","").replace(" ","")
    email = f.get("email","").strip()
    line_id = f.get("line","").strip()
    raddr = f.get("raddr","").strip()
    live_same = f.get("sameck","")
    laddr = f.get("laddr","").strip()
    cmpname = f.get("cmpname","").strip()
    cnum = f.get("cnum","").strip()
    caddr = f.get("caddr","").strip()
    c1name = f.get("c1name","").strip()
    c1tel = f.get("c1tel","").strip().replace("-","").replace(" ","")
    c2name = f.get("c2name","").strip()
    c2tel = f.get("c2tel","").strip().replace("-","").replace(" ","")

    import re as _re
    errs = []
    if not name: errs.append("姓名不可空白")
    if not id_no: errs.append("身分證字號不可空白")
    elif not _re.match(r"^[A-Z][A-Z0-9][0-9]{8}$", id_no): errs.append("身分證/居留證格式錯誤")
    if not phone: errs.append("行動電話不可空白")
    elif len(phone)!=10 or not phone.startswith("09"): errs.append("行動電話需為10碼（09開頭）")
    if not email: errs.append("Email為必填")
    if not line_id: errs.append("LINE ID為必填")
    if not raddr: errs.append("戶籍地址不可空白")
    if not live_same and not laddr: errs.append("居住地址不可空白（或勾選同戶籍）")
    lyear_str = f.get("lyear","").strip()
    if not lyear_str: errs.append("居住年數不可空白")
    if not f.get("lstatus","").strip(): errs.append("居住狀況不可空白")
    # 居住年數 vs 年齡檢查
    if lyear_str and f.get("birth","").strip():
        try:
            ly = int(float(lyear_str))
            birth = f.get("birth","").strip()
            bm = _re.match(r"^(\d{2,4})/(\d{1,2})/(\d{1,2})$", birth)
            if bm:
                by = int(bm.group(1))
                if by < 200:  # 民國年
                    by += 1911
                from datetime import datetime as _dt
                age = _dt.now().year - by
                if age > 0 and ly > age:
                    errs.append(f"居住年數({ly})不能大於年齡({age})")
        except Exception:
            pass
    if not cmpname: errs.append("公司名稱不可空白")
    is_pack_work = cmpname and ("包工作" in cmpname or cmpname == "包工")
    if not is_pack_work:
        if not cnum: errs.append("公司電話不可空白")
        if not caddr: errs.append("公司地址不可空白")
    if not c1name: errs.append("聯絡人1姓名必填")
    if not c1tel: errs.append("聯絡人1電話必填")
    elif len(c1tel)!=10: errs.append("聯絡人1電話需為10碼")
    if not c2name: errs.append("聯絡人2姓名必填")
    if not c2tel: errs.append("聯絡人2電話必填")
    elif len(c2tel)!=10: errs.append("聯絡人2電話需為10碼")
    # 電話重複
    phones = {"申請人":phone,"聯絡人1":c1tel,"聯絡人2":c2tel}
    pkeys = list(phones.keys())
    for i in range(len(pkeys)):
        for j in range(i+1,len(pkeys)):
            if phones[pkeys[i]] and phones[pkeys[j]] and phones[pkeys[i]]==phones[pkeys[j]]:
                errs.append(f"{pkeys[i]}與{pkeys[j]}電話相同（{phones[pkeys[i]]}），請確認")
    if errs:
        err_html = "<ul>" + "".join(f"<li>{h(e)}</li>" for e in errs) + "</ul>"
        return HTMLResponse(f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><style>body{{font-family:'Microsoft JhengHei',sans-serif;background:#ece8e2;display:flex;justify-content:center;align-items:center;min-height:100vh;margin:0;}}.box{{background:#faf7f4;border:1px solid #ddd5ca;border-radius:12px;padding:32px;max-width:480px;}}.title{{font-size:18px;font-weight:700;color:#991b1b;margin-bottom:12px;}}.btn{{background:#6a5e4e;color:#fff;border:none;padding:10px 24px;border-radius:8px;font-size:14px;cursor:pointer;text-decoration:none;display:inline-block;margin-top:16px;}}</style></head><body><div class="box"><div class="title">⚠️ 請修正以下欄位</div>{err_html}<a href="javascript:history.back()" class="btn">← 返回修正</a></div></body></html>""", status_code=400)
    source_group_id = f.get("grp","")
    live_same = "1" if f.get("sameck") else "0"
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT * FROM customers WHERE id_no=? AND status IN ('ACTIVE','PENDING')", (id_no,))
    existing = cur.fetchone()
    now = now_iso()
    if existing:
        case_id = existing["case_id"]
        conn.close()
        existing_name = existing["customer_name"]
        return HTMLResponse(f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
<style>body{{font-family:'Microsoft JhengHei',sans-serif;background:#ece8e2;display:flex;justify-content:center;align-items:center;min-height:100vh;margin:0;}}
.box{{background:#faf7f4;border:1px solid #ddd5ca;border-radius:12px;padding:32px;max-width:480px;text-align:center;}}
.icon{{font-size:48px;margin-bottom:16px;}}
.title{{font-size:20px;font-weight:700;color:#991b1b;margin-bottom:8px;}}
.msg{{font-size:14px;color:#4a3e30;margin-bottom:24px;line-height:1.6;}}
.btn{{background:#6a5e4e;color:#fff;border:none;padding:10px 24px;border-radius:8px;font-size:14px;cursor:pointer;text-decoration:none;display:inline-block;margin:4px;}}
.btn2{{background:#e8e2da;color:#4a3e30;border:1px solid #ddd5ca;padding:10px 24px;border-radius:8px;font-size:14px;cursor:pointer;text-decoration:none;display:inline-block;margin:4px;}}
</style></head><body>
<div class="box">
<div class="icon">⚠️</div>
<div class="title">此客戶已存在！</div>
<div class="msg">身分證號 <strong>{h(id_no)}</strong><br>客戶「<strong>{h(existing_name)}</strong>」已在系統中<br>（案件編號：{h(case_id)}）<br><br>資料已自動更新為最新填寫內容。</div>
<a href="/report" class="btn">前往日報查看</a>
<a href="/new-customer" class="btn2">繼續新增其他客戶</a>
</div></body></html>""", status_code=200)
    else:
        case_id = short_id()
        cur.execute("""INSERT INTO customers
            (case_id,customer_name,id_no,source_group_id,company,last_update,status,created_at,updated_at)
            VALUES (?,?,?,?,?,?,'PENDING',?,?)""",
            (case_id, name, id_no, source_group_id, "", "新增客戶："+name, now, now))
    extra = dict(
        birth_date=f.get("birth",""), phone=f.get("phone",""), email=f.get("email",""),
        line_id=f.get("line",""), marriage=f.get("marry",""), education=f.get("edu",""),
        id_issue_date=f.get("iddate",""), id_issue_place=f.get("idplace",""), id_issue_type=f.get("idtype",""),
        reg_city=f.get("rcity",""), reg_district=f.get("rdist",""), reg_address=f.get("raddr",""), reg_phone=f.get("rphone",""),
        live_same_as_reg=live_same,
        live_city=f.get("lcity","") if live_same=="0" else f.get("rcity",""),
        live_district=f.get("ldist","") if live_same=="0" else f.get("rdist",""),
        live_address=f.get("laddr","") if live_same=="0" else f.get("raddr",""),
        live_phone=f.get("lphone",""), live_status=f.get("lstatus",""),
        live_years=f.get("lyear",""), live_months=f.get("lmon",""),
        company_name_detail=f.get("cmpname",""), company_phone_area=f.get("carea",""),
        company_phone_num=f.get("cnum",""), company_phone_ext=f.get("cext",""),
        company_role=f.get("crole",""), company_years=f.get("cyear",""),
        company_months=f.get("cmon",""),
        company_salary=f.get("csal",""), company_city=f.get("ccity",""),
        company_district=f.get("cdist",""), company_address=f.get("caddr",""),
        contact1_name=f.get("c1name",""), contact1_relation=f.get("c1rel",""),
        contact1_phone=f.get("c1tel",""), contact1_known=f.get("c1know",""),
        contact2_name=f.get("c2name",""), contact2_relation=f.get("c2rel",""),
        contact2_phone=f.get("c2tel",""), contact2_known=f.get("c2know",""),
        eval_fund_need=f.get("efund",""), eval_sent_3m=f.get("esent",""),
        eval_labor_ins=f.get("elabor",""), eval_salary_transfer=f.get("esal",""),
        eval_alert=f.get("eprivate",""), eval_credit_card=f.get("ecard",""),
        eval_property=f.get("eprop",""), eval_fine=f.get("efine",""),
        eval_fuel_tax=f.get("efuel",""), eval_late=f.get("elate",""), eval_late_days=f.get("elateday",""), eval_note=f.get("enote",""),
        debt_list=f.get("debt_list",""),
        carrier=f.get("carrier",""), fb=f.get("fb",""),
        sales_name=f.get("sales",""), source_group_id=source_group_id, created_by_role=role,
    )
    flds = ", ".join(k+" = ?" for k in extra)
    vals = list(extra.values()) + [now_iso(), case_id]
    cur.execute("UPDATE customers SET "+flds+", updated_at=? WHERE case_id=?", vals)
    conn.commit(); conn.close()
    gname = get_group_name(source_group_id)
    # 行政A填完 → 顯示提示頁，告知業務去 LINE 建相簿
    return HTMLResponse(f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
<style>body{{font-family:'Microsoft JhengHei',sans-serif;background:#ece8e2;display:flex;justify-content:center;align-items:center;min-height:100vh;margin:0;}}
.box{{background:#faf7f4;border:1px solid #ddd5ca;border-radius:12px;padding:32px;max-width:480px;text-align:center;}}
.title{{font-size:20px;font-weight:700;color:#2c2820;margin-bottom:12px;}}
.msg{{font-size:14px;color:#4a3e40;margin-bottom:8px;line-height:1.8;}}
.note{{background:#fef9c3;border:1px solid #fde047;border-radius:8px;padding:12px;font-size:13px;color:#854d0e;margin-bottom:20px;text-align:left;}}
.btn{{background:#6a5e4e;color:#fff;border:none;padding:10px 24px;border-radius:8px;font-size:14px;cursor:pointer;text-decoration:none;display:inline-block;margin:4px;}}
.btn2{{background:#e8e2da;color:#4a3e30;border:1px solid #ddd5ca;padding:10px 24px;border-radius:8px;font-size:14px;text-decoration:none;display:inline-block;margin:4px;}}
</style></head><body>
<div class="box">
<div style="font-size:48px;margin-bottom:12px;">✅</div>
<div class="title">客戶資料已儲存！</div>
<div class="msg">客戶「<strong>{h(name)}</strong>」資料已完成填寫</div>
<div class="note">
⚠️ 請通知業務：<br>
1. 將客戶證件/勞保/財力證明上傳到 LINE 群組相簿<br>
2. 在 LINE 群組打格式建立客戶：<br>
<strong>{datetime.now().strftime('%y/%m/%d')}-{h(name)}-身分證號/...</strong><br>
完成後客戶才會出現在日報！
</div>
<a href="/new-customer" class="btn">繼續新增客戶</a>
<a href="/pending-customers" class="btn2">查看待確認客戶</a>
</div></body></html>""", status_code=200)

# =========================
# 客戶PDF列印頁面
# =========================
@app.get("/customer-pdf", response_class=HTMLResponse)
def customer_pdf(request: Request, case_id: str = ""):
    from fastapi.responses import RedirectResponse
    role = check_auth(request)
    if not role: return RedirectResponse("/login")
    if not case_id: return RedirectResponse("/pending-customers")
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT * FROM customers WHERE case_id=?", (case_id,))
    row = cur.fetchone(); conn.close()
    if not row: return RedirectResponse("/pending-customers")
    r = dict(row)
    def v(k): return h(r.get(k, "") or "")
    gname = h(get_group_name(r.get("source_group_id", "")))
    created = h((r.get("created_at", "") or "")[:10])

    # Parse debt list
    import json as _json
    try:
        debt_data = _json.loads(r.get("debt_list", "") or "[]") if r.get("debt_list") else []
    except Exception:
        debt_data = []
    debt_html = ""
    if debt_data:
        debt_html = '<table style="width:100%;border-collapse:collapse;font-size:13px;margin-top:8px;">'
        debt_html += '<tr style="background:#e8e2da;font-weight:600;"><th style="padding:6px;border:1px solid #bbb;">貸款商家</th><th style="padding:6px;border:1px solid #bbb;">金額</th><th style="padding:6px;border:1px solid #bbb;">期數/已繳</th><th style="padding:6px;border:1px solid #bbb;">月繳</th><th style="padding:6px;border:1px solid #bbb;">剩餘</th><th style="padding:6px;border:1px solid #bbb;">日期</th><th style="padding:6px;border:1px solid #bbb;">動保</th></tr>'
        for d in debt_data:
            debt_html += f'<tr><td style="padding:5px;border:1px solid #bbb;">{h(d.get("co",""))}</td><td style="padding:5px;border:1px solid #bbb;">{h(d.get("lo",""))}</td><td style="padding:5px;border:1px solid #bbb;">{h(d.get("pe",""))}/{h(d.get("pa",""))}</td><td style="padding:5px;border:1px solid #bbb;">{h(d.get("mo",""))}</td><td style="padding:5px;border:1px solid #bbb;">{h(d.get("re",""))}</td><td style="padding:5px;border:1px solid #bbb;">{h(d.get("da",""))}</td><td style="padding:5px;border:1px solid #bbb;">{h(d.get("dy",""))}</td></tr>'
        debt_html += '</table>'

    lsame = r.get("live_same_as_reg", "") == "1"
    live_addr = f'{v("reg_city")}{v("reg_district")}{v("reg_address")}' if lsame else f'{v("live_city")}{v("live_district")}{v("live_address")}'
    reg_addr = f'{v("reg_city")}{v("reg_district")}{v("reg_address")}'
    company_addr = f'{v("company_city")}{v("company_district")}{v("company_address")}'

    return f"""<!DOCTYPE html><html lang="zh-TW"><head><meta charset="UTF-8">
<title>客戶資料 - {v("customer_name")}</title>
<style>
@media print {{ @page {{ size: A4 portrait; margin: 10mm 12mm; }} .no-print {{ display: none !important; }} }}
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: 'Microsoft JhengHei', 'PingFang TC', sans-serif; background: #fff; color: #1a1a1a; font-size: 13px; padding: 20px; }}
.header {{ background: #3a3530; color: #fff; padding: 16px 20px; border-radius: 8px; margin-bottom: 16px; display: flex; justify-content: space-between; align-items: center; }}
.header-name {{ font-size: 22px; font-weight: 700; }}
.header-sub {{ font-size: 12px; color: #c8bfb5; margin-top: 4px; }}
table {{ width: 100%; border-collapse: collapse; margin-bottom: 16px; }}
th, td {{ border: 1px solid #bbb; padding: 7px 10px; font-size: 13px; line-height: 1.5; }}
th {{ background: #f0ebe4; color: #3a3020; font-weight: 700; width: 100px; white-space: nowrap; text-align: left; }}
td {{ background: #fff; }}
.sec {{ background: #3a3530; color: #fff; font-size: 12px; font-weight: 700; padding: 6px 10px; }}
.sec td {{ background: #3a3530; color: #fff; font-weight: 700; }}
</style>
</head><body>
<div class="no-print" style="text-align:center;margin-bottom:16px;">
  <button onclick="window.print()" style="background:#4e7055;color:#fff;border:none;padding:10px 28px;border-radius:6px;font-size:14px;cursor:pointer;font-weight:600;">列印 / 存 PDF</button>
  <button onclick="history.back()" style="background:#6a5e4e;color:#fff;border:none;padding:10px 28px;border-radius:6px;font-size:14px;cursor:pointer;font-weight:600;margin-left:8px;">返回</button>
</div>
<div class="header">
  <div><div class="header-name">{v("customer_name")}</div><div class="header-sub">群組：{gname}　建立日期：{created}</div></div>
  <div style="font-size:13px;color:#c8bfb5;font-weight:600;">客戶資料表</div>
</div>
<table>
<colgroup><col class="c-th"><col class="c-td"><col class="c-th"><col class="c-td"></colgroup>
<tr class="sec"><td colspan="4">基本資料</td></tr>
<tr><th>姓名</th><td>{v("customer_name")}</td><th>身分證</th><td>{v("id_no")}</td></tr>
<tr><th>出生日期</th><td>{v("birth_date")}</td><th>行動電話</th><td>{v("phone")}</td></tr>
<tr><th>Email</th><td>{v("email")}</td><th>LINE ID</th><td>{v("line_id")}</td></tr>
<tr><th>婚姻</th><td>{v("marriage")}</td><th>學歷</th><td>{v("education")}</td></tr>
<tr class="sec"><td colspan="4">身分證發證</td></tr>
<tr><th>發證日期</th><td>{v("id_issue_date")}</td><th>發證地</th><td>{v("id_issue_place")}</td></tr>
<tr><th>換補發</th><td colspan="3">{v("id_issue_type")}</td></tr>
<tr class="sec"><td colspan="4">地址資料</td></tr>
<tr><th>戶籍地址</th><td colspan="3">{reg_addr}</td></tr>
<tr><th>戶籍電話</th><td>{v("reg_phone")}</td><th>現住電話</th><td>{v("live_phone")}</td></tr>
<tr><th>住家地址</th><td colspan="3">{live_addr}</td></tr>
<tr><th>居住狀況</th><td>{v("live_status")}</td><th>居住時間</th><td>{v("live_years")}年{v("live_months")}月</td></tr>
<tr class="sec"><td colspan="4">職業資料</td></tr>
<tr><th>公司名稱</th><td colspan="3">{v("company_name_detail")}</td></tr>
<tr><th>公司電話</th><td>{v("company_phone_area")}-{v("company_phone_num")}</td><th>職稱</th><td>{v("company_role")}</td></tr>
<tr><th>年資</th><td>{v("company_years")}年{v("company_months")}月</td><th>月薪</th><td>{v("company_salary")}萬</td></tr>
<tr><th>公司地址</th><td colspan="3">{company_addr}</td></tr>
<tr><th>行業</th><td colspan="3">{v("company_industry")}</td></tr>
<tr class="sec"><td colspan="4">聯絡人</td></tr>
<tr><th>聯絡人1</th><td>{v("contact1_name")}（{v("contact1_relation")}）</td><th>電話</th><td>{v("contact1_phone")}</td></tr>
<tr><th>知情</th><td colspan="3">{v("contact1_known")}</td></tr>
<tr><th>聯絡人2</th><td>{v("contact2_name")}（{v("contact2_relation")}）</td><th>電話</th><td>{v("contact2_phone")}</td></tr>
<tr><th>知情</th><td colspan="3">{v("contact2_known")}</td></tr>
</table>
<div style="page-break-before:always;margin-top:20px;"></div>
<table>
<colgroup><col class="c-th"><col class="c-td"><col class="c-th"><col class="c-td"></colgroup>
<tr class="sec"><td colspan="4">貸款諮詢</td></tr>
<tr><th>資金需求</th><td>{v("eval_fund_need")}</td><th>近三月送件</th><td>{v("eval_sent_3m")}</td></tr>
<tr><th>當鋪私設</th><td>{v("eval_alert")}</td><th>勞保</th><td>{v("eval_labor_ins")}</td></tr>
<tr><th>薪轉</th><td>{v("eval_salary_transfer")}</td><th>遲繳</th><td>{v("eval_late")} {v("eval_late_days")}天</td></tr>
<tr><th>罰單</th><td>{v("eval_fine")}</td><th>燃料稅</th><td>{v("eval_fuel_tax")}</td></tr>
<tr><th>信用卡</th><td>{v("eval_credit_card")}</td><th>動產</th><td>{v("eval_property")}</td></tr>
<tr><th>證照</th><td>{v("eval_license")}</td><th>車輛</th><td>{v("eval_vehicle")}</td></tr>
<tr><th>法學</th><td colspan="3">{v("eval_law")}</td></tr>
{"<tr><th>備註</th><td colspan='3'>" + v("eval_note") + "</td></tr>" if r.get("eval_note") else ""}
</table>
{"<div style='font-size:13px;font-weight:700;color:#3a3530;margin-bottom:8px;'>負債明細</div>" + debt_html if debt_html else ""}
</body></html>"""


def _build_customer_pdf_body(r: dict) -> str:
    """組出單一客戶在 PDF 內的內容（header + 兩頁表格），給單筆與批次共用"""
    def v(k): return h(r.get(k, "") or "")
    gname = h(get_group_name(r.get("source_group_id", "")))
    created = h((r.get("created_at", "") or "")[:10])

    import json as _json
    try:
        debt_data = _json.loads(r.get("debt_list", "") or "[]") if r.get("debt_list") else []
    except Exception:
        debt_data = []
    debt_html = ""
    if debt_data:
        debt_html = '<table style="width:100%;border-collapse:collapse;font-size:13px;margin-top:8px;">'
        debt_html += '<tr style="background:#e8e2da;font-weight:600;"><th style="padding:6px;border:1px solid #bbb;">貸款商家</th><th style="padding:6px;border:1px solid #bbb;">金額</th><th style="padding:6px;border:1px solid #bbb;">期數/已繳</th><th style="padding:6px;border:1px solid #bbb;">月繳</th><th style="padding:6px;border:1px solid #bbb;">剩餘</th><th style="padding:6px;border:1px solid #bbb;">日期</th><th style="padding:6px;border:1px solid #bbb;">動保</th></tr>'
        for d in debt_data:
            debt_html += f'<tr><td style="padding:5px;border:1px solid #bbb;">{h(d.get("co",""))}</td><td style="padding:5px;border:1px solid #bbb;">{h(d.get("lo",""))}</td><td style="padding:5px;border:1px solid #bbb;">{h(d.get("pe",""))}/{h(d.get("pa",""))}</td><td style="padding:5px;border:1px solid #bbb;">{h(d.get("mo",""))}</td><td style="padding:5px;border:1px solid #bbb;">{h(d.get("re",""))}</td><td style="padding:5px;border:1px solid #bbb;">{h(d.get("da",""))}</td><td style="padding:5px;border:1px solid #bbb;">{h(d.get("dy",""))}</td></tr>'
        debt_html += '</table>'

    lsame = r.get("live_same_as_reg", "") == "1"
    live_addr = f'{v("reg_city")}{v("reg_district")}{v("reg_address")}' if lsame else f'{v("live_city")}{v("live_district")}{v("live_address")}'
    reg_addr = f'{v("reg_city")}{v("reg_district")}{v("reg_address")}'
    company_addr = f'{v("company_city")}{v("company_district")}{v("company_address")}'
    note_html = "<tr><th>備註</th><td colspan='3'>" + v("eval_note") + "</td></tr>" if r.get("eval_note") else ""
    debt_section = "<div style='font-size:13px;font-weight:700;color:#3a3530;margin-bottom:8px;'>負債明細</div>" + debt_html if debt_html else ""

    return f"""<div class="header">
  <div><div class="header-name">{v("customer_name")}</div><div class="header-sub">群組：{gname}　建立日期：{created}</div></div>
  <div style="font-size:13px;color:#c8bfb5;font-weight:600;">客戶資料表</div>
</div>
<table>
<colgroup><col class="c-th"><col class="c-td"><col class="c-th"><col class="c-td"></colgroup>
<tr class="sec"><td colspan="4">基本資料</td></tr>
<tr><th>姓名</th><td>{v("customer_name")}</td><th>身分證</th><td>{v("id_no")}</td></tr>
<tr><th>出生日期</th><td>{v("birth_date")}</td><th>行動電話</th><td>{v("phone")}</td></tr>
<tr><th>Email</th><td>{v("email")}</td><th>LINE ID</th><td>{v("line_id")}</td></tr>
<tr><th>婚姻</th><td>{v("marriage")}</td><th>學歷</th><td>{v("education")}</td></tr>
<tr class="sec"><td colspan="4">身分證發證</td></tr>
<tr><th>發證日期</th><td>{v("id_issue_date")}</td><th>發證地</th><td>{v("id_issue_place")}</td></tr>
<tr><th>換補發</th><td colspan="3">{v("id_issue_type")}</td></tr>
<tr class="sec"><td colspan="4">地址資料</td></tr>
<tr><th>戶籍地址</th><td colspan="3">{reg_addr}</td></tr>
<tr><th>戶籍電話</th><td>{v("reg_phone")}</td><th>現住電話</th><td>{v("live_phone")}</td></tr>
<tr><th>住家地址</th><td colspan="3">{live_addr}</td></tr>
<tr><th>居住狀況</th><td>{v("live_status")}</td><th>居住時間</th><td>{v("live_years")}年{v("live_months")}月</td></tr>
<tr class="sec"><td colspan="4">職業資料</td></tr>
<tr><th>公司名稱</th><td colspan="3">{v("company_name_detail")}</td></tr>
<tr><th>公司電話</th><td>{v("company_phone_area")}-{v("company_phone_num")}</td><th>職稱</th><td>{v("company_role")}</td></tr>
<tr><th>年資</th><td>{v("company_years")}年{v("company_months")}月</td><th>月薪</th><td>{v("company_salary")}萬</td></tr>
<tr><th>公司地址</th><td colspan="3">{company_addr}</td></tr>
<tr><th>行業</th><td colspan="3">{v("company_industry")}</td></tr>
<tr class="sec"><td colspan="4">聯絡人</td></tr>
<tr><th>聯絡人1</th><td>{v("contact1_name")}（{v("contact1_relation")}）</td><th>電話</th><td>{v("contact1_phone")}</td></tr>
<tr><th>知情</th><td colspan="3">{v("contact1_known")}</td></tr>
<tr><th>聯絡人2</th><td>{v("contact2_name")}（{v("contact2_relation")}）</td><th>電話</th><td>{v("contact2_phone")}</td></tr>
<tr><th>知情</th><td colspan="3">{v("contact2_known")}</td></tr>
</table>
<div style="page-break-before:always;margin-top:20px;"></div>
<table>
<colgroup><col class="c-th"><col class="c-td"><col class="c-th"><col class="c-td"></colgroup>
<tr class="sec"><td colspan="4">貸款諮詢</td></tr>
<tr><th>資金需求</th><td>{v("eval_fund_need")}</td><th>近三月送件</th><td>{v("eval_sent_3m")}</td></tr>
<tr><th>當鋪私設</th><td>{v("eval_alert")}</td><th>勞保</th><td>{v("eval_labor_ins")}</td></tr>
<tr><th>薪轉</th><td>{v("eval_salary_transfer")}</td><th>遲繳</th><td>{v("eval_late")} {v("eval_late_days")}天</td></tr>
<tr><th>罰單</th><td>{v("eval_fine")}</td><th>燃料稅</th><td>{v("eval_fuel_tax")}</td></tr>
<tr><th>信用卡</th><td>{v("eval_credit_card")}</td><th>動產</th><td>{v("eval_property")}</td></tr>
<tr><th>證照</th><td>{v("eval_license")}</td><th>車輛</th><td>{v("eval_vehicle")}</td></tr>
<tr><th>法學</th><td colspan="3">{v("eval_law")}</td></tr>
{note_html}
</table>
{debt_section}"""


_PDF_STYLE = """<style>
@media print { @page { size: A4 portrait; margin: 10mm 12mm; } .no-print { display: none !important; } }
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: 'Microsoft JhengHei', 'PingFang TC', sans-serif; background: #eee; color: #1a1a1a; font-size: 14px; -webkit-font-smoothing: antialiased; margin: 0; padding: 0; }
#pdf-wrap { padding: 20px 0 0 0; }
#pdf-content { width: 210mm; min-height: 297mm; padding: 10mm 12mm; background: #fff; box-sizing: border-box; margin: 0 auto; }
@media print { body { background: #fff; } #pdf-wrap { padding: 0; } #pdf-content { padding: 0; width: auto; min-height: auto; margin: 0; } }
.header { background: #3a3530; color: #fff; padding: 10px 16px; border-radius: 6px; margin-bottom: 10px; display: flex; justify-content: space-between; align-items: center; }
.header-name { font-size: 18px; font-weight: 700; line-height: 1.3; }
.header-sub { font-size: 11px; color: #c8bfb5; margin-top: 2px; line-height: 1.3; }
table { width: 100%; border-collapse: collapse; margin-bottom: 10px; table-layout: fixed; }
col.c-th { width: 15%; }
col.c-td { width: 35%; }
th, td { border: 1px solid #bbb; padding: 6px 9px; font-size: 14px; line-height: 1.4; vertical-align: middle; word-break: break-all; }
th { background: #f0ebe4; color: #3a3020; font-weight: 700; white-space: nowrap; text-align: left; }
td { background: #fff; }
.sec td { background: #3a3530; color: #fff; font-size: 11px; font-weight: 700; padding: 4px 8px; }
</style>"""


@app.get("/customer-pdf-batch", response_class=HTMLResponse)
def customer_pdf_batch(request: Request, ids: str = ""):
    from fastapi.responses import RedirectResponse
    role = check_auth(request)
    if not role: return RedirectResponse("/login")
    id_list = [i.strip() for i in (ids or "").split(",") if i.strip()]
    if not id_list:
        return RedirectResponse("/pending-customers")
    conn = get_conn(); cur = conn.cursor()
    placeholders = ",".join(["?"] * len(id_list))
    cur.execute(f"SELECT * FROM customers WHERE case_id IN ({placeholders})", id_list)
    rows = cur.fetchall(); conn.close()
    # 依使用者選取順序排序
    row_map = {dict(r)["case_id"]: dict(r) for r in rows}
    ordered = [row_map[i] for i in id_list if i in row_map]
    if not ordered:
        return RedirectResponse("/pending-customers")

    bodies = []
    for idx, r in enumerate(ordered):
        if idx > 0:
            bodies.append('<div style="page-break-before:always;"></div>')
        bodies.append(_build_customer_pdf_body(r))
    body_html = "".join(bodies)

    # 第一筆客戶姓名當檔名
    first_name = (ordered[0].get("customer_name") or "客戶資料").strip()
    if len(ordered) > 1:
        pdf_filename = f"{first_name}等{len(ordered)}筆客戶資料.pdf"
    else:
        pdf_filename = f"{first_name}_客戶資料.pdf"

    return f"""<!DOCTYPE html><html lang="zh-TW"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>客戶資料批次列印（{len(ordered)} 筆）</title>
{_PDF_STYLE}
<script src="https://cdnjs.cloudflare.com/ajax/libs/html2pdf.js/0.10.1/html2pdf.bundle.min.js"></script>
</head><body>
<div class="no-print" style="text-align:center;margin-bottom:16px;">
  <button id="pdf-btn" onclick="downloadPDF()" style="background:#4e7055;color:#fff;border:none;padding:10px 28px;border-radius:6px;font-size:14px;cursor:pointer;font-weight:600;">下載 PDF（共 {len(ordered)} 筆）</button>
  <button onclick="window.close()" style="background:#6a5e4e;color:#fff;border:none;padding:10px 28px;border-radius:6px;font-size:14px;cursor:pointer;font-weight:600;margin-left:8px;">關閉</button>
</div>
<div id="pdf-wrap">
<div id="pdf-content">
{body_html}
</div>
</div>
<script>
function downloadPDF() {{
  var btn = document.getElementById('pdf-btn');
  var originalText = btn.innerText;
  if (typeof html2pdf === 'undefined') {{
    window.print();
    return;
  }}
  btn.disabled = true;
  btn.innerText = '產生中，請稍候…';
  btn.style.opacity = '0.6';
  btn.style.cursor = 'wait';
  var element = document.getElementById('pdf-content');
  var opt = {{
    margin: 0,
    filename: {json.dumps(pdf_filename, ensure_ascii=False)},
    image: {{ type: 'jpeg', quality: 0.95 }},
    html2canvas: {{
      scale: 2, useCORS: true, letterRendering: true, backgroundColor: '#ffffff',
      scrollX: 0, scrollY: -window.scrollY,
      onclone: function(doc) {{
        var el = doc.getElementById('pdf-content');
        if (el) {{
          el.style.margin = '0';
          el.style.position = 'static';
          el.style.left = '0';
          el.style.transform = 'none';
        }}
        doc.body.style.padding = '0';
        doc.body.style.margin = '0';
        doc.body.style.background = '#fff';
        doc.documentElement.style.margin = '0';
        doc.documentElement.style.padding = '0';
      }}
    }},
    jsPDF: {{ unit: 'mm', format: 'a4', orientation: 'portrait' }},
    pagebreak: {{ mode: ['css', 'legacy'] }}
  }};
  var ua = navigator.userAgent || '';
  var isMobile = /iPhone|iPad|iPod|Android|Line/i.test(ua);
  var resetBtn = function() {{
    btn.disabled = false;
    btn.innerText = originalText;
    btn.style.opacity = '1';
    btn.style.cursor = 'pointer';
  }};
  var worker = html2pdf().from(element).set(opt);
  if (isMobile) {{
    // 手機（iOS/Android/LINE 內建瀏覽器）：產生 blob URL 在新分頁開啟
    // 讓使用者用系統分享鍵存檔或傳送，不靠 programmatic download
    worker.toPdf().get('pdf').then(function(pdf) {{
      var blob = pdf.output('blob');
      var url = URL.createObjectURL(blob);
      var opened = window.open(url, '_blank');
      if (!opened) {{
        // 跳窗被擋 → 直接換頁顯示 PDF
        window.location.href = url;
      }}
      resetBtn();
    }}).catch(function(err) {{
      console.error('PDF 產生失敗：', err);
      alert('PDF 產生失敗：' + (err && err.message ? err.message : err));
      resetBtn();
    }});
  }} else {{
    // 電腦：直接下載
    worker.save().then(resetBtn).catch(function(err) {{
      console.error('PDF 產生失敗：', err);
      alert('PDF 產生失敗，改用瀏覽器列印功能');
      window.print();
      resetBtn();
    }});
  }}
}}
</script>
</body></html>"""


# =========================
# 下載 Excel 申請書
# =========================
@app.post("/adminb/save-signature")
async def adminb_save_signature(request: Request):
    role = check_auth(request)
    if not role or role not in ("admin", "adminB"):
        return JSONResponse({"ok": False, "error": "無權限"}, status_code=403)
    try:
        data = await request.json()
        case_id = data.get("case_id", "")
        sig_type = data.get("type", "")
        sig_data = data.get("data", "")
        if not case_id or not sig_type or not sig_data:
            return JSONResponse({"ok": False, "error": "缺少參數"})
        # sig_type 白名單檢查（防靜默接受非法值）
        if sig_type not in ("applicant", "legal_rep", "both"):
            return JSONResponse({"ok": False, "error": "sig_type 必須是 applicant/legal_rep/both"}, status_code=400)
        conn = get_conn(); cur = conn.cursor()
        if sig_type == "both":
            cur.execute("UPDATE customers SET signature_applicant=?, signature_legal_rep=? WHERE case_id=?",
                        (sig_data, sig_data, case_id))
        elif sig_type == "applicant":
            cur.execute("UPDATE customers SET signature_applicant=? WHERE case_id=?", (sig_data, case_id))
        else:  # legal_rep
            cur.execute("UPDATE customers SET signature_legal_rep=? WHERE case_id=?", (sig_data, case_id))
        conn.commit(); conn.close()
        return JSONResponse({"ok": True})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)})


@app.get("/adminb/download-qiaomei")
def adminb_download_qiaomei(request: Request, case_id: str = ""):
    from fastapi.responses import StreamingResponse
    role = check_auth(request)
    if not role or role not in ("admin", "adminB"):
        return JSONResponse({"error": "無權限"}, status_code=403)
    if not case_id:
        return JSONResponse({"error": "缺少 case_id"}, status_code=400)
    try:
        conn = get_conn(); cur = conn.cursor()
        cur.execute("SELECT * FROM customers WHERE case_id=?", (case_id,))
        row = cur.fetchone(); conn.close()
        if not row:
            return JSONResponse({"error": "找不到客戶"}, status_code=404)
        r = dict(row)
        try:
            pdf_bytes = _fill_qiaomei_pdf(r)
        except Exception as gen_e:
            import traceback
            tb = traceback.format_exc()
            print(f"喬美 PDF 詳細錯誤:\n{tb}")
            return JSONResponse({"error": f"PDF 生成失敗：{str(gen_e)}"}, status_code=500)
        if not pdf_bytes:
            return JSONResponse({"error": "PDF 生成失敗：範本檔案不存在或內容為空"}, status_code=500)
        from urllib.parse import quote
        fname = quote(f"{r.get('customer_name','')}_喬美申請書.pdf")
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname}"}
        )
    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"error": f"下載失敗：{str(e)}"}, status_code=500)


def _fill_qiaomei_pdf(r: dict) -> bytes:
    """填入客戶資料到喬美 PDF 範本（reportlab 疊加文字層）"""
    try:
        from pypdf import PdfReader, PdfWriter
        from reportlab.pdfgen import canvas
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
        from reportlab.lib.pagesizes import A4
        import base64

        # 註冊中文字型
        try:
            pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
            font_name = 'STSong-Light'
        except Exception:
            font_name = 'Helvetica'

        _base = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.join(_base, "申請書", "喬美3-14萬.pdf")
        if not os.path.exists(template_path):
            return b""

        def v(k): return (r.get(k, "") or "").strip()

        # 公司電話組合（mobile 不顯示）
        co_area = v("company_phone_area")
        co_num = v("company_phone_num")
        if co_area == "mobile":
            co_area = ""
        # 手機（10碼 09 開頭）：不加區碼橫線，純數字
        combined = (co_area + co_num) if co_area else co_num
        if combined.startswith("09") and len(combined.replace("-","")) == 10:
            co_phone = combined.replace("-", "")
        else:
            co_phone = (co_area + "-" + co_num) if co_area and co_num else co_num

        live_same = v("live_same_as_reg") == "1"
        reg_addr = v("reg_city") + v("reg_district") + v("reg_address")
        live_addr = reg_addr if live_same else (v("live_city") + v("live_district") + v("live_address"))

        # 欄位座標表（PDF 點坐標，原點左下，y 反轉）
        # PDF page1 size: 612 x 859
        # 這些座標基於對 PDF 標籤位置的分析，可能需要微調
        page_h = 859
        def yp(top): return page_h - top  # PDF y 反轉

        # === 解析日期為年/月/日 ===
        def parse_ymd(d):
            if not d:
                return ("", "", "")
            parts = d.replace("-", "/").split("/")
            if len(parts) != 3:
                return ("", "", "")
            try:
                y = int(parts[0])
                if y >= 1911:
                    y -= 1911
                return (str(y), str(int(parts[1])), str(int(parts[2])))
            except Exception:
                return ("", "", "")
        b_y, b_m, b_d = parse_ymd(v("birth_date"))
        i_y, i_m, i_d = parse_ymd(v("id_issue_date"))

        from datetime import datetime as _dt
        now = _dt.now()
        ap_y = str(now.year - 1911)
        ap_m = str(now.month)
        ap_d = str(now.day)

        id_no_str = v("id_no").upper()
        qm_model = v("product_model")
        qm_imei = v("product_imei")

        # === 精準座標表（從用戶填好的範本 PDF 提取）===
        # 格式：(x, top, value)
        fields_p1 = [
            # 申請人姓名
            (105, 78, v("customer_name")),
            # 出生日期 年/月/日 (避開範本字 年=249.4 月=265.2 日=284.2)
            (240, 80, b_y),
            (258, 80, b_m),
            (275, 80, b_d),
            # 發證日期 年/月/日 (避開 年=115.4 月=137 日=160.3)
            (98, 129, i_y),
            (122, 129, i_m),
            (145, 129, i_d),
            # 發證地（短碼）
            (241, 128, v("id_issue_place")),
            # 親屬姓名 / 關係 / 電話
            (367, 147, v("contact1_name")),
            (416, 151, v("contact1_relation")),
            (504, 150, v("contact1_phone")),
            # 親友姓名 / 關係+電話
            (367, 196, v("contact2_name")),
            (416, 199, v("contact2_relation")),
            (501, 199, v("contact2_phone")),
            # 戶籍地址
            (107, 201, reg_addr),
            # 住宅地址（同戶籍時清空）
            (151, 223, live_addr if not live_same else ""),
            # E-mail
            (107, 250, v("email")),
            # 戶籍電話
            (101, 276, v("reg_phone")),
            # 行動電話
            (105, 300, v("phone")),
            # LINE ID
            (112, 373, v("line_id")),
            # 公司名稱（長字串自動換較小字體，由下方迴圈處理）
            # (98, 410, company_name_detail) — 用 _LONG_FIELDS 處理
            (188, 412, co_phone),
            (260, 413, v("company_phone_ext")),
            # 公司地址
            (105, 429, v("company_city") + v("company_district") + v("company_address")),
            # 職稱
            (105, 450, v("company_role")),
            # 年資 年 / 月（位置：年=229, 月=267）
            (229, 452, v("company_years")),
            (267, 451, v("company_months")),
            # 月薪
            (110, 472, v("company_salary")),
            # 居住時間 年/月
            (119.8, 326, v("live_years")),
            (174.8, 326, v("live_months")),
            # 信用卡：發卡銀行 / 卡號 / 有效日期年/月 (同卡號行 y=323) / 額度 (下行 y=353)
            (371, 285, v("adminb_credit_bank")),
            (359, 323, v("adminb_credit_no")),
            # 有效日期年：在「年」字 (538.1) 之前
            (505, 323, v("adminb_credit_exp").split("/")[0] if "/" in v("adminb_credit_exp") else ""),
            # 有效日期月：在「年」(538) 和「月」(569) 之間
            (548, 323, v("adminb_credit_exp").split("/")[1] if "/" in v("adminb_credit_exp") else v("adminb_credit_exp")),
            # 額度：在「萬」字 (423.9) 之前的空白
            (380, 353, v("adminb_credit_limit")),
            # 商品名稱（手機型號）+ IMEI
            (93, 565, qm_model),
            (80, 588, qm_imei),
        ]

        sig_app = r.get("signature_applicant", "") or ""
        sig_leg = r.get("signature_legal_rep", "") or ""

        def draw_signature(c, sig_data, x, y, w, h):
            if not sig_data or not sig_data.startswith("data:image"):
                return
            try:
                _, b64 = sig_data.split(",", 1)
                img_bytes = base64.b64decode(b64)
                from reportlab.lib.utils import ImageReader
                img = ImageReader(io.BytesIO(img_bytes))
                c.drawImage(img, x, y, width=w, height=h, mask='auto')
            except Exception as ex:
                print(f"draw_signature error: {ex}")

        # === Page 1 疊加層 ===
        # 用 PDF 實際 mediabox（612.288 x 858.898）
        from pypdf import PdfReader as _PR
        _r = _PR(template_path)
        p1_w = float(_r.pages[0].mediabox.width)
        p1_h = float(_r.pages[0].mediabox.height)
        def yp1(top): return p1_h - top  # PDF y 反轉

        overlay1 = io.BytesIO()
        c1 = canvas.Canvas(overlay1, pagesize=(p1_w, p1_h))
        DEFAULT_FONT_SIZE = 10

        # 一般欄位：pdfplumber top → PDF baseline ≈ top + ascent (≈7.5 for 10pt)
        ASCENT = 7.5
        c1.setFont(font_name, DEFAULT_FONT_SIZE)
        for x, top, val in fields_p1:
            if not val:
                continue
            c1.drawString(x, yp1(top + ASCENT), str(val))

        # === 公司名稱：超長自動縮字體（避免覆蓋電話欄位）===
        co_name = v("company_name_detail")
        if co_name:
            # 公司名稱欄寬約 90pt（x=98 到 188）
            max_w = 88
            fs = DEFAULT_FONT_SIZE
            # CJK 字符約 1 字 = 字體大小寬度
            while fs > 6 and c1.stringWidth(co_name, font_name, fs) > max_w:
                fs -= 0.5
            c1.setFont(font_name, fs)
            c1.drawString(98, yp1(410 + ASCENT), co_name)
            c1.setFont(font_name, DEFAULT_FONT_SIZE)

        # === 身分證字號 10 格（每格中心填一字）===
        # 範本中顯示位置 y=106，x=107 起，每格 19pt
        if id_no_str:
            c1.setFont(font_name, 11)
            for i, ch in enumerate(id_no_str[:10]):
                cx = 107 + i * 19
                c1.drawString(cx, yp1(106 + 8), ch)
            c1.setFont(font_name, DEFAULT_FONT_SIZE)

        # === 勾選方塊（畫 ✓ 勾號；rect 7x7pt）===
        def tick_box(box_x, box_top):
            cx = box_x + 3.5
            ct = box_top + 3.5
            cy = yp1(ct)
            c1.setLineWidth(1.4)
            # 勾號：左下短斜→右上長斜
            c1.line(cx - 3.5, cy, cx - 1, cy - 3)
            c1.line(cx - 1, cy - 3, cx + 4, cy + 4)

        # 換補發：範本沒有 rect，直接在標籤前畫 ✓
        def tick_inline(x, top):
            cy = yp1(top + 3)
            c1.setLineWidth(1.4)
            c1.line(x, cy, x + 2.5, cy - 3)
            c1.line(x + 2.5, cy - 3, x + 7, cy + 4)

        # 換補發 (3 vertical rect at x=268.8): 初=122.2, 換=129.3, 補=136.5
        issue_kind = v("id_issue_type") or v("id_issue_kind")
        if "初" in issue_kind:
            tick_box(268.8, 122.2)
        elif "換" in issue_kind:
            tick_box(268.8, 129.3)
        elif "補" in issue_kind:
            tick_box(268.8, 136.5)

        # 婚姻 (y=153.5): 已婚=105.9, 未婚=142.5, 離婚=177.2, 喪偶=212.6
        marriage = v("marriage")
        if "已婚" in marriage:
            tick_box(105.9, 153.5)
        elif "未婚" in marriage:
            tick_box(142.5, 153.5)
        elif "離" in marriage:
            tick_box(177.2, 153.5)
        elif "喪" in marriage:
            tick_box(212.6, 153.5)

        # 教育 (y=172.7 上排: 碩士=105.9, 大學=153.5; y=183.0 下排: 高中=105.9, 國中=153.5)
        edu = v("education")
        if "研究" in edu or "碩" in edu:
            tick_box(105.9, 172.7)
        elif any(k in edu for k in ("大學","專科")):
            tick_box(153.5, 172.7)
        elif "高中" in edu or "高職" in edu:
            tick_box(105.9, 183.0)
        else:
            tick_box(153.5, 183.0)

        # 居住狀況 (y=342.6: 自有=105.9, 父母/配偶=152.4, 親屬=197.8;
        #          y=353.3: 租屋=105.9, 宿舍=134.5, 借住=162.8)
        live = v("live_status")
        if "自有" in live:
            tick_box(105.9, 342.6)
        elif "父母" in live or "配偶" in live:
            tick_box(152.4, 342.6)
        elif "親屬" in live or "親" in live:
            tick_box(197.8, 342.6)
        elif "租" in live:
            tick_box(105.9, 353.3)
        elif "宿舍" in live:
            tick_box(134.5, 353.3)
        elif "借" in live:
            tick_box(162.8, 353.3)

        # 同戶籍勾選 (y=226.0 x=105.9)
        if live_same:
            tick_box(105.9, 226.0)

        # 申請人正楷簽名（簽名框 rect: x=25.1 top=787.2 268.6x46.7，標籤在 y=791）
        # 簽名圖放在標籤右側，框內
        draw_signature(c1, sig_app, 130, yp1(830), 150, 38)
        # 法定代理人不簽

        c1.showPage()
        c1.save()
        overlay1.seek(0)

        # === Page 2 疊加層 ===
        p2_w = float(_r.pages[1].mediabox.width) if len(_r.pages) >= 2 else 541.68
        p2_h = float(_r.pages[1].mediabox.height) if len(_r.pages) >= 2 else 745.68
        def yp2(top): return p2_h - top
        overlay2 = io.BytesIO()
        c2 = canvas.Canvas(overlay2, pagesize=(p2_w, p2_h))
        # 立約定書人簽名 → 標籤右側（微調對齊）
        draw_signature(c2, sig_app, 320, yp2(728), 110, 18)
        # 日期：今天民國年（範本位置 y=719, x=433/472/505）
        c2.setFont(font_name, 10)
        c2.drawString(433, yp2(719 + 10), ap_y)
        c2.drawString(472, yp2(719 + 10), ap_m)
        c2.drawString(505, yp2(719 + 10), ap_d)
        c2.showPage()
        c2.save()
        overlay2.seek(0)

        # 用 pikepdf 合併（pypdf 對此範本會把所有字元重複輸出）
        # 必須傳 Rectangle 強制使用 mediabox 而非 trimbox（範本 trimbox 會造成 ~10pt 偏移）
        import pikepdf
        src = pikepdf.open(template_path)
        ov1_pdf = pikepdf.open(overlay1)
        rect1 = pikepdf.Rectangle(0, 0, p1_w, p1_h)
        src.pages[0].add_overlay(ov1_pdf.pages[0], rect1)
        if len(src.pages) >= 2:
            ov2_pdf = pikepdf.open(overlay2)
            rect2 = pikepdf.Rectangle(0, 0, p2_w, p2_h)
            src.pages[1].add_overlay(ov2_pdf.pages[0], rect2)

        out = io.BytesIO()
        src.save(out)
        out.seek(0)
        return out.getvalue()
    except Exception as e:
        import traceback; traceback.print_exc()
        print(f"喬美 PDF 生成錯誤: {e}")
        raise  # 讓上層 catch 到具體錯誤


@app.get("/adminb/download-excel")
def adminb_download_excel(request: Request, case_id: str = ""):
    from fastapi.responses import StreamingResponse
    try:
        return _do_download_excel(request, case_id)
    except Exception as e:
        print(f"Download Excel error: {e}")
        import traceback; traceback.print_exc()
        return JSONResponse({"error": f"下載失敗：{str(e)}"}, status_code=500)


def _fill_excel_multi_sheet(template_path: str, sheet_cell_maps: dict) -> bytes:
    """
    支援多 sheet 填入。
    sheet_cell_maps: {"進件表格": {"B5": "值", ...}, "擔保品資訊": {"B2": "", ...}}
    空字串 "" = 清空儲存格；None = 不動原值；__FORMULA_RECALC__ = 清快取。
    """
    import zipfile
    import re as _re

    with open(template_path, "rb") as f:
        original_bytes = f.read()

    orig_zip = zipfile.ZipFile(io.BytesIO(original_bytes), 'r')
    try:
        # 讀 workbook.xml 取得 sheet name → xml path 對照
        wb_xml = orig_zip.read('xl/workbook.xml').decode('utf-8')
        rels_xml = orig_zip.read('xl/_rels/workbook.xml.rels').decode('utf-8')
        rel_map = dict(_re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', rels_xml))

        sheet_name_to_path = {}
        first_visible_path = None
        for m in _re.finditer(r'<sheet\b[^/]*?/>', wb_xml):
            block = m.group(0)
            nm = _re.search(r'name="([^"]+)"', block)
            rm = _re.search(r'r:id="([^"]+)"', block)
            sm = _re.search(r'state="([^"]+)"', block)
            if not (nm and rm):
                continue
            sname = nm.group(1)
            rid = rm.group(1)
            state = sm.group(1) if sm else "visible"
            target = rel_map.get(rid, "")
            if not target:
                continue
            path = 'xl/' + target.lstrip('/')
            sheet_name_to_path[sname] = path
            if state == "visible" and first_visible_path is None:
                first_visible_path = path

        # sharedStrings
        ss_xml_name = 'xl/sharedStrings.xml'
        if ss_xml_name not in orig_zip.namelist():
            # 沒有 sharedStrings 檔時，無法寫入（下拉選單 xlsx 應一定有）
            return original_bytes
        ss_xml = orig_zip.read(ss_xml_name).decode('utf-8')
        si_blocks = list(_re.finditer(r'(<si>)(.*?)(</si>)', ss_xml, _re.DOTALL))
        si_list = [m.group(0) for m in si_blocks]

        # 每個 sheet 計算要改的內容
        new_sheet_xmls = {}  # path -> new xml
        for sheet_name, cell_map in sheet_cell_maps.items():
            if not cell_map:
                continue
            path = sheet_name_to_path.get(sheet_name)
            if not path:
                # 若找不到，嘗試匹配常見舊名
                for alias in (sheet_name, ):
                    if alias in sheet_name_to_path:
                        path = sheet_name_to_path[alias]
                        break
                if not path:
                    continue

            sheet_xml = new_sheet_xmls.get(path) or orig_zip.read(path).decode('utf-8')

            # 解析該 sheet 的 cell → ss_idx 對照
            cell_to_ss_idx = {}
            for m in _re.finditer(r'<c\s+r="([A-Z]+\d+)"[^>]*\s+t="s"[^>]*>\s*<v>(\d+)</v>', sheet_xml):
                cell_to_ss_idx[m.group(1)] = int(m.group(2))

            ss_cell_changes = {}
            direct_changes = {}
            formula_recalc = []
            for cell_ref, new_value in cell_map.items():
                if new_value is None:
                    continue
                if new_value == "__FORMULA_RECALC__":
                    formula_recalc.append(cell_ref)
                    continue
                if not isinstance(new_value, str):
                    new_value = str(new_value)
                if cell_ref in cell_to_ss_idx:
                    ss_cell_changes[cell_ref] = new_value
                else:
                    direct_changes[cell_ref] = new_value

            # shared string 改動：新增 ss 並修改 sheet xml 引用
            for cell_ref, new_value in ss_cell_changes.items():
                escaped = new_value.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                new_idx = len(si_list)
                si_list.append(f'<si><t>{escaped}</t></si>')
                old_idx = cell_to_ss_idx[cell_ref]
                cell_pattern = _re.compile(
                    r'(<c\s+r="' + _re.escape(cell_ref) + r'"[^>]*>)\s*<v>' + str(old_idx) + r'</v>\s*(</c>)')
                mm = cell_pattern.search(sheet_xml)
                if mm:
                    sheet_xml = sheet_xml[:mm.start()] + mm.group(1) + f'<v>{new_idx}</v>' + mm.group(2) + sheet_xml[mm.end():]

            # direct value 改動
            def _is_numeric(s):
                if not s: return False
                try: float(s); return True
                except: return False
            for cell_ref, new_value in direct_changes.items():
                escaped_val = new_value.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;') if new_value else ""
                pattern1 = _re.compile(r'(<c\s+r="' + _re.escape(cell_ref) + r'"[^>]*>)\s*<v>[^<]*</v>\s*(</c>)')
                mm = pattern1.search(sheet_xml)
                if mm:
                    if new_value:
                        # 非數字值要用 shared string 寫入，避免 Excel 解析數字失敗
                        if not _is_numeric(new_value):
                            open_tag = mm.group(1)
                            # 移除舊的 t 屬性
                            open_tag_clean = _re.sub(r'\s+t="[^"]*"', '', open_tag)
                            # 在 > 前插入 t="s"
                            new_open = open_tag_clean[:-1] + ' t="s">'
                            new_idx = len(si_list)
                            si_list.append(f'<si><t>{escaped_val}</t></si>')
                            sheet_xml = sheet_xml[:mm.start()] + new_open + f'<v>{new_idx}</v>' + mm.group(2) + sheet_xml[mm.end():]
                        else:
                            sheet_xml = sheet_xml[:mm.start()] + mm.group(1) + f'<v>{escaped_val}</v>' + mm.group(2) + sheet_xml[mm.end():]
                    else:
                        attrs_match = _re.search(r'<c\s+r="' + _re.escape(cell_ref) + r'"([^>]*)>', mm.group(1))
                        if attrs_match:
                            attrs = _re.sub(r'\s+t="[^"]*"', '', attrs_match.group(1))
                            new_cell = f'<c r="{cell_ref}"{attrs}/>'
                            sheet_xml = sheet_xml[:mm.start()] + new_cell + sheet_xml[mm.end():]
                    continue
                # self-closing 空格：<c r="B7" s="77"/> → 用 shared string 寫入（不用 inlineStr 避免 Excel 修復提示）
                pattern2 = _re.compile(r'<c\s+r="' + _re.escape(cell_ref) + r'"([^/]*)/>')
                mm2 = pattern2.search(sheet_xml)
                if mm2 and new_value:
                    attrs = mm2.group(1).strip()
                    new_idx = len(si_list)
                    si_list.append(f'<si><t>{escaped_val}</t></si>')
                    new_cell = f'<c r="{cell_ref}" {attrs} t="s"><v>{new_idx}</v></c>'
                    sheet_xml = sheet_xml[:mm2.start()] + new_cell + sheet_xml[mm2.end():]

            # 公式快取清除
            for cell_ref in formula_recalc:
                pattern = _re.compile(
                    r'<c\s+r="' + _re.escape(cell_ref) + r'"([^>]*)>(.*?)<v>[^<]*</v>(.*?)</c>',
                    _re.DOTALL
                )
                mm = pattern.search(sheet_xml)
                if mm:
                    attrs = _re.sub(r'\s+t="[^"]*"', '', mm.group(1))
                    inner = mm.group(2) + mm.group(3)
                    new_cell = f'<c r="{cell_ref}"{attrs}>{inner}</c>'
                    sheet_xml = sheet_xml[:mm.start()] + new_cell + sheet_xml[mm.end():]

            # 若什麼都沒改，跳過
            if ss_cell_changes or direct_changes or formula_recalc:
                new_sheet_xmls[path] = sheet_xml

        # 重組 sharedStrings
        added = len(si_list) - len(si_blocks)
        if added > 0 and si_blocks:
            first_si = si_blocks[0].start()
            last_si_end = si_blocks[-1].end()
            new_si_content = ''.join(si_list)
            header_xml = ss_xml[:first_si]
            orig_count_m = _re.search(r'count="(\d+)"', header_xml)
            if orig_count_m:
                new_count = int(orig_count_m.group(1)) + added
                header_xml = _re.sub(r'count="\d+"', f'count="{new_count}"', header_xml)
            header_xml = _re.sub(r'uniqueCount="\d+"', f'uniqueCount="{len(si_list)}"', header_xml)
            new_ss_xml = header_xml + new_si_content + ss_xml[last_si_end:]
        else:
            new_ss_xml = ss_xml

        # 重新打包 ZIP
        output_buf = io.BytesIO()
        output_zip = zipfile.ZipFile(output_buf, 'w', zipfile.ZIP_DEFLATED)
        for item in orig_zip.infolist():
            if item.filename == ss_xml_name and added > 0:
                output_zip.writestr(item, new_ss_xml.encode('utf-8'))
            elif item.filename in new_sheet_xmls:
                output_zip.writestr(item, new_sheet_xmls[item.filename].encode('utf-8'))
            else:
                output_zip.writestr(item, orig_zip.read(item.filename))
        output_zip.close()
        output_buf.seek(0)
        return output_buf.getvalue()
    finally:
        orig_zip.close()


def build_sheet_cell_maps_from_mapping(plan_name: str, mapping: dict, r: dict) -> dict:
    """
    根據儲存的 mapping JSON 與客戶資料，產生每個 sheet 的 cell_map。
    mapping 格式: {sheet_name: {cell: field_key}}
    回傳: {sheet_name: {cell: 處理後的值}}
    """
    out = {}
    for sheet_name, cell_field_map in mapping.items():
        if not isinstance(cell_field_map, dict):
            continue
        sheet_cm = {}
        for cell_ref, field_key in cell_field_map.items():
            val = compute_field_value(field_key, r, plan_name)
            # None = 不動；否則寫入（含空字串 = 清空）
            if val is not None:
                sheet_cm[cell_ref] = val
        if sheet_cm:
            out[sheet_name] = sheet_cm
    return out


def _fill_excel_template(template_path: str, cell_map: dict) -> bytes:
    """
    Fill customer data into Excel template by modifying ONLY sharedStrings.xml.
    All other files (styles, formulas, calcChain, comments, VML, validations) stay untouched.

    cell_map: {"C4": "王小明", "E4": "A123456789", ...}
    Empty string values will clear the cell's text.
    """
    import zipfile
    import re as _re

    with open(template_path, "rb") as f:
        original_bytes = f.read()

    orig_zip = zipfile.ZipFile(io.BytesIO(original_bytes), 'r')
    try:
        return _fill_excel_inner(orig_zip, original_bytes, cell_map, _re)
    finally:
        orig_zip.close()


def _fill_excel_inner(orig_zip, original_bytes, cell_map, _re):
    import zipfile

    # Step 1: Find the FIRST visible sheet's XML file via workbook.xml + rels
    sheet_xml_name = None
    try:
        wb_xml = orig_zip.read('xl/workbook.xml').decode('utf-8')
        rels_xml = orig_zip.read('xl/_rels/workbook.xml.rels').decode('utf-8')
        # Get first sheet's r:id from workbook.xml
        first_sheet = _re.search(r'<sheet\s+name="[^"]*"\s+sheetId="\d+"[^>]*r:id="([^"]+)"', wb_xml)
        if first_sheet:
            rid = first_sheet.group(1)
            # Find the target file for this r:id in rels
            rel_match = _re.search(r'Id="' + _re.escape(rid) + r'"[^>]*Target="([^"]+)"', rels_xml)
            if rel_match:
                target = rel_match.group(1)
                sheet_xml_name = 'xl/' + target if not target.startswith('/') else target.lstrip('/')
    except Exception:
        pass
    # Fallback: if above fails, scan all sheets for the one that has our target cells
    if not sheet_xml_name or sheet_xml_name not in orig_zip.namelist():
        target_cells = set(cell_map.keys())
        for name in sorted(orig_zip.namelist()):
            if _re.match(r'xl/worksheets/sheet\d+\.xml$', name):
                xml_content = orig_zip.read(name).decode('utf-8')
                # Check if this sheet has any of our target cells
                found = sum(1 for c in target_cells if f'r="{c}"' in xml_content)
                if found > len(target_cells) // 2:  # More than half match
                    sheet_xml_name = name
                    break

    if not sheet_xml_name:
        return original_bytes

    sheet_xml = orig_zip.read(sheet_xml_name).decode('utf-8')

    # Parse cells: find <c r="XX" ... t="s"><v>INDEX</v></c>
    # Build map: cell_ref -> shared_string_index
    cell_to_ss_idx = {}
    for m in _re.finditer(r'<c\s+r="([A-Z]+\d+)"[^>]*\s+t="s"[^>]*>\s*<v>(\d+)</v>', sheet_xml):
        cell_ref = m.group(1)
        ss_idx = int(m.group(2))
        cell_to_ss_idx[cell_ref] = ss_idx

    # Step 2: Parse sharedStrings.xml
    ss_xml_name = 'xl/sharedStrings.xml'
    if ss_xml_name not in orig_zip.namelist():
        return original_bytes

    ss_xml = orig_zip.read(ss_xml_name).decode('utf-8')

    # Find all <si>...</si> blocks
    si_blocks = list(_re.finditer(r'(<si>)(.*?)(</si>)', ss_xml, _re.DOTALL))

    # Step 3: 分類 cell_map
    # 策略：不修改現有 shared string（避免破壞下拉選單），
    # 而是新增新的 shared string 並修改 sheet XML 中的引用索引
    ss_cell_changes = {}  # cell_ref -> new_value
    direct_changes = {}   # cell_ref -> new_value
    formula_recalc = []   # cells whose cached <v> should be cleared (force Excel recalc)
    for cell_ref, new_value in cell_map.items():
        if new_value is None:
            continue
        if new_value == "__FORMULA_RECALC__":
            formula_recalc.append(cell_ref)
            continue
        # 強制轉字串（int/float 也要能寫入）
        if not isinstance(new_value, str):
            new_value = str(new_value) if new_value != 0 else ("0" if new_value == 0 else "")
        if cell_ref in cell_to_ss_idx:
            ss_cell_changes[cell_ref] = new_value if new_value else ""
        else:
            direct_changes[cell_ref] = new_value if new_value else ""

    if not ss_cell_changes and not direct_changes and not formula_recalc:
        return original_bytes

    # Step 4: 新增 shared strings + 修改 sheet XML 引用
    if not si_blocks:
        # sharedStrings.xml 沒有 <si> 區塊，跳過 shared string 修改
        return original_bytes
    si_list = [m.group(0) for m in si_blocks]
    new_sheet_xml = sheet_xml
    for cell_ref, new_value in ss_cell_changes.items():
        escaped = new_value.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        # 新增一個 shared string
        new_idx = len(si_list)
        si_list.append(f'<si><t>{escaped}</t></si>')
        # 修改 sheet XML 中該 cell 的 <v> 指向新索引
        old_idx = cell_to_ss_idx[cell_ref]
        cell_pattern = _re.compile(
            r'(<c\s+r="' + _re.escape(cell_ref) + r'"[^>]*>)\s*<v>' + str(old_idx) + r'</v>\s*(</c>)')
        m = cell_pattern.search(new_sheet_xml)
        if m:
            new_sheet_xml = new_sheet_xml[:m.start()] + m.group(1) + f'<v>{new_idx}</v>' + m.group(2) + new_sheet_xml[m.end():]
    # Step 4b: Modify sheet XML for direct-value cells
    if not ss_cell_changes:
        new_sheet_xml = sheet_xml  # 只有在 Step 4 沒修改時才重新賦值
    def _is_numeric_inner(s):
        if not s: return False
        try: float(s); return True
        except: return False
    for cell_ref, new_value in direct_changes.items():
        escaped_val = new_value.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;') if new_value else ""
        # Case 1: 已有值 <c r="G18" s="48"><v>5.4</v></c>
        pattern1 = _re.compile(r'(<c\s+r="' + _re.escape(cell_ref) + r'"[^>]*>)\s*<v>[^<]*</v>\s*(</c>)')
        m = pattern1.search(new_sheet_xml)
        if m:
            if new_value:
                # 非數字值改用 shared string 避免 Excel 解析數字失敗
                if not _is_numeric_inner(new_value):
                    open_tag = m.group(1)
                    open_tag_clean = _re.sub(r'\s+t="[^"]*"', '', open_tag)
                    new_open = open_tag_clean[:-1] + ' t="s">'
                    new_idx = len(si_list)
                    si_list.append(f'<si><t>{escaped_val}</t></si>')
                    new_sheet_xml = new_sheet_xml[:m.start()] + new_open + f'<v>{new_idx}</v>' + m.group(2) + new_sheet_xml[m.end():]
                else:
                    new_sheet_xml = new_sheet_xml[:m.start()] + m.group(1) + f'<v>{escaped_val}</v>' + m.group(2) + new_sheet_xml[m.end():]
            else:
                # 清空：移除 <v> 標籤
                attrs_match = _re.search(r'<c\s+r="' + _re.escape(cell_ref) + r'"([^>]*)>', m.group(1))
                if attrs_match:
                    attrs = _re.sub(r'\s+t="[^"]*"', '', attrs_match.group(1))
                    new_cell = f'<c r="{cell_ref}"{attrs}/>'
                    new_sheet_xml = new_sheet_xml[:m.start()] + new_cell + new_sheet_xml[m.end():]
            continue
        # Case 2: 空的 self-closing <c r="B7" s="77"/> → 用 shared string 寫入（不用 inlineStr 避免 Excel 修復提示）
        pattern2 = _re.compile(r'<c\s+r="' + _re.escape(cell_ref) + r'"([^/]*)/>')
        m2 = pattern2.search(new_sheet_xml)
        if m2 and new_value:  # 只在有值時寫入
            attrs = m2.group(1).strip()
            new_idx = len(si_list)
            si_list.append(f'<si><t>{escaped_val}</t></si>')
            new_cell = f'<c r="{cell_ref}" {attrs} t="s"><v>{new_idx}</v></c>'
            new_sheet_xml = new_sheet_xml[:m2.start()] + new_cell + new_sheet_xml[m2.end():]

    # Step 4c: 清除公式儲存格的快取值（強制 Excel 重算）
    # <c r="C39" s="99" t="str"><f>C11</f><v>陳耀晨</v></c>
    # → <c r="C39" s="99"><f>C11</f></c>
    if formula_recalc:
        if not ss_cell_changes and not direct_changes:
            new_sheet_xml = sheet_xml
        for cell_ref in formula_recalc:
            # 匹配公式儲存格：含 <f>...</f> 和 <v>...</v>
            pattern = _re.compile(
                r'<c\s+r="' + _re.escape(cell_ref) + r'"([^>]*)>(.*?)<v>[^<]*</v>(.*?)</c>',
                _re.DOTALL
            )
            m = pattern.search(new_sheet_xml)
            if m:
                attrs = m.group(1)
                # 移除 t="..." 屬性（避免類型衝突）
                attrs = _re.sub(r'\s+t="[^"]*"', '', attrs)
                inner = m.group(2) + m.group(3)
                new_cell = f'<c r="{cell_ref}"{attrs}>{inner}</c>'
                new_sheet_xml = new_sheet_xml[:m.start()] + new_cell + new_sheet_xml[m.end():]

    # Step 4d: 更新 sharedStrings（在所有 cell 處理完後才重建，確保 direct_changes 新增的 si 也包含在內）
    first_si = si_blocks[0].start()
    last_si_end = si_blocks[-1].end()
    new_si_content = ''.join(si_list)
    new_unique = len(si_list)
    added = new_unique - len(si_blocks)
    header_xml = ss_xml[:first_si]
    orig_count_m = _re.search(r'count="(\d+)"', header_xml)
    if orig_count_m:
        new_count = int(orig_count_m.group(1)) + added
        header_xml = _re.sub(r'count="\d+"', f'count="{new_count}"', header_xml)
    header_xml = _re.sub(r'uniqueCount="\d+"', f'uniqueCount="{new_unique}"', header_xml)
    new_ss_xml = header_xml + new_si_content + ss_xml[last_si_end:]

    # Step 5: Repackage ZIP
    output_buf = io.BytesIO()
    output_zip = zipfile.ZipFile(output_buf, 'w', zipfile.ZIP_DEFLATED)

    for item in orig_zip.infolist():
        if item.filename == ss_xml_name:
            output_zip.writestr(item, new_ss_xml.encode('utf-8'))
        elif item.filename == sheet_xml_name and (ss_cell_changes or direct_changes or formula_recalc):
            output_zip.writestr(item, new_sheet_xml.encode('utf-8'))
        else:
            output_zip.writestr(item, orig_zip.read(item.filename))

    output_zip.close()
    output_buf.seek(0)
    return output_buf.getvalue()


def _do_download_excel(request: Request, case_id: str):
    from fastapi.responses import StreamingResponse
    role = check_auth(request)
    if not role or role not in ("admin", "adminB"):
        return JSONResponse({"error": "無權限"}, status_code=403)
    if not case_id:
        return JSONResponse({"error": "缺少 case_id"}, status_code=400)

    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT * FROM customers WHERE case_id=?", (case_id,))
    row = cur.fetchone(); conn.close()
    if not row:
        return JSONResponse({"error": "找不到客戶"}, status_code=404)

    r = dict(row)
    plans = (r.get("adminb_selected_plans", "") or "").split(",")
    plans = [p.strip() for p in plans if p.strip()]

    if not plans:
        return JSONResponse({"error": "尚未勾選任何方案，請先勾選並儲存"}, status_code=400)

    _base = os.path.dirname(os.path.abspath(__file__))
    PLAN_TEMPLATE_MAP = {
        "亞太商品": os.path.join(_base, "申請書", "亞太商品範本.xlsx"),
        "亞太機車15萬": os.path.join(_base, "申請書", "亞太15萬機車範本.xlsx"),
        "亞太機車25萬": os.path.join(_base, "申請書", "亞太15萬機車範本.xlsx"),
        "亞太工會機車": os.path.join(_base, "申請書", "亞太工會範本.xlsx"),
        "和裕機車": os.path.join(_base, "申請書", "和裕維力貸機車範本).xlsx"),
        "和裕商品": os.path.join(_base, "申請書", "和裕維力貸商品範本.xlsx"),
        "第一": os.path.join(_base, "申請書", "第一申請書範本.xlsx"),
        "貸救補": os.path.join(_base, "申請書", "貸就補範本.xlsx"),
        "21機車12萬": os.path.join(_base, "申請書", "21機車申請書範本xlsx.xlsx"),
        "21機車25萬": os.path.join(_base, "申請書", "21機25萬範本.xlsx"),
        "21商品": os.path.join(_base, "申請書", "21商品範本.xlsx"),
        # TXT 填寫表方案
        "麻吉機車": os.path.join(_base, "申請書", "麻吉機車申請書.txt"),
        "麻吉手機": os.path.join(_base, "申請書", "麻吉手機申請書.txt"),
        "手機分期": os.path.join(_base, "申請書", "手機分期.txt"),
        "分貝機車": os.path.join(_base, "申請書", "分貝機車.txt"),
        "分貝汽車": os.path.join(_base, "申請書", "分貝汽車.txt"),
        "21汽車":   os.path.join(_base, "申請書", "21汽車申請書.txt"),
    }

    # 自訂範本覆蓋：若使用者在 /admin/templates 上傳過該方案的範本，優先使用
    for _plan_name in list(PLAN_TEMPLATE_MAP.keys()):
        if not PLAN_TEMPLATE_MAP[_plan_name].endswith(".xlsx"):
            continue
        _custom = get_custom_template_path(_plan_name)
        if os.path.isfile(_custom):
            PLAN_TEMPLATE_MAP[_plan_name] = _custom

    files_to_zip = []
    name = r.get("customer_name", "")

    # Build cell data mapping for each plan type
    def _build_cell_map(plan_name, r):
        """Build {cell_ref: value} map. Empty string = clear cell."""
        CITY_TO_CODE = {
            "台北市": "北市", "新北市": "新北市", "桃園市": "桃市", "台中市": "中市",
            "台南市": "南市", "高雄市": "高市", "基隆市": "基市", "新竹市": "竹市",
            "新竹縣": "竹縣", "苗栗縣": "苗縣", "彰化縣": "彰縣", "南投縣": "投縣",
            "雲林縣": "雲縣", "嘉義市": "嘉市", "嘉義縣": "嘉縣", "屏東縣": "屏縣",
            "宜蘭縣": "宜縣", "花蓮縣": "花縣", "台東縣": "東縣", "澎湖縣": "澎縣",
            "金門縣": "金門", "連江縣": "連江",
        }

        def v(key): return (r.get(key, "") or "").strip()

        name = v("customer_name")
        id_no = v("id_no")
        birth = v("birth_date")
        phone = v("phone")
        email = v("email")
        line_id = v("line_id")
        marriage = v("marriage")
        education = v("education")
        id_date = v("id_issue_date")
        id_place = v("id_issue_place")
        id_type = v("id_issue_type")
        reg_addr = v("reg_city") + v("reg_district") + v("reg_address")
        live_same = v("live_same_as_reg") == "1"
        live_addr = reg_addr if live_same else (v("live_city") + v("live_district") + v("live_address"))
        live_status = v("live_status")
        live_years = v("live_years")
        company = v("company_name_detail") or v("company")
        co_phone_area = v("company_phone_area")
        co_phone_num = v("company_phone_num")
        # 手機（mobile）→ 區碼留空，只用號碼
        if co_phone_area == "mobile":
            co_phone_area = ""
        co_phone = (co_phone_area + "-" + co_phone_num) if co_phone_area and co_phone_num else co_phone_num
        co_role = v("company_role")
        co_years = v("company_years")
        co_salary = v("company_salary")
        co_addr = v("company_city") + v("company_district") + v("company_address")
        c1_name = v("contact1_name")
        c1_rel = v("contact1_relation")
        c1_phone = v("contact1_phone")
        c1_known = v("contact1_known")
        c2_name = v("contact2_name")
        c2_rel = v("contact2_relation")
        c2_phone = v("contact2_phone")
        c2_known = v("contact2_known")

        if plan_name == "貸救補":
            # F5 發證地下拉
            valid_lj_place = ["北市","新北市","北縣","基市","宜縣","桃市","桃縣","竹市","竹縣",
                              "苗縣","中市","中縣","嘉市","彰縣","投縣","雲縣","嘉縣","南市",
                              "南縣","高市","高縣","屏縣","東縣","花縣","澎縣","連江","金門"]
            lj_place_map = {"台北市":"北市","桃園市":"桃市","台中市":"中市","台南市":"南市",
                            "高雄市":"高市","基隆市":"基市","新竹市":"竹市","新竹縣":"竹縣",
                            "苗栗縣":"苗縣","彰化縣":"彰縣","南投縣":"投縣","雲林縣":"雲縣",
                            "嘉義市":"嘉市","嘉義縣":"嘉縣","屏東縣":"屏縣","宜蘭縣":"宜縣",
                            "花蓮縣":"花縣","台東縣":"東縣","澎湖縣":"澎縣","金門縣":"金門",
                            "連江縣":"連江","新北市":"新北市"}
            lj_place = lj_place_map.get(id_place, id_place)
            lj_place = lj_place if lj_place in valid_lj_place else ""

            # G5 換補發
            id_type_val = id_type if id_type in ("初發", "換發", "補發") else ""

            # 日期民國 7 位（無斜線）
            def to_roc7(d):
                if not d: return ""
                parts = d.replace("-", "/").split("/")
                if len(parts) != 3: return ""
                try:
                    y = int(parts[0])
                    if y >= 1911:
                        y -= 1911
                    return f"{str(y).zfill(3)}{parts[1].zfill(2)}{parts[2].zfill(2)}"
                except Exception:
                    return ""
            birth_roc7 = to_roc7(birth)
            id_date_roc7 = to_roc7(id_date)

            # 現住電話：優先 live_phone，無則用 reg_phone
            live_phone_lj = v("live_phone") or v("reg_phone")

            # 月薪 E11 純數字（4萬→40000）
            try:
                sal_c = co_salary.replace("萬","").replace(",","").strip() if co_salary else ""
                sn = float(sal_c) if sal_c else 0
                if 0 < sn < 1000:
                    e11_val = int(sn * 10000)
                elif sn >= 1000:
                    e11_val = int(sn)
                else:
                    e11_val = ""
            except Exception:
                e11_val = ""

            # 年資 G11 年數 / I11 月數（純數字）
            try:
                g11_val = int(float(co_years)) if co_years else 0
                co_mos_lj = v("company_months") or "0"
                i11_val = int(float(co_mos_lj)) if co_mos_lj and co_mos_lj != "0" else 0
            except Exception:
                g11_val = 0
                i11_val = 0

            return {
                "C4": name, "E4": id_no, "J4": phone,
                "C5": birth_roc7, "E5": id_date_roc7,
                "F5": lj_place, "G5": id_type_val,
                "J5": live_phone_lj,
                "C7": reg_addr,
                "C8": reg_addr if live_same else live_addr,
                "C9": co_addr,
                "C10": company, "G10": co_phone_area + co_phone_num,
                "J10": v("company_phone_ext"),
                "C11": co_role, "E11": e11_val,
                "G11": g11_val, "I11": i11_val,
                "C18": c1_name, "E18": c1_rel, "H18": c1_phone,
                "C19": c2_name, "E19": c2_rel, "H19": c2_phone,
                # 商品名稱/型號（J7/J8）從貸就補補充資料 lj_pname/lj_pmodel
                "J7": v("adminb_product_name"),
                "J8": v("adminb_product_model"),
            }

        elif plan_name in ("和裕機車", "和裕商品"):
            # === 發證地（F13）必須匹配下拉 ===
            id_place_code = CITY_TO_CODE.get(id_place, id_place)

            # === 婚姻（C13）===
            marriage_val = marriage if marriage in ("已婚", "未婚") else ""

            # === 學歷（C14）必須匹配下拉 ===
            valid_edu_hr = ["高中職", "專科、大學", "研究所以上", "其他"]
            edu_map = {"高中/職": "高中職", "高中": "高中職", "高職": "高中職",
                       "專科/大學": "專科、大學", "大學": "專科、大學", "專科": "專科、大學",
                       "研究所以上": "研究所以上", "研究所": "研究所以上"}
            edu_val = edu_map.get(education, education)
            edu_val = edu_val if edu_val in valid_edu_hr else ""

            # === 公司電話（H17）格式：02-27189890 或 0989-615422 ===
            if co_phone_area and co_phone_num:
                co_phone_fmt = co_phone_area + "-" + co_phone_num
            elif co_phone_num:
                co_phone_fmt = co_phone_num
            else:
                co_phone_fmt = ""

            # === 年資（I18）格式：N年M月（年月字保留，缺少數字以空白代替）===
            co_mos = v("company_months") or "0"
            try:
                yr_int = int(float(co_years)) if co_years else 0
                mo_int = int(float(co_mos)) if co_mos and co_mos != "0" else 0
                yr_str = str(yr_int) if yr_int > 0 else ""
                mo_str = str(mo_int) if mo_int > 0 else ""
                if yr_str or mo_str:
                    years_fmt = f"{yr_str}年{mo_str}月"
                else:
                    years_fmt = ""
            except Exception:
                years_fmt = co_years

            # === 月薪（I19）格式：整數顯示整數（5萬），小數顯示小數（4.5萬）===
            try:
                sal_raw = float(co_salary) if co_salary else 0
                if sal_raw >= 1000:
                    sal_wan = round(sal_raw / 10000, 1)
                else:
                    sal_wan = sal_raw
                if sal_wan:
                    if sal_wan == int(sal_wan):
                        sal_fmt = f"{int(sal_wan)}萬"  # 5萬（非 5.0萬）
                    else:
                        sal_fmt = f"{sal_wan}萬"        # 4.5萬
                else:
                    sal_fmt = ""
            except Exception:
                sal_fmt = co_salary

            # === 行業（G19）必須匹配下拉，從 adminB ===
            valid_hr_ind = ["服務業","餐飲業","科技業","軍人","運輸業","倉儲業","金融業","製造業",
                           "營造業","電商網拍業","農狩林牧業","礦業","漁業","證券期貨業","保險業",
                           "不動產業","公教人員","水電燃氣業","通信業","社團個人服務","其它"]
            # 優先用和裕專屬欄位 adminb_hr_industry，fallback 到 adminb_industry（亞太用）並轉換
            hr_industry = v("adminb_hr_industry") or v("adminb_industry") or ""
            ind_map = {"製造業":"製造業","餐飲與服務業":"服務業","建築與營造":"營造業",
                       "軍警與公教":"公教人員","科技與資訊":"科技業","運輸與物流":"運輸業",
                       "金融與保險業":"金融業","批發與零售業":"服務業","醫療與教育":"服務業",
                       "農林漁牧業":"農狩林牧業","自由職業":"其它"}
            hr_ind_val = ind_map.get(hr_industry, hr_industry)
            hr_ind_val = hr_ind_val if hr_ind_val in valid_hr_ind else ""  # 無值清空

            # === 手機電信（C20）：不在下拉清單就選「其他」 ===
            carrier_raw = v("carrier") or ""
            valid_carriers = ["中華電信", "遠傳電信", "台灣大哥大", "其他"]
            if carrier_raw in valid_carriers:
                carrier_val = carrier_raw
            elif carrier_raw:
                carrier_val = "其他"
            else:
                carrier_val = ""

            # === 聯絡人關係（C24/F24）：和裕沒有下拉選單，直接填原始值 ===
            c1_rel_hr = c1_rel or ""
            c2_rel_hr = c2_rel or ""

            # === 聯絡人知情（D22/G22）智能判別 ===
            valid_known = ["知情", "保密"]
            def map_known(raw):
                if not raw: return ""
                if raw in valid_known: return raw
                if "知情" in raw or "可知" in raw: return "知情"
                if "保密" in raw or "不知" in raw or "無可" in raw: return "保密"
                return ""
            c1_known_val = map_known(c1_known)
            c2_known_val = map_known(c2_known)

            # === 聯絡人電話格式：0955-389338 ===
            def fmt_phone(p):
                p = (p or "").replace("-", "").replace(" ", "")
                if len(p) == 10 and p.startswith("09"):
                    return p[:4] + "-" + p[4:]
                return p
            c1_ph_fmt = fmt_phone(c1_phone)
            c2_ph_fmt = fmt_phone(c2_phone)

            # === 行動電話（F14）也要格式化 ===
            phone_fmt = fmt_phone(phone)

            result = {
                "C11": name,
                "C39": "__FORMULA_RECALC__",  # 戶名是公式 =C11，清快取值強制重算
                "F11": id_no,
                "C12": birth, "F12": (id_date + " " + id_type) if id_date else "",
                "C13": marriage_val, "F13": id_place_code,
                "C14": edu_val, "F14": phone_fmt,
                "C15": reg_addr, "G15": v("reg_phone") or "",
                "C16": live_addr, "G16": v("live_phone") or "", "H16": "同戶籍" if live_same else "",
                "C17": line_id, "F17": "家用",  # 資金用途固定家用
                "H17": co_phone_fmt,
                "C18": company, "G18": co_role, "I18": years_fmt,
                "C19": co_addr, "I19": sal_fmt,
                "C23": c1_name, "F23": c2_name,
                "C24": c1_rel_hr, "F24": c2_rel_hr,
                "C25": c1_ph_fmt, "F25": c2_ph_fmt,
                "D22": c1_known_val, "G22": c2_known_val,
                # 撥款資訊（無填寫則清空，戶名 C39 不動有公式）
                "C37": v("adminb_bank") or "",
                "C38": v("adminb_branch") or "",
                # 商品資訊（無填寫則清空）
                "C42": v("adminb_product") or "",
                "F42": v("adminb_model") or "",
                # 行業/電信
                "G19": hr_ind_val,
                "C20": carrier_val,
            }
            return result

        elif plan_name in ("亞太商品", "亞太機車15萬", "亞太工會機車", "亞太機車25萬"):
            # === 日期轉換：民國→西元 ===
            def roc_to_ad(date_str):
                """民國日期轉西元：086/12/15 → 1989/12/15"""
                if not date_str:
                    return ""
                parts = date_str.replace("-", "/").split("/")
                if len(parts) == 3:
                    try:
                        y = int(parts[0])
                        if y < 200:  # 是民國年
                            y += 1911
                        return f"{y}/{parts[1].zfill(2)}/{parts[2].zfill(2)}"
                    except Exception:
                        pass
                return date_str  # 已是西元或無法轉換

            birth_ad = roc_to_ad(birth)
            id_date_ad = roc_to_ad(id_date)

            # === 公司電話區碼（B18）：手機時選 0 ===
            if co_phone_area == "mobile" or (co_phone_num and co_phone_num.startswith("09")):
                co_phone_area = "0"

            # === 資金用途（B5）從 adminB 補充資料 ===
            fund_use = v("adminb_fund_use")
            valid_funds = ["I-1教育費","I-2醫藥費","I-3出國旅遊","I-4創業","II-1購買交通工具","II-2購買手機","II-3購買3C產品","III-1交友","III-2健身&醫美","III-3美容課程","IV-1個人理財投資(含不動產、裝修、理財商品)","V-1生活周轉金","V-2整合負債(償還銀行/融資等)"]
            fund_val = fund_use if fund_use in valid_funds else ""

            # === 婚姻（B10）===
            marriage_val = marriage if marriage in ("未婚", "已婚") else ""

            # === 教育程度（D10）必須完全匹配下拉 ===
            valid_edu_at = ["小學/國中", "高中/職", "專科/大學", "研究所以上"]
            if education in valid_edu_at:
                edu_val = education
            else:
                # 轉換常見格式
                edu_at_map = {"國中": "小學/國中", "國小": "小學/國中", "小學": "小學/國中",
                              "高中": "高中/職", "高職": "高中/職", "高中職": "高中/職",
                              "大學": "專科/大學", "專科": "專科/大學", "專科、大學": "專科/大學",
                              "研究所": "研究所以上", "碩士": "研究所以上", "博士": "研究所以上",
                              "其他": "小學/國中"}
                edu_val = edu_at_map.get(education, "")

            # === 發證狀態（F11）===
            id_type_val = id_type if id_type in ("初發", "補發", "換發") else ""

            # === 居住狀況（B14）：智能判別對照 ===
            valid_live = ("自有", "配偶", "親屬", "租屋", "宿舍")
            def map_live_status(raw):
                if not raw: return ""
                if raw in valid_live: return raw
                # 自有相關
                for k in ["自有","本人","名下","自宅","自有房屋"]:
                    if k in raw: return "自有"
                # 配偶相關
                for k in ["配偶","老公","老婆","太太","先生","夫","妻"]:
                    if k in raw: return "配偶"
                # 親屬（含父母/兄弟姊妹/家人/父母名下/租屋/宿舍）
                for k in ["父母","媽媽","爸爸","母親","父親","親屬","親戚","家人","兄","弟","姊","妹","祖","外公","外婆","租","宿舍","借住","寄住"]:
                    if k in raw: return "親屬"
                return ""
            live_status_val = map_live_status(live_status)

            # === 行業（E17）從 adminB，無則不填（D17 是標籤「行業類別」不能動） ===
            industry = v("adminb_industry")
            valid_industries = ["餐飲與服務業","製造業","建築與營造","軍警與公教","科技與資訊","運輸與物流","金融與保險業","批發與零售業","醫療與教育","農林漁牧業","自由職業","其他"]
            industry_val = industry if industry in valid_industries else None  # None = 不動原值

            # === 職務（G17）從 adminB，無選填則清空 ===
            valid_roles = ["行政與內勤","勞力與現場","銷售與業務","財務與專業","技術與工程","教學與醫護","管理與經營","自營與自由"]
            at_role = v("adminb_role")
            role_val = at_role if at_role in valid_roles else ""  # 不在下拉清單中則清空

            # === 聯絡人關係（D21）：常用詞轉換 ===
            valid_rels = ["父母","配偶","子女","兄姊","弟妹","祖父母","旁系血親","姻親","朋友","其他"]
            def map_relation(raw):
                if not raw: return ""
                if raw in valid_rels: return raw
                # 配偶相關
                for k in ["夫妻","妻","夫","老婆","老公","太太","先生","配偶"]:
                    if k in raw: return "配偶"
                # 父母相關
                for k in ["媽媽","爸爸","母親","父親","媽","爸","母","父"]:
                    if k in raw: return "父母"
                # 子女
                for k in ["兒子","女兒","兒","女","子女"]:
                    if k in raw: return "子女"
                # 兄姊
                for k in ["哥哥","姊姊","姐姐","哥","姊","姐","兄"]:
                    if k in raw: return "兄姊"
                # 弟妹
                for k in ["弟弟","妹妹","弟","妹"]:
                    if k in raw: return "弟妹"
                # 祖父母
                for k in ["祖父","祖母","爺爺","奶奶","外公","外婆","祖父母"]:
                    if k in raw: return "祖父母"
                # 旁系血親
                for k in ["叔","伯","姑","舅","姨","堂","表","侄","甥","旁系"]:
                    if k in raw: return "旁系血親"
                # 姻親
                for k in ["公公","婆婆","岳父","岳母","媳","婿","姻"]:
                    if k in raw: return "姻親"
                # 朋友
                for k in ["朋友","友","同事","同學"]:
                    if k in raw: return "朋友"
                return ""
            c1_rel_val = map_relation(c1_rel)

            # === 公司電話區碼（B18）===
            valid_areas = ["0","02","03","037","04","049","05","06","07","08","089","082","083"]
            co_area = co_phone_area if co_phone_area in valid_areas else ""

            # === 年資（G18）：5年6個月→5.6 ===
            try:
                yrs = float(co_years) if co_years else 0
                mos_raw = v("company_months") or "0"
                mos = float(mos_raw) if mos_raw and mos_raw != "0" else 0
                if mos > 0:
                    years_decimal = str(int(yrs)) + "." + str(int(mos))
                else:
                    years_decimal = str(int(yrs)) if yrs == int(yrs) else str(yrs)
            except Exception:
                years_decimal = co_years

            # === 月薪（H18）：58000→5.8萬 ===
            try:
                sal_raw = float(co_salary) if co_salary else 0
                if sal_raw >= 1000:  # 如果是元，轉換為萬
                    sal_wan = round(sal_raw / 10000, 1)
                else:  # 已經是萬
                    sal_wan = sal_raw
                salary_str = str(sal_wan) if sal_wan else ""
            except Exception:
                salary_str = co_salary

            # === 城市：B12/B13/B19 用完整名稱（桃園市非桃市）===
            reg_city_full = v("reg_city")  # 直接用完整名稱
            live_city_full = reg_city_full if live_same else v("live_city")
            co_city_full = v("company_city")

            # === 發證地（D11）用短碼 ===
            id_place_code = CITY_TO_CODE.get(id_place, id_place)

            result = {
                "B5": fund_val if fund_val else None,  # None = 不動
                "B9": name, "D9": id_no, "F9": birth_ad,
                "B10": marriage_val, "D10": edu_val,
                "B11": id_date_ad, "D11": id_place_code, "F11": id_type_val,
                "B12": reg_city_full, "C12": v("reg_district"), "D12": v("reg_address"),
                "B13": live_city_full,
                "C13": v("live_district") if not live_same else v("reg_district"),
                "D13": v("live_address") if not live_same else v("reg_address"),
                "B14": live_status_val, "D14": live_years,
                "B15": phone,
                "B16": email,
                "B17": company,
                "B18": co_area, "C18": co_phone_num, "E18": v("company_phone_ext"),
                "G18": years_decimal, "H18": salary_str,
                "B19": co_city_full, "C19": v("company_district"), "D19": v("company_address"),
                "B21": c1_name, "D21": c1_rel_val,
                "B25": c1_phone,
            }
            # 行業/職務：有值填入，無值清空
            result["E17"] = industry_val if industry_val else ""
            result["G17"] = role_val if role_val else ""
            # 車輛資料：只有機車範本才填，亞太商品不填寫車輛欄位
            if plan_name in ("亞太機車15萬", "亞太工會機車", "亞太機車25萬"):
                result["B7"] = v("adminb_vehicle_type") or ""   # 車輛型式
                result["D7"] = v("adminb_engine_no") or ""      # 引擎號碼
                result["F7"] = v("adminb_displacement") or ""   # 排氣量
                result["H7"] = v("adminb_color") or ""          # 顏色
                result["K2"] = v("adminb_brand") or ""          # 廠牌
                result["K3"] = v("vehicle_plate") or ""         # 牌照號碼
            return result

        elif plan_name == "第一":
            # 換發類型（J6）
            id_type_val = id_type if id_type in ("初發", "換發", "補發") else ""

            # 換證地點（G6）下拉匹配
            valid_dy1_place = ["北市","新北市","北縣","基市","宜縣","桃市","桃縣","竹市","竹縣",
                               "苗縣","中市","中縣","嘉市","彰縣","投縣","雲縣","嘉縣","南市",
                               "南縣","高市","高縣","屏縣","東縣","花縣","澎縣","連江","金門"]
            dy1_place_map = {"台北市":"北市","桃園市":"桃市","台中市":"中市","台南市":"南市",
                             "高雄市":"高市","基隆市":"基市","新竹市":"竹市","新竹縣":"竹縣",
                             "苗栗縣":"苗縣","彰化縣":"彰縣","南投縣":"投縣","雲林縣":"雲縣",
                             "嘉義市":"嘉市","嘉義縣":"嘉縣","屏東縣":"屏縣","宜蘭縣":"宜縣",
                             "花蓮縣":"花縣","台東縣":"東縣","澎湖縣":"澎縣","金門縣":"金門",
                             "連江縣":"連江","新北市":"新北市"}
            raw_place = id_place
            dy1_place = dy1_place_map.get(raw_place, raw_place)
            dy1_place = dy1_place if dy1_place in valid_dy1_place else ""

            # 日期轉民國無斜線 7 位 (113/12/24 → 1131224, 87/02/24 → 0870224)
            def to_roc_7digit(d):
                if not d: return ""
                parts = d.replace("-", "/").split("/")
                if len(parts) != 3: return ""
                try:
                    y = int(parts[0])
                    if y >= 1911:
                        y -= 1911
                    return f"{str(y).zfill(3)}{parts[1].zfill(2)}{parts[2].zfill(2)}"
                except Exception:
                    return ""
            birth_roc = to_roc_7digit(birth)
            id_date_roc = to_roc_7digit(id_date)

            # 公司電話：區碼+號碼+分機 (0223570707#722865)
            co_ext = v("company_phone_ext")
            if co_phone_area and co_phone_num:
                dy1_co_phone = co_phone_area + co_phone_num
            else:
                dy1_co_phone = co_phone_num
            if dy1_co_phone and co_ext:
                dy1_co_phone += "#" + co_ext

            # 年資：M8 年數、O8 月數（純數字）
            try:
                m8_val = int(float(co_years)) if co_years else 0
                co_mos_dy1 = v("company_months") or "0"
                o8_val = int(float(co_mos_dy1)) if co_mos_dy1 and co_mos_dy1 != "0" else 0
            except Exception:
                m8_val = 0
                o8_val = 0

            # 月薪：M9 純數字（4.5萬→45000）
            try:
                sal_c = co_salary.replace("萬","").replace(",","").strip() if co_salary else ""
                sn = float(sal_c) if sal_c else 0
                if 0 < sn < 1000:
                    m9_val = str(int(sn * 10000))
                elif sn >= 1000:
                    m9_val = str(int(sn))
                else:
                    m9_val = ""
            except Exception:
                m9_val = ""

            # 關係智能判別（第一下拉清單）
            valid_dy1_rels = ["父母","夫妻","兄弟姐妹","子女","朋友","同事","祖父母","外祖父母",
                              "孫子女","姪子女","岳父母","女婿","堂兄弟姊","表兄弟姊","伯父母",
                              "叔/嬸","舅/舅媽","姨/姨丈","姑/姑丈","公婆媳","大伯小叔","姑嫂","妯娌"]
            def map_dy1_rel(raw):
                if not raw: return ""
                if raw in valid_dy1_rels: return raw
                # 夫妻
                for k in ["夫妻","配偶","老公","老婆","太太","先生","夫","妻"]:
                    if k in raw: return "夫妻"
                # 父母
                for k in ["媽媽","爸爸","母親","父親","媽","爸","母","父"]:
                    if k in raw: return "父母"
                # 子女
                for k in ["兒子","女兒","兒","女","子女"]:
                    if k in raw: return "子女"
                # 兄弟姐妹
                for k in ["哥哥","姊姊","姐姐","弟弟","妹妹","哥","姊","姐","兄","弟","妹"]:
                    if k in raw: return "兄弟姐妹"
                # 外祖父母
                for k in ["外公","外婆","外祖"]:
                    if k in raw: return "外祖父母"
                # 祖父母
                for k in ["祖父","祖母","爺爺","奶奶"]:
                    if k in raw: return "祖父母"
                # 岳父母
                for k in ["岳父","岳母"]:
                    if k in raw: return "岳父母"
                # 公婆媳
                for k in ["公公","婆婆","媳"]:
                    if k in raw: return "公婆媳"
                # 朋友/同事
                if "同事" in raw: return "同事"
                if "朋友" in raw or "友" in raw or "同學" in raw: return "朋友"
                return ""
            c1_rel_dy1 = map_dy1_rel(c1_rel)
            c2_rel_dy1 = map_dy1_rel(c2_rel)

            return {
                "B5": name, "G5": id_no,
                "B6": id_date_roc, "G6": dy1_place, "J6": id_type_val,
                "B7": birth_roc, "G7": phone,
                "B8": v("reg_phone"), "G8": v("live_phone"),
                "B9": reg_addr,
                "B10": "同上" if live_same else live_addr,
                "M5": company,
                "T6": dy1_co_phone,
                "M7": co_addr,
                "M8": m8_val, "O8": o8_val,
                "M9": m9_val,
                "M13": c1_name, "Q13": c1_rel_dy1, "T13": c1_phone,
                "M14": c2_name, "Q14": c2_rel_dy1, "T14": c2_phone,
            }

        elif plan_name in ("21機車12萬", "21機車25萬", "21商品"):
            # 21 發證地短碼
            CITY_TO_21CODE = {
                "台北市": "北市", "新北市": "北縣", "桃園市": "桃縣", "台中市": "中縣",
                "台南市": "南縣", "高雄市": "高縣", "基隆市": "基市", "新竹市": "竹市",
                "新竹縣": "竹縣", "苗栗縣": "苗縣", "彰化縣": "彰縣", "南投縣": "投縣",
                "雲林縣": "雲縣", "嘉義市": "嘉市", "嘉義縣": "嘉縣", "屏東縣": "屏縣",
                "宜蘭縣": "宜縣", "花蓮縣": "花縣", "台東縣": "東縣", "澎湖縣": "澎縣",
                "金門縣": "金門", "連江縣": "連江",
            }
            id_place_code = CITY_TO_21CODE.get(id_place, id_place) if id_place else ""
            # 補換發（F4 用短碼，實際是 G4）：初發/補發/換發
            id_type_val = id_type if id_type in ("初發", "補發", "換發") else ""
            # 21 地址城市用全名且「台」改「臺」
            def city_to_21full(c):
                if not c: return ""
                return c.replace("台", "臺")
            reg_city_21 = city_to_21full(v("reg_city"))
            live_city_21 = reg_city_21 if live_same else city_to_21full(v("live_city"))

            # 從範本載入 D 欄下拉選項（郵遞區號+鄉鎮區）做智能匹配
            def match_21_district(city_full, district):
                """把區名匹配到「郵遞區號  區名」格式"""
                if not district:
                    return ""
                try:
                    import openpyxl as _opx
                    import warnings as _w
                    _w.filterwarnings("ignore")
                    _tpl = PLAN_TEMPLATE_MAP.get(plan_name)
                    if not _tpl or not os.path.exists(_tpl):
                        return district
                    _wb = _opx.load_workbook(_tpl, data_only=True)
                    if '工作表3' not in _wb.sheetnames:
                        return district
                    _ws = _wb['工作表3']
                    # 找城市對應的欄位
                    city_col = None
                    for c_idx in range(1, 30):
                        col_letter = _opx.utils.get_column_letter(c_idx)
                        if _ws[f'{col_letter}1'].value == city_full:
                            city_col = col_letter
                            break
                    if not city_col:
                        return district
                    # 遍歷該欄尋找匹配
                    for r_idx in range(2, 40):
                        v_cell = _ws[f'{city_col}{r_idx}'].value
                        if v_cell and district in str(v_cell):
                            return str(v_cell)
                    return district
                except Exception:
                    return district

            reg_district_21 = match_21_district(reg_city_21, v("reg_district"))
            live_district_21 = reg_district_21 if live_same else match_21_district(live_city_21, v("live_district"))

            # 日期轉民國格式（帶斜線）
            def to_roc_slash(d):
                if not d: return ""
                parts = d.replace("-", "/").split("/")
                if len(parts) != 3: return d
                try:
                    y = int(parts[0])
                    if y >= 1911:  # 西元
                        y -= 1911
                    return f"{str(y).zfill(3)}/{parts[1].zfill(2)}/{parts[2].zfill(2)}"
                except Exception:
                    return d
            birth_roc = to_roc_slash(birth)
            id_date_roc = to_roc_slash(id_date)

            # 關係智能判別（21 下拉清單）
            valid_21_rels = ["配偶","父母","子女","兄弟姐妹","祖父母","外祖父母","孫子女","外孫子女","配偶之父母","配偶之兄弟姐妹","其他親屬","負責人"]
            def map_21_rel(raw):
                if not raw: return ""
                if raw in valid_21_rels: return raw
                # 配偶
                for k in ["夫妻","配偶","老公","老婆","太太","先生","夫","妻"]:
                    if k in raw: return "配偶"
                # 父母
                for k in ["媽媽","爸爸","母親","父親","媽","爸","母","父"]:
                    if k in raw: return "父母"
                # 子女
                for k in ["兒子","女兒","兒","女","子女"]:
                    if k in raw: return "子女"
                # 兄弟姐妹
                for k in ["哥哥","姊姊","姐姐","弟弟","妹妹","哥","姊","姐","兄","弟","妹"]:
                    if k in raw: return "兄弟姐妹"
                # 外祖父母
                for k in ["外公","外婆","外祖"]:
                    if k in raw: return "外祖父母"
                # 祖父母
                for k in ["祖父","祖母","爺爺","奶奶"]:
                    if k in raw: return "祖父母"
                # 配偶之父母
                for k in ["公公","婆婆","岳父","岳母"]:
                    if k in raw: return "配偶之父母"
                # 其他親屬（含朋友/同事/同學）
                for k in ["朋友","友","同事","同學","叔","伯","姑","舅","姨","堂","表"]:
                    if k in raw: return "其他親屬"
                return ""
            c1_rel_21 = map_21_rel(c1_rel)
            c2_rel_21 = map_21_rel(c2_rel)

            # 年資 G10：偵測儲存格類型，數值型填純數字，文字型填「N年」
            # 月份 H10：固定文字「N月」或「月」
            try:
                yr_int = int(float(co_years)) if co_years else 0
                co_mos_21 = v("company_months") or "0"
                mo_int = int(float(co_mos_21)) if co_mos_21 and co_mos_21 != "0" else 0
                # G10：25萬範本是數值型，其他是文字型
                if plan_name == "21機車25萬":
                    g10_val = str(yr_int) if yr_int > 0 else "0"
                else:
                    g10_val = f"{yr_int}年"
                h10_val = f"{mo_int}月" if mo_int > 0 else "月"
            except Exception:
                g10_val = co_years
                h10_val = "月"

            # 月薪 E10：填純數字（4.5萬→45000）
            try:
                sal_clean = co_salary.replace("萬","").replace(",","").strip() if co_salary else ""
                sal_num = float(sal_clean) if sal_clean else 0
                if sal_num > 0 and sal_num < 1000:  # 萬為單位
                    sal_e10 = str(int(sal_num * 10000))
                elif sal_num >= 1000:
                    sal_e10 = str(int(sal_num))
                else:
                    sal_e10 = ""
            except Exception:
                sal_e10 = co_salary

            return {
                "C3": name, "E3": id_no, "J3": phone,
                "C4": birth_roc, "E4": id_date_roc,
                "F4": id_place_code, "G4": id_type_val,
                "J4": v("live_phone"),
                # 戶籍地址：C6=縣市、D6=鄉鎮區、E6=詳細地址
                "C6": reg_city_21,
                "D6": reg_district_21,
                "E6": v("reg_address"),
                # 住家地址：同戶籍時用戶籍資料
                "C7": reg_city_21 if live_same else live_city_21,
                "D7": reg_district_21 if live_same else live_district_21,
                "E7": v("reg_address") if live_same else v("live_address"),
                "C8": email,
                "C9": company, "G9": co_phone_area + co_phone_num,
                "C10": co_role, "E10": sal_e10,
                "G10": g10_val, "H10": h10_val,
                "C17": c1_name, "E17": c1_rel_21, "H17": c1_phone, "K17": "保密" if c1_known == "保密" else "",
                "C18": c2_name, "E18": c2_rel_21, "H18": c2_phone, "K18": "保密" if c2_known == "保密" else "",
            }

        return {}  # Unknown plan - no data fill

    def _build_txt_content(template_path: str, r: dict) -> str:
        """讀範本逐行，依 label 關鍵字填入客戶資料；無對應的保留範本預設值"""
        def v(k): return (r.get(k, "") or "").strip()

        # ===== 組合常用欄位 =====
        reg_addr = v("reg_city") + v("reg_district") + v("reg_address")
        live_same = v("live_same_as_reg") == "1"
        live_addr = reg_addr if live_same else (v("live_city") + v("live_district") + v("live_address"))
        co_area = v("company_phone_area")
        if co_area == "mobile":
            co_area = ""
        co_num = v("company_phone_num")
        co_phone = (co_area + "-" + co_num) if co_area and co_num else co_num
        co_addr = v("company_city") + v("company_district") + v("company_address")
        # 麻吉商品廠牌/型號 優先用 adminb_mj_*，沒填則 fallback 到 adminb_product_*
        mj_brand = v("adminb_mj_brand") or v("adminb_product_name")
        mj_model = v("adminb_mj_model") or v("adminb_product_model")
        product_brand_model = (mj_brand + " " + mj_model).strip()
        # 居住時間
        live_time = ""
        if v("live_years") or v("live_months"):
            live_time = f'{v("live_years")}年{v("live_months")}月'
        # 年資
        co_years_full = v("company_years")
        if v("company_months"):
            co_years_full += f'年{v("company_months")}月'
        elif co_years_full:
            co_years_full += "年"

        # ===== label 關鍵字 → 值 對應 =====
        # 比對時用 startswith / in 容錯，越長 / 越精確的 key 排前面
        LABEL_MAP = [
            # 21汽車 專屬
            ("專案名稱", v("adminb_21car_project")),
            ("貸款金額.期數.月付金", f'{v("adminb_21car_amount")}/{v("adminb_21car_period")}/{v("adminb_21car_monthly")}'.strip("/")),
            ("貸款成數", "100%"),
            ("實際售價", v("adminb_21car_price")),
            ("車價參考來源", v("adminb_21car_ref_src")),
            ("天書", v("adminb_21car_ref_src")),
            ("參考車價金額", v("adminb_21car_ref_price")),
            ("選擇利率", v("adminb_21car_rate") or "16%"),
            ("是否有信用卡", v("adminb_21car_hascc")),
            # 個人資料
            ("申請人姓名", v("customer_name")),
            ("姓名", v("customer_name")),
            ("名字", v("customer_name")),
            ("身分證字號", v("id_no")),
            ("身分證號", v("id_no")),
            ("身分證", v("id_no")),
            ("出生年月日", v("birth_date")),
            ("出生日期", v("birth_date")),
            ("發證日期", v("id_issue_date")),
            ("發證地", v("id_issue_place")),
            ("發證狀態", v("id_issue_type")),
            ("換補發類別", v("id_issue_type")),
            # 聯絡
            ("E-mail", v("email")),
            ("Email", v("email")),
            ("電子信箱", v("email")),
            ("電子帳單", v("email")),
            ("本人手機電話", v("phone")),
            ("行動電話【電信業者】", f'{v("phone")} {v("carrier")}'.strip()),
            ("行動電話", v("phone")),
            ("門號電信業者", v("carrier")),
            ("電信業者", v("carrier")),
            ("LINE ID", v("line_id")),
            ("LINEID", v("line_id")),
            # 地址 / 電話
            ("戶籍地址", reg_addr),
            ("現居地地址", live_addr),
            ("現居地址", live_addr),
            ("現住地址", live_addr),
            ("住宅地址", live_addr),
            ("戶籍電話", v("reg_phone")),
            ("現住電話", v("live_phone")),
            ("住家電話", v("live_phone")),
            ("住宅電話", v("live_phone")),
            # 居住
            ("居住時間", live_time),
            ("居住狀況", v("live_status")),
            ("最高學歷", v("education")),
            ("教育程度", v("education")),
            ("婚姻狀態", v("marriage")),
            ("婚姻", v("marriage")),
            # 公司
            ("公司名稱", v("company_name_detail") or v("company")),
            ("公司地址", co_addr),
            ("公司分機", v("company_phone_ext")),
            ("公司電話", co_phone),
            ("公司職稱", v("company_role")),
            ("職稱", v("company_role")),
            ("年資【做多久】", co_years_full),
            ("工作幾年", co_years_full),
            ("年資", co_years_full),
            ("月薪多少", v("company_salary")),
            ("月薪", v("company_salary")),
            ("是否有信用卡", v("adminb_credit_bank")),
            # 商品
            ("商品廠牌/型號", product_brand_model),
            ("商品廠牌", product_brand_model),
            # 雜項
            ("可照會時間", v("adminb_contact_time")),
            ("資金用途", v("adminb_21car_fund") or v("adminb_fund_use")),
        ]

        # 聯絡人狀態：當前是否在「親屬/1聯絡人」或「朋友/2聯絡人」段
        contact_state = {"idx": 0}  # 0=未進入, 1=聯絡人1, 2=聯絡人2

        def find_value(label_text):
            """匹配 label，回傳 (是否找到, 值)。"""
            t = label_text.strip()
            nonlocal contact_state
            # 「1聯絡人」「2聯絡人」「親屬聯絡人」「朋友聯絡人」單獨當作姓名欄位
            if t == "1聯絡人" or "親屬聯絡人" in t:
                contact_state["idx"] = 1
                return True, v("contact1_name")
            if t == "2聯絡人" or "朋友聯絡人" in t:
                contact_state["idx"] = 2
                return True, v("contact2_name")

            # 聯絡人段內欄位
            if contact_state["idx"] == 1:
                if "聯絡人姓名" in t: return True, v("contact1_name")
                if t == "姓名": return True, v("contact1_name")
                if "電話" in t: return True, v("contact1_phone")
                if "關係" in t or "稱謂" in t: return True, v("contact1_relation")
                if "知情" in t: return True, v("contact1_known")
            if contact_state["idx"] == 2:
                if "聯絡人姓名" in t: return True, v("contact2_name")
                if t == "姓名": return True, v("contact2_name")
                if "電話" in t: return True, v("contact2_phone")
                if "關係" in t or "稱謂" in t: return True, v("contact2_relation")
                if "知情" in t: return True, v("contact2_known")

            # 一般欄位
            for key, val in LABEL_MAP:
                if key in t:
                    return True, val
            return False, ""

        def detect_state_marker(line):
            """處理沒有 : 的標記行，更新 contact_state。"""
            nonlocal contact_state
            s = line.strip()
            # 1.（...） / 1.姓名 / 親屬聯絡人 / 1聯絡人
            if s.startswith("1.") or "親屬聯絡人" in s or s == "1聯絡人":
                contact_state["idx"] = 1
            elif s.startswith("2.") or "朋友聯絡人" in s or s == "2聯絡人":
                contact_state["idx"] = 2

        # 讀範本，逐行處理
        import re as _re
        with open(template_path, "r", encoding="utf-8") as f:
            content = f.read()

        def process_line(line):
            # 拆分成 [text, sep, text, sep, ...] 保留分隔符
            parts = _re.split(r'([:：])', line)
            n_seps = sum(1 for p in parts if p in (":", "："))
            if n_seps == 0:
                return line

            if n_seps == 1:
                # 單一 label:value — 整行覆寫
                label = parts[0]
                sep = parts[1]
                found, val = find_value(label.strip())
                if found:
                    return f"{label}{sep}{val}"
                return line  # 沒對應 → 保留範本預設

            # 多個 label:value 在同一行（如聯絡人姓名/關係/電話）
            # 值區域 = 下一段的「前置空白」，剩下的是下一個 label
            out = ""
            i = 0
            while i < len(parts):
                if i + 1 < len(parts) and parts[i + 1] in (":", "："):
                    label = parts[i]
                    sep = parts[i + 1]
                    next_seg = parts[i + 2] if i + 2 < len(parts) else ""
                    ws_m = _re.match(r'^[ \t]*', next_seg)
                    ws = ws_m.group(0) if ws_m else ""
                    found, val = find_value(label.strip())
                    if found:
                        out += f"{label}{sep}{val}{ws}"
                    else:
                        out += f"{label}{sep}{ws}"
                    if i + 2 < len(parts):
                        parts[i + 2] = next_seg[len(ws):]
                    i += 2
                else:
                    out += parts[i]
                    i += 1
            return out

        out_lines = []
        for line in content.splitlines():
            detect_state_marker(line)
            out_lines.append(process_line(line))
        return "\n".join(out_lines)

    for plan in plans:
        template_path = PLAN_TEMPLATE_MAP.get(plan)
        if not template_path or not os.path.exists(template_path):
            continue

        try:
            if template_path.lower().endswith(".txt"):
                # TXT 填寫表方案
                txt_content = _build_txt_content(template_path, r)
                files_to_zip.append((f"{name}_{plan}.txt", txt_content.encode("utf-8")))
            else:
                # 若存在自訂映射 → 使用動態多 sheet 填入（階段 2）
                custom_mapping = load_template_mapping(plan)
                if custom_mapping:
                    # 取得 plan-aware 處理過的 cell 值（民國年/月薪萬/縣市短碼/智能判別等）
                    processed_cells = _build_cell_map(plan, r) or {}
                    reverse_map = _build_reverse_field_map(plan)
                    # 組成多 sheet cell_map
                    sheet_cell_maps = {}
                    for sheet_name, cell_field_map in custom_mapping.items():
                        if not isinstance(cell_field_map, dict):
                            continue
                        sheet_cm = {}
                        for cell_ref, field_key in cell_field_map.items():
                            val = compute_field_value(field_key, r, plan, processed_cells, reverse_map)
                            if val is not None:
                                sheet_cm[cell_ref] = val
                        if sheet_cm:
                            sheet_cell_maps[sheet_name] = sheet_cm
                    # 亞太機車：擔保品 B7=年、C7=月，若來源是「YYYY/MM」格式自動拆開
                    if plan in ("亞太機車15萬", "亞太機車25萬", "亞太工會機車"):
                        for sname, cmap in sheet_cell_maps.items():
                            if "擔保品" in sname and "B7" in cmap:
                                val = str(cmap["B7"] or "")
                                mm = re.match(r"^(\d{3,4})[/\-年.](\d{1,2})", val)
                                if mm:
                                    cmap["B7"] = mm.group(1)
                                    cmap["C7"] = str(int(mm.group(2)))
                    filled_bytes = _fill_excel_multi_sheet(template_path, sheet_cell_maps)
                else:
                    # 未設定自訂映射 → 先嘗試用 DEFAULT_MAPPINGS 多 sheet 填入
                    cell_map = _build_cell_map(plan, r)
                    if plan in DEFAULT_MAPPINGS and cell_map:
                        dm = DEFAULT_MAPPINGS[plan]
                        sheet_cell_maps = {}
                        for sheet_name, field_map in dm.items():
                            sheet_cm = {}
                            for cell_ref, field_key in field_map.items():
                                val = compute_field_value(field_key, r, plan, cell_map, _build_reverse_field_map(plan))
                                if val is not None:
                                    sheet_cm[cell_ref] = val
                            if sheet_cm:
                                sheet_cell_maps[sheet_name] = sheet_cm
                        # 亞太機車：擔保品 B7=年、C7=月，拆開「YYYY/MM」格式
                        if plan in ("亞太機車15萬", "亞太機車25萬", "亞太工會機車"):
                            for sname, cmap in sheet_cell_maps.items():
                                if "擔保品" in sname and "B7" in cmap:
                                    val = str(cmap["B7"] or "")
                                    mm = re.match(r"^(\d{3,4})[/\-年.](\d{1,2})", val)
                                    if mm:
                                        cmap["B7"] = mm.group(1)
                                        cmap["C7"] = str(int(mm.group(2)))
                        if sheet_cell_maps:
                            filled_bytes = _fill_excel_multi_sheet(template_path, sheet_cell_maps)
                        else:
                            filled_bytes = _fill_excel_template(template_path, cell_map)
                    elif cell_map:
                        filled_bytes = _fill_excel_template(template_path, cell_map)
                    else:
                        with open(template_path, "rb") as f:
                            filled_bytes = f.read()
                files_to_zip.append((f"{name}_{plan}.xlsx", filled_bytes))
        except Exception as e:
            print(f"Generate error for {plan}: {e}")
            import traceback; traceback.print_exc()
            continue

    if not files_to_zip:
        return JSONResponse({"error": "沒有可下載的範本檔案，請確認已勾選方案並儲存"}, status_code=400)

    from urllib.parse import quote

    if len(files_to_zip) == 1:
        fname, data = files_to_zip[0]
        encoded_fname = quote(fname)
        mime = "text/plain; charset=utf-8" if fname.lower().endswith(".txt") \
            else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return StreamingResponse(
            io.BytesIO(data),
            media_type=mime,
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_fname}"}
        )
    else:
        zip_buf = io.BytesIO()
        import zipfile
        with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            for fname, data in files_to_zip:
                zf.writestr(fname, data)
        zip_buf.seek(0)
        zip_fname = quote(f"{name}_申請書.zip")
        return StreamingResponse(
            zip_buf,
            media_type="application/zip",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{zip_fname}"}
        )


# =========================
# 密碼管理頁面（管理員專用）
# =========================
@app.get("/admin/passwords", response_class=HTMLResponse)
def passwords_page(request: Request):
    from fastapi.responses import RedirectResponse
    role = check_auth(request)
    if role != "admin": return RedirectResponse("/login")
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT group_id, group_name FROM groups WHERE group_type='SALES_GROUP' AND is_active=1 ORDER BY group_name")
    sales_groups = cur.fetchall(); conn.close()
    grp_rows = "".join(f"""
        <div style="display:grid;grid-template-columns:120px 1fr 100px;gap:10px;align-items:end;padding:10px 0;border-bottom:1px solid #e5e7eb">
          <div style="font-size:14px;font-weight:500">{h(g["group_name"])}</div>
          <input type="password" id="gpw_{h(g["group_id"])}" class="input" placeholder="輸入新密碼">
          <button onclick="changePw('group','{h(g["group_id"])}','gpw_{h(g["group_id"])}')" class="btn btn-primary">更新</button>
        </div>""" for g in sales_groups)
    return f"""<!DOCTYPE html><html><head>{PAGE_CSS}<title>密碼管理</title></head><body>
    {make_topnav("admin","passwords")}
    <div class="page" style="max-width:680px">
      <h2 style="font-size:18px;font-weight:600;margin-bottom:20px">🔑 密碼管理</h2>
      <div id="msg"></div>
      <div style="background:#fff;border:1px solid #e5e7eb;border-radius:10px;padding:20px;margin-bottom:16px">
        <div style="font-size:13px;font-weight:600;color:#6b7280;margin-bottom:14px;padding-bottom:8px;border-bottom:1px solid #f3f4f6">系統角色密碼</div>
        <div style="display:grid;grid-template-columns:120px 1fr 100px;gap:10px;align-items:end;padding:10px 0;border-bottom:1px solid #f3f4f6">
          <div style="font-size:14px;font-weight:500">管理員</div>
          <input type="password" id="admin_pw" class="input" placeholder="輸入新密碼">
          <button onclick="changePw('admin','','admin_pw')" class="btn btn-primary">更新</button>
        </div>
        <div style="display:grid;grid-template-columns:120px 1fr 100px;gap:10px;align-items:end;padding:10px 0;border-bottom:1px solid #f3f4f6">
          <div style="font-size:14px;font-weight:500">行政B</div>
          <input type="password" id="adminB_pw" class="input" placeholder="輸入新密碼">
          <button onclick="changePw('adminB','','adminB_pw')" class="btn btn-primary">更新</button>
        </div>
        <div style="display:grid;grid-template-columns:120px 1fr 100px;gap:10px;align-items:end;padding:10px 0;border-bottom:1px solid #f3f4f6">
          <div style="font-size:14px;font-weight:500">行政A</div>
          <input type="password" id="report_pw" class="input" placeholder="輸入新密碼">
          <button onclick="changePw('normal','','report_pw')" class="btn btn-primary">更新</button>
        </div>
        <div style="display:grid;grid-template-columns:120px 1fr 100px;gap:10px;align-items:end;padding:10px 0">
          <div style="font-size:14px;font-weight:500">VBA密鑰</div>
          <input type="password" id="vba_secret" class="input" placeholder="輸入新密鑰">
          <button onclick="changePw('vba','','vba_secret')" class="btn btn-primary">更新</button>
        </div>
      </div>
      <div style="background:#fff;border:1px solid #e5e7eb;border-radius:10px;padding:20px;margin-bottom:16px">
        <div style="font-size:13px;font-weight:600;color:#6b7280;margin-bottom:14px;padding-bottom:8px;border-bottom:1px solid #f3f4f6">業務群組密碼</div>
        {grp_rows if grp_rows else '<div style="color:#9ca3af;font-size:13px;padding:10px 0">尚無業務群組</div>'}
      </div>
      <div style="background:#fff;border:1px solid #e5e7eb;border-radius:10px;padding:20px;margin-bottom:16px">
        <div style="font-size:13px;font-weight:600;color:#6b7280;margin-bottom:14px;padding-bottom:8px;border-bottom:1px solid #f3f4f6">🔓 解鎖帳號</div>
        <div id="locked-list" style="margin-bottom:14px"></div>
        <button class="btn" onclick="loadLocked()" style="font-size:12px">🔄 查看目前鎖定帳號</button>
      </div>
      <div style="background:#fffbeb;border:1px solid #fde68a;border-radius:8px;padding:12px 16px;font-size:13px;color:#92400e">
        ⚠️ 更新密碼後立即生效，請通知相關人員！
      </div>
    </div>
    <script>
    async function loadLocked(){{
      const r = await fetch('/admin/locked-accounts');
      const d = await r.json();
      const el = document.getElementById('locked-list');
      if(!d.locked || d.locked.length === 0){{
        el.innerHTML = '<div style="color:#6b7280;font-size:13px">目前沒有鎖定帳號 ✅</div>';
        return;
      }}
      el.innerHTML = d.locked.map(x => `
        <div style="display:flex;align-items:center;justify-content:space-between;padding:8px 0;border-bottom:1px solid #f3f4f6">
          <div>
            <span style="font-size:13px;font-weight:500">${{x.identifier}}</span>
            <span style="font-size:12px;color:#dc2626;margin-left:8px">失敗${{x.attempts}}次｜鎖定到 ${{x.locked_until}}</span>
          </div>
          <button onclick="unlockAccount('${{x.identifier}}')" class="btn" style="font-size:12px;color:#dc2626;border-color:#fca5a5">解鎖</button>
        </div>`).join('');
    }}
    async function unlockAccount(identifier){{
      const r = await fetch('/admin/unlock-account', {{
        method:'POST',
        headers:{{'Content-Type':'application/json'}},
        body:JSON.stringify({{identifier}})
      }});
      const d = await r.json();
      if(d.ok) loadLocked();
    }}
    window.onload = loadLocked;
    async function changePw(role, groupId, inputId) {{
      const pw = document.getElementById(inputId).value.trim();
      if (!pw) {{ alert("請輸入新密碼"); return; }}
      if (pw.length < 6) {{ alert("密碼至少6個字元"); return; }}
      const r = await fetch("/admin/update-password", {{
        method: "POST",
        headers: {{"Content-Type":"application/json"}},
        body: JSON.stringify({{role, group_id: groupId, password: pw}})
      }});
      const d = await r.json();
      const msg = document.getElementById("msg");
      if (d.ok) {{
        msg.innerHTML = '<div style="background:#dcfce7;color:#15803d;padding:10px 14px;border-radius:6px;margin-bottom:14px">✅ ' + d.message + '</div>';
        document.getElementById(inputId).value = "";
      }} else {{
        msg.innerHTML = '<div style="background:#fef2f2;color:#dc2626;padding:10px 14px;border-radius:6px;margin-bottom:14px">❌ ' + d.message + '</div>';
      }}
      setTimeout(() => msg.innerHTML = "", 3000);
    }}
    </script>
    </body></html>"""


@app.post("/admin/update-password")
async def update_password(request: Request):
    role = check_auth(request)
    if role != "admin":
        return JSONResponse({"ok": False, "message": "無權限"}, status_code=403)
    data = await request.json()
    pw_role = data.get("role","")
    group_id = data.get("group_id","")
    password = data.get("password","").strip()
    if not password or len(password) < 6:
        return JSONResponse({"ok": False, "message": "密碼至少6個字元"})
    hashed = hash_pw(password)
    if pw_role == "admin":
        set_setting("admin_pw", hashed)
        return JSONResponse({"ok": True, "message": "管理員密碼已更新"})
    elif pw_role == "adminB":
        set_setting("adminB_pw", hashed)
        return JSONResponse({"ok": True, "message": "行政B密碼已更新"})
    elif pw_role == "normal":
        set_setting("report_pw", hashed)
        return JSONResponse({"ok": True, "message": "行政A密碼已更新"})
    elif pw_role == "vba":
        set_setting("vba_secret", hashed)
        return JSONResponse({"ok": True, "message": "VBA密鑰已更新"})
    elif pw_role == "group" and group_id:
        conn = get_conn(); cur = conn.cursor()
        cur.execute("UPDATE groups SET password_hash=? WHERE group_id=?", (hashed, group_id))
        conn.commit(); conn.close()
        return JSONResponse({"ok": True, "message": "群組密碼已更新"})
    return JSONResponse({"ok": False, "message": "未知角色"})


@app.get("/admin/logs", response_class=HTMLResponse)
def admin_logs_page(request: Request, page: int = 1, q: str = ""):
    from fastapi.responses import RedirectResponse
    role = check_auth(request)
    if role != "admin":
        return RedirectResponse("/login")
    per_page = 50
    offset = (page - 1) * per_page
    conn = get_conn(); cur = conn.cursor()
    if q:
        cur.execute("SELECT COUNT(*) as c FROM case_logs WHERE customer_name LIKE ? OR message_text LIKE ?", (f"%{q}%", f"%{q}%"))
        total = cur.fetchone()["c"]
        cur.execute("SELECT * FROM case_logs WHERE customer_name LIKE ? OR message_text LIKE ? ORDER BY created_at DESC LIMIT ? OFFSET ?", (f"%{q}%", f"%{q}%", per_page, offset))
    else:
        cur.execute("SELECT COUNT(*) as c FROM case_logs")
        total = cur.fetchone()["c"]
        cur.execute("SELECT * FROM case_logs ORDER BY created_at DESC LIMIT ? OFFSET ?", (per_page, offset))
    logs = cur.fetchall(); conn.close()
    total_pages = (total + per_page - 1) // per_page
    rows_html = ""
    for log in logs:
        dt = (log["created_at"] or "")[:19].replace("T", " ")
        gname = get_group_name(log["from_group_id"])
        txt = (log["message_text"] or "")[:80]
        rows_html += f'<tr><td style="white-space:nowrap">{h(dt)}</td><td>{h(log["customer_name"])}</td><td>{h(gname)}</td><td style="word-break:break-all">{h(txt)}</td></tr>'
    if not rows_html:
        rows_html = '<tr><td colspan="4" style="text-align:center;color:#999;padding:20px">沒有操作紀錄</td></tr>'
    # 分頁
    q_param = f"&q={h(q)}" if q else ""
    pag = ""
    if total_pages > 1:
        pag = '<div style="display:flex;gap:6px;justify-content:center;margin-top:14px;flex-wrap:wrap">'
        if page > 1:
            pag += f'<a href="/admin/logs?page={page-1}{q_param}" class="btn" style="background:#e8e2da;color:#4a3e30;font-size:12px">上一頁</a>'
        pag += f'<span style="padding:6px 12px;font-size:12px;color:#6a5e4e">{page}/{total_pages}（共{total}筆）</span>'
        if page < total_pages:
            pag += f'<a href="/admin/logs?page={page+1}{q_param}" class="btn" style="background:#e8e2da;color:#4a3e30;font-size:12px">下一頁</a>'
        pag += '</div>'
    return f"""<!DOCTYPE html><html><head>{PAGE_CSS}<title>操作紀錄</title>
    <style>
    table{{width:100%;border-collapse:collapse;font-size:12px}}
    th{{background:#f0ebe4;color:#4a3e30;padding:8px;text-align:left;font-weight:600;position:sticky;top:0}}
    td{{padding:6px 8px;border-bottom:1px solid #ece8e2}}
    tr:hover{{background:#faf7f4}}
    </style></head><body>
    {make_topnav(role, "logs")}
    <div class="page">
      <h2 style="font-size:18px;font-weight:600;margin-bottom:14px">📝 操作紀錄</h2>
      <form method="get" action="/admin/logs" style="margin-bottom:12px;display:flex;gap:8px">
        <input name="q" value="{h(q)}" placeholder="搜尋客戶名或操作內容..." class="input" style="flex:1" autofocus>
        <button type="submit" class="btn btn-primary" style="font-size:12px">🔍 搜尋</button>
        {"<a href='/admin/logs' class='btn' style='background:#e8e2da;color:#4a3e30;font-size:12px'>清除</a>" if q else ""}
      </form>
      <div style="overflow-x:auto;background:#fff;border:1px solid #ddd5ca;border-radius:10px">
        <table><thead><tr><th>時間</th><th>客戶</th><th>來源群組</th><th>操作內容</th></tr></thead>
        <tbody>{rows_html}</tbody></table>
      </div>
      {pag}
    </div></body></html>"""


@app.get("/admin/download-db")
async def download_db(request: Request):
    role = check_auth(request)
    if role != "admin":
        return RedirectResponse("/login")
    import shutil
    backup_path = DB_PATH + ".backup"
    shutil.copy2(DB_PATH, backup_path)
    today = datetime.now().strftime("%Y%m%d_%H%M")
    return FileResponse(backup_path, filename=f"loan_system_{today}.db",
                        media_type="application/octet-stream")


@app.get("/admin/gdrive-backup", response_class=HTMLResponse)
async def gdrive_backup_page(request: Request):
    from fastapi.responses import RedirectResponse
    role = check_auth(request)
    if role != "admin":
        return RedirectResponse("/login", status_code=303)
    qs = str(request.url.query or "")
    flash = ""
    if "ok=1" in qs:
        flash = '<div style="background:#dcfce7;color:#15803d;padding:10px 14px;border-radius:6px;font-size:13px;margin-bottom:14px">✅ 備份完成</div>'
    elif "err=" in qs:
        import urllib.parse as _up
        msg = _up.unquote(qs.split("err=",1)[1].split("&",1)[0])
        flash = f'<div style="background:#fee;color:#b91c1c;padding:10px 14px;border-radius:6px;font-size:13px;margin-bottom:14px">❌ 備份失敗：{h(msg)}</div>'
    status_txt = "✅ 已啟用" if BACKUP_ENABLED else "❌ 未啟用（BACKUP_ENABLED 環境變數）"
    cfg_ok = BACKUP_ENABLED and GOOGLE_SERVICE_ACCOUNT_JSON and GOOGLE_DRIVE_FOLDER_ID
    files = list_gdrive_backups() if cfg_ok else []
    def _fmt_size(n):
        try:
            n = int(n)
        except Exception:
            return "-"
        for u in ("B","KB","MB","GB"):
            if n < 1024: return f"{n:.1f}{u}"
            n /= 1024
        return f"{n:.1f}TB"
    rows_html = "".join(
        f'<tr><td style="padding:8px 12px">{h(f.get("name",""))}</td>'
        f'<td style="padding:8px 12px;text-align:right">{_fmt_size(f.get("size",0))}</td>'
        f'<td style="padding:8px 12px;color:#666">{h((f.get("createdTime","") or "")[:19].replace("T"," "))}</td></tr>'
        for f in files
    ) or '<tr><td colspan="3" style="padding:16px;text-align:center;color:#999">尚無備份（按上方按鈕手動執行一次）</td></tr>'
    return f"""<!DOCTYPE html><html><head>{PAGE_CSS}<title>Google Drive 備份</title></head><body>
    {make_topnav(role,"gdrive-backup")}
    <div class="page">
      <div class="card" style="max-width:900px;margin:0 auto;padding:20px">
        {flash}
        <h2 style="margin:0 0 12px;font-size:18px">💾 Google Drive 自動備份</h2>
        <div style="background:#f7f4ef;padding:12px 14px;border-radius:8px;margin-bottom:14px;font-size:13px;line-height:1.8">
          <div>狀態：<b>{status_txt}</b></div>
          <div>每日自動執行時間：<b>03:00（台灣時間）</b></div>
          <div>保留策略：<b>近 {BACKUP_KEEP_DAILY} 天每日 + 近 {BACKUP_KEEP_MONTHLY} 個月每月 1 份</b></div>
        </div>
        <form method="post" action="/admin/gdrive-backup" style="margin-bottom:16px">
          <button type="submit" class="btn btn-primary" style="padding:10px 18px;font-size:14px">▶ 立即備份一次</button>
          <span style="color:#999;font-size:12px;margin-left:10px">（手動測試用，按下後約 5-15 秒完成）</span>
        </form>
        <h3 style="margin:20px 0 8px;font-size:15px">Drive 上現存的備份</h3>
        <table style="width:100%;border-collapse:collapse;font-size:13px">
          <thead><tr style="background:#ece8e2">
            <th style="padding:8px 12px;text-align:left">檔名</th>
            <th style="padding:8px 12px;text-align:right">大小</th>
            <th style="padding:8px 12px;text-align:left">建立時間(UTC)</th>
          </tr></thead>
          <tbody>{rows_html}</tbody>
        </table>
      </div>
    </div></body></html>"""


@app.post("/admin/gdrive-backup")
async def gdrive_backup_trigger(request: Request):
    from fastapi.responses import RedirectResponse
    role = check_auth(request)
    if role != "admin":
        return RedirectResponse("/login", status_code=303)
    result = do_backup_now()
    if result.get("ok"):
        return RedirectResponse("/admin/gdrive-backup?ok=1", status_code=303)
    import urllib.parse as _up
    return RedirectResponse(f"/admin/gdrive-backup?err={_up.quote(str(result.get('reason',''))[:200])}", status_code=303)


# =========================
# 申請書範本管理（階段 1：上傳替換，座標不變時可用）
# =========================
@app.get("/admin/templates", response_class=HTMLResponse)
def admin_templates_page(request: Request):
    role = check_auth(request)
    if role not in ("admin", "adminB"):
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/login", status_code=303)

    qs = str(request.url.query or "")
    flash = ""
    if "ok=1" in qs:
        flash = '<div style="background:#dcfce7;color:#15803d;padding:10px 14px;border-radius:6px;font-size:13px;margin-bottom:16px;">✅ 範本已上傳並套用。下次下載申請書會使用新範本。</div>'
    elif "reset=1" in qs:
        flash = '<div style="background:#e0f2fe;color:#075985;padding:10px 14px;border-radius:6px;font-size:13px;margin-bottom:16px;">↩️ 已還原為系統內建範本。</div>'
    elif "err=" in qs:
        _em = re.search(r"err=([^&]+)", qs)
        _msg = _em.group(1) if _em else "上傳失敗"
        try:
            from urllib.parse import unquote
            _msg = unquote(_msg)
        except Exception:
            pass
        flash = f'<div style="background:#fef2f2;color:#dc2626;padding:10px 14px;border-radius:6px;font-size:13px;margin-bottom:16px;">❌ {h(_msg)}</div>'

    _base = os.path.dirname(os.path.abspath(__file__))
    rows_html = []
    # 記錄已顯示過的檔名（亞太機車15萬/25萬共用一份內建），避免重複說明
    for plan_name, builtin_file in APPLICATION_PLAN_LIST:
        custom_path = get_custom_template_path(plan_name)
        has_custom = os.path.isfile(custom_path)
        builtin_path = os.path.join(_base, "申請書", builtin_file)
        builtin_exists = os.path.isfile(builtin_path)

        mapping_path = get_template_mapping_path(plan_name)
        has_mapping = os.path.isfile(mapping_path)
        inherited_src = "" if has_mapping else get_inherited_mapping_source(plan_name)
        if has_mapping:
            mapping_badge = '<span style="background:#dcfce7;color:#15803d;padding:1px 8px;border-radius:8px;font-size:11px;margin-left:4px;">已設定映射</span>'
        elif inherited_src:
            mapping_badge = f'<span style="background:#fef3c7;color:#92400e;padding:1px 8px;border-radius:8px;font-size:11px;margin-left:4px;">繼承「{h(inherited_src)}」</span>'
        else:
            mapping_badge = '<span style="background:#f3f4f6;color:#6b7280;padding:1px 8px;border-radius:8px;font-size:11px;margin-left:4px;">未設定映射</span>'

        if has_custom:
            st = os.stat(custom_path)
            status_badge = '<span style="background:#dcfce7;color:#15803d;padding:2px 10px;border-radius:10px;font-size:12px;font-weight:500;">● 自訂範本</span>' + mapping_badge
            size_kb = max(1, st.st_size // 1024)
            updated = datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M")
            info = f"{size_kb} KB　·　更新於 {updated}"
            actions = (
                f'<a href="/admin/templates/edit?plan={h(plan_name)}" '
                f'style="display:inline-block;background:#eef2ff;color:#4338ca;border:1px solid #c7d2fe;padding:5px 12px;border-radius:4px;text-decoration:none;font-size:12px;margin-right:6px;font-weight:500;">🔧 編輯映射</a>'
                f'<a href="/admin/templates/download?plan={h(plan_name)}" '
                f'style="color:#4e7055;margin-right:10px;font-size:13px;text-decoration:none;">📥 下載</a>'
                f'<button onclick="upFor(\'{h(plan_name)}\')" '
                f'style="background:#4e7055;color:#fff;border:none;padding:5px 12px;border-radius:4px;cursor:pointer;font-size:12px;margin-right:6px;">上傳新範本</button>'
                f'<button onclick="rstFor(\'{h(plan_name)}\')" '
                f'style="background:#fff;color:#dc2626;border:1px solid #fecaca;padding:5px 12px;border-radius:4px;cursor:pointer;font-size:12px;">還原內建</button>'
            )
        else:
            status_badge = '<span style="background:#f3f4f6;color:#6b7280;padding:2px 10px;border-radius:10px;font-size:12px;">○ 內建範本</span>' + mapping_badge
            info = f'<span style="color:#9ca3af;">（尚未上傳自訂範本）</span>'
            if not builtin_exists:
                info += ' <span style="color:#dc2626;font-size:12px;">⚠ 內建檔案遺失</span>'
            actions = (
                f'<a href="/admin/templates/edit?plan={h(plan_name)}" '
                f'style="display:inline-block;background:#eef2ff;color:#4338ca;border:1px solid #c7d2fe;padding:5px 12px;border-radius:4px;text-decoration:none;font-size:12px;margin-right:6px;font-weight:500;">🔧 編輯映射</a>'
                f'<a href="/admin/templates/download?plan={h(plan_name)}" '
                f'style="color:#4e7055;margin-right:10px;font-size:13px;text-decoration:none;">📥 下載內建</a>'
                f'<button onclick="upFor(\'{h(plan_name)}\')" '
                f'style="background:#4e7055;color:#fff;border:none;padding:5px 12px;border-radius:4px;cursor:pointer;font-size:12px;">上傳範本</button>'
            )

        rows_html.append(
            f'<tr style="border-bottom:1px solid #f1f1f1;">'
            f'<td style="padding:12px 14px;font-weight:500;color:#3a3530;">{h(plan_name)}</td>'
            f'<td style="padding:12px 14px;">{status_badge}</td>'
            f'<td style="padding:12px 14px;color:#6b7280;font-size:13px;">{info}</td>'
            f'<td style="padding:12px 14px;text-align:right;white-space:nowrap;">{actions}</td>'
            f'</tr>'
        )

    nav = make_topnav(role, "templates")
    body = f"""<!DOCTYPE html><html><head>{PAGE_CSS}<title>申請書範本管理</title></head><body>
{nav}
<div class="container" style="max-width:1100px;margin:24px auto;padding:0 16px;">
  <h2 style="font-size:20px;font-weight:600;margin-bottom:6px;color:#3a3530;">📄 申請書範本管理</h2>
  <p style="color:#6b7280;font-size:13px;margin-bottom:16px;line-height:1.6;">
    上傳自訂範本後，<b>行政B下載</b>申請書會優先使用你的版本。填入邏輯（欄位對應、下拉選單匹配、格式處理）完全不變。
  </p>
  {flash}
  <div style="background:#fffbea;border:1px solid #fde68a;padding:12px 16px;border-radius:6px;font-size:13px;color:#854d0e;margin-bottom:20px;line-height:1.7;">
    ⚠️ <b>階段 1 限制</b>：目前版本要求「儲存格座標與原範本一致」，你可以改：
    <br>&nbsp;&nbsp;✅ 標題文字、欄寬、字型、顏色、邊框、下拉選項　　✅ 新增說明欄、Logo
    <br>&nbsp;&nbsp;❌ 不要搬動填入欄位的座標（例如原本 E17 填行業類別，不要搬到 F20）
    <br>若需要「欄位可搬動」的完整功能（自動掃描 + 人工標註映射 + 異動比對），請等階段 2。
  </div>
  <table style="width:100%;border-collapse:collapse;background:#fff;box-shadow:0 1px 3px rgba(0,0,0,0.05);border-radius:6px;overflow:hidden;">
    <thead><tr style="background:#f9fafb;font-size:13px;color:#374151;">
      <th style="padding:11px 14px;text-align:left;font-weight:600;">方案</th>
      <th style="padding:11px 14px;text-align:left;font-weight:600;">目前範本</th>
      <th style="padding:11px 14px;text-align:left;font-weight:600;">檔案資訊</th>
      <th style="padding:11px 14px;text-align:right;font-weight:600;">操作</th>
    </tr></thead>
    <tbody>{''.join(rows_html)}</tbody>
  </table>
  <p style="margin-top:20px;color:#9ca3af;font-size:12px;line-height:1.7;">
    上傳限制：副檔名 .xlsx / 檔案大小 10MB 以內。<br>
    儲存位置：{h(TEMPLATES_DIR)}（與資料庫同目錄，部署重啟後仍保留）
  </p>
</div>

<form id="upForm" method="post" action="/admin/templates/upload" enctype="multipart/form-data" style="display:none;">
  <input type="hidden" name="plan" id="upPlan">
  <input type="file" name="file" id="upFile" accept=".xlsx,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet">
</form>
<form id="rstForm" method="post" action="/admin/templates/reset" style="display:none;">
  <input type="hidden" name="plan" id="rstPlan">
</form>
<script>
function upFor(plan) {{
  document.getElementById('upPlan').value = plan;
  var fi = document.getElementById('upFile');
  fi.value = '';
  fi.onchange = function() {{
    if (!fi.files.length) return;
    var f = fi.files[0];
    var nm = f.name.toLowerCase();
    if (!nm.endsWith('.xlsx')) {{ alert('只接受 .xlsx 檔案'); return; }}
    if (f.size > 10 * 1024 * 1024) {{ alert('檔案超過 10MB 限制'); return; }}
    if (confirm('確定把「' + plan + '」的範本替換為「' + f.name + '」嗎？\\n\\n覆蓋後，原本的自訂範本會被取代。')) {{
      document.getElementById('upForm').submit();
    }}
  }};
  fi.click();
}}
function rstFor(plan) {{
  if (!confirm('確定要還原「' + plan + '」為系統內建範本嗎？\\n\\n你上傳的自訂檔案會被刪除。')) return;
  document.getElementById('rstPlan').value = plan;
  document.getElementById('rstForm').submit();
}}
</script>
</body></html>"""
    return HTMLResponse(body)


@app.post("/admin/templates/upload")
async def admin_templates_upload(request: Request):
    role = check_auth(request)
    if role not in ("admin", "adminB"):
        return JSONResponse({"error": "無權限"}, status_code=403)

    from fastapi.responses import RedirectResponse
    from urllib.parse import quote

    valid_plans = {p for p, _ in APPLICATION_PLAN_LIST}
    try:
        form = await request.form()
    except Exception as e:
        return RedirectResponse(f"/admin/templates?err={quote('表單解析失敗：' + str(e))}", status_code=303)

    plan_name = (form.get("plan") or "").strip()
    uploaded = form.get("file")

    if plan_name not in valid_plans:
        return RedirectResponse(f"/admin/templates?err={quote('方案名稱無效')}", status_code=303)

    if not uploaded or not hasattr(uploaded, "filename") or not uploaded.filename:
        return RedirectResponse(f"/admin/templates?err={quote('未選擇檔案')}", status_code=303)

    if not uploaded.filename.lower().endswith(".xlsx"):
        return RedirectResponse(f"/admin/templates?err={quote('僅接受 .xlsx 檔案')}", status_code=303)

    content = await uploaded.read()
    if len(content) > 10 * 1024 * 1024:
        return RedirectResponse(f"/admin/templates?err={quote('檔案超過 10MB 限制')}", status_code=303)
    if len(content) < 200:
        return RedirectResponse(f"/admin/templates?err={quote('檔案太小，疑似不是合法 xlsx')}", status_code=303)

    # 驗證 xlsx 合法性：必須是 zip 且含 xl/workbook.xml
    import zipfile as _zf
    try:
        zf = _zf.ZipFile(io.BytesIO(content))
        names = zf.namelist()
        if "xl/workbook.xml" not in names:
            return RedirectResponse(f"/admin/templates?err={quote('檔案不是合法 xlsx（缺 xl/workbook.xml）')}", status_code=303)
    except Exception as e:
        return RedirectResponse(f"/admin/templates?err={quote('檔案解析失敗：' + str(e))}", status_code=303)

    # NOTE: 之前會在此自動清除「擔保品資訊」分頁的填色/條件格式，
    # 但發現會破壞 x14 擴充（下拉選單/公式變純文字），關閉自動處理。
    # 若需要白底請直接在 Excel 範本內設定白底後再上傳。

    target_path = get_custom_template_path(plan_name)
    try:
        os.makedirs(TEMPLATES_DIR, exist_ok=True)
        with open(target_path, "wb") as f:
            f.write(content)
    except Exception as e:
        return RedirectResponse(f"/admin/templates?err={quote('寫入失敗：' + str(e))}", status_code=303)

    # 上傳成功 → 導向編輯頁讓使用者檢查 / 調整映射
    return RedirectResponse(f"/admin/templates/edit?plan={quote(plan_name)}&uploaded=1", status_code=303)


@app.post("/admin/templates/reset")
async def admin_templates_reset(request: Request):
    role = check_auth(request)
    if role not in ("admin", "adminB"):
        return JSONResponse({"error": "無權限"}, status_code=403)

    from fastapi.responses import RedirectResponse
    from urllib.parse import quote

    valid_plans = {p for p, _ in APPLICATION_PLAN_LIST}
    form = await request.form()
    plan_name = (form.get("plan") or "").strip()
    if plan_name not in valid_plans:
        return RedirectResponse(f"/admin/templates?err={quote('方案名稱無效')}", status_code=303)

    target_path = get_custom_template_path(plan_name)
    if os.path.isfile(target_path):
        try:
            os.remove(target_path)
        except Exception as e:
            return RedirectResponse(f"/admin/templates?err={quote('刪除失敗：' + str(e))}", status_code=303)

    return RedirectResponse("/admin/templates?reset=1", status_code=303)


@app.get("/admin/templates/download")
def admin_templates_download(request: Request, plan: str = ""):
    role = check_auth(request)
    if role not in ("admin", "adminB"):
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/login", status_code=303)

    valid_plans = {p: f for p, f in APPLICATION_PLAN_LIST}
    if plan not in valid_plans:
        return JSONResponse({"error": "方案名稱無效"}, status_code=400)

    custom_path = get_custom_template_path(plan)
    xlsx_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if os.path.isfile(custom_path):
        return FileResponse(custom_path, filename=f"{plan}.xlsx", media_type=xlsx_mime)

    # fallback：回傳內建範本
    _base = os.path.dirname(os.path.abspath(__file__))
    builtin_path = os.path.join(_base, "申請書", valid_plans[plan])
    if os.path.isfile(builtin_path):
        return FileResponse(builtin_path, filename=f"{plan}_內建.xlsx", media_type=xlsx_mime)
    return JSONResponse({"error": "範本不存在"}, status_code=404)


# =========================
# 範本編輯頁（階段 2：映射設定）
# =========================
def _get_template_xlsx_bytes(plan_name: str) -> bytes:
    """取得目前使用中的範本 bytes（自訂優先，否則內建）"""
    custom_path = get_custom_template_path(plan_name)
    if os.path.isfile(custom_path):
        with open(custom_path, "rb") as f:
            return f.read()
    _base = os.path.dirname(os.path.abspath(__file__))
    filename = dict(APPLICATION_PLAN_LIST).get(plan_name)
    if filename:
        builtin_path = os.path.join(_base, "申請書", filename)
        if os.path.isfile(builtin_path):
            with open(builtin_path, "rb") as f:
                return f.read()
    return b""


@app.get("/admin/templates/edit", response_class=HTMLResponse)
def admin_templates_edit(request: Request, plan: str = ""):
    role = check_auth(request)
    if role not in ("admin", "adminB"):
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/login", status_code=303)

    valid_plans = {p for p, _ in APPLICATION_PLAN_LIST}
    if plan not in valid_plans:
        return HTMLResponse("<h3>方案名稱無效</h3>", status_code=400)

    xlsx_bytes = _get_template_xlsx_bytes(plan)
    if not xlsx_bytes:
        return HTMLResponse(f"<h3>找不到「{h(plan)}」的範本檔案</h3>", status_code=404)

    scan = scan_xlsx_structure(xlsx_bytes)
    if scan.get("error"):
        return HTMLResponse(f"<h3>掃描範本失敗：{h(scan['error'])}</h3>", status_code=500)

    # 讀取現有映射（含 sibling fallback）
    mapping = load_template_mapping(plan)
    inherited_from = get_inherited_mapping_source(plan)
    is_new_mapping = not mapping
    if is_new_mapping:
        mapping = get_default_mapping(plan)

    # 建立欄位下拉選項（中文）
    field_options_html = ['<option value="">（不填入資料）</option>']
    # 分組：把同類欄位放一起
    groups = [
        ("個人基本", ["customer_name","id_no","birth_date","phone","email","line_id","marriage","education","carrier","fb"]),
        ("身分證", ["id_issue_date","id_issue_place","id_issue_type"]),
        ("戶籍地址", ["reg_city","reg_district","reg_address","reg_phone","reg_full_address"]),
        ("現居地址", ["live_city","live_district","live_address","live_phone","live_same_as_reg","live_full_address","live_status","live_years","live_months"]),
        ("公司職務", ["company","company_name_detail","company_role","company_phone_area","company_phone_num","company_phone_ext","company_full_phone","company_years","company_months","company_salary","company_city","company_district","company_address","company_full_address"]),
        ("聯絡人", ["contact1_name","contact1_relation","contact1_phone","contact1_known","contact2_name","contact2_relation","contact2_phone","contact2_known"]),
        ("車輛", ["adminb_brand","vehicle_plate","adminb_vehicle_type","adminb_engine_no","adminb_body_no","adminb_mfg_date","adminb_displacement","adminb_color"]),
        ("亞太專用", ["adminb_fund_use","adminb_industry","adminb_role"]),
        ("和裕專用", ["adminb_hr_industry","adminb_hr_role","adminb_bank","adminb_branch","adminb_product","adminb_model","adminb_contact_time"]),
        ("貸救補", ["adminb_product_name","adminb_product_model"]),
        ("特殊操作", ["__CLEAR__","__KEEP__"]),
    ]
    for grp_name, keys in groups:
        field_options_html.append(f'<optgroup label="{grp_name}">')
        for k in keys:
            label = FIELD_LABELS.get(k, k)
            field_options_html.append(f'<option value="{h(k)}">{h(label)}</option>')
        field_options_html.append("</optgroup>")
    field_options_str = "".join(field_options_html)

    # 產生每個 sheet 的儲存格清單
    # 只讀 sheet（隱藏表、下拉參照表）不列儲存格、但顯示摘要讓使用者知道存在
    READONLY_SHEET_KEYWORDS = ("參照表", "對照表", "代碼表")
    sheets_html = []
    for si, sheet in enumerate(scan["sheets"]):
        sheet_name = sheet["name"]
        is_hidden = sheet["state"] != "visible"
        is_readonly_named = any(kw in sheet_name for kw in READONLY_SHEET_KEYWORDS)
        is_readonly = is_hidden or is_readonly_named
        state_badge = ' <span style="color:#9ca3af;font-size:11px;">(hidden)</span>' if is_hidden else ''

        # 該 sheet 的現有映射
        sheet_mapping = mapping.get(sheet_name, {})

        # 只讀分頁：不列儲存格
        if is_readonly:
            reason = "隱藏分頁（系統資料對照）" if is_hidden else "下拉選單參照表"
            sheets_html.append(
                f'<div class="sheet-block" data-sheet="{h(sheet_name)}" style="background:#fef2f2;border:1px solid #fecaca;border-radius:8px;margin-bottom:18px;overflow:hidden;">'
                f'<div style="padding:12px 16px;display:flex;justify-content:space-between;align-items:center;">'
                f'<div><span style="font-weight:600;color:#991b1b;">🔒 {h(sheet_name)}{state_badge}</span>'
                f'<span style="color:#64748b;font-size:12px;margin-left:10px;">{reason}，不可動也不填入資料</span></div>'
                f'<div style="font-size:12px;color:#64748b;">共 {len(sheet["cells"])} 格（不列出）</div>'
                f'</div></div>'
            )
            continue

        # 排序儲存格（依欄字母、再依列）
        def _cell_sort_key(c):
            col = c["col"]
            # A→1, B→2, Z→26, AA→27
            n = 0
            for ch in col:
                n = n * 26 + (ord(ch) - ord('A') + 1)
            return (c["row"], n)
        cells_sorted = sorted(sheet["cells"], key=_cell_sort_key)

        rows_html = []
        empty_cells_with_dropdown = []
        for c in cells_sorted:
            ref = c["ref"]
            val = c["value"] or ""
            ctype = c["type"]
            has_dd = "dropdown" in c
            current_field = sheet_mapping.get(ref, "")

            # 類型標記
            if ctype == "formula":
                type_tag = '<span style="background:#fef3c7;color:#92400e;padding:1px 6px;border-radius:3px;font-size:10px;">公式</span>'
            elif ctype == "empty":
                type_tag = '<span style="background:#e0f2fe;color:#075985;padding:1px 6px;border-radius:3px;font-size:10px;">空白</span>'
            elif has_dd:
                type_tag = '<span style="background:#dbeafe;color:#1d4ed8;padding:1px 6px;border-radius:3px;font-size:10px;">下拉</span>'
            elif val:
                type_tag = '<span style="background:#f3f4f6;color:#6b7280;padding:1px 6px;border-radius:3px;font-size:10px;">內容</span>'
            else:
                type_tag = ''

            dd_preview = ""
            if has_dd and c.get("dropdown"):
                opts = c["dropdown"][:4]
                more = "..." if len(c["dropdown"]) > 4 else ""
                dd_preview = f'<div style="color:#64748b;font-size:11px;margin-top:2px;">選項：{h(", ".join(opts))}{more}</div>'

            val_display = h(val[:60] + ("..." if len(val) > 60 else "")) if val else '<span style="color:#cbd5e1;">（空）</span>'

            # 選中的欄位標記
            selected_options = field_options_str.replace(
                f'value="{h(current_field)}"',
                f'value="{h(current_field)}" selected'
            ) if current_field else field_options_str

            rows_html.append(
                f'<tr>'
                f'<td style="padding:6px 10px;font-family:monospace;font-weight:600;color:#475569;">{ref}</td>'
                f'<td style="padding:6px 10px;">{type_tag}</td>'
                f'<td style="padding:6px 10px;font-size:13px;max-width:300px;overflow:hidden;">{val_display}{dd_preview}</td>'
                f'<td style="padding:6px 10px;">'
                f'<select name="map_{h(sheet_name)}_{ref}" class="map-sel" style="width:100%;padding:3px 6px;font-size:12px;border:1px solid #cbd5e1;border-radius:4px;">{selected_options}</select>'
                f'</td>'
                f'</tr>'
            )

        sheet_note = ""

        sheets_html.append(
            f'<div class="sheet-block" data-sheet="{h(sheet_name)}" style="background:#fff;border:1px solid #e5e7eb;border-radius:8px;margin-bottom:18px;overflow:hidden;">'
            f'<div style="background:#f8fafc;padding:10px 16px;border-bottom:1px solid #e5e7eb;display:flex;justify-content:space-between;align-items:center;">'
            f'<div style="font-weight:600;color:#1e293b;">📊 工作表：{h(sheet_name)}{state_badge}</div>'
            f'<div style="font-size:12px;color:#64748b;">共 {len(cells_sorted)} 格</div>'
            f'</div>'
            f'<div style="padding:10px 16px;">{sheet_note}'
            f'<table style="width:100%;border-collapse:collapse;font-size:13px;">'
            f'<thead><tr style="background:#f9fafb;border-bottom:1px solid #e5e7eb;">'
            f'<th style="padding:6px 10px;text-align:left;font-weight:600;color:#6b7280;font-size:11px;width:70px;">座標</th>'
            f'<th style="padding:6px 10px;text-align:left;font-weight:600;color:#6b7280;font-size:11px;width:60px;">類型</th>'
            f'<th style="padding:6px 10px;text-align:left;font-weight:600;color:#6b7280;font-size:11px;">目前內容</th>'
            f'<th style="padding:6px 10px;text-align:left;font-weight:600;color:#6b7280;font-size:11px;width:280px;">對應系統欄位</th>'
            f'</tr></thead>'
            f'<tbody>{"".join(rows_html)}</tbody>'
            f'</table></div></div>'
        )

    nav = make_topnav(role, "templates")
    if inherited_from:
        default_badge = f'<span style="background:#fef3c7;color:#92400e;padding:2px 10px;border-radius:10px;font-size:12px;font-weight:500;">繼承自「{h(inherited_from)}」的映射</span>'
    elif is_new_mapping:
        default_badge = '<span style="background:#dbeafe;color:#1d4ed8;padding:2px 10px;border-radius:10px;font-size:12px;font-weight:500;">套用系統預設映射</span>'
    else:
        default_badge = '<span style="background:#dcfce7;color:#15803d;padding:2px 10px;border-radius:10px;font-size:12px;font-weight:500;">已儲存自訂映射</span>'

    body = f"""<!DOCTYPE html><html><head>{PAGE_CSS}<title>{h(plan)} - 範本編輯</title></head><body>
{nav}
<div class="container" style="max-width:1300px;margin:20px auto;padding:0 16px;">
  <div style="margin-bottom:16px;">
    <a href="/admin/templates" style="color:#4e7055;text-decoration:none;font-size:13px;">← 返回範本列表</a>
  </div>
  <div style="background:#fff;padding:16px 20px;border:1px solid #e5e7eb;border-radius:8px;margin-bottom:20px;">
    <div style="display:flex;justify-content:space-between;align-items:center;">
      <div>
        <h2 style="font-size:18px;font-weight:600;color:#1e293b;margin:0 0 4px 0;">📄 {h(plan)} 範本編輯</h2>
        <div style="color:#64748b;font-size:13px;">範本內含 {len(scan['sheets'])} 個工作表，共 {sum(len(s['cells']) for s in scan['sheets'])} 個儲存格</div>
      </div>
      <div>{default_badge}</div>
    </div>
  </div>
  <div style="background:#eef2ff;border:1px solid #c7d2fe;padding:12px 16px;border-radius:6px;font-size:13px;color:#3730a3;margin-bottom:20px;line-height:1.7;">
    💡 <b>使用說明：</b>下表列出範本的每個儲存格。在「對應系統欄位」下拉選擇要填入什麼資料（例如「姓名」「身分證」），空白則不填。
    <br>📌 <b>特殊選項：</b>「清空」= 下載時強制清空（清除範本示範值）、「保留原值」= 不動。
    <br>⚠️ 標題文字儲存格通常<b>不要選</b>對應欄位（否則會蓋掉標題）。
  </div>
  <form method="post" action="/admin/templates/save-mapping">
    <input type="hidden" name="plan" value="{h(plan)}">
    {''.join(sheets_html)}
    <div style="position:sticky;bottom:0;background:#fff;border-top:2px solid #4e7055;padding:14px 0;margin-top:20px;text-align:center;">
      <a href="/admin/templates" style="display:inline-block;padding:10px 24px;margin-right:10px;background:#fff;color:#64748b;border:1px solid #cbd5e1;border-radius:6px;text-decoration:none;font-size:14px;">取消</a>
      <button type="submit" style="padding:10px 32px;background:#4e7055;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:14px;font-weight:600;">💾 儲存映射</button>
    </div>
  </form>
</div>
</body></html>"""
    return HTMLResponse(body)


@app.post("/admin/templates/save-mapping")
async def admin_templates_save_mapping(request: Request):
    role = check_auth(request)
    if role not in ("admin", "adminB"):
        return JSONResponse({"error": "無權限"}, status_code=403)

    from fastapi.responses import RedirectResponse
    from urllib.parse import quote

    valid_plans = {p for p, _ in APPLICATION_PLAN_LIST}
    form = await request.form()
    plan_name = (form.get("plan") or "").strip()
    if plan_name not in valid_plans:
        return RedirectResponse(f"/admin/templates?err={quote('方案名稱無效')}", status_code=303)

    # 解析 map_{sheet_name}_{cell} 格式的 form field
    mapping = {}
    for key, val in form.multi_items() if hasattr(form, 'multi_items') else form.items():
        if not key.startswith("map_"):
            continue
        val = (val or "").strip()
        if not val:
            continue
        # key = "map_{sheet_name}_{cell}"，cell 固定是 [A-Z]+\d+
        m = re.match(r"^map_(.+)_([A-Z]+\d+)$", key)
        if not m:
            continue
        sheet_name = m.group(1)
        cell_ref = m.group(2)
        mapping.setdefault(sheet_name, {})[cell_ref] = val

    try:
        save_template_mapping(plan_name, mapping)
    except Exception as e:
        return RedirectResponse(f"/admin/templates?err={quote('儲存失敗：' + str(e))}", status_code=303)

    return RedirectResponse(f"/admin/templates/edit?plan={quote(plan_name)}&saved=1", status_code=303)


# ── PDF export JS（必須在 init_db 之前定義，因為 /new-customer 頁面需要引用）──
_PDF_EXPORT_JS = '''
function exportPDF(){
  var grpMap=__GRP_MAP__;
  var grpRaw=gv('grp');
  var grpName=grpMap[grpRaw]||grpRaw;
  var name=gv('cname')||'-';
  var sa=document.getElementById('sameck').checked;
  var raddr=gv('rcity')+gv('rdist')+gv('raddr');
  var laddr=sa?raddr:gv('lcity')+gv('ldist')+gv('laddr');
  var ca=gv('carea'),cn2=gv('cnum'),ce=gv('cext');
  var ctel=(ca&&ca!=='mobile')?ca+'-'+cn2+(ce?' 分機'+ce:''):cn2+(ce?' 分機'+ce:'');
  var caddr=gv('ccity')+gv('cdist')+gv('caddr');
  var fine=parseFloat(gv('efine'))||0;
  var fuel=parseFloat(gv('efuel'))||0;
  var tot=fine+fuel;
  var today=new Date().toLocaleDateString('zh-TW');
  var lyear=gv('lyear'),lmon=gv('lmon');
  var ltime=(lyear||lmon)?(lyear||'0')+'年'+(lmon||'0')+'月':'-';
  var cyear=gv('cyear'),cmon=gv('cmon');
  var ctime=(cyear||cmon)?(cyear||'0')+'年'+(cmon||'0')+'月':'-';
  function r2(l1,v1,l2,v2){return '<tr><th>'+l1+'</th><td>'+(v1||'-')+'</td><th>'+l2+'</th><td>'+(v2||'-')+'</td></tr>';}
  function r1(l,v){return '<tr><th>'+l+'</th><td colspan="3">'+(v||'-')+'</td></tr>';}
  function s3(t,rows){return '<tr class="sh"><th colspan="4">'+t+'</th></tr>'+rows;}
  var trows='';
  trows+=s3('基本資料',
    r2('客戶姓名',gv('cname'),'身分證字號',gv('idno').toUpperCase())+
    r2('出生年月日',gv('birth'),'行動電話',gv('phone'))+
    r2('電信業者',gv('carrier'),'Email',gv('email'))+
    r2('LINE ID',gv('line'),'客戶FB',gv('fb'))+
    r2('婚姻狀態',gv('marry'),'最高學歷',gv('edu')));
  trows+=s3('身分證發證',
    r2('發證日期',gv('iddate'),'發證地',gv('idplace'))+
    r2('換補發類別',gv('idtype'),'',''));
  trows+=s3('地址資料',
    r1('戶籍地址',raddr)+
    r2('戶籍電話',gv('rphone'),'現住電話',gv('lphone'))+
    r1('住家地址',laddr)+
    r2('居住狀況',gv('lstatus'),'居住時間',ltime));
  trows+=s3('職業資料',
    r1('公司名稱',gv('cmpname'))+
    r2('公司電話',ctel,'職稱',gv('crole'))+
    r2('年資',ctime,'月薪',gv('csal')+'萬')+
    r1('公司地址',caddr));
  trows+=s3('聯絡人資料',
    r2('聯絡人1',gv('c1name')+'（'+gv('c1rel')+'）','電話1',gv('c1tel'))+
    r2('知情1',gv('c1know'),'聯絡人2',gv('c2name')+'（'+gv('c2rel')+'）')+
    r2('電話2',gv('c2tel'),'知情2',gv('c2know')));
  trows+=s3('貸款諮詢事項',
    r2('資金需求',gv('efund'),'近三月送件',gv('esent'))+
    r2('當鋪私設',gv('eprivate'),'勞保狀態',gv('elabor'))+
    r2('有無薪轉',gv('esal'),'有無證照',gv('elicense'))+
    r2('貸款遲繳',gv('elate'),'遲繳天數',gv('elateday'))+
    r2('罰單欠費','$'+fine.toLocaleString(),'燃料稅','$'+fuel.toLocaleString())+
    r2('欠費總額','$'+tot.toLocaleString(),'名下信用卡',gv('ecard'))+
    r2('動產/不動產',gv('eprop'),'法學',gv('elaw'))+
    (gv('enote')?r1('備註',gv('enote')):''));
  var drows='';
  for(var ii=1;ii<=dc;ii++){
    var eli=document.getElementById('d'+ii);if(!eli)continue;
    var co2=document.getElementById('dc'+ii);co2=co2?co2.value:'';if(!co2)continue;
    var lo2=document.getElementById('dl'+ii);lo2=lo2?lo2.value:'';
    var dp2=document.getElementById('dp'+ii);dp2=dp2?dp2.value.trim():'';
    var mo2=document.getElementById('dm'+ii);mo2=mo2?mo2.value:'';
    var dr2=document.getElementById('dr'+ii);dr2=dr2?dr2.textContent:'-';
    drows+='<tr><td>'+co2+'</td><td>'+lo2+'</td><td>'+dp2+'</td><td>'+mo2+'</td><td>'+dr2+'</td></tr>';
  }
  if(drows){trows+='<tr class="sh"><th colspan="4">負債明細</th></tr>'+
    '<tr><td colspan="4"><table style="width:100%;border-collapse:collapse;font-size:12px;">'+
    '<tr style="background:#e8e2da;font-weight:600;"><th style="padding:5px;border:1px solid #bbb;">貸款商家</th>'+
    '<th style="padding:5px;border:1px solid #bbb;">貸款金額</th>'+
    '<th style="padding:5px;border:1px solid #bbb;">期數/已繳</th>'+
    '<th style="padding:5px;border:1px solid #bbb;">月繳</th>'+
    '<th style="padding:5px;border:1px solid #bbb;">剩餘</th></tr>'+
    drows+'</table></td></tr>';}
  var css='*{box-sizing:border-box;margin:0;padding:0;}body{font-family:"Microsoft JhengHei","PingFang TC",sans-serif;background:#fff;color:#1a1a1a;font-size:13px;}'+
    '.hdr{background:#3a3530;color:#fff;padding:14px 20px;display:flex;justify-content:space-between;align-items:center;margin-bottom:14px;}'+
    '.hdr-name{font-size:22px;font-weight:700;}.hdr-sub{font-size:12px;color:#c8bfb5;margin-top:4px;}'+
    'table.main{width:100%;border-collapse:collapse;margin-bottom:16px;}'+
    'table.main th,table.main td{border:1px solid #bbb;padding:6px 10px;vertical-align:top;font-size:13px;line-height:1.5;}'+
    'table.main th{background:#f0ebe4;color:#3a3020;font-weight:600;width:90px;white-space:nowrap;}'+
    'table.main td{background:#fff;}tr.sh th{background:#3a3530;color:#fff;font-size:12px;font-weight:700;padding:5px 10px;}'+
    '@media print{@page{size:A4 portrait;margin:10mm 12mm;}.np{display:none!important;}}';
  var html='<!DOCTYPE html><html lang="zh-TW"><head><meta charset="UTF-8">'+
    '<title>客戶資料表 - '+name+'</title><style>'+css+'</style></head><body>'+
    '<div class="hdr"><div><div class="hdr-name">'+name+'</div>'+
    '<div class="hdr-sub">群組：'+grpName+'　日期：'+today+'</div></div>'+
    '<div style="font-size:13px;color:#c8bfb5;font-weight:600;">客戶資料表</div></div>'+
    '<div class="np" style="text-align:center;margin-bottom:12px;">'+
    '<button onclick="window.print()" style="background:#4e7055;color:#fff;border:none;padding:9px 28px;border-radius:6px;font-size:14px;cursor:pointer;font-weight:600;">🖨 列印 / 存PDF</button></div>'+
    '<table class="main">'+trows+'</table></body></html>';
  var w=window.open('','_blank','width=900,height=750');
  w.document.write(html);w.document.close();
}
'''


# =========================
# Google Drive 自動備份
# =========================
def _gdrive_service():
    """建立 Google Drive API client（lazy import，套件沒裝也不會炸）"""
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    sa_info = json.loads(GOOGLE_SERVICE_ACCOUNT_JSON)
    creds = service_account.Credentials.from_service_account_info(
        sa_info, scopes=["https://www.googleapis.com/auth/drive.file"]
    )
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def _snapshot_db(dest_path: str):
    """用 SQLite backup API 做一致性快照（避免複製到寫到一半的檔）"""
    src = sqlite3.connect(DB_PATH)
    try:
        dst = sqlite3.connect(dest_path)
        try:
            with dst:
                src.backup(dst)
        finally:
            dst.close()
    finally:
        src.close()


def _prune_gdrive_backups(service) -> dict:
    """依保留策略清掉過期備份。
    策略：30 天內每日都留；超過 30 天只留每月最早一份；總共最多留 BACKUP_KEEP_MONTHLY 個月。"""
    files = service.files().list(
        q=f"'{GOOGLE_DRIVE_FOLDER_ID}' in parents and trashed=false and name contains 'loan-backup-'",
        fields="files(id,name,createdTime)", pageSize=1000
    ).execute().get("files", [])
    today = datetime.now().date()
    keep_ids = set()
    monthly_candidates = {}
    name_re = re.compile(r"loan-backup-(\d{4}-\d{2}-\d{2})\.db")
    for f in files:
        m = name_re.match(f["name"])
        if not m:
            keep_ids.add(f["id"])
            continue
        try:
            fdate = datetime.strptime(m.group(1), "%Y-%m-%d").date()
        except Exception:
            keep_ids.add(f["id"])
            continue
        age = (today - fdate).days
        if age <= BACKUP_KEEP_DAILY:
            keep_ids.add(f["id"])
        else:
            mk = fdate.strftime("%Y-%m")
            cur = monthly_candidates.get(mk)
            if cur is None or fdate < cur[1]:
                monthly_candidates[mk] = (f["id"], fdate)
    for mk in sorted(monthly_candidates.keys(), reverse=True)[:BACKUP_KEEP_MONTHLY]:
        keep_ids.add(monthly_candidates[mk][0])
    deleted = 0
    for f in files:
        if f["id"] not in keep_ids:
            try:
                service.files().delete(fileId=f["id"]).execute()
                deleted += 1
            except Exception as e:
                print(f"[backup-prune] delete {f['name']} failed: {e}")
    return {"total": len(files), "kept": len(keep_ids), "deleted": deleted}


def do_backup_now() -> dict:
    """執行一次備份：快照 → 上傳 Drive → 清理過期檔。回傳 dict 結果。"""
    if not BACKUP_ENABLED:
        return {"ok": False, "reason": "BACKUP_ENABLED 未設定為 true"}
    if not GOOGLE_SERVICE_ACCOUNT_JSON or not GOOGLE_DRIVE_FOLDER_ID:
        return {"ok": False, "reason": "缺少 GOOGLE_SERVICE_ACCOUNT_JSON 或 GOOGLE_DRIVE_FOLDER_ID"}
    tmp_path = os.path.join(os.path.dirname(DB_PATH) or ".", ".backup_gdrive_tmp.db")
    try:
        _snapshot_db(tmp_path)
        service = _gdrive_service()
        from googleapiclient.http import MediaFileUpload
        today = datetime.now().strftime("%Y-%m-%d")
        fname = f"loan-backup-{today}.db"
        existing = service.files().list(
            q=f"name='{fname}' and '{GOOGLE_DRIVE_FOLDER_ID}' in parents and trashed=false",
            fields="files(id,name)"
        ).execute().get("files", [])
        media = MediaFileUpload(tmp_path, mimetype="application/x-sqlite3", resumable=False)
        if existing:
            service.files().update(fileId=existing[0]["id"], media_body=media).execute()
            action = "updated"
        else:
            service.files().create(
                body={"name": fname, "parents": [GOOGLE_DRIVE_FOLDER_ID]},
                media_body=media, fields="id"
            ).execute()
            action = "created"
        prune = _prune_gdrive_backups(service)
        print(f"[backup] {action} {fname}, prune={prune}")
        return {"ok": True, "file": fname, "action": action, "prune": prune, "ts": now_iso()}
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"[backup] failed: {e}")
        return {"ok": False, "reason": str(e)}
    finally:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass


def list_gdrive_backups() -> list:
    """列出 Drive 上的備份檔（給管理頁顯示用）"""
    if not BACKUP_ENABLED or not GOOGLE_SERVICE_ACCOUNT_JSON or not GOOGLE_DRIVE_FOLDER_ID:
        return []
    try:
        service = _gdrive_service()
        files = service.files().list(
            q=f"'{GOOGLE_DRIVE_FOLDER_ID}' in parents and trashed=false",
            fields="files(id,name,size,createdTime)",
            orderBy="name desc", pageSize=100
        ).execute().get("files", [])
        return files
    except Exception as e:
        print(f"[backup-list] failed: {e}")
        return []


def _start_backup_scheduler():
    if not BACKUP_ENABLED:
        print("[backup] scheduler disabled (BACKUP_ENABLED != true)")
        return
    if not GOOGLE_SERVICE_ACCOUNT_JSON or not GOOGLE_DRIVE_FOLDER_ID:
        print("[backup] scheduler disabled (missing env vars)")
        return
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        sched = BackgroundScheduler(daemon=True)
        # 19:00 UTC = 03:00 Asia/Taipei
        sched.add_job(do_backup_now, "cron", hour=19, minute=0, id="daily_gdrive_backup")
        sched.start()
        print("[backup] scheduler started: daily 19:00 UTC (03:00 Taipei)")
    except Exception as e:
        print(f"[backup] scheduler start failed: {e}")


# =========================
# 啟動（模組載入時就初始化 DB，確保任何情況下都能正常運作）
# =========================
init_db()
seed_groups()
_start_backup_scheduler()


if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=int(os.getenv("PORT", 10000)))
